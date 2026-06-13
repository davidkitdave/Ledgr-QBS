import io
import csv
import openpyxl
from google.genai import types
from google.adk.plugins.base_plugin import BasePlugin
from google.adk.agents.callback_context import CallbackContext
from google.adk.models.llm_request import LlmRequest
from google.adk.models.llm_response import LlmResponse

class ExcelParserPlugin(BasePlugin):
    async def before_model_callback(self, *, callback_context: CallbackContext, llm_request: LlmRequest) -> LlmResponse | None:
        for content in llm_request.contents:
            if not content.parts:
                continue
            for i, part in enumerate(content.parts):
                if part.inline_data and part.inline_data.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                    wb = openpyxl.load_workbook(io.BytesIO(part.inline_data.data))
                    csv_data = io.StringIO()
                    writer = csv.writer(csv_data)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        writer.writerow([f"--- Sheet: {sheet_name} ---"])
                        for row in sheet.iter_rows(values_only=True):
                            writer.writerow([str(cell) if cell is not None else "" for cell in row])
                    content.parts[i] = types.Part.from_text(text=f"Uploaded Excel File Contents:\n{csv_data.getvalue()}")
        return None

