"""Mutate mixin for ``SlackLedgerStore`` — amend / remove / month-clear on the channel's FY workbook.

See :mod:`ledgr_slack.ledger_store_base` for the full architecture; this module
only adds the mutation path.
"""

from __future__ import annotations

import logging
from typing import Any, Optional

from openpyxl import Workbook

from ledgr_slack.ledger_doc_identity import (
    ledger_row_signature,
    sheet_lacks_invoice_identity_column,
)
from ledgr_slack.ledger_store_base import SlackLedgerStoreBase

logger = logging.getLogger(__name__)

#: Sheet titles for the invoice ledger workbook (mirrors LedgerExporter).
_INVOICE_SHEETS = ("Purchase", "Sales")


class _SlackLedgerStoreMutateMixin:
    def read_rows(
        self,
        client_id: str,
        fy: str,
        slack_client: Any,
        channel_id: str,
    ) -> list[dict]:
        """Download the current FY workbook and return all data rows as dicts.

        Fetches the workbook pointed to by Firestore ``clients/{client_id}/ledgers/{fy}``.
        The workbook contains no internal dedupe column (dedupe is Firestore-side),
        so all columns are returned as-is.

        Returns an empty list when no pointer exists yet (ledger not started).

        Args:
            client_id: The client whose ledger pointer to look up.
            fy: The financial-year label (e.g. ``"2026"``).
            slack_client: A Slack WebClient (or fake) used to download the file.
            channel_id: Unused here but kept for symmetry with ``append_rows``
                (e.g. for logging or future per-channel locking).

        Returns:
            A list of dicts keyed by sheet column headers, one entry per data
            row across ALL sheets in the workbook (Purchase + Sales for invoice
            workbooks; per-account sheets for bank workbooks).  Each dict also
            carries ``"_sheet"`` so callers can distinguish row origin.
        """
        pointer = self.get_pointer(client_id, fy)
        if not pointer or not pointer.get("slack_file_id"):
            return []

        data = self._download_workbook(slack_client, pointer["slack_file_id"])
        wb = self._load_workbook(data)

        rows: list[dict] = []
        for sheet in wb.worksheets:
            if sheet.max_row < 1:
                continue
            headers = [c.value for c in sheet[1]]

            for ws_row_num, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                row_dict: dict = {"_sheet": sheet.title, "_row": ws_row_num}
                for idx, header in enumerate(headers):
                    if header is not None:
                        row_dict[header] = row[idx] if idx < len(row) else None
                # Skip entirely-blank rows (all values None).
                if any(v is not None for k, v in row_dict.items() if k not in ("_sheet", "_row")):
                    rows.append(row_dict)

        return rows

    # ------------------------------------------------------------------ #
    # Workbook mutation helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _is_bank_sheet(sheet_name: str) -> bool:
        """Return True when the sheet is a bank/account sheet rather than an invoice sheet.

        Invoice workbooks only ever have "Purchase" and "Sales" sheets (defined by
        ``_INVOICE_SHEETS``).  Any other sheet title belongs to a bank workbook
        (e.g. "OCBC - 0001", "DBS - 9002").  Checking by *inclusion* in the
        known-invoice set is more robust than pattern-matching account names —
        it never needs updating when new banks appear, and it cannot be fooled by a
        vendor description that happens to look like an account label.
        """
        return sheet_name not in _INVOICE_SHEETS

    def _download_current_workbook(
        self, slack_client: Any, client_id: str, fy: str
    ) -> tuple[dict, bytes]:
        """Return (pointer, workbook_bytes) or raise ValueError if pointer is absent."""
        pointer = self.get_pointer(client_id, fy)
        if not pointer or not pointer.get("slack_file_id"):
            raise ValueError(
                f"no ledger pointer for client={client_id!r} fy={fy!r}; "
                "cannot mutate a workbook that doesn't exist yet"
            )
        data = self._download_workbook(slack_client, pointer["slack_file_id"])
        return pointer, data

    def _upload_and_reroute(
        self,
        slack_client: Any,
        wb: "Workbook",
        pointer: dict,
        client_id: str,
        fy: str,
        channel_id: str,
        *,
        seen_doc_keys_override: Optional[list] = None,
    ) -> str:
        """Serialize *wb*, upload as a new Slack file, update the Firestore pointer.

        By default leaves ``seen_doc_keys`` intact (mutation does not un-see
        source docs).  Pass ``seen_doc_keys_override`` to replace the stored set
        — used by :meth:`remove_rows_for_month` to purge the cleared month's keys
        so re-dropped documents are not silently deduped.
        Returns the new ``slack_file_id``.
        """
        prev_file_id: Optional[str] = pointer.get("slack_file_id")

        # Re-use the same filename convention as append_rows.
        kind = pointer.get("kind", "invoice")
        # Rebuild the client-scoped prefix from the persisted client_name.
        # Legacy pointers written before this field was stored have no client_name;
        # fall back to no prefix so they continue working and self-heal on next append.
        stored_name: str = pointer.get("client_name") or ""
        prefix = f"{stored_name.strip()} - " if stored_name.strip() else ""
        if kind == "bank":
            filename = f"{prefix}BankStatement_FY{fy}.xlsx"
        else:
            filename = f"{prefix}Ledger_FY{fy}.xlsx"

        new_bytes = self._to_bytes(wb)
        result = slack_client.files_upload_v2(
            channel=channel_id,
            filename=filename,
            file=new_bytes,
            title=filename,
        )
        new_file_id = self._extract_uploaded_file_id(result)
        if new_file_id:
            # Use the caller-supplied override when provided (month-clear purge);
            # otherwise preserve the pointer's existing seen_doc_keys.
            persisted_keys = (
                seen_doc_keys_override
                if seen_doc_keys_override is not None
                else pointer.get("seen_doc_keys")
            )
            self._set_pointer(
                client_id,
                fy,
                new_file_id,
                seen_doc_keys=persisted_keys,
                channel_id=channel_id,
                kind=kind,
            )

        if prev_file_id and new_file_id and prev_file_id != new_file_id:
            try:
                slack_client.files_delete(file=prev_file_id)
            except Exception:  # noqa: BLE001
                logger.warning(
                    "Could not delete superseded ledger file %s after mutation (non-fatal).",
                    prev_file_id,
                )

        return new_file_id or ""

    @staticmethod
    def _header_col_map(ws) -> dict[str, int]:
        """Return {column_header: 1-based column index} from worksheet row 1."""
        return {
            cell.value: cell.column
            for cell in ws[1]
            if cell.value is not None
        }

    @staticmethod
    def _validate_mutation_args(
        wb: "Workbook",
        sheet: str,
        row: int,
    ) -> Any:
        """Validate sheet + row args; return the worksheet or raise ValueError."""
        if sheet not in wb.sheetnames:
            raise ValueError(
                f"sheet not found: {sheet!r}; available sheets are {wb.sheetnames}"
            )
        ws = wb[sheet]
        max_row = ws.max_row
        # Row 1 is always the header; data rows start at 2.
        if row < 2 or row > max_row:
            raise ValueError(
                f"row {row} out of range for sheet {sheet!r} "
                f"(valid data rows: 2–{max_row})"
            )
        return ws

    # ------------------------------------------------------------------ #
    # Public mutation API
    # ------------------------------------------------------------------ #

    def amend_row(
        self,
        client_id: str,
        fy: str,
        slack_client: Any,
        channel_id: str,
        *,
        sheet: str,
        row: int,
        updates: dict,
    ) -> dict:
        """Update one or more cells in an invoice ledger row.

        Downloads the current workbook, mutates exactly the cells named in
        ``updates``, uploads a new file version, and updates the Firestore
        pointer — the same accumulate-the-record contract as ``append_rows``.
        ``seen_doc_keys`` is left intact.

        Args:
            client_id: Client whose ledger pointer to look up.
            fy: Financial-year label.
            slack_client: Slack WebClient (or fake) for download + upload.
            channel_id: Channel the new workbook version is posted to.
            sheet: Worksheet title to mutate.  Must be an invoice sheet
                ("Purchase" or "Sales"); bank sheets are refused.
            row: 1-based worksheet row number (≥ 2; row 1 is the header).
            updates: ``{column_header: new_value}`` dict.

        Returns:
            ``{"sheet", "row", "before": {col: old}, "after": {col: new}}``.

        Raises:
            ValueError: if no pointer exists; the sheet is a bank sheet; the
                sheet name is not found; ``row`` is out of range; or an
                ``updates`` key is not a column header in that sheet.
        """
        lock = self._lock_for(client_id, fy)
        with lock:  # fast same-process serialize
            token = self._lease.acquire(client_id, fy)  # cross-instance serialize
            try:
                pointer, data = self._download_current_workbook(slack_client, client_id, fy)
                wb = self._load_workbook(data)

                # Sheet-existence check before bank guard: gives a clear "not found"
                # error for misspelled names rather than a misleading "bank" error.
                if sheet not in wb.sheetnames:
                    raise ValueError(
                        f"sheet not found: {sheet!r}; available sheets are {wb.sheetnames}"
                    )

                if self._is_bank_sheet(sheet):
                    raise ValueError(
                        f"bank-statement rows are read-only from chat; "
                        f"balances are derived (sheet={sheet!r}). "
                        "Amend invoice ledger rows (Purchase / Sales) only."
                    )

                ws = self._validate_mutation_args(wb, sheet, row)
                col_map = self._header_col_map(ws)

                # Validate all update keys before touching any cell.
                for col_name in updates:
                    if col_name not in col_map:
                        raise ValueError(
                            f"unknown column {col_name!r} in sheet {sheet!r}; "
                            f"known headers: {sorted(col_map)}"
                        )

                before: dict = {}
                after: dict = {}
                for col_name, new_value in updates.items():
                    col_idx = col_map[col_name]
                    cell = ws.cell(row=row, column=col_idx)
                    before[col_name] = cell.value
                    cell.value = new_value
                    after[col_name] = new_value

                self._upload_and_reroute(slack_client, wb, pointer, client_id, fy, channel_id)
            finally:
                self._lease.release(client_id, fy, token)

        return {"sheet": sheet, "row": row, "before": before, "after": after}

    def remove_row(
        self,
        client_id: str,
        fy: str,
        slack_client: Any,
        channel_id: str,
        *,
        sheet: str,
        row: int,
    ) -> dict:
        """Delete one row from an invoice ledger sheet.

        Subsequent rows shift up by one (openpyxl ``delete_rows``).  Uploads a
        new file version and updates the Firestore pointer.  ``seen_doc_keys``
        is left intact — removing a line does not un-see the source document.

        Args:
            client_id: Client whose ledger pointer to look up.
            fy: Financial-year label.
            slack_client: Slack WebClient (or fake) for download + upload.
            channel_id: Channel the new workbook version is posted to.
            sheet: Worksheet title.  Must be an invoice sheet ("Purchase" or
                "Sales"); bank sheets are refused.
            row: 1-based worksheet row number (≥ 2; row 1 is the header).

        Returns:
            ``{"sheet", "row", "removed": {col: value}}``.

        Raises:
            ValueError: if no pointer exists; the sheet is a bank sheet; the
                sheet name is not found; or ``row`` is out of range.
        """
        lock = self._lock_for(client_id, fy)
        with lock:  # fast same-process serialize
            token = self._lease.acquire(client_id, fy)  # cross-instance serialize
            try:
                pointer, data = self._download_current_workbook(slack_client, client_id, fy)
                wb = self._load_workbook(data)

                # Sheet-existence check before bank guard (same ordering as amend_row).
                if sheet not in wb.sheetnames:
                    raise ValueError(
                        f"sheet not found: {sheet!r}; available sheets are {wb.sheetnames}"
                    )

                if self._is_bank_sheet(sheet):
                    raise ValueError(
                        f"bank-statement rows are read-only from chat; "
                        f"balances are derived (sheet={sheet!r}). "
                        "Remove invoice ledger rows (Purchase / Sales) only."
                    )

                ws = self._validate_mutation_args(wb, sheet, row)
                col_map = self._header_col_map(ws)

                # Capture the row's values before deletion for the return dict.
                removed: dict = {
                    col_name: ws.cell(row=row, column=col_idx).value
                    for col_name, col_idx in col_map.items()
                }

                ws.delete_rows(row, 1)

                self._upload_and_reroute(slack_client, wb, pointer, client_id, fy, channel_id)
            finally:
                self._lease.release(client_id, fy, token)

        return {"sheet": sheet, "row": row, "removed": removed}

    # ------------------------------------------------------------------ #
    # Month-clear mutation (Step 7 / C-3)
    # ------------------------------------------------------------------ #

    @staticmethod
    def _parse_row_date(value) -> Optional[tuple[int, int]]:
        """Return ``(year, month)`` from a Date cell (``DD/MM/YYYY`` or date obj).

        Returns ``None`` when the cell is absent or unparseable.
        """
        if value is None:
            return None
        # datetime.date / datetime.datetime objects.
        month = getattr(value, "month", None)
        year = getattr(value, "year", None)
        if month and year:
            return (int(year), int(month))
        # String "DD/MM/YYYY" (or "DD/MM/YY").
        parts = str(value).strip().split("/")
        if len(parts) == 3:
            try:
                m = int(parts[1])
                y = int(parts[2])
                if y < 100:
                    y += 2000
                return (y, m)
            except ValueError:
                return None
        return None

    def remove_rows_for_month(
        self,
        client_id: str,
        fy: str,
        slack_client: Any,
        channel_id: str,
        *,
        year: int,
        month: int,
        sheets: tuple[str, ...] = ("Purchase", "Sales"),
    ) -> dict:
        """Delete all invoice rows for ``(year, month)`` across Purchase + Sales.

        Downloads the current workbook, finds every data row whose ``Date``
        column parses to ``(year, month)``, deletes them BOTTOM-UP (highest row
        index first so openpyxl row-shifts don't invalidate earlier indices), and
        purges the reconstructed ``doc_key`` values from the Firestore pointer's
        ``seen_doc_keys`` so re-dropped documents are not silently deduped.
        Uploads the trimmed workbook and updates the pointer.

        Args:
            client_id: Client whose ledger pointer to look up.
            fy: Financial-year label.
            slack_client: Slack WebClient (or fake) for download + upload.
            channel_id: Channel the new workbook version is posted to.
            year: 4-digit year to clear.
            month: Month number 1–12 to clear.
            sheets: Invoice sheet names to search.  Bank sheets are refused.

        Returns:
            ``{"removed": [row_desc, ...], "purged_keys": [...], "sheets": {sheet: count}}``.

        Raises:
            ValueError: if no pointer/workbook exists; ``year`` / ``month``
                are out of range; or any entry in ``sheets`` is a bank sheet.
        """
        if not 1 <= month <= 12:
            raise ValueError(f"month must be 1–12, got {month!r}")
        if year < 1:
            raise ValueError(f"year must be a positive integer, got {year!r}")

        for sheet_name in sheets:
            if self._is_bank_sheet(sheet_name):
                raise ValueError(
                    f"bank-statement sheets are read-only from chat; "
                    f"only Purchase / Sales sheets can be cleared (sheet={sheet_name!r})."
                )

        lock = self._lock_for(client_id, fy)
        with lock:  # fast same-process serialize
            token = self._lease.acquire(client_id, fy)  # cross-instance serialize
            try:
                pointer, data = self._download_current_workbook(slack_client, client_id, fy)
                wb = self._load_workbook(data)
                software = str(pointer.get("software") or "qbs")
                try:
                    exporter = self._exporter_for(software)
                except ValueError:
                    exporter = None

                existing_keys: set = self._get_seen_doc_keys(pointer)
                purged_keys: list[str] = []
                removed_descs: list[str] = []
                sheet_counts: dict[str, int] = {}

                for sheet_name in sheets:
                    if sheet_name not in wb.sheetnames:
                        sheet_counts[sheet_name] = 0
                        continue

                    ws = wb[sheet_name]
                    if ws.max_row < 2:
                        sheet_counts[sheet_name] = 0
                        continue

                    col_map = self._header_col_map(ws)
                    doc_type = "sales" if sheet_name == "Sales" else "purchase"
                    date_header = None
                    inv_header = None
                    if exporter is not None and hasattr(exporter, "column_for_field"):
                        date_header = exporter.column_for_field("invoice_date", doc_type)
                        inv_header = self._invoice_identity_column(exporter, sheet_name)
                    date_col = (
                        (col_map.get(date_header) if date_header else None)
                        or col_map.get("Invoice Date")
                        or col_map.get("*InvoiceDate")
                        or col_map.get("DocDate")
                        or col_map.get("DOCDATE")
                        or col_map.get("Date")
                    )
                    inv_col = (
                        (col_map.get(inv_header) if inv_header else None)
                        or col_map.get("Invoice Number")
                        or col_map.get("*InvoiceNumber")
                        or col_map.get("SupplierInvoiceNo")
                        or col_map.get("DOCNO(20)")
                    )

                    # Issue #34: AutoCount sales (AR) has no readable invoice-
                    # identity column, so its appended doc_key is a row signature
                    # (sheet:DocDate:code:Amount), NOT sheet:invoice_number. Detect
                    # that case and resolve the signature columns from the SAME
                    # exporter mapping the append side used, so the reconstructed
                    # key matches and the purge fires. Every other ERP keeps the
                    # invoice_number path below.
                    use_row_signature = (
                        inv_col is None
                        and exporter is not None
                        and sheet_lacks_invoice_identity_column(exporter, doc_type)
                    )
                    sig_code_col = None
                    sig_amount_col = None
                    if use_row_signature:
                        code_field = (
                            "debtor_code" if doc_type == "sales" else "creditor_code"
                        )
                        code_header = exporter.column_for_field(code_field, doc_type)
                        amount_header = exporter.column_for_field("sub_total", doc_type)
                        sig_code_col = (
                            col_map.get(code_header) if code_header else None
                        )
                        sig_amount_col = (
                            col_map.get(amount_header) if amount_header else None
                        )

                    # Collect matching row numbers in ascending order, then delete bottom-up.
                    matching_rows: list[int] = []
                    for row_num in range(2, ws.max_row + 1):
                        date_val = (
                            ws.cell(row=row_num, column=date_col).value
                            if date_col else None
                        )
                        parsed = self._parse_row_date(date_val)
                        if parsed is not None and parsed == (year, month):
                            matching_rows.append(row_num)

                    count = len(matching_rows)
                    sheet_counts[sheet_name] = count

                    # Reconstruct doc_keys for the matching rows BEFORE deletion.
                    for row_num in matching_rows:
                        key = None
                        if use_row_signature:
                            # AutoCount sales: rebuild the row signature from the
                            # Excel cells via the shared helper (#34). Page-range
                            # keys never apply here (AutoCount sales has no
                            # invoice_number identity), so a bare signature match
                            # against existing_keys is sufficient.
                            key = ledger_row_signature(
                                sheet_name,
                                ws.cell(row=row_num, column=date_col).value
                                if date_col else None,
                                ws.cell(row=row_num, column=sig_code_col).value
                                if sig_code_col else None,
                                ws.cell(row=row_num, column=sig_amount_col).value
                                if sig_amount_col else None,
                            )
                        else:
                            inv_num = (
                                ws.cell(row=row_num, column=inv_col).value
                                if inv_col else None
                            )
                            # Mirror nodes._doc_key: f"{sheet}:{invoice_number}"
                            # (no index suffix for a single-row-per-doc batch).
                            if inv_num is not None:
                                key = f"{sheet_name}:{str(inv_num).strip()}"
                        if key is not None and key in existing_keys:
                            purged_keys.append(key)
                        removed_descs.append(
                            f"{sheet_name} row {row_num}"
                        )

                    # Delete bottom-up so row shifts don't corrupt earlier indices.
                    for row_num in sorted(matching_rows, reverse=True):
                        ws.delete_rows(row_num, 1)

                # Rebuild the surviving seen_doc_keys (purge the cleared month's keys).
                surviving_keys = list(existing_keys - set(purged_keys))

                self._upload_and_reroute(
                    slack_client, wb, pointer, client_id, fy, channel_id,
                    seen_doc_keys_override=surviving_keys,
                )
            finally:
                self._lease.release(client_id, fy, token)

        return {
            "removed": removed_descs,
            "purged_keys": purged_keys,
            "sheets": sheet_counts,
        }

    @staticmethod
    def _extract_uploaded_file_id(result: Any) -> Optional[str]:
        """Pull the uploaded file id out of a ``files_upload_v2`` response.

        ``files_upload_v2`` returns ``{"files": [{"id": ...}]}`` (a list) or the
        legacy ``{"file": {"id": ...}}`` shape; handle both + a plain dict.
        """
        if result is None:
            return None
        data = result.data if hasattr(result, "data") else result
        if not isinstance(data, dict):
            return None
        files = data.get("files")
        if isinstance(files, list) and files:
            first = files[0]
            # v2 nests as {"files": [{"file": {...}}]} in some SDK versions.
            if isinstance(first, dict):
                if "id" in first:
                    return first["id"]
                nested = first.get("file")
                if isinstance(nested, dict):
                    return nested.get("id")
        f = data.get("file")
        if isinstance(f, dict):
            return f.get("id")
        return None
