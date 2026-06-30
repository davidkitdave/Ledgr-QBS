"""Per-client context: load client profile and entity memory into typed objects.

Multi-tenant by construction — nothing here hardcodes any client's account numbers.
The client profile (region, accounting_software, base_currency, gst/tax_registered,
fye_month) comes from the **per-channel Firestore profile** created at onboarding
(see ``FirestoreClientStore``).

Firestore layout (spec §1 — 2026-06-12, ADR-0036):
  clients/{client_id}          -> profile fields + category_mapping map
  clients/{client_id}/entity_memory/{n} -> vendor memory (optional)
  channels/{channel_id}        -> { client_id }
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Protocol

from openpyxl import load_workbook


# --------------------------------------------------------------------------- #
# Dataclasses
# --------------------------------------------------------------------------- #
@dataclass
class EntityMemoryEntry:
    name: str
    reg_no: Optional[str] = None
    mapping_code: Optional[str] = None        # account code/name
    role: Optional[str] = None
    tax_code: Optional[str] = None
    creditor_code: Optional[str] = None       # ERP creditor/vendor code (purchases)


@dataclass
class ClientContext:
    client_id: Optional[str] = None
    client_name: Optional[str] = None
    client_uen: Optional[str] = None      # UEN / company registration number
    channel_id: Optional[str] = None
    slack_team_id: Optional[str] = None
    firm_id: Optional[str] = None
    status: Optional[str] = None
    region: str = ""
    accounting_software: str = "QBS Ledger"
    base_currency: str = ""
    tax_registered: Optional[bool] = None
    partial_exempt: bool = False
    fye_month: Optional[int] = None
    category_mapping: dict[str, Optional[str]] = field(default_factory=dict)  # category -> account_code | null
    entity_memory: list[EntityMemoryEntry] = field(default_factory=list)
    tax_codes: list[dict] = field(default_factory=list)                       # client ERP tax-code master
    sys_config: dict[str, str] = field(default_factory=dict)                  # kept for back-compat; not populated from Firestore

    def to_state(self) -> dict:
        """Serializable dict for ``session.state`` (basic types only)."""
        return {
            "client_id": self.client_id,
            "client_name": self.client_name,
            "client_uen": self.client_uen,
            "firm_id": self.firm_id,
            "slack_team_id": self.slack_team_id,
            "region": self.region,
            "tax_registered": self.tax_registered,
            "partial_exempt": self.partial_exempt,
            "software": self.accounting_software,
            "base_currency": self.base_currency,
            "fye_month": self.fye_month,
            "category_mapping": dict(self.category_mapping),
            "entity_memory": [
                {
                    "name": e.name,
                    "reg_no": e.reg_no,
                    "mapping_code": e.mapping_code,
                    "role": e.role,
                    "tax_code": e.tax_code,
                    "creditor_code": e.creditor_code,
                }
                for e in self.entity_memory
            ],
            "tax_codes": list(self.tax_codes),
        }


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _s(v) -> Optional[str]:
    """Coerce a cell to a stripped non-empty string or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _truthy(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ("true", "yes", "y", "1")


def _row_is_empty(row) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def _header_index(header: tuple, *names: str) -> Optional[int]:
    """Find the first column whose header matches any of ``names`` (case-insensitive)."""
    want = {n.strip().lower() for n in names}
    for i, h in enumerate(header):
        if h is not None and str(h).strip().lower() in want:
            return i
    return None


# --------------------------------------------------------------------------- #
# Workbook loader
# --------------------------------------------------------------------------- #
def load_client_setup(xlsx_path: str | Path, client_id: Optional[str] = None) -> ClientContext:
    """Parse a Client Setup workbook into a :class:`ClientContext`.

    Parses only the client's uploaded COA / Category_Mapping / Entity_Memory data.
    Profile metadata (region, accounting_software, base_currency, tax_registered,
    fye_month) now comes from the per-channel Firestore profile created at onboarding
    and is NOT read from this workbook.

    The caller should supply ``client_id`` explicitly; it is set directly on the
    returned context. Missing sheets are handled gracefully. ``Account code`` may be
    blank (QBS keys by Description). Only enabled ``Category_Mapping`` rows are kept.
    Fully-empty rows in COA / Entity_Memory are skipped.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheets = set(wb.sheetnames)
        ctx = ClientContext()
        ctx.client_id = client_id

        # --- Category_Mapping (optional) ---
        if "Category_Mapping" in sheets:
            rows = list(wb["Category_Mapping"].iter_rows(values_only=True))
            if rows:
                header = rows[0]
                i_cat = _header_index(header, "Category") or 0
                i_acc = _header_index(header, "Account Code", "Account code")
                i_en = _header_index(header, "Enabled")
                for row in rows[1:]:
                    if _row_is_empty(row):
                        continue
                    cat = _s(row[i_cat] if i_cat < len(row) else None)
                    if not cat:
                        continue
                    enabled = _truthy(row[i_en]) if (i_en is not None and i_en < len(row)) else True
                    if not enabled:
                        continue
                    acc = _s(row[i_acc]) if (i_acc is not None and i_acc < len(row)) else None
                    ctx.category_mapping[cat] = acc  # keep None (unmapped)

        # --- Entity_Memory (optional / may be empty) ---
        if "Entity_Memory" in sheets:
            rows = list(wb["Entity_Memory"].iter_rows(values_only=True))
            if rows:
                header = rows[0]
                i_name = _header_index(header, "Name") or 0
                i_reg = _header_index(header, "Reg No / Tax ID", "Reg No", "Tax ID")
                i_map = _header_index(header, "Mapping Code")
                i_role = _header_index(header, "Role (Debtor / Creditor)", "Role")
                i_tax = _header_index(header, "Tax Code")
                i_cred = _header_index(header, "Creditor Code", "Vendor Code")
                for row in rows[1:]:
                    if _row_is_empty(row):
                        continue
                    name = _s(row[i_name] if i_name < len(row) else None)
                    if not name:
                        continue

                    def cell(idx):
                        return row[idx] if idx is not None and idx < len(row) else None

                    ctx.entity_memory.append(EntityMemoryEntry(
                        name=name,
                        reg_no=_s(cell(i_reg)),
                        mapping_code=_s(cell(i_map)),
                        role=_s(cell(i_role)),
                        tax_code=_s(cell(i_tax)),
                        creditor_code=_s(cell(i_cred)),
                    ))

        # --- Tax_Codes (optional) ---
        if "Tax_Codes" in sheets:
            rows = list(wb["Tax_Codes"].iter_rows(values_only=True))
            if rows:
                header = rows[0]
                i_code = _header_index(header, "Code") or 0
                i_desc = _header_index(header, "Description")
                i_treat = _header_index(header, "Treatment")
                for row in rows[1:]:
                    if _row_is_empty(row):
                        continue
                    code = _s(row[i_code] if i_code < len(row) else None)
                    if not code:
                        continue
                    entry = {"code": code}
                    if i_desc is not None and i_desc < len(row):
                        desc = _s(row[i_desc])
                        if desc:
                            entry["description"] = desc
                    if i_treat is not None and i_treat < len(row):
                        treat = _s(row[i_treat])
                        if treat:
                            entry["treatment"] = treat
                    ctx.tax_codes.append(entry)

        return ctx
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# Plain-dict (state) accessors
# --------------------------------------------------------------------------- #


def entity_memory_from_state(state: dict) -> list[EntityMemoryEntry]:
    out: list[EntityMemoryEntry] = []
    for d in state.get("entity_memory") or []:
        out.append(EntityMemoryEntry(
            name=d.get("name") or "",
            reg_no=d.get("reg_no"),
            mapping_code=d.get("mapping_code"),
            role=d.get("role"),
            tax_code=d.get("tax_code"),
            creditor_code=d.get("creditor_code"),
        ))
    return out


def tax_codes_from_state(state: dict) -> list[dict]:
    raw = state.get("tax_codes") or []
    if isinstance(raw, dict):
        return [{"code": code, "description": desc} for code, desc in raw.items()]
    return [dict(entry) for entry in raw]


def category_mapping_from_state(state: dict) -> dict[str, Optional[str]]:
    return dict(state.get("category_mapping") or {})


def _profile_tax_registered(profile: dict) -> Optional[bool]:
    """Map profile gst_registered / tax_registered to Optional[bool] (None = unknown)."""
    if "gst_registered" in profile:
        return bool(profile["gst_registered"])
    if "tax_registered" in profile:
        val = profile["tax_registered"]
        if val is None:
            return None
        return bool(val)
    return None


def _jurisdiction_helpers():
    from ledgr_slack.jurisdiction import REGION_REGISTRY, _norm_region
    return REGION_REGISTRY, _norm_region


def _profile_region_and_currency(profile: dict) -> tuple[str, str]:
    """Resolve region + base_currency from a profile dict (Firestore/onboarding)."""
    REGION_REGISTRY, _norm_region = _jurisdiction_helpers()
    raw_region = profile.get("region")
    if raw_region:
        region = _norm_region(raw_region)
        currency = REGION_REGISTRY.get(region, {}).get("currency") or ""
        if not currency:
            stored = profile.get("base_currency")
            currency = str(stored).strip().upper() if stored else ""
        return region, currency

    if profile.get("legacy_profile") or profile.get("legacy"):
        stored_currency = profile.get("base_currency")
        currency = str(stored_currency).strip().upper() if stored_currency else ""
        return "", currency

    default_region = os.environ.get("LEDGR_DEFAULT_REGION", "").strip().upper()
    if default_region in REGION_REGISTRY:
        return default_region, REGION_REGISTRY[default_region]["currency"]

    stored_currency = profile.get("base_currency")
    currency = str(stored_currency).strip().upper() if stored_currency else ""
    return "", currency


def _state_region_and_currency(state: dict) -> tuple[str, str]:
    """Resolve region + base_currency from session state."""
    REGION_REGISTRY, _norm_region = _jurisdiction_helpers()
    raw_region = state.get("region") or state.get("client_region") or ""
    region = _norm_region(raw_region) if raw_region else ""
    if not region:
        default_region = os.environ.get("LEDGR_DEFAULT_REGION", "").strip().upper()
        if default_region in REGION_REGISTRY:
            region = default_region

    if region and region in REGION_REGISTRY:
        return region, REGION_REGISTRY[region]["currency"]

    raw_currency = state.get("base_currency")
    currency = str(raw_currency).strip().upper() if raw_currency else ""
    return region, currency


def _state_tax_registered(state: dict) -> Optional[bool]:
    if "tax_registered" not in state:
        return None
    val = state.get("tax_registered")
    if val is None:
        return None
    return bool(val)


def client_context_from_state(state: dict) -> ClientContext:
    """Rebuild a :class:`ClientContext` from a plain ``to_state()`` dict."""
    region, base_currency = _state_region_and_currency(state)
    return ClientContext(
        client_id=state.get("client_id"),
        client_name=state.get("client_name"),
        client_uen=state.get("client_uen"),
        firm_id=state.get("firm_id") or state.get("slack_team_id"),
        slack_team_id=state.get("slack_team_id"),
        region=region,
        accounting_software=state.get("software") or "QBS Ledger",
        base_currency=base_currency,
        tax_registered=_state_tax_registered(state),
        partial_exempt=bool(state.get("partial_exempt", False)),
        fye_month=state.get("fye_month"),
        category_mapping=category_mapping_from_state(state),
        entity_memory=entity_memory_from_state(state),
        tax_codes=tax_codes_from_state(state),
    )


# --------------------------------------------------------------------------- #
# Client store + before_agent_callback
# --------------------------------------------------------------------------- #
class ProfileStore(Protocol):
    """Full read+write store protocol for client profiles."""

    def get(self, client_id: Optional[str]) -> Optional[ClientContext]: ...
    def get_by_channel(self, channel_id: Optional[str]) -> Optional[ClientContext]: ...
    def save_profile(self, profile: dict) -> None: ...
    def set_channel(self, channel_id: str, client_id: str) -> None: ...
    def set_status(self, client_id: str, status: str) -> None: ...


class ClientStore(Protocol):
    def get(self, client_id: Optional[str]) -> Optional[ClientContext]:
        ...


class ChannelClientStore(Protocol):
    """Store that can resolve a client by Slack channel_id (reverse index)."""

    def get_by_channel(self, channel_id: Optional[str]) -> Optional[ClientContext]:
        ...


class InMemoryClientStore:
    """A simple in-process store of :class:`ClientContext` keyed by ``client_id``.

    Supports optional channel-id → client-id mapping for local dev and tests
    (mirrors the ``channels/{channel_id}`` reverse index in Firestore).
    """

    def __init__(self, contexts: Optional[dict[str, ClientContext]] = None):
        self._by_id: dict[str, ClientContext] = dict(contexts or {})
        self._by_channel: dict[str, str] = {}  # channel_id -> client_id

    def add(self, ctx: ClientContext, channel_id: Optional[str] = None) -> None:
        if ctx.client_id:
            self._by_id[ctx.client_id] = ctx
        if channel_id and ctx.client_id:
            self._by_channel[channel_id] = ctx.client_id

    def get(self, client_id: Optional[str]) -> Optional[ClientContext]:
        if not client_id:
            return None
        return self._by_id.get(client_id)

    def get_by_channel(self, channel_id: Optional[str]) -> Optional[ClientContext]:
        if not channel_id:
            return None
        client_id = self._by_channel.get(channel_id)
        return self.get(client_id)

    # ---- write methods (ProfileStore protocol) ----

    def save_profile(self, profile: dict) -> None:
        """Build a ClientContext from a spec §1 profile dict and store it."""
        client_id = profile["client_id"]
        region, base_currency = _profile_region_and_currency(profile)
        ctx = ClientContext(
            client_id=client_id,
            client_name=profile.get("client_name"),
            client_uen=profile.get("client_uen"),
            channel_id=profile.get("channel_id"),
            slack_team_id=profile.get("slack_team_id"),
            firm_id=profile.get("firm_id"),
            fye_month=int(profile["fye_month"]) if profile.get("fye_month") is not None else None,
            region=region,
            accounting_software=profile.get("accounting_software") or "QBS Ledger",
            base_currency=base_currency,
            tax_registered=_profile_tax_registered(profile),
            partial_exempt=bool(profile.get("partial_exempt", False)),
            status=profile.get("status"),
            category_mapping=dict(profile.get("category_mapping") or {}),
        )
        self._by_id[client_id] = ctx
        channel_id = profile.get("channel_id")
        if channel_id:
            self._by_channel[channel_id] = client_id

    def set_channel(self, channel_id: str, client_id: str) -> None:
        self._by_channel[channel_id] = client_id

    def set_status(self, client_id: str, status: str) -> None:
        ctx = self._by_id.get(client_id)
        if ctx is not None:
            ctx.status = status

    def add_correction(self, *, client_id: str, vendor: str,
                       account_code: Optional[str] = None,
                       tax_code: Optional[str] = None,
                       creditor_code: Optional[str] = None) -> None:
        """Upsert a per-client vendor -> {account_code, tax_code, creditor_code} Correction.

        Mirrors :meth:`FirestoreClientStore.add_correction` for the in-process
        store so tests / local dev can exercise the HITL learning hook
        without touching GCP. Merges into the existing ``entity_memory`` entry
        for the same vendor, or appends a new one.
        """
        if not (client_id and vendor):
            return
        ctx = self._by_id.get(client_id)
        if ctx is None:
            return
        for e in ctx.entity_memory:
            if e.name == vendor:
                if account_code:
                    e.mapping_code = account_code
                if tax_code:
                    e.tax_code = tax_code
                if creditor_code:
                    e.creditor_code = creditor_code
                return
        ctx.entity_memory.append(EntityMemoryEntry(
            name=vendor,
            mapping_code=account_code,
            tax_code=tax_code,
            creditor_code=creditor_code,
        ))

    # ---------------------------------------------------------------------- #
    # Familiarity store (ADR-0017 §6, Lever 4)
    # ---------------------------------------------------------------------- #

    def _familiarity_key(self, doc_type: str, vendor: Optional[str] = None) -> str:
        return f"{doc_type}:{vendor}" if vendor else doc_type

    def record_familiarity(self, *, client_id: str, doc_type: str,
                           vendor: Optional[str] = None,
                           direction: Optional[str] = None) -> None:
        """Increment seen_count for the doc_type (and doc_type:vendor) familiarity key(s).

        Mirrors :meth:`FirestoreClientStore.record_familiarity` for the
        in-process store so tests can exercise the Lever 4 hook without GCP.
        When ``vendor`` is supplied, BOTH the bare ``doc_type`` key and the
        compound ``doc_type:vendor`` key are incremented.
        """
        if not (client_id and doc_type):
            return
        if client_id not in self._by_id:
            return
        if not hasattr(self, "_familiarity"):
            self._familiarity: dict[str, dict[str, dict]] = {}
        per_client = self._familiarity.setdefault(client_id, {})

        # Always increment the bare doc_type key.
        bare = self._familiarity_key(doc_type)
        entry = per_client.setdefault(bare, {"seen_count": 0})
        entry["seen_count"] = entry.get("seen_count", 0) + 1
        if direction:
            entry["last_direction"] = direction

        # Also increment the compound key when vendor is known.
        if vendor:
            compound = self._familiarity_key(doc_type, vendor)
            v_entry = per_client.setdefault(compound, {"seen_count": 0})
            v_entry["seen_count"] = v_entry.get("seen_count", 0) + 1
            if direction:
                v_entry["last_direction"] = direction

    def reset_familiarity(self, *, client_id: str, doc_type: str,
                          vendor: Optional[str] = None) -> None:
        """Set seen_count to 0 for the matching familiarity key(s).

        When ``vendor`` is supplied, only the compound ``doc_type:vendor`` key
        is reset (the bare ``doc_type`` key is left intact). When ``vendor`` is
        absent, only the bare ``doc_type`` key is reset.

        Mirrors :meth:`FirestoreClientStore.reset_familiarity`.
        """
        if not (client_id and doc_type):
            return
        if not hasattr(self, "_familiarity"):
            return
        per_client = self._familiarity.get(client_id, {})
        key = self._familiarity_key(doc_type, vendor)
        if key in per_client:
            per_client[key]["seen_count"] = 0

    def get_familiarity(self, client_id: str, doc_type: str,
                        vendor: Optional[str] = None) -> int:
        """Return the current seen_count for the given key (0 if absent).

        Used by tests; not part of the production store contract.
        """
        if not hasattr(self, "_familiarity"):
            return 0
        per_client = self._familiarity.get(client_id, {})
        key = self._familiarity_key(doc_type, vendor)
        return per_client.get(key, {}).get("seen_count", 0)

    def get_familiarity_map(self, client_id: str) -> dict:
        """Return the full {key: {seen_count: n, ...}} map for a client.

        This is the format injected into ``state[FAMILIARITY_KEY]`` by the
        profile callback so ``detect_struggle`` can read it without touching
        the store directly.
        """
        if not hasattr(self, "_familiarity"):
            return {}
        return dict(self._familiarity.get(client_id, {}))

    @classmethod
    def from_setup_dir(cls, directory: str | Path) -> "InMemoryClientStore":
        """Load every ``*Client Setup*.xlsx`` under ``directory`` (recursively),
        keyed by the workbook's parent folder name (e.g. ``path.parent.name``).

        The client identity is derived from the parent folder rather than a
        ``Sys_Config`` sheet (which has been removed); the caller is responsible
        for naming each client's folder with its stable ``client_id``."""
        store = cls()
        for path in sorted(Path(directory).rglob("*Client Setup*.xlsx")):
            if path.name.startswith("~$"):  # skip Excel lock files
                continue
            try:
                store.add(load_client_setup(path, client_id=path.parent.name))
            except Exception:
                # Be defensive: one bad workbook shouldn't break the whole load.
                continue
        return store


def make_load_client_callback(store: ClientStore):
    """Build an ADK ``before_agent_callback`` that injects the client's context into state.

    The returned ``load_client_context(callback_context)`` reads
    ``callback_context.state['client_id']``, loads that client's :class:`ClientContext`
    from ``store``, and writes its ``to_state()`` keys into ``callback_context.state``.

    Duck-typed: ``callback_context`` only needs a ``.state`` mapping, so it is unit-testable
    with a stub having ``.state = {}``. If there is no ``client_id`` or the client is not
    found, state is left as-is (never crashes). Returns ``None`` (ADK convention: proceed).
    """

    def load_client_context(callback_context):
        try:
            state = getattr(callback_context, "state", None)
            if state is None:
                return None
            client_id = state.get("client_id")
            ctx = store.get(client_id)
            if ctx is None:
                return None
            for k, v in ctx.to_state().items():
                state[k] = v
        except Exception:
            # Defensive: a loader failure must not abort the agent run.
            pass
        return None

    return load_client_context


def make_load_client_by_channel_callback(store: ChannelClientStore):
    """Build an ADK ``before_agent_callback`` that resolves a client by Slack channel.

    The returned ``load_client_by_channel(callback_context)`` reads
    ``callback_context.state.get('channel_id')``, resolves the client via
    ``store.get_by_channel(channel_id)``, and writes its ``to_state()`` keys into
    ``callback_context.state``.

    Duck-typed: ``callback_context`` only needs a ``.state`` mapping. If
    ``channel_id`` is absent or the channel has no registered client, state is
    left as-is (never crashes). Returns ``None`` always (ADK convention: proceed).
    """

    def load_client_by_channel(callback_context):
        try:
            state = getattr(callback_context, "state", None)
            if state is None:
                return None
            channel_id = state.get("channel_id")
            ctx = store.get_by_channel(channel_id)
            if ctx is None:
                return None
            for k, v in ctx.to_state().items():
                state[k] = v
        except Exception:
            # Defensive: a loader failure must not abort the agent run.
            pass
        return None

    return load_client_by_channel


# --------------------------------------------------------------------------- #
# Firestore store (production)
# --------------------------------------------------------------------------- #
class FirestoreClientStore:
    """Production client store backed by Firestore.

    Reads the spec §1 client profile document and its ``coa`` / ``entity_memory``
    subcollections into a :class:`ClientContext`. ``google.cloud.firestore`` is
    imported lazily so importing this module never requires the dependency and no
    Firestore call is made unless ``get()`` is invoked.

    A ``client`` injection seam is available for hermetic testing: pass
    ``client=<fake>`` to bypass the real ``firestore.Client`` construction
    entirely (``_firestore()`` returns the injected object directly).

    Expected Firestore layout (spec §1 — 2026-06-12):
        clients/{client_id}          -> { client_id, channel_id, slack_team_id,
                                          client_name, fye_month:int,
                                          accounting_software, gst_registered:bool,
                                          region, base_currency, status, firm_id?,
                                          category_mapping: { category -> account_code | null } }
        clients/{client_id}/coa/{n}          -> { code, description, account_type,
                                                   financial_statement, nature, keywords }
        clients/{client_id}/entity_memory/{n}-> { name, reg_no, mapping_code, role, tax_code }
        channels/{channel_id}        -> { client_id }     # reverse index
    """

    def __init__(self, project: Optional[str] = None, database: Optional[str] = None,
                 collection: str = "clients", client=None):
        from ledgr_slack.config import _ns
        self._project = project
        self._database = database
        self._collection = _ns(collection)
        self._channels_collection = _ns("channels")
        self._injected_client = client  # test seam: if set, _firestore() returns it directly
        self._client = None  # lazy real client

    def _firestore(self):
        # If a client was injected (e.g. in tests), return it without touching GCP.
        if self._injected_client is not None:
            return self._injected_client
        if self._client is None:
            from google.cloud import firestore  # lazy import — never loaded in tests
            kwargs: dict = {}
            if self._project:
                kwargs["project"] = self._project
            if self._database:
                kwargs["database"] = self._database
            self._client = firestore.Client(**kwargs)
        return self._client

    def get(self, client_id: Optional[str]) -> Optional[ClientContext]:
        """Load a :class:`ClientContext` from the Firestore client profile doc.

        Reads profile fields directly from the client doc (spec §1).
        ``category_mapping`` is a MAP FIELD on the client doc (not a subcollection).
        ``coa`` and ``entity_memory`` remain subcollections.
        ``ClientContext.tax_registered`` is the internal name for the profile's
        ``gst_registered`` field; falls back to legacy ``tax_registered`` key.
        """
        if not client_id:
            return None
        db = self._firestore()
        doc_ref = db.collection(self._collection).document(client_id)
        snap = doc_ref.get()
        if not snap.exists:
            return None
        data = snap.to_dict() or {}

        region, base_currency = _profile_region_and_currency(data)

        ctx = ClientContext(
            client_id=client_id,
            client_name=data.get("client_name"),
            client_uen=data.get("client_uen"),
            channel_id=data.get("channel_id"),
            slack_team_id=data.get("slack_team_id"),
            firm_id=data.get("firm_id"),
            fye_month=int(data["fye_month"]) if data.get("fye_month") is not None else None,
            region=region,
            accounting_software=data.get("accounting_software") or "QBS Ledger",
            base_currency=base_currency,
            tax_registered=_profile_tax_registered(data),
            partial_exempt=bool(data.get("partial_exempt", False)),
            status=data.get("status"),
            # category_mapping is a map field on the client doc (spec §1), not a subcollection.
            category_mapping=dict(data.get("category_mapping") or {}),
            tax_codes=list(data.get("tax_codes") or []),
            # sys_config kept as empty dict for back-compat; profile fields are now explicit above.
            sys_config={},
        )

        # entity_memory subcollection
        for d in (c.to_dict() or {} for c in doc_ref.collection("entity_memory").stream()):
            name = d.get("name")
            if not name:
                continue
            ctx.entity_memory.append(EntityMemoryEntry(
                name=name,
                reg_no=d.get("reg_no"),
                mapping_code=d.get("mapping_code"),
                role=d.get("role"),
                tax_code=d.get("tax_code"),
                creditor_code=d.get("creditor_code"),
            ))

        return ctx

    def get_by_channel(self, channel_id: Optional[str]) -> Optional[ClientContext]:
        """Resolve a client via the ``channels/{channel_id}`` reverse-index doc.

        Returns ``None`` if the channel doc is missing or has no ``client_id``.
        """
        if not channel_id:
            return None
        db = self._firestore()
        snap = db.collection(self._channels_collection).document(channel_id).get()
        if not snap.exists:
            return None
        data = snap.to_dict() or {}
        client_id = data.get("client_id")
        if not client_id:
            return None
        return self.get(client_id)

    # ---- write methods (ProfileStore protocol) ----

    def save_profile(self, profile: dict) -> None:
        """Write/merge the profile dict into ``clients/{client_id}``."""
        db = self._firestore()
        db.collection(self._collection).document(profile["client_id"]).set(profile, merge=True)

    def set_channel(self, channel_id: str, client_id: str) -> None:
        """Write the reverse-index ``channels/{channel_id}`` doc."""
        db = self._firestore()
        db.collection(self._channels_collection).document(channel_id).set({"client_id": client_id})

    def set_status(self, client_id: str, status: str) -> None:
        """Merge-update the status field on ``clients/{client_id}``."""
        db = self._firestore()
        db.collection(self._collection).document(client_id).set({"status": status}, merge=True)

    def append_processing_log(
        self, *, client_id: str, file_id: str, entry: dict
    ) -> None:
        """Write one document-delivery record under ``clients/{id}/processing_log/{file_id}``."""
        if not (client_id and file_id and entry):
            return
        db = self._firestore()
        (
            db.collection(self._collection)
            .document(client_id)
            .collection("processing_log")
            .document(file_id)
            .set(entry, merge=True)
        )

    def list_processing_log(
        self, client_id: Optional[str], *, limit: int = 20
    ) -> list[dict]:
        """Return recent processing-log entries for a client (newest first)."""
        if not client_id:
            return []
        cap = max(1, min(int(limit), 50))
        db = self._firestore()
        entries: list[dict] = []
        for snap in (
            db.collection(self._collection)
            .document(client_id)
            .collection("processing_log")
            .stream()
        ):
            row = snap.to_dict() or {}
            doc_id = (
                row.get("file_id")
                or getattr(snap, "id", None)
                or getattr(getattr(snap, "reference", None), "_doc_id", None)
            )
            if doc_id:
                row["file_id"] = doc_id
            entries.append(row)
        entries.sort(key=lambda r: str(r.get("delivered_at") or ""), reverse=True)
        return entries[:cap]

    def add_correction(self, *, client_id: str, vendor: str,
                       account_code: Optional[str] = None,
                       tax_code: Optional[str] = None,
                       creditor_code: Optional[str] = None) -> None:
        """Upsert a per-client vendor -> {account_code, tax_code, creditor_code} Correction.

        ADR-0004: when a human edits an extracted invoice's account_code or
        tax_code, that mapping is persisted as a Correction under
        ``clients/{client_id}/entity_memory/{vendor}`` so the next document
        from the same vendor auto-applies it. ``entity_memory`` is the
        spec-defined learning collection (already wired into categorizer
        ``vendor_name`` resolution).
        """
        if not (client_id and vendor):
            return
        db = self._firestore()
        ref = (
            db.collection(self._collection)
            .document(client_id)
            .collection("entity_memory")
            .document(vendor)
        )
        patch: dict[str, str] = {"name": vendor}
        if account_code:
            patch["mapping_code"] = account_code
        if tax_code:
            patch["tax_code"] = tax_code
        if creditor_code:
            patch["creditor_code"] = creditor_code
        ref.set(patch, merge=True)

    # ---------------------------------------------------------------------- #
    # Familiarity store (ADR-0017 §6, Lever 4)
    # ---------------------------------------------------------------------- #

    @staticmethod
    def _familiarity_key(doc_type: str, vendor: Optional[str] = None) -> str:
        return f"{doc_type}:{vendor}" if vendor else doc_type

    def record_familiarity(self, *, client_id: str, doc_type: str,
                           vendor: Optional[str] = None,
                           direction: Optional[str] = None) -> None:
        """Increment seen_count for the doc_type (and optionally doc_type:vendor) key.

        Uses a Firestore atomic increment so concurrent calls are safe.
        When ``vendor`` is supplied, BOTH the bare ``doc_type`` key and the
        compound ``doc_type:vendor`` key are incremented.

        Subcollection layout:
            clients/{client_id}/familiarity/{key} -> {seen_count, last_seen_at,
                                                       last_direction}
        """
        if not (client_id and doc_type):
            return
        from datetime import datetime, timezone
        from google.cloud import firestore as _fs  # lazy import — not loaded in tests

        db = self._firestore()
        now_iso = datetime.now(timezone.utc).isoformat()
        fam_coll = (
            db.collection(self._collection)
            .document(client_id)
            .collection("familiarity")
        )

        patch: dict = {"seen_count": _fs.Increment(1), "last_seen_at": now_iso}
        if direction:
            patch["last_direction"] = direction

        # Always update the bare doc_type key.
        fam_coll.document(self._familiarity_key(doc_type)).set(patch, merge=True)

        # Also update the compound key when vendor is known.
        if vendor:
            fam_coll.document(self._familiarity_key(doc_type, vendor)).set(
                patch, merge=True
            )

    def reset_familiarity(self, *, client_id: str, doc_type: str,
                          vendor: Optional[str] = None) -> None:
        """Set seen_count to 0 for the matching familiarity key.

        When ``vendor`` is supplied, only the compound ``doc_type:vendor`` key
        is reset (the bare ``doc_type`` key is left intact). When ``vendor`` is
        absent, only the bare ``doc_type`` key is reset.
        """
        if not (client_id and doc_type):
            return
        db = self._firestore()
        key = self._familiarity_key(doc_type, vendor)
        (
            db.collection(self._collection)
            .document(client_id)
            .collection("familiarity")
            .document(key)
            .set({"seen_count": 0}, merge=True)
        )

    def get_familiarity_map(self, client_id: str) -> dict:
        """Load the full familiarity map for ``client_id`` from Firestore.

        Returns ``{key: {seen_count: n, ...}}`` — the format injected into
        ``state[FAMILIARITY_KEY]`` by the profile callback.
        """
        if not client_id:
            return {}
        db = self._firestore()
        result: dict = {}
        for snap in (
            db.collection(self._collection)
            .document(client_id)
            .collection("familiarity")
            .stream()
        ):
            row = snap.to_dict() or {}
            result[snap.id] = row
        return result
