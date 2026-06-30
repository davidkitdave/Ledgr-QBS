"""``SlackLedgerStore`` — the channel-hosted FY ledger workbook (Slack = file store).

Architecture
------------
The ADK graph is **Slack-agnostic**: nodes only prepare a serializable
``state["ledger_rows"]`` payload. This store — owned by the runner layer — is
the only place that talks to Slack about the ledger workbook. Each drop:

1. Look up the Firestore pointer ``clients/{client_id}/ledgers/{fy}`` to find the
   channel's current workbook ``slack_file_id`` (if any).
2. If a pointer exists, **download the current workbook bytes** from Slack (via
   the parked SSRF-hardened downloader) and open it; else start a **fresh
   workbook** with the right exporter's sheet layout.
3. **Append** the new rows into the correct sheet, **idempotently** — the set of
   already-seen doc keys is stored in Firestore on the pointer doc (``seen_doc_keys``
   array field) so the same document never double-appends. The Excel workbook itself
   contains NO dedupe column — it is clean and human-readable.
4. **Re-upload** the updated workbook via ``files_upload_v2``, **update the
   Firestore pointer** with the new ``slack_file_id`` and the updated
   ``seen_doc_keys``, then **delete the previous Slack file** so only ONE growing
   ledger file exists in the channel at any time.

Concurrency: two drops racing the same FY workbook are serialized **per client**
by an in-process lock keyed on ``(client_id, fy)``. (A multi-instance Cloud Run
deployment would back this with a Firestore transaction on the pointer doc; noted
for the deploy step.)

Both the Slack client and the Firestore client are **injectable** so the whole
store is unit-testable with fakes.
"""

from __future__ import annotations

import io
import logging
import os
import re
import threading
import urllib.parse
import urllib.request
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

from ledgr_slack.config import _ns
from ledgr_slack.lease_lock import FirestoreLeaseLock

from ledgr_slack.export.exporters import (
    BankStatementExporter,
    get_exporter,
)

logger = logging.getLogger(__name__)

#: Firestore collection holding client profiles (pointer lives in a subcollection).
_CLIENTS_COLLECTION = "clients"
_LEDGER_STASH_COLLECTION = "dedup_stash"
#: Subcollection name for the per-FY ledger pointer docs.
_LEDGERS_SUBCOLLECTION = "ledgers"

#: Sheet titles for the invoice ledger workbook (mirrors LedgerExporter).
_INVOICE_SHEETS = ("Purchase", "Sales")


def _is_slack_host(host: str) -> bool:
    host = (host or "").lower()
    return host == "slack.com" or host.endswith(".slack.com")


def _invoice_date_sort_key(row: dict) -> tuple:
    """Return a sortable (year, month, day, invoice_number) key for an invoice row.

    Reads the real exporter column names defensively:
    - QBS:  ``"Invoice Date"``  /  ``"Invoice Number"``
    - Xero: ``"*InvoiceDate"``  /  ``"*InvoiceNumber"``
    - Fallback ``"Date"`` for test/legacy rows.

    Dates are parsed from DD/MM/YYYY (the format ``_fmt_date`` produces) or
    YYYY-MM-DD.  Blank / unparseable dates sort LAST via a (9999, 12, 31)
    sentinel so malformed rows never crash and never displace real ones.
    ``doc_key`` is intentionally excluded — it lives on the batch, not the row.
    """
    raw_date = (
        row.get("Invoice Date")
        or row.get("*InvoiceDate")
        or row.get("Date")
        or ""
    )
    inv = str(row.get("Invoice Number") or row.get("*InvoiceNumber") or "")

    # Parse DD/MM/YYYY or YYYY-MM-DD; anything else → sentinel.
    _LAST = (9999, 12, 31)
    s = str(raw_date).strip()
    if not s:
        return (*_LAST, inv)
    # DD/MM/YYYY or DD-MM-YYYY or DD/MM/YY
    m = re.fullmatch(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            # Validate ranges without importing datetime at module level.
            if 1 <= mo <= 12 and 1 <= d <= 31:
                return (y, mo, d, inv)
        except Exception:  # noqa: BLE001
            pass
        return (*_LAST, inv)
    # YYYY-MM-DD
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)), inv)
    return (*_LAST, inv)


class SlackLedgerStoreBase:
    """Fetch → append → re-upload the channel's FY ledger workbook.

    Args:
        db: A Firestore client (or compatible fake) holding the ledger pointer.
        opener: Optional ``urllib`` opener used to stream workbook bytes from
            Slack; defaults to a plain opener. Injected in tests is unnecessary
            because the fake Slack client returns bytes directly (see below).
    """

    def __init__(
        self, db: Any, *, opener: Optional[Any] = None, lease: Optional[Any] = None
    ) -> None:
        self._db = db
        self._opener = opener or urllib.request.build_opener()
        # Per-(client_id, fy) locks serialize concurrent writes to the same workbook
        # WITHIN a single process. Cross-instance serialization is layered on top by
        # the Firestore lease lock (WS5b), taken inside this in-process lock.
        self._locks: dict[tuple[str, str], threading.Lock] = {}
        self._locks_guard = threading.Lock()
        self._lease = lease or FirestoreLeaseLock(
            db,
            instance_id=os.environ.get("K_REVISION", "local"),
            # A fake db (tests) may carry its own firestore namespace so the lease
            # stays hermetic; production db has no such attr → real lazy import.
            firestore_ns=getattr(db, "firestore_ns", None),
        )

    # ------------------------------------------------------------------ #
    # Firestore pointer
    # ------------------------------------------------------------------ #

    def _pointer_ref(self, client_id: str, fy: str) -> Any:
        return (
            self._db.collection(_ns(_CLIENTS_COLLECTION))
            .document(client_id)
            .collection(_LEDGERS_SUBCOLLECTION)
            .document(str(fy))
        )

    def get_pointer(self, client_id: str, fy: str) -> Optional[dict]:
        """Return the ledger pointer doc for ``(client_id, fy)`` or ``None``."""
        snap = self._pointer_ref(client_id, fy).get()
        if not snap.exists:
            return None
        return snap.to_dict()

    def latest_fy(self, client_id: str) -> Optional[str]:
        """Return the highest FY label that has a ledger pointer, or ``None``."""
        coll = (
            self._db.collection(_ns(_CLIENTS_COLLECTION))
            .document(client_id)
            .collection(_LEDGERS_SUBCOLLECTION)
        )
        fys = [snap.id for snap in coll.stream()]
        if not fys:
            return None
        fys.sort()
        return fys[-1]

    def fy_pointers(self, client_id: str) -> list[dict]:
        """Return all FY pointer docs for ``client_id`` as ``{fy, ...}`` dicts.

        Each entry contains at least the ``fy`` label. Used by chat tooling to
        surface every FY the client has ledgers for, so the agent can name the
        correct one when the user asks "show me last year's books".
        """
        coll = (
            self._db.collection(_ns(_CLIENTS_COLLECTION))
            .document(client_id)
            .collection(_LEDGERS_SUBCOLLECTION)
        )
        out: list[dict] = []
        for snap in coll.stream():
            data = snap.to_dict() or {}
            data["fy"] = snap.id
            out.append(data)
        return out

    def _count_rows_in_workbook(self, slack_client: Any, slack_file_id: str) -> int:
        """Return the non-blank row count across all sheets in a workbook.

        Downloads the workbook bytes (same SSRF-hardened path as
        :meth:`read_rows`) and counts rows containing at least one non-None
        cell. Returns 0 on any failure (network, missing file, malformed
        workbook) so the caller can fall back gracefully.
        """
        try:
            data = self._download_workbook(slack_client, slack_file_id)
        except Exception:  # noqa: BLE001
            return 0
        try:
            wb = self._load_workbook(data)
        except Exception:  # noqa: BLE001
            return 0
        total = 0
        for sheet in wb.worksheets:
            if sheet.max_row < 2:
                continue
            headers = [c.value for c in sheet[1]]
            if not any(h is not None for h in headers):
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    total += 1
        return total

    def best_fy_for_chat(
        self,
        client_id: str,
        slack_client: Any,
    ) -> tuple[Optional[str], list[dict]]:
        """Pick the FY label that has the most data for chat answers.

        Returns ``(best_fy, summaries)`` where ``summaries`` is a list of
        ``{fy, row_count, has_data}`` for every FY pointer the client has,
        ordered by ``row_count`` desc then FY label desc. Ties go to the
        highest FY label (matches the previous ``latest_fy`` behaviour for
        empty ledgers, but only after data is checked).

        If no pointers exist, ``best_fy`` is ``None`` and ``summaries`` is
        empty. Pointers with no ``slack_file_id`` are reported as
        ``row_count=0`` without a network call.
        """
        pointers = self.fy_pointers(client_id)
        if not pointers:
            return None, []
        summaries: list[dict] = []
        for ptr in pointers:
            fy = str(ptr.get("fy") or "")
            slack_file_id = ptr.get("slack_file_id")
            if not slack_file_id:
                summaries.append({"fy": fy, "row_count": 0, "has_data": False})
                continue
            count = self._count_rows_in_workbook(slack_client, slack_file_id)
            summaries.append(
                {"fy": fy, "row_count": count, "has_data": count > 0}
            )
        # Best = most rows; tie-break by highest FY label.
        summaries_sorted = sorted(
            summaries,
            key=lambda s: (s.get("row_count", 0), s.get("fy", "")),
            reverse=True,
        )
        best_fy = None
        for s in summaries_sorted:
            if s.get("has_data"):
                best_fy = s.get("fy")
                break
        if best_fy is None and summaries_sorted:
            # No FY has data; fall back to the highest FY label (matches old
            # ``latest_fy`` behaviour so the agent still gets a stable key
            # to report to the user).
            best_fy = max(
                (s.get("fy", "") for s in summaries_sorted),
                default=None,
            )
        return best_fy, summaries_sorted

    def _set_pointer(
        self,
        client_id: str,
        fy: str,
        slack_file_id: str,
        seen_doc_keys: Optional[list] = None,
        **extra: Any,
    ) -> None:
        doc: dict = {"slack_file_id": slack_file_id, "fy": str(fy), "client_id": client_id}
        if seen_doc_keys is not None:
            doc["seen_doc_keys"] = list(seen_doc_keys)
        doc.update(extra)
        self._pointer_ref(client_id, fy).set(doc, merge=True)

    def _get_seen_doc_keys(self, pointer: Optional[dict]) -> set:
        """Return the set of already-processed doc keys from the Firestore pointer."""
        if not pointer:
            return set()
        return set(pointer.get("seen_doc_keys") or [])

    # ------------------------------------------------------------------ #
    # Per-client serialization
    # ------------------------------------------------------------------ #

    def _lock_for(self, client_id: str, fy: str) -> threading.Lock:
        # In-process serialization only; cross-instance serialization is handled by
        # the wrapping Firestore lease lock (WS5b) at each write call site.
        # Keys on (client_id, fy) to match the workbook identity — two different
        # Slack channels mapping to the same client share ONE lock and cannot race.
        key = (client_id, str(fy))
        with self._locks_guard:
            lock = self._locks.get(key)
            if lock is None:
                lock = threading.Lock()
                self._locks[key] = lock
            return lock

    # ------------------------------------------------------------------ #
    # Slack workbook download
    # ------------------------------------------------------------------ #

    def _download_workbook(self, slack_client: Any, slack_file_id: str) -> bytes:
        """Download a workbook's bytes from Slack by file id (SSRF-hardened).

        Mirrors the parked ``app.slack_app.slack_download_file`` host checks but
        returns bytes in-memory (no temp file) since workbooks are small.
        """
        info = slack_client.files_info(file=slack_file_id)
        file_meta = info["file"]
        url = file_meta.get("url_private_download") or file_meta.get("url_private")
        host = urllib.parse.urlparse(url).hostname or ""
        if not _is_slack_host(host):
            raise ValueError(f"refusing to download from non-slack host: {host!r}")
        token = getattr(slack_client, "token", None)
        req = urllib.request.Request(
            url, headers={"Authorization": f"Bearer {token}"}
        )
        with self._opener.open(req) as resp:
            return resp.read()

    # ------------------------------------------------------------------ #
    # Workbook helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _exporter_for(software: str):
        """Return the invoice exporter matching the client's accounting software."""
        from ledgr_slack.export.exporters import normalize_software_key

        key = normalize_software_key(software)
        if key is not None:
            return get_exporter(key)
        if software and str(software).strip():
            raise ValueError(f"unknown software: {software!r}")
        raise ValueError("software not set")

    @staticmethod
    def _invoice_identity_column(exporter: Any, sheet_name: str) -> str:
        """Workbook column used to match invoice rows on replace (MAP5).

        AutoCount AP uses ``SupplierInvoiceNo`` because ``DocNo`` is always
        the constant ``<<New>>``; QBS uses ``Invoice Number``; Xero uses
        ``*InvoiceNumber``.
        """
        doc_type = "sales" if sheet_name == "Sales" else "purchase"
        if hasattr(exporter, "column_for_field"):
            for field in ("invoice_number", "supplier_invoice_no"):
                col = exporter.column_for_field(field, doc_type)
                if col:
                    return col
        return "Invoice Number"

    def _fresh_invoice_workbook(self, software: str) -> Workbook:
        """Create an empty invoice workbook (Purchase + Sales sheets, headers only).

        No dedupe column is written — the Excel is clean/human-readable. Dedupe
        state lives in Firestore on the pointer doc (``seen_doc_keys`` field).
        """
        exporter = self._exporter_for(software)
        wb = Workbook()
        for i, (title, cols) in enumerate(
            (("Purchase", exporter.purchase_cols), ("Sales", exporter.sales_cols))
        ):
            sheet = wb.active if i == 0 else wb.create_sheet(title)
            sheet.title = title
            sheet.append(list(cols))
        return wb

    def _fresh_bank_workbook(self) -> Workbook:
        """Create an empty bank workbook (header-only single placeholder sheet).

        No dedupe column — dedupe state is in Firestore ``seen_doc_keys``.
        """
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Bank"
        sheet.append(list(BankStatementExporter.BANK_COLS))
        return wb

    @staticmethod
    def _slack_file_unavailable(exc: BaseException) -> bool:
        """True when Slack reports the file id is gone (deleted / not found)."""
        err = str(exc).lower()
        return any(
            token in err
            for token in (
                "file_deleted",
                "file_not_found",
                "not_found",
                "missing_file",
                "no_file",
            )
        )

    @staticmethod
    def _workbook_has_transaction_data(wb: Workbook) -> bool:
        """True when at least one sheet has real transaction rows (not B/F or TOTALS)."""
        skip = frozenset({"BALANCE B/F", "TOTALS"})
        for sheet in wb.worksheets:
            if sheet.max_row < 2:
                continue
            headers = [c.value for c in sheet[1]]
            if not any(h for h in headers if h):
                continue
            desc_idx = None
            for i, h in enumerate(headers):
                if h == "Description":
                    desc_idx = i
                    break
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                if desc_idx is not None and desc_idx < len(row):
                    desc = row[desc_idx]
                    if desc is not None and str(desc).strip() not in skip:
                        return True
                elif any(
                    cell is not None and str(cell).strip()
                    for cell in row
                ):
                    return True
        return False

    def _stash_ref(self, client_id: str, stash_key: str):
        doc_id = urllib.parse.quote(stash_key, safe="")
        return (
            self._db.collection(_ns(_CLIENTS_COLLECTION))
            .document(client_id)
            .collection(_LEDGER_STASH_COLLECTION)
            .document(doc_id)
        )

    def stash_bank_dedup_replace(
        self,
        *,
        stash_key: str,
        client_id: str,
        fy: str,
        kind: str,
        software: str,
        client_name: str,
        batches: list[dict],
    ) -> None:
        """Store incoming bank batches for a deferred dedup Replace action."""
        if not client_id or not batches:
            return
        self._stash_ref(client_id, stash_key).set({
            "stash_key": stash_key,
            "client_id": client_id,
            "fy": str(fy),
            "kind": kind,
            "software": software,
            "client_name": client_name,
            "batches": batches,
        })

    def consume_bank_dedup_replace(self, stash_key: str) -> Optional[dict]:
        """Load and delete a stashed dedup-replace payload."""
        parts = stash_key.split("|", 3)
        client_id = parts[0] if parts else ""
        if not client_id:
            return None
        ref = self._stash_ref(client_id, stash_key)
        snap = ref.get()
        if not snap.exists:
            return None
        data = snap.to_dict() or {}
        ref.delete()
        return data

    def purge_seen_doc_keys(
        self,
        client_id: str,
        fy: str,
        doc_keys: list[str],
    ) -> int:
        """Remove ``doc_keys`` from the FY pointer so batches can re-append."""
        pointer = self.get_pointer(client_id, fy)
        if not pointer:
            return 0
        seen = self._get_seen_doc_keys(pointer)
        purged = 0
        for dk in doc_keys:
            if dk in seen:
                seen.discard(dk)
                purged += 1
        if purged == 0:
            return 0
        slack_file_id = pointer.get("slack_file_id") or ""
        extra = {k: v for k, v in pointer.items()
                 if k not in ("slack_file_id", "seen_doc_keys", "fy", "client_id")}
        self._set_pointer(
            client_id,
            fy,
            slack_file_id,
            seen_doc_keys=list(seen),
            **extra,
        )
        return purged

    def _upload_workbook_bytes(
        self,
        *,
        wb: Workbook,
        slack_client: Any,
        channel_id: str,
        client_id: str,
        fy: str,
        filename: str,
        seen_doc_keys: set,
        kind: str,
        client_name: str,
        prev_file_id: Optional[str],
        software: str = "",
    ) -> tuple[Optional[str], Any]:
        """Upload workbook bytes to Slack and update the Firestore pointer."""
        new_bytes = self._to_bytes(wb)
        upload_result = slack_client.files_upload_v2(
            channel=channel_id,
            filename=filename,
            file=new_bytes,
            title=filename,
        )
        new_file_id = self._extract_uploaded_file_id(upload_result)
        if new_file_id:
            # Persist ``software`` on the pointer so remove_rows_for_month can
            # resolve the right exporter at clear time (issue #34): the AutoCount
            # sales row-signature purge needs to know the workbook is AutoCount.
            # Only write a non-empty value so we never blank an onboarding seed.
            extra: dict[str, Any] = {}
            if software:
                extra["software"] = software
            self._set_pointer(
                client_id,
                fy,
                new_file_id,
                seen_doc_keys=list(seen_doc_keys),
                channel_id=channel_id,
                kind=kind,
                client_name=client_name,
                **extra,
            )
        if prev_file_id and new_file_id and prev_file_id != new_file_id:
            try:
                slack_client.files_delete(file=prev_file_id)
            except Exception:  # noqa: BLE001
                logger.warning(
                    "Could not delete superseded ledger file %s (non-fatal).",
                    prev_file_id,
                )
        return new_file_id, upload_result

    @staticmethod
    def _to_bytes(wb: Workbook) -> bytes:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _load_workbook(data: bytes) -> Workbook:
        return load_workbook(io.BytesIO(data))

    @staticmethod
    def _append_rows_to_sheet(sheet, cols: list[str], rows: list[dict]) -> int:
        """Append ``rows`` to ``sheet`` in ``cols`` order (no dedupe column).

        Dedupe is now handled by the Firestore ``seen_doc_keys`` set before this is
        called, so this method always appends. Returns the number of rows appended.

        Incoming rows are sorted by (parsed_date, invoice_number) via
        ``_invoice_date_sort_key`` before appending so the per-batch output is
        deterministic regardless of which concurrent doc completes first (fan-out,
        Step 5).  This is a sort of the *incoming* rows only — rows already in the
        sheet are not touched.  The bank path achieves the same by rebuilding the
        entire sheet via ``_merge_bank_statement``.

        Column names handled (see ``_invoice_date_sort_key`` for details):
        - QBS:  ``"Invoice Date"``  /  ``"Invoice Number"``
        - Xero: ``"*InvoiceDate"``  /  ``"*InvoiceNumber"``
        - Fallback ``"Date"`` for test / legacy rows.
        """
        sorted_rows = sorted(rows, key=_invoice_date_sort_key)
        for row in sorted_rows:
            sheet.append([row.get(c, "") for c in cols])
        return len(sorted_rows)

    # Mapping from old 8-col header names to current BANK_COLS names.
    _LEGACY_COL_MAP: dict[str, str] = {
        "Stated Balance": "Balance",
        "Check": "Math_Check",
    }

    @classmethod
    def _read_bank_blocks(cls, sheet, cols: list[str]) -> list[dict]:
        """Read an existing bank sheet back into normalized month-blocks.

        Returns a list of ``{"stated_bf", "currency", "transactions": [...]}`` blocks
        the exporter rebuilds from. ``TOTALS`` rows are dropped (regenerated on
        rebuild). Dedupe is now fully Firestore-side, so no key tracking here.

        Hardening applied on every read:

        1. **Legacy header migration**: the old 8-col layout used ``Stated Balance``
           and ``Check`` instead of ``Balance`` and ``Math_Check``. We remap those
           column names on the fly so old workbooks are read without data loss.
        2. **Formula / None Balance recompute**: when a Balance cell is a formula
           string (starts with ``"="``) or ``None`` — both possible from workbooks
           written with the old formula-chain style — we RECOMPUTE the running
           balance deterministically from ``stated_bf + Σ(deposit − withdrawal)``.
           The stored Balance value is NEVER trusted; recompute is the single source
           of truth on every rebuild.
        """
        raw_header = [c.value for c in sheet[1]] if sheet.max_row >= 1 else []
        if not raw_header or not any(h for h in raw_header if h):
            return []

        # Remap legacy column names to current canonical names.
        header = [cls._LEGACY_COL_MAP.get(h, h) for h in raw_header]
        if "Description" not in header:
            return []
        col_idx = {name: i for i, name in enumerate(header)}

        value_rows: list[dict] = []
        for raw in sheet.iter_rows(min_row=2, values_only=True):
            row: dict = {}
            for name in cols:
                i = col_idx.get(name)
                row[name] = raw[i] if i is not None and i < len(raw) else None
            value_rows.append(row)

        # Recompute running balances — never trust stored Balance cells.
        # A stored Balance may be a formula string (old layout) or None (data_only
        # read of a formula cell that was never evaluated). We walk block-by-block,
        # seeding each block from its BALANCE B/F opening (which is always a literal
        # number on the B/F marker row itself), then recomputing every txn balance.
        cls._recompute_balances(value_rows)

        return BankStatementExporter.rows_to_blocks(value_rows)

    @staticmethod
    def _is_formula_or_missing(value) -> bool:
        """Return True when a Balance cell cannot be trusted as a numeric value.

        Covers None, formula/empty strings (``=...``), and any other non-numeric
        text — e.g. a stray currency code (``"SGD"``) or a ``BALANCE B/F`` label
        that landed in the Balance column of an older sheet. Recompute then
        carries the running balance forward instead of crashing on
        ``float(value)`` (live bug: ``ValueError: could not convert 'SGD'``).
        """
        if value is None:
            return True
        if isinstance(value, bool):  # bool is an int subclass — not a balance
            return True
        if isinstance(value, (int, float)):
            return False
        s = str(value).strip()
        if not s or s.startswith("="):
            return True
        try:
            float(s)
            return False
        except (TypeError, ValueError):
            return True

    @classmethod
    def _recompute_balances(cls, value_rows: list[dict]) -> None:
        """Recompute Balance on every row in-place from stated_bf + Σ(dep − wd).

        This is the single source of truth: we NEVER use a stored Balance value
        from the workbook. For ``BALANCE B/F`` marker rows we use the stated
        opening (the literal number on that row) as the seed; for transaction rows
        we chain from the prior row's recomputed balance. If the B/F opening
        itself is a formula/None we carry forward the last known balance so the
        chain never breaks silently.
        """
        running: Optional[float] = None
        for row in value_rows:
            desc = row.get("Description")
            if desc == BankStatementExporter.OPENING_MARKER:
                bal = row.get("Balance")
                if cls._is_formula_or_missing(bal):
                    # Carry forward — the B/F opening was a formula; best we can do.
                    row["Balance"] = running
                else:
                    running = float(bal)
                    row["Balance"] = running
            elif desc == BankStatementExporter.TOTALS_MARKER:
                # TOTALS rows have no meaningful Balance; keep blank.
                row["Balance"] = None
            else:
                # Transaction row: recompute from running + deposit − withdrawal.
                if running is not None:
                    dep = row.get("Deposit")
                    wd = row.get("Withdrawal")
                    dep_f = float(dep) if dep not in (None, "") else 0.0
                    wd_f = float(wd) if wd not in (None, "") else 0.0
                    running = round(running + dep_f - wd_f, 2)
                    row["Balance"] = running
                else:
                    # No prior B/F seed — can't recompute; leave as-is (pass through).
                    bal = row.get("Balance")
                    if cls._is_formula_or_missing(bal):
                        row["Balance"] = None
                    else:
                        running = float(bal)
                        row["Balance"] = running

    @classmethod
    def _migrate_legacy_header(cls, sheet, cols: list[str]) -> None:
        """Rewrite row 1 in-place if it contains legacy column names.

        Old 8-col layout used ``Stated Balance`` and ``Check``; current layout
        uses ``Balance`` and ``Math_Check``. If any cell in the header row
        matches a known legacy name, replace it with the canonical current name
        so ``rebuild_account_sheet`` (which always writes the current ``cols``
        on a fresh sheet) doesn't leave a mismatched header behind.
        """
        if sheet.max_row < 1:
            return
        for cell in sheet[1]:
            if cell.value in cls._LEGACY_COL_MAP:
                cell.value = cls._LEGACY_COL_MAP[cell.value]

    @classmethod
    def _ensure_bank_header(cls, sheet, cols: list[str]) -> None:
        """Ensure row 1 carries the canonical ``BANK_COLS`` header.

        ``openpyxl``'s ``create_sheet`` leaves row 1 as a blank placeholder
        (``[None]``), which is truthy in Python — so the old ``if not header``
        guard never fired and secondary currency tabs were written without
        column names or ``Math_Check`` formulas.
        """
        raw = [c.value for c in sheet[1]] if sheet.max_row >= 1 else []
        normalized = [cls._LEGACY_COL_MAP.get(h, h) for h in raw if h]
        if "Description" in normalized and "Balance" in normalized:
            cls._migrate_legacy_header(sheet, cols)
            return
        for i, name in enumerate(cols, start=1):
            sheet.cell(row=1, column=i, value=name)

    def _merge_bank_statement(
        self, sheet, cols: list[str], rows: list[dict]
    ) -> int:
        """Merge one statement's value rows into the account sheet's continuous chain.

        Reads the existing sheet back into sorted month-blocks, appends this
        statement's blocks, re-sorts by date, and REBUILDS the whole sheet.
        Dedupe (doc_key already seen?) is checked in Firestore before this is called.
        Returns the count of newly-added value rows.
        """
        self._ensure_bank_header(sheet, cols)

        existing_blocks = self._read_bank_blocks(sheet, cols)

        new_blocks = BankStatementExporter.rows_to_blocks(rows)

        # Collapse duplicate statement blocks (safety net + one-shot cleanup of
        # any pre-existing duplication already in the sheet, e.g. the doc_key
        # format transition that duplicated September in the Sample Bank Client workbook).
        before = len(existing_blocks) + len(new_blocks)
        deduped_blocks = BankStatementExporter.dedupe_blocks(existing_blocks + new_blocks)
        existing_sig = {
            BankStatementExporter._block_signature(b) for b in existing_blocks
        }
        # "Added" = value rows in genuinely new (not-already-present) blocks.
        added = 0
        for block in deduped_blocks:
            if BankStatementExporter._block_signature(block) not in existing_sig:
                added += len(block["transactions"]) + 1  # +1 for BALANCE B/F row
        if len(deduped_blocks) < before:
            logger.info(
                "bank merge: collapsed %d duplicate statement block(s) on %s",
                before - len(deduped_blocks), sheet.title,
            )

        all_blocks = BankStatementExporter.sort_blocks(deduped_blocks)
        BankStatementExporter.rebuild_account_sheet(sheet, all_blocks, cols)
        return added

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

