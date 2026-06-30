"""Append mixin for ``SlackLedgerStore`` — fetch → append → re-upload of the channel's FY workbook.

See :mod:`ledgr_slack.ledger_store_base` for the full architecture; this module
only adds the append path.
"""

from __future__ import annotations

import logging
from typing import Any, Optional

from openpyxl import Workbook

from ledgr_slack.export.exporters import BankStatementExporter

from ledgr_slack.ledger_store_base import SlackLedgerStoreBase

logger = logging.getLogger(__name__)

#: Sheet titles for the invoice ledger workbook (mirrors LedgerExporter).
_INVOICE_SHEETS = ("Purchase", "Sales")


class _SlackLedgerStoreAppendMixin:
    def append_rows(
        self,
        *,
        client_id: str,
        fy: str,
        slack_client: Any,
        channel_id: str,
        batches: list[dict],
        software: str = "qbs",
        kind: str = "invoice",
        client_name: str = "",
        replace: bool = False,
    ) -> dict:
        """Append a run's rows to the channel FY workbook (fetch → append → upload).

        Args:
            client_id: The client whose ledger pointer is updated.
            fy: Financial-year label (the workbook is per-FY).
            slack_client: A Slack WebClient (or fake) for download + upload.
            channel_id: The channel the workbook is uploaded to.
            batches: A list of ``{"sheet": str, "doc_key": str, "rows": [dict]}``
                payloads (as produced by ``consolidate_node`` into
                ``state["ledger_rows"]``). ``sheet`` is "Purchase"/"Sales" for the
                invoice workbook or the bank account label for the bank workbook.
            software: The client's accounting software ("qbs"/"xero") — selects the
                exporter column layout for a fresh invoice workbook.
            kind: "invoice" or "bank" — selects workbook layout + filename.
            replace: When ``True``, for each INVOICE batch, existing rows whose
                ``Invoice Number`` matches the batch's invoice number(s) are deleted
                BEFORE the batch is appended, and the batch's ``doc_key`` is removed
                from ``seen_doc_keys`` so the re-append is not silently deduped.
                Bank batches are unaffected (they use the merge path regardless).
                ``False`` (default) preserves today's exact dedup behaviour.

        Returns:
            ``{"slack_file_id", "appended", "deduped", "filename"}``.
            When ``replace=True`` the dict also includes
            ``"batch_replace_counts": [{sheet, doc_key, replaced, appended}, ...]``
            so callers can warn when no old rows were matched (identity changed).
        """
        lock = self._lock_for(client_id, fy)
        with lock:  # fast same-process serialize
            token = self._lease.acquire(client_id, fy)  # cross-instance serialize
            try:
                return self._append_rows_locked(
                    client_id=client_id,
                    fy=fy,
                    slack_client=slack_client,
                    channel_id=channel_id,
                    batches=batches,
                    software=software,
                    kind=kind,
                    client_name=client_name,
                    replace=replace,
                )
            finally:
                self._lease.release(client_id, fy, token)

    def _append_rows_locked(
        self,
        *,
        client_id: str,
        fy: str,
        slack_client: Any,
        channel_id: str,
        batches: list[dict],
        software: str,
        kind: str,
        client_name: str = "",
        replace: bool = False,
    ) -> dict:
        pointer = self.get_pointer(client_id, fy)

        # Capture the PREVIOUS file id before we overwrite the pointer.
        prev_file_id: Optional[str] = pointer.get("slack_file_id") if pointer else None

        # Read the Firestore-side set of already-processed doc keys.
        seen_doc_keys: set = self._get_seen_doc_keys(pointer)

        # When the user deleted the workbook message from Slack the pointer can
        # still reference a dead file id while seen_doc_keys blocks re-append.
        if prev_file_id:
            try:
                slack_client.files_info(file=prev_file_id)
            except Exception as exc:  # noqa: BLE001
                if self._slack_file_unavailable(exc):
                    logger.warning(
                        "Workbook %s unavailable in Slack — resetting FY%s pointer "
                        "and clearing seen_doc_keys so the drop can re-append.",
                        prev_file_id, fy,
                    )
                    prev_file_id = None
                    seen_doc_keys = set()
                else:
                    raise

        if prev_file_id:
            try:
                data = self._download_workbook(slack_client, prev_file_id)
                wb = self._load_workbook(data)
            except Exception as exc:
                if self._slack_file_unavailable(exc):
                    logger.warning(
                        "Previous workbook %s gone from Slack — starting fresh FY%s.",
                        prev_file_id, fy,
                    )
                    prev_file_id = None
                    seen_doc_keys = set()
                    wb = self._fresh_bank_workbook() if kind == "bank" else self._fresh_invoice_workbook(software)
                else:
                    raise
        elif kind == "bank":
            wb = self._fresh_bank_workbook()
        else:
            wb = self._fresh_invoice_workbook(software)

        # Client-scoped filename: "<Client> - BankStatement_FY<fy>.xlsx" /
        # "<Client> - Ledger_FY<fy>.xlsx" (matches the reference workbook naming).
        # Falls back to the bare name when the profile has no client_name.
        prefix = f"{client_name.strip()} - " if client_name.strip() else ""
        if kind == "bank":
            filename = f"{prefix}BankStatement_FY{fy}.xlsx"
            cols = list(BankStatementExporter.BANK_COLS)
        else:
            filename = f"{prefix}Ledger_FY{fy}.xlsx"
            exporter = self._exporter_for(software)
            sheet_cols = {
                "Purchase": list(exporter.purchase_cols),
                "Sales": list(exporter.sales_cols),
            }

        appended = 0
        deduped = 0
        batch_replace_counts: list[dict] = []

        for batch in batches:
            sheet_name = batch["sheet"]
            doc_key = str(batch["doc_key"])
            rows = batch.get("rows") or []

            # ------------------------------------------------------------------ #
            # replace=True path (INVOICE sheets only).
            # For each invoice batch: find existing rows whose Invoice Number
            # matches the batch's invoice numbers, delete them bottom-up, then
            # remove the doc_key from seen so the re-append is NOT deduped.
            # Bank batches are skipped — they use the merge path below regardless.
            # ------------------------------------------------------------------ #
            replaced_count = 0
            if replace and kind == "invoice" and sheet_name in _INVOICE_SHEETS:
                if not rows:
                    deduped += 1
                    batch_replace_counts.append(
                        {"sheet": sheet_name, "doc_key": doc_key,
                         "replaced": 0, "appended": 0}
                    )
                    continue
                identity_col = self._invoice_identity_column(exporter, sheet_name)
                # Collect supplier/reference invoice numbers carried by the batch.
                batch_inv_nums: set[str] = {
                    str(r.get(identity_col, "")).strip()
                    for r in rows
                    if r.get(identity_col) not in (None, "")
                }

                if batch_inv_nums and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.max_row >= 2:
                        col_map = self._header_col_map(ws)
                        inv_col = (
                            col_map.get(identity_col)
                            or col_map.get("Invoice Number")
                            or col_map.get("*InvoiceNumber")
                        )
                        if inv_col is not None:
                            # Collect matching row indices (ascending) then delete bottom-up.
                            matching_rows: list[int] = []
                            for row_num in range(2, ws.max_row + 1):
                                cell_val = ws.cell(row=row_num, column=inv_col).value
                                if cell_val is not None and str(cell_val).strip() in batch_inv_nums:
                                    matching_rows.append(row_num)

                            replaced_count = len(matching_rows)
                            for row_num in sorted(matching_rows, reverse=True):
                                ws.delete_rows(row_num, 1)

                # Remove this batch's doc_key from seen so the subsequent append
                # is NOT skipped by the dedup guard below.
                seen_doc_keys.discard(doc_key)

            # ------------------------------------------------------------------ #
            # Firestore-side dedupe: skip entirely if this doc was already processed.
            # (replace=True already discarded the key above, so this only fires for
            #  replace=False or for bank batches where the key is genuinely new.)
            # ------------------------------------------------------------------ #
            logger.debug("append_rows: doc_key=%r match=%s", doc_key, doc_key in seen_doc_keys)
            if doc_key in seen_doc_keys:
                deduped += 1
                if replace:
                    batch_replace_counts.append(
                        {"sheet": sheet_name, "doc_key": doc_key,
                         "replaced": 0, "appended": 0}
                    )
                continue

            sheet = self._get_or_create_sheet(wb, sheet_name, kind)
            if kind == "bank":
                # Bank: REBUILD the account sheet as one continuous, date-sorted chain
                # (merge existing rows + this statement, sort months, re-thread the
                # running balance), rather than blindly appending a self-seeding block.
                n = self._merge_bank_statement(sheet, cols, rows)
            else:
                cols_for_sheet = sheet_cols.get(sheet_name, list(exporter.purchase_cols))
                n = self._append_rows_to_sheet(sheet, cols_for_sheet, rows)

            if n == 0:
                deduped += 1
                if replace:
                    batch_replace_counts.append(
                        {"sheet": sheet_name, "doc_key": doc_key,
                         "replaced": replaced_count, "appended": 0}
                    )
            else:
                appended += n
                seen_doc_keys.add(doc_key)
                if replace:
                    batch_replace_counts.append(
                        {"sheet": sheet_name, "doc_key": doc_key,
                         "replaced": replaced_count, "appended": n}
                    )

        # Skip the upload when nothing new was appended — unless every batch was
        # deduped and we still hold a populated workbook. That happens when the
        # user deleted the Slack message but Firestore seen_doc_keys still block
        # re-append; re-post the existing workbook so the channel regains a file.
        if appended == 0:
            result: dict = {
                "slack_file_id": prev_file_id or "",
                "appended": 0,
                "deduped": deduped,
                "filename": filename,
            }
            if replace:
                result["batch_replace_counts"] = batch_replace_counts
            if deduped > 0 and self._workbook_has_transaction_data(wb):
                new_file_id, _ = self._upload_workbook_bytes(
                    wb=wb,
                    slack_client=slack_client,
                    channel_id=channel_id,
                    client_id=client_id,
                    fy=fy,
                    filename=filename,
                    seen_doc_keys=seen_doc_keys,
                    kind=kind,
                    client_name=client_name,
                    prev_file_id=prev_file_id,
                    software=software,
                )
                if new_file_id:
                    result["slack_file_id"] = new_file_id
                    result["reshared"] = True
            return result

        new_file_id, _ = self._upload_workbook_bytes(
            wb=wb,
            slack_client=slack_client,
            channel_id=channel_id,
            client_id=client_id,
            fy=fy,
            filename=filename,
            seen_doc_keys=seen_doc_keys,
            kind=kind,
            client_name=client_name,
            prev_file_id=prev_file_id,
            software=software,
        )

        result = {
            "slack_file_id": new_file_id,
            "appended": appended,
            "deduped": deduped,
            "filename": filename,
        }
        if replace:
            result["batch_replace_counts"] = batch_replace_counts
        return result

    @staticmethod
    def _get_or_create_sheet(wb: Workbook, sheet_name: str, kind: str):
        """Return the named sheet, creating it (with bank header when needed) if absent.

        For the bank workbook the first fresh sheet is titled "Bank"; the first
        real account batch renames it instead of leaving an empty placeholder.
        Additional bank account/currency tabs get a canonical header row immediately.
        """
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        # Reuse an empty default "Bank" placeholder for the first bank account.
        if kind == "bank" and "Bank" in wb.sheetnames and wb["Bank"].max_row <= 1:
            sheet = wb["Bank"]
            sheet.title = sheet_name
            return sheet
        sheet = wb.create_sheet(sheet_name)
        if kind == "bank":
            sheet.append(list(BankStatementExporter.BANK_COLS))
        return sheet

    # ------------------------------------------------------------------ #
    # Public read API
    # ------------------------------------------------------------------ #

