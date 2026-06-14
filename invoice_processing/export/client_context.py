"""Per-client context: load a client's uploaded COA/Category_Mapping/Entity_Memory
into typed objects and inject them into ADK ``session.state`` via a
``before_agent_callback``.

Multi-tenant by construction — nothing here hardcodes any client's account numbers.
The client profile (region, accounting_software, base_currency, gst/tax_registered,
fye_month) comes from the **per-channel Firestore profile** created at onboarding
(see ``FirestoreClientStore``). This module parses only the client's uploaded
COA / Category_Mapping / Entity_Memory workbook data; the categorizer reads them back
out of plain ``state`` (see ``categorizer.py``).

Firestore layout (spec §1 — 2026-06-12):
  clients/{client_id}          -> { client_id, channel_id, slack_team_id, client_name,
                                    fye_month:int, accounting_software, gst_registered:bool,
                                    region, base_currency, status, firm_id?,
                                    category_mapping: { category -> account_code | null } }
  clients/{client_id}/coa/{n}          -> { code, description, account_type,
                                             financial_statement, nature, keywords }
  clients/{client_id}/entity_memory/{n}-> { name, reg_no, mapping_code, role, tax_code }
  channels/{channel_id}        -> { client_id }     # reverse index: channel -> client

Client Setup workbook schema (verified against a sample client's files):
- ``COA``: ``Account code | Description | Account type | Financial Statement | Nature |
  AI Search Keywords``. In SG/QBS files ``Account code`` is blank — the account is keyed
  by ``Description``; in MY files it is a code like ``200-010``.
- ``Category_Mapping`` (may be ABSENT): ``Category | Account Code | Enabled | Notes`` —
  universal category -> this client's account code (often unmapped/None).
- ``Entity_Memory`` (may be empty): ``Name | Reg No / Tax ID | Mapping Code |
  Role (Debtor / Creditor) | Tax Code`` — learned vendor -> account + tax.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Protocol

from openpyxl import load_workbook


# --------------------------------------------------------------------------- #
# Dataclasses
# --------------------------------------------------------------------------- #
@dataclass
class CoaAccount:
    code: Optional[str]                       # may be None (QBS keys by name)
    description: str
    account_type: Optional[str] = None
    financial_statement: Optional[str] = None
    nature: Optional[str] = None
    keywords: Optional[str] = None            # "AI Search Keywords"

    @property
    def key(self) -> str:                     # code if present else description
        return (self.code or self.description or "").strip()


@dataclass
class EntityMemoryEntry:
    name: str
    reg_no: Optional[str] = None
    mapping_code: Optional[str] = None        # account code/name
    role: Optional[str] = None
    tax_code: Optional[str] = None


@dataclass
class ClientContext:
    client_id: Optional[str] = None
    client_name: Optional[str] = None
    client_uen: Optional[str] = None      # UEN / company registration number
    channel_id: Optional[str] = None
    slack_team_id: Optional[str] = None
    firm_id: Optional[str] = None
    status: Optional[str] = None
    region: str = "SINGAPORE"
    accounting_software: str = "QBS Ledger"
    base_currency: str = "SGD"
    tax_registered: bool = True
    fye_month: Optional[int] = None
    coa: list[CoaAccount] = field(default_factory=list)
    category_mapping: dict[str, Optional[str]] = field(default_factory=dict)  # category -> account_code | null
    entity_memory: list[EntityMemoryEntry] = field(default_factory=list)
    sys_config: dict[str, str] = field(default_factory=dict)                  # kept for back-compat; not populated from Firestore

    def to_state(self) -> dict:
        """Serializable dict for ``session.state`` (basic types only)."""
        return {
            "client_id": self.client_id,
            "client_name": self.client_name,
            "client_uen": self.client_uen,
            "region": self.region,
            "tax_registered": self.tax_registered,
            "software": self.accounting_software,
            "base_currency": self.base_currency,
            "fye_month": self.fye_month,
            "coa": [
                {
                    "key": c.key,
                    "code": c.code,
                    "description": c.description,
                    "account_type": c.account_type,
                    "financial_statement": c.financial_statement,
                    "nature": c.nature,
                    "keywords": c.keywords,
                }
                for c in self.coa
            ],
            "category_mapping": dict(self.category_mapping),
            "entity_memory": [
                {
                    "name": e.name,
                    "reg_no": e.reg_no,
                    "mapping_code": e.mapping_code,
                    "role": e.role,
                    "tax_code": e.tax_code,
                }
                for e in self.entity_memory
            ],
        }


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _s(v) -> Optional[str]:
    """Coerce a cell to a stripped non-empty string or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _truthy(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ("true", "yes", "y", "1")


def _row_is_empty(row) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def _header_index(header: tuple, *names: str) -> Optional[int]:
    """Find the first column whose header matches any of ``names`` (case-insensitive)."""
    want = {n.strip().lower() for n in names}
    for i, h in enumerate(header):
        if h is not None and str(h).strip().lower() in want:
            return i
    return None


# --------------------------------------------------------------------------- #
# Workbook loader
# --------------------------------------------------------------------------- #
def load_client_setup(xlsx_path: str | Path, client_id: Optional[str] = None) -> ClientContext:
    """Parse a Client Setup workbook into a :class:`ClientContext`.

    Parses only the client's uploaded COA / Category_Mapping / Entity_Memory data.
    Profile metadata (region, accounting_software, base_currency, tax_registered,
    fye_month) now comes from the per-channel Firestore profile created at onboarding
    and is NOT read from this workbook.

    The caller should supply ``client_id`` explicitly; it is set directly on the
    returned context. Missing sheets are handled gracefully. ``Account code`` may be
    blank (QBS keys by Description). Only enabled ``Category_Mapping`` rows are kept.
    Fully-empty rows in COA / Entity_Memory are skipped.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheets = set(wb.sheetnames)
        ctx = ClientContext()
        ctx.client_id = client_id

        # --- COA ---
        if "COA" in sheets:
            rows = list(wb["COA"].iter_rows(values_only=True))
            if rows:
                header = rows[0]
                i_code = _header_index(header, "Account code", "Account Code") or 0
                i_desc = _header_index(header, "Description")
                i_type = _header_index(header, "Account type", "Account Type")
                i_fs = _header_index(header, "Financial Statement")
                i_nat = _header_index(header, "Nature")
                i_kw = _header_index(header, "AI Search Keywords", "Keywords")
                for row in rows[1:]:
                    if _row_is_empty(row):
                        continue

                    def cell(idx):
                        return row[idx] if idx is not None and idx < len(row) else None

                    code = _s(cell(i_code))
                    desc = _s(cell(i_desc)) or ""
                    if not code and not desc:
                        continue
                    ctx.coa.append(CoaAccount(
                        code=code,
                        description=desc,
                        account_type=_s(cell(i_type)),
                        financial_statement=_s(cell(i_fs)),
                        nature=_s(cell(i_nat)),
                        keywords=_s(cell(i_kw)),
                    ))

        # --- Category_Mapping (optional) ---
        if "Category_Mapping" in sheets:
            rows = list(wb["Category_Mapping"].iter_rows(values_only=True))
            if rows:
                header = rows[0]
                i_cat = _header_index(header, "Category") or 0
                i_acc = _header_index(header, "Account Code", "Account code")
                i_en = _header_index(header, "Enabled")
                for row in rows[1:]:
                    if _row_is_empty(row):
                        continue
                    cat = _s(row[i_cat] if i_cat < len(row) else None)
                    if not cat:
                        continue
                    enabled = _truthy(row[i_en]) if (i_en is not None and i_en < len(row)) else True
                    if not enabled:
                        continue
                    acc = _s(row[i_acc]) if (i_acc is not None and i_acc < len(row)) else None
                    ctx.category_mapping[cat] = acc  # keep None (unmapped)

        # --- Entity_Memory (optional / may be empty) ---
        if "Entity_Memory" in sheets:
            rows = list(wb["Entity_Memory"].iter_rows(values_only=True))
            if rows:
                header = rows[0]
                i_name = _header_index(header, "Name") or 0
                i_reg = _header_index(header, "Reg No / Tax ID", "Reg No", "Tax ID")
                i_map = _header_index(header, "Mapping Code")
                i_role = _header_index(header, "Role (Debtor / Creditor)", "Role")
                i_tax = _header_index(header, "Tax Code")
                for row in rows[1:]:
                    if _row_is_empty(row):
                        continue
                    name = _s(row[i_name] if i_name < len(row) else None)
                    if not name:
                        continue

                    def cell(idx):
                        return row[idx] if idx is not None and idx < len(row) else None

                    ctx.entity_memory.append(EntityMemoryEntry(
                        name=name,
                        reg_no=_s(cell(i_reg)),
                        mapping_code=_s(cell(i_map)),
                        role=_s(cell(i_role)),
                        tax_code=_s(cell(i_tax)),
                    ))

        return ctx
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# Plain-dict (state) accessors — the categorizer reads these out of session.state
# --------------------------------------------------------------------------- #
def coa_from_state(state: dict) -> list[CoaAccount]:
    out: list[CoaAccount] = []
    for d in state.get("coa") or []:
        out.append(CoaAccount(
            code=d.get("code"),
            description=d.get("description") or d.get("key") or "",
            account_type=d.get("account_type"),
            financial_statement=d.get("financial_statement"),
            nature=d.get("nature"),
            keywords=d.get("keywords"),
        ))
    return out


def entity_memory_from_state(state: dict) -> list[EntityMemoryEntry]:
    out: list[EntityMemoryEntry] = []
    for d in state.get("entity_memory") or []:
        out.append(EntityMemoryEntry(
            name=d.get("name") or "",
            reg_no=d.get("reg_no"),
            mapping_code=d.get("mapping_code"),
            role=d.get("role"),
            tax_code=d.get("tax_code"),
        ))
    return out


def category_mapping_from_state(state: dict) -> dict[str, Optional[str]]:
    return dict(state.get("category_mapping") or {})


def client_context_from_state(state: dict) -> ClientContext:
    """Rebuild a :class:`ClientContext` from a plain ``to_state()`` dict."""
    return ClientContext(
        client_id=state.get("client_id"),
        client_name=state.get("client_name"),
        client_uen=state.get("client_uen"),
        region=state.get("region") or "SINGAPORE",
        accounting_software=state.get("software") or "QBS Ledger",
        base_currency=state.get("base_currency") or "SGD",
        tax_registered=bool(state.get("tax_registered", True)),
        fye_month=state.get("fye_month"),
        coa=coa_from_state(state),
        category_mapping=category_mapping_from_state(state),
        entity_memory=entity_memory_from_state(state),
    )


# --------------------------------------------------------------------------- #
# Client store + before_agent_callback
# --------------------------------------------------------------------------- #
class ProfileStore(Protocol):
    """Full read+write store protocol for client profiles."""

    def get(self, client_id: Optional[str]) -> Optional[ClientContext]: ...
    def get_by_channel(self, channel_id: Optional[str]) -> Optional[ClientContext]: ...
    def save_profile(self, profile: dict) -> None: ...
    def set_channel(self, channel_id: str, client_id: str) -> None: ...
    def save_coa(self, client_id: str, coa_rows: list[dict]) -> None: ...
    def set_status(self, client_id: str, status: str) -> None: ...


class ClientStore(Protocol):
    def get(self, client_id: Optional[str]) -> Optional[ClientContext]:
        ...


class ChannelClientStore(Protocol):
    """Store that can resolve a client by Slack channel_id (reverse index)."""

    def get_by_channel(self, channel_id: Optional[str]) -> Optional[ClientContext]:
        ...


class InMemoryClientStore:
    """A simple in-process store of :class:`ClientContext` keyed by ``client_id``.

    Supports optional channel-id → client-id mapping for local dev and tests
    (mirrors the ``channels/{channel_id}`` reverse index in Firestore).
    """

    def __init__(self, contexts: Optional[dict[str, ClientContext]] = None):
        self._by_id: dict[str, ClientContext] = dict(contexts or {})
        self._by_channel: dict[str, str] = {}  # channel_id -> client_id

    def add(self, ctx: ClientContext, channel_id: Optional[str] = None) -> None:
        if ctx.client_id:
            self._by_id[ctx.client_id] = ctx
        if channel_id and ctx.client_id:
            self._by_channel[channel_id] = ctx.client_id

    def get(self, client_id: Optional[str]) -> Optional[ClientContext]:
        if not client_id:
            return None
        return self._by_id.get(client_id)

    def get_by_channel(self, channel_id: Optional[str]) -> Optional[ClientContext]:
        if not channel_id:
            return None
        client_id = self._by_channel.get(channel_id)
        return self.get(client_id)

    # ---- write methods (ProfileStore protocol) ----

    def save_profile(self, profile: dict) -> None:
        """Build a ClientContext from a spec §1 profile dict and store it."""
        client_id = profile["client_id"]
        tax_registered = bool(
            profile.get("gst_registered", profile.get("tax_registered", True))
        )
        ctx = ClientContext(
            client_id=client_id,
            client_name=profile.get("client_name"),
            client_uen=profile.get("client_uen"),
            channel_id=profile.get("channel_id"),
            slack_team_id=profile.get("slack_team_id"),
            firm_id=profile.get("firm_id"),
            fye_month=int(profile["fye_month"]) if profile.get("fye_month") is not None else None,
            region=profile.get("region") or "SINGAPORE",
            accounting_software=profile.get("accounting_software") or "QBS Ledger",
            base_currency=profile.get("base_currency") or "SGD",
            tax_registered=tax_registered,
            status=profile.get("status"),
            category_mapping=dict(profile.get("category_mapping") or {}),
        )
        self._by_id[client_id] = ctx
        channel_id = profile.get("channel_id")
        if channel_id:
            self._by_channel[channel_id] = client_id

    def set_channel(self, channel_id: str, client_id: str) -> None:
        self._by_channel[channel_id] = client_id

    def save_coa(self, client_id: str, coa_rows: list[dict]) -> None:
        ctx = self._by_id.get(client_id)
        if ctx is None:
            return
        # REPLACE semantics: a re-upload fully supersedes any prior COA.
        ctx.coa.clear()
        for row in coa_rows:
            ctx.coa.append(CoaAccount(
                code=row.get("code"),
                description=row.get("description") or "",
                account_type=row.get("account_type"),
                financial_statement=row.get("financial_statement"),
                nature=row.get("nature"),
                keywords=row.get("keywords"),
            ))

    def set_status(self, client_id: str, status: str) -> None:
        ctx = self._by_id.get(client_id)
        if ctx is not None:
            ctx.status = status

    def add_correction(self, *, client_id: str, vendor: str,
                       account_code: Optional[str] = None,
                       tax_code: Optional[str] = None) -> None:
        """Upsert a per-client vendor -> {account_code, tax_code} Correction.

        Mirrors :meth:`FirestoreClientStore.add_correction` for the in-process
        store so tests / local dev can exercise the HITL learning hook
        without touching GCP. Merges into the existing ``entity_memory`` entry
        for the same vendor, or appends a new one.
        """
        if not (client_id and vendor):
            return
        ctx = self._by_id.get(client_id)
        if ctx is None:
            return
        for e in ctx.entity_memory:
            if e.name == vendor:
                if account_code:
                    e.mapping_code = account_code
                if tax_code:
                    e.tax_code = tax_code
                return
        ctx.entity_memory.append(EntityMemoryEntry(
            name=vendor, mapping_code=account_code, tax_code=tax_code,
        ))

    @classmethod
    def from_setup_dir(cls, directory: str | Path) -> "InMemoryClientStore":
        """Load every ``*Client Setup*.xlsx`` under ``directory`` (recursively),
        keyed by the workbook's parent folder name (e.g. ``path.parent.name``).

        The client identity is derived from the parent folder rather than a
        ``Sys_Config`` sheet (which has been removed); the caller is responsible
        for naming each client's folder with its stable ``client_id``."""
        store = cls()
        for path in sorted(Path(directory).rglob("*Client Setup*.xlsx")):
            if path.name.startswith("~$"):  # skip Excel lock files
                continue
            try:
                store.add(load_client_setup(path, client_id=path.parent.name))
            except Exception:
                # Be defensive: one bad workbook shouldn't break the whole load.
                continue
        return store


def make_load_client_callback(store: ClientStore):
    """Build an ADK ``before_agent_callback`` that injects the client's context into state.

    The returned ``load_client_context(callback_context)`` reads
    ``callback_context.state['client_id']``, loads that client's :class:`ClientContext`
    from ``store``, and writes its ``to_state()`` keys into ``callback_context.state``.

    Duck-typed: ``callback_context`` only needs a ``.state`` mapping, so it is unit-testable
    with a stub having ``.state = {}``. If there is no ``client_id`` or the client is not
    found, state is left as-is (never crashes). Returns ``None`` (ADK convention: proceed).
    """

    def load_client_context(callback_context):
        try:
            state = getattr(callback_context, "state", None)
            if state is None:
                return None
            client_id = state.get("client_id")
            ctx = store.get(client_id)
            if ctx is None:
                return None
            for k, v in ctx.to_state().items():
                state[k] = v
        except Exception:
            # Defensive: a loader failure must not abort the agent run.
            pass
        return None

    return load_client_context


def make_load_client_by_channel_callback(store: ChannelClientStore):
    """Build an ADK ``before_agent_callback`` that resolves a client by Slack channel.

    The returned ``load_client_by_channel(callback_context)`` reads
    ``callback_context.state.get('channel_id')``, resolves the client via
    ``store.get_by_channel(channel_id)``, and writes its ``to_state()`` keys into
    ``callback_context.state``.

    Duck-typed: ``callback_context`` only needs a ``.state`` mapping. If
    ``channel_id`` is absent or the channel has no registered client, state is
    left as-is (never crashes). Returns ``None`` always (ADK convention: proceed).
    """

    def load_client_by_channel(callback_context):
        try:
            state = getattr(callback_context, "state", None)
            if state is None:
                return None
            channel_id = state.get("channel_id")
            ctx = store.get_by_channel(channel_id)
            if ctx is None:
                return None
            for k, v in ctx.to_state().items():
                state[k] = v
        except Exception:
            # Defensive: a loader failure must not abort the agent run.
            pass
        return None

    return load_client_by_channel


# --------------------------------------------------------------------------- #
# Firestore store (production)
# --------------------------------------------------------------------------- #
class FirestoreClientStore:
    """Production client store backed by Firestore.

    Reads the spec §1 client profile document and its ``coa`` / ``entity_memory``
    subcollections into a :class:`ClientContext`. ``google.cloud.firestore`` is
    imported lazily so importing this module never requires the dependency and no
    Firestore call is made unless ``get()`` is invoked.

    A ``client`` injection seam is available for hermetic testing: pass
    ``client=<fake>`` to bypass the real ``firestore.Client`` construction
    entirely (``_firestore()`` returns the injected object directly).

    Expected Firestore layout (spec §1 — 2026-06-12):
        clients/{client_id}          -> { client_id, channel_id, slack_team_id,
                                          client_name, fye_month:int,
                                          accounting_software, gst_registered:bool,
                                          region, base_currency, status, firm_id?,
                                          category_mapping: { category -> account_code | null } }
        clients/{client_id}/coa/{n}          -> { code, description, account_type,
                                                   financial_statement, nature, keywords }
        clients/{client_id}/entity_memory/{n}-> { name, reg_no, mapping_code, role, tax_code }
        channels/{channel_id}        -> { client_id }     # reverse index
    """

    def __init__(self, project: Optional[str] = None, database: Optional[str] = None,
                 collection: str = "clients", client=None):
        self._project = project
        self._database = database
        self._collection = collection
        self._injected_client = client  # test seam: if set, _firestore() returns it directly
        self._client = None  # lazy real client

    def _firestore(self):
        # If a client was injected (e.g. in tests), return it without touching GCP.
        if self._injected_client is not None:
            return self._injected_client
        if self._client is None:
            from google.cloud import firestore  # lazy import — never loaded in tests
            kwargs: dict = {}
            if self._project:
                kwargs["project"] = self._project
            if self._database:
                kwargs["database"] = self._database
            self._client = firestore.Client(**kwargs)
        return self._client

    def get(self, client_id: Optional[str]) -> Optional[ClientContext]:
        """Load a :class:`ClientContext` from the Firestore client profile doc.

        Reads profile fields directly from the client doc (spec §1).
        ``category_mapping`` is a MAP FIELD on the client doc (not a subcollection).
        ``coa`` and ``entity_memory`` remain subcollections.
        ``ClientContext.tax_registered`` is the internal name for the profile's
        ``gst_registered`` field; falls back to legacy ``tax_registered`` key.
        """
        if not client_id:
            return None
        db = self._firestore()
        doc_ref = db.collection(self._collection).document(client_id)
        snap = doc_ref.get()
        if not snap.exists:
            return None
        data = snap.to_dict() or {}

        # tax_registered is the internal name; profile stores it as gst_registered.
        # Fall back to legacy tax_registered key; default True if absent.
        tax_registered = bool(
            data.get("gst_registered", data.get("tax_registered", True))
        )

        ctx = ClientContext(
            client_id=client_id,
            client_name=data.get("client_name"),
            client_uen=data.get("client_uen"),
            channel_id=data.get("channel_id"),
            slack_team_id=data.get("slack_team_id"),
            firm_id=data.get("firm_id"),
            fye_month=int(data["fye_month"]) if data.get("fye_month") is not None else None,
            region=data.get("region") or "SINGAPORE",
            accounting_software=data.get("accounting_software") or "QBS Ledger",
            base_currency=data.get("base_currency") or "SGD",
            tax_registered=tax_registered,
            status=data.get("status"),
            # category_mapping is a map field on the client doc (spec §1), not a subcollection.
            category_mapping=dict(data.get("category_mapping") or {}),
            # sys_config kept as empty dict for back-compat; profile fields are now explicit above.
            sys_config={},
        )

        # coa subcollection
        for d in (c.to_dict() or {} for c in doc_ref.collection("coa").stream()):
            code = d.get("code")
            description = d.get("description") or ""
            ctx.coa.append(CoaAccount(
                code=code,
                description=description,
                account_type=d.get("account_type"),
                financial_statement=d.get("financial_statement"),
                nature=d.get("nature"),
                keywords=d.get("keywords"),
            ))

        # entity_memory subcollection
        for d in (c.to_dict() or {} for c in doc_ref.collection("entity_memory").stream()):
            name = d.get("name")
            if not name:
                continue
            ctx.entity_memory.append(EntityMemoryEntry(
                name=name,
                reg_no=d.get("reg_no"),
                mapping_code=d.get("mapping_code"),
                role=d.get("role"),
                tax_code=d.get("tax_code"),
            ))

        return ctx

    def get_by_channel(self, channel_id: Optional[str]) -> Optional[ClientContext]:
        """Resolve a client via the ``channels/{channel_id}`` reverse-index doc.

        Returns ``None`` if the channel doc is missing or has no ``client_id``.
        """
        if not channel_id:
            return None
        db = self._firestore()
        snap = db.collection("channels").document(channel_id).get()
        if not snap.exists:
            return None
        data = snap.to_dict() or {}
        client_id = data.get("client_id")
        if not client_id:
            return None
        return self.get(client_id)

    # ---- write methods (ProfileStore protocol) ----

    def save_profile(self, profile: dict) -> None:
        """Write/merge the profile dict into ``clients/{client_id}``."""
        db = self._firestore()
        db.collection(self._collection).document(profile["client_id"]).set(profile, merge=True)

    def set_channel(self, channel_id: str, client_id: str) -> None:
        """Write the reverse-index ``channels/{channel_id}`` doc."""
        db = self._firestore()
        db.collection("channels").document(channel_id).set({"client_id": client_id})

    def save_coa(self, client_id: str, coa_rows: list[dict]) -> None:
        """REPLACE the COA: delete every existing ``coa/{n}`` doc, then write the
        new rows by index.

        Writing by index without first clearing orphans higher-index docs when a
        smaller COA replaces a larger one (e.g. a 5-row COA replaced by 2 rows
        would leave docs 2,3,4 behind). Streaming + deleting first guarantees the
        subcollection holds exactly the new rows.
        """
        db = self._firestore()
        coa_col = db.collection(self._collection).document(client_id).collection("coa")
        # Delete existing docs first (REPLACE, not append/orphan).
        for snap in coa_col.stream():
            snap.reference.delete()
        for i, row in enumerate(coa_rows):
            coa_col.document(str(i)).set(row)

    def set_status(self, client_id: str, status: str) -> None:
        """Merge-update the status field on ``clients/{client_id}``."""
        db = self._firestore()
        db.collection(self._collection).document(client_id).set({"status": status}, merge=True)

    def add_correction(self, *, client_id: str, vendor: str,
                       account_code: Optional[str] = None,
                       tax_code: Optional[str] = None) -> None:
        """Upsert a per-client vendor -> {account_code, tax_code} Correction.

        ADR-0004: when a human edits an extracted invoice's account_code or
        tax_code, that mapping is persisted as a Correction under
        ``clients/{client_id}/entity_memory/{vendor}`` so the next document
        from the same vendor auto-applies it. ``entity_memory`` is the
        spec-defined learning collection (already wired into categorizer
        ``vendor_name`` resolution).
        """
        if not (client_id and vendor):
            return
        db = self._firestore()
        ref = (
            db.collection(self._collection)
            .document(client_id)
            .collection("entity_memory")
            .document(vendor)
        )
        patch: dict[str, str] = {"name": vendor}
        if account_code:
            patch["mapping_code"] = account_code
        if tax_code:
            patch["tax_code"] = tax_code
        ref.set(patch, merge=True)
