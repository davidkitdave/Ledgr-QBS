"""Accounting-ledger exporters — emit the client's `<Client> - Ledger_FY<year>.xlsx`
workbook (sheets: Sys_Config, Purchase, Sales), matching the native formats observed in
the client data.

Two target formats, chosen by the client's `ACCOUNTING_SOFTWARE` setting:
- **QBS Ledger** (`QbsLedgerExporter`): native QBS columns; no tax-code column (SR/ZR is
  implied by the Tax Amount being 9% vs 0).
- **Xero Ledger** (`XeroLedgerExporter`): Xero import columns + `Source File ID` / `[AI Status]`,
  carrying the explicit `*TaxType` string.

Each invoice line becomes one row, so a mixed SR/ZR telco bill yields two rows. The tax
treatment + amount come from the shared `TaxClassifier`; account codes come from the COA
categorization layer (later phase). Adding a target = one subclass with its column lists +
row mapper.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import yaml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .categorizer import UNMAPPED_ACCOUNT_CODE
from .code_resolver import (
    resolve_creditor_code,
    resolve_rate_for_line,
    resolve_tax_code,
)
from .client_context import EntityMemoryEntry
from .models import BankStatement, InvoiceLine, NormalizedInvoice
from .tax_classifier import TaxClassifier, classify_invoice

_ERP_PROFILES_DIR = Path(__file__).resolve().parent.parent / "shared_libraries" / "erp_profiles"


def _fmt_date(d) -> str:
    if d is None:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")
    return str(d)


def _num(x: Optional[float]) -> Optional[float]:
    return None if x is None else round(float(x), 2)


@dataclass(frozen=True)
class ExportAccountCodeValidation:
    """Result of zero-tolerance COA validation at the export boundary."""

    account_code: str
    flagged: bool
    reason: str | None = None


def validate_export_account_code(
    account_code: str | None,
    *,
    coa_keys: set[str],
) -> ExportAccountCodeValidation:
    """Blank and flag account codes that are not in the client's COA key set.

    ``UNMAPPED`` is treated as abstention (blank + flagged). Empty input is
    left blank without a not-in-COA flag. This is the last-line defense before
    rows are written to the workbook — enum-constraint and categorizer
    post-validation may already have nulled bad codes upstream.
    """
    code = (account_code or "").strip()
    if not code:
        return ExportAccountCodeValidation(account_code="", flagged=False)
    if code == UNMAPPED_ACCOUNT_CODE or code not in coa_keys:
        return ExportAccountCodeValidation(
            account_code="",
            flagged=True,
            reason=f"account_code_not_in_coa: {code}",
        )
    return ExportAccountCodeValidation(account_code=code, flagged=False)


def _doc_sign(inv: NormalizedInvoice) -> int:
    """Return -1 for credit notes (amounts must be negative on export), +1 for all others.

    Keyed on ``inv.document_kind`` (the classify doc type: "credit_note", "invoice",
    "receipt", ...) — NOT ``inv.doc_type``, which is the direction ("purchase"/"sales").
    """
    return -1 if (inv.document_kind or "").strip().lower() == "credit_note" else 1


def _line_net_amount(line: InvoiceLine, inv: NormalizedInvoice) -> float:
    """Exportable line subtotal (sign-adjusted for credit notes).

    GST-registered clients: ex-GST net. Non-registered clients absorb irrecoverable
    input GST into the line cost (no separate tax column on export).
    Credit notes: result is negated so amounts reduce the books rather than add to them.
    """
    net = float(line.net_amount or 0.0)
    if not inv.our_gst_registered:
        raw = round(net + float(line.gst_amount or 0.0), 2)
    else:
        raw = round(net, 2)
    return round(raw * _doc_sign(inv), 2)


def _tax_amount(line: InvoiceLine, inv: NormalizedInvoice, clf: TaxClassifier) -> float:
    if not inv.our_gst_registered:
        return 0.0
    if inv.tax_visible_on_document is False:
        return 0.0
    if line.tax_treatment != "SR":
        return 0.0
    if line.gst_amount:
        return round(float(line.gst_amount) * _doc_sign(inv), 2)
    return 0.0


def _invoice_total(inv: NormalizedInvoice, clf: TaxClassifier) -> float:
    """Invoice-level grand total (sign-adjusted for credit notes).

    Prefer the authoritative doc total carried from extraction; otherwise fall back
    to Σ line net + Σ line tax. Credit notes: result is negated.
    """
    sign = _doc_sign(inv)
    if inv.doc_total is not None:
        return round(float(inv.doc_total) * sign, 2)
    net = sum((line.net_amount or 0.0) for line in inv.lines)
    tax = sum(_tax_amount(line, inv, clf) for line in inv.lines)
    # tax already carries the sign via _tax_amount; net needs sign applied here
    # (we sum raw net_amount values, so multiply by sign separately)
    return round(net * sign + tax, 2)


class LedgerExporter:
    """Base: write a Ledger_FY workbook (Sys_Config + Purchase + Sales)."""

    system: str = ""           # key into sg_gst.yaml code_map
    software_name: str = ""    # Sys_Config SOFTWARE value
    purchase_cols: list[str] = []
    sales_cols: list[str] = []

    # Per-class column → logical-field map. Subclasses can override to declare
    # which logical fields (e.g. "sub_total", "currency", "account_code") are
    # exposed under which column name. This is the source-of-truth that
    # ``column_for_field`` (and any preview/note/import-readiness code that
    # wants to look up the real column for a logical field) reads from.
    #
    # ProfileLedgerExporter ignores this attribute — its YAML `purchase_fields`
    # / `sales_fields` map is authoritative (see ProfileLedgerExporter.column_for_field).
    _LOGICAL_FIELDS: dict[str, str] = {}

    def __init__(self, classifier: Optional[TaxClassifier] = None):
        self.clf = classifier or TaxClassifier()
        self._coa_keys: set[str] | None = None

    def configure_client_context(
        self,
        *,
        tax_codes: list[dict] | dict[str, str] | None = None,
        entity_memory: list[EntityMemoryEntry] | None = None,
        coa_keys: set[str] | None = None,
    ) -> None:
        """Attach client COA keys for zero-tolerance export validation."""
        self._coa_keys = coa_keys

    def _sanitize_row_account_code(self, row: dict, doc_type: str) -> dict:
        if not self._coa_keys:
            return row
        col = self.column_for_field("account_code", doc_type)
        if not col:
            return row
        validated = validate_export_account_code(row.get(col, ""), coa_keys=self._coa_keys)
        row[col] = validated.account_code
        return row

    def required_fields(self, doc_type: str) -> list[str]:
        """Column names that must be non-empty in every exported row for this
        software. Subclasses override; used by the pipeline to flag (not drop)
        documents that would export half-filled rows."""
        return []

    def column_for_field(self, field_name: str, doc_type: str) -> str | None:
        """Return the actual column name for a logical *field_name* in *doc_type*.

        Used by delivery notes, import-readiness, and any preview surface that
        needs to look up the real column a logical field is written to (e.g. to
        render a "reconciles to $X" total, we need the real column for
        ``sub_total`` — which is "Amount" for AutoCount, "_AMOUNT" for SQL
        Account, "Sub Total" for QBS purchase, "Amount" for QBS sales, and None
        for Xero because Xero stores per-unit amount, not per-line net).

        Returns ``None`` when the field is not emitted for this doc_type (e.g.
        ``currency`` on a profile-driven purchase sheet that has no currency
        column). Callers must handle ``None`` (typically: skip the field rather
        than guess a literal column name).
        """
        cols = self.sales_cols if doc_type == "sales" else self.purchase_cols
        for col in cols:
            if self._LOGICAL_FIELDS.get(col) == field_name:
                return col
        return None

    # subclasses implement the per-line row dicts
    def _purchase_row(self, inv: NormalizedInvoice, line: InvoiceLine) -> dict:
        raise NotImplementedError

    def _sales_row(self, inv: NormalizedInvoice, line: InvoiceLine) -> dict:
        raise NotImplementedError

    def _ensure_classified(self, invoices: list[NormalizedInvoice]) -> None:
        for inv in invoices:
            if any(ln.tax_treatment is None for ln in inv.lines):
                classify_invoice(inv, self.clf)

    def rows(self, invoices: list[NormalizedInvoice], doc_type: str) -> list[dict]:
        self._ensure_classified(invoices)
        builder = self._sales_row if doc_type == "sales" else self._purchase_row
        out = []
        for inv in invoices:
            for line in inv.lines:
                out.append(self._sanitize_row_account_code(builder(inv, line), doc_type))
        return out

    def write_workbook(
        self,
        path: str | Path,
        purchases: Optional[list[NormalizedInvoice]] = None,
        sales: Optional[list[NormalizedInvoice]] = None,
    ) -> Path:
        """Write a Ledger workbook with just Purchase + Sales sheets (no Sys_Config)."""
        purchases, sales = purchases or [], sales or []
        wb = Workbook()
        for i, (title, cols, invs, doc) in enumerate([
            ("Purchase", self.purchase_cols, purchases, "purchase"),
            ("Sales", self.sales_cols, sales, "sales"),
        ]):
            sheet = wb.active if i == 0 else wb.create_sheet(title)
            sheet.title = title
            sheet.append(cols)
            for row in self.rows(invs, doc):
                sheet.append([row.get(c, "") for c in cols])
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path


class QbsLedgerExporter(LedgerExporter):
    """Native QBS Ledger format (no tax-code column; Tax Amount carries SR 9% / ZR 0)."""

    system = "qbs"
    software_name = "QBS Ledger"
    purchase_cols = [
        "Invoice Number", "Invoice Date", "Vendor Name", "Entity Tax ID", "Description",
        "Source Amount", "Currency", "Currency Rate", "Sub Total", "Tax Amount",
        "Total Amount", "Account Code / COA",
    ]
    sales_cols = [
        "Invoice Date", "Invoice Number", "Customer Name", "Description", "Source Amount",
        "Currency", "Currency Rate", "Amount", "Tax Amount", "Total", "Account Code / COA",
    ]
    # Logical-field map: column name → logical name. Used by column_for_field to
    # find the real column a logical field is written to (e.g. compose_confident_note
    # needs to render "reconciles to $X" using whichever column carries the line
    # net for the current doc_type). "Amount" on the sales sheet is the line net
    # (not a per-unit amount) so it maps to sub_total, same as "Sub Total" on
    # the purchase sheet.
    _LOGICAL_FIELDS = {
        "Invoice Number": "invoice_number",
        "Invoice Date": "invoice_date",
        "Vendor Name": "vendor_name",
        "Customer Name": "customer_name",
        "Entity Tax ID": "entity_tax_id",
        "Description": "description",
        "Source Amount": "source_amount",
        "Currency": "currency",
        "Currency Rate": "currency_rate",
        "Sub Total": "sub_total",
        "Amount": "sub_total",  # QBS sales: "Amount" carries the line net
        "Tax Amount": "tax_amount",
        "Total Amount": "total",
        "Total": "total",
        "Account Code / COA": "account_code",
    }

    def required_fields(self, doc_type: str) -> list[str]:
        if doc_type == "sales":
            return [
                "Invoice Number", "Invoice Date", "Customer Name", "Amount", "Total",
                "Account Code / COA",
            ]
        return [
            "Invoice Number", "Invoice Date", "Vendor Name", "Sub Total", "Total Amount",
            "Account Code / COA",
        ]

    def _purchase_row(self, inv, line):
        net = _line_net_amount(line, inv)
        tax = _tax_amount(line, inv, self.clf)
        # Currency Rate: the printed rate exactly as extracted, or blank when
        # the document prints none.  Never silently 1.0.
        fx = inv.fx_rate if inv.fx_rate is not None else ""
        return {
            "Invoice Number": inv.invoice_number or "",
            "Invoice Date": _fmt_date(inv.invoice_date),
            "Vendor Name": inv.supplier.name or "",
            "Entity Tax ID": inv.supplier.gst_regno or "",
            "Description": line.description,
            "Source Amount": net,
            "Currency": inv.currency,
            "Currency Rate": fx,
            "Sub Total": net,
            "Tax Amount": tax,
            "Total Amount": round(net + tax, 2),
            "Account Code / COA": line.account_code or "",
        }

    def _sales_row(self, inv, line):
        net = _line_net_amount(line, inv)
        tax = _tax_amount(line, inv, self.clf)
        # Currency Rate: the printed rate exactly as extracted, or blank when
        # the document prints none.  Never silently 1.0.
        fx = inv.fx_rate if inv.fx_rate is not None else ""
        return {
            "Invoice Date": _fmt_date(inv.invoice_date),
            "Invoice Number": inv.invoice_number or "",
            "Customer Name": inv.customer.name or "",
            "Description": line.description,
            "Source Amount": net,
            "Currency": inv.currency,
            "Currency Rate": fx,
            "Amount": net,
            "Tax Amount": tax,
            "Total": round(net + tax, 2),
            "Account Code / COA": line.account_code or "",
        }


class XeroLedgerExporter(LedgerExporter):
    """Xero Ledger format: Xero import columns + Source File ID / [AI Status], explicit *TaxType."""

    system = "xero"
    software_name = "Xero Ledger"
    _XERO_PURCHASE = [
        "*ContactName", "EmailAddress", "POAddressLine1", "POAddressLine2", "POAddressLine3",
        "POAddressLine4", "POCity", "PORegion", "POPostalCode", "POCountry", "*InvoiceNumber",
        "*InvoiceDate", "*DueDate", "Total", "InventoryItemCode", "Description", "*Quantity",
        "*UnitAmount", "*AccountCode", "*TaxType", "TaxAmount", "TrackingName1", "TrackingOption1",
        "TrackingName2", "TrackingOption2", "Currency",
    ]
    _XERO_SALES = [
        "*ContactName", "EmailAddress", "POAddressLine1", "POAddressLine2", "POAddressLine3",
        "POAddressLine4", "POCity", "PORegion", "POPostalCode", "POCountry", "*InvoiceNumber",
        "Reference", "*InvoiceDate", "*DueDate", "Total", "InventoryItemCode", "*Description",
        "*Quantity", "*UnitAmount", "Discount", "*AccountCode", "*TaxType", "TaxAmount",
        "TrackingName1", "TrackingOption1", "TrackingName2", "TrackingOption2", "Currency",
        "BrandingTheme",
    ]
    purchase_cols = list(_XERO_PURCHASE)
    sales_cols = list(_XERO_SALES)
    # Logical-field map for column_for_field. Note: Xero's "*UnitAmount" is a
    # PER-UNIT amount, not a per-line net — so it maps to "unit_amount", not
    # "sub_total". Anything that wants a "reconciles to $X" total must compute
    # it (qty × *UnitAmount) or accept that Xero has no per-line sub_total.
    _LOGICAL_FIELDS = {
        "*ContactName": "contact_name",
        "*InvoiceNumber": "invoice_number",
        "Reference": "reference",
        "*InvoiceDate": "invoice_date",
        "*DueDate": "due_date",
        "Total": "total",
        "InventoryItemCode": "inventory_item_code",
        "Description": "description",
        "*Description": "description",
        "*Quantity": "quantity",
        "*UnitAmount": "unit_amount",
        "Discount": "discount",
        "*AccountCode": "account_code",
        "*TaxType": "tax_code",
        "TaxAmount": "tax_amount",
        "Currency": "currency",
    }

    def required_fields(self, doc_type: str) -> list[str]:
        """Xero-required columns = the `*`-marked headers."""
        cols = self.sales_cols if doc_type == "sales" else self.purchase_cols
        return [c for c in cols if c.startswith("*")]

    def _xero_common(self, inv, line):
        party = inv.counterparty
        tax_type = self.clf.tax_code(line.tax_treatment, inv.doc_type, "xero")
        qty = line.quantity if line.quantity is not None else 1
        # *UnitAmount must satisfy Quantity × UnitAmount = the line's post-discount net,
        # so the invoice ties out on Xero import. Prefer the effective amount derived from
        # net_amount; the raw unit_amount is the pre-discount sticker price and over-states
        # discounted lines. Fall back to unit_amount only when net_amount is absent.
        line_net = _line_net_amount(line, inv)
        if line.net_amount is not None or (not inv.our_gst_registered and line.gst_amount):
            unit = line_net / qty if qty else line_net
        else:
            # Fallback: no net_amount; use the raw unit_amount but still apply the
            # credit-note sign so *UnitAmount is negative on a credit note.
            unit = (line.unit_amount * _doc_sign(inv)) if line.unit_amount is not None else line.unit_amount
        return {
            "*ContactName": party.name or "",
            "POCountry": party.country or "",
            "*InvoiceNumber": inv.invoice_number or "",
            "*InvoiceDate": _fmt_date(inv.invoice_date),
            "*DueDate": _fmt_date(inv.due_date or inv.invoice_date),
            "Total": _invoice_total(inv, self.clf),
            "*Quantity": _num(qty),
            "*UnitAmount": _num(unit),
            "*AccountCode": line.account_code or "",
            "*TaxType": tax_type,
            "TaxAmount": _tax_amount(line, inv, self.clf),
            "Currency": inv.currency,
        }

    def _purchase_row(self, inv, line):
        row = self._xero_common(inv, line)
        row["Description"] = line.description
        return row

    def _sales_row(self, inv, line):
        row = self._xero_common(inv, line)
        row["*Description"] = line.description
        return row


def _load_erp_profile(profile_name: str) -> dict[str, Any]:
    path = _ERP_PROFILES_DIR / profile_name
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


class ProfileLedgerExporter(LedgerExporter):
    """Generic exporter driven by a declarative ERP profile YAML.

    Profile schema extensions supported here:
    - ``purchase_sheet`` / ``sales_sheet``: worksheet names (default "Purchase"/"Sales").
    - ``purchase_constants`` / ``sales_constants``: column→value pairs applied after
      field mapping so fixed ERP values (DocNo=<<New>>, JournalType, _QTY, _UOM …)
      are always present regardless of the invoice data.
    - ``_row_context`` now carries ``supplier_invoice_no``, ``unit_price``, ``qty``,
      ``uom``, ``source_file_id``, ``ai_status``, ``ai_note`` for ERP columns that
      need them.
    """

    def __init__(
        self,
        profile: dict[str, Any],
        classifier: Optional[TaxClassifier] = None,
    ):
        super().__init__(classifier)
        self._profile = profile
        self.system = profile["system"]
        self.software_name = profile["software_name"]
        self.purchase_cols = list(profile["purchase_cols"])
        self.sales_cols = list(profile["sales_cols"])
        self._purchase_fields = dict(profile.get("purchase_fields") or {})
        self._sales_fields = dict(profile.get("sales_fields") or {})
        self._purchase_constants: dict[str, Any] = dict(profile.get("purchase_constants") or {})
        self._sales_constants: dict[str, Any] = dict(profile.get("sales_constants") or {})
        self._purchase_sheet: str = profile.get("purchase_sheet") or "Purchase"
        self._sales_sheet: str = profile.get("sales_sheet") or "Sales"
        self._client_tax_codes: list[dict] | dict[str, str] | None = None
        self._entity_memory: list[EntityMemoryEntry] = []

    def configure_client_context(
        self,
        *,
        tax_codes: list[dict] | dict[str, str] | None = None,
        entity_memory: list[EntityMemoryEntry] | None = None,
        coa_keys: set[str] | None = None,
    ) -> None:
        super().configure_client_context(
            tax_codes=tax_codes,
            entity_memory=entity_memory,
            coa_keys=coa_keys,
        )
        self._client_tax_codes = tax_codes
        self._entity_memory = list(entity_memory or [])

    def required_fields(self, doc_type: str) -> list[str]:
        key = "required_sales" if doc_type == "sales" else "required_purchase"
        return list(self._profile.get(key) or [])

    def _field_map(self, doc_type: str) -> dict[str, str]:
        return self._sales_fields if doc_type == "sales" else self._purchase_fields

    def _constants(self, doc_type: str) -> dict[str, Any]:
        return self._sales_constants if doc_type == "sales" else self._purchase_constants

    def _row_context(
        self,
        inv: NormalizedInvoice,
        line: InvoiceLine,
        doc_type: str,
    ) -> dict[str, Any]:
        rate = resolve_rate_for_line(self.clf, line, inv)
        tax_code = resolve_tax_code(
            line.tax_treatment,
            rate=rate,
            doc_type=doc_type,
            software=self.system,
            client_tax_codes=self._client_tax_codes,
            classifier=self.clf,
        )
        party = inv.counterparty
        creditor_code = ""
        debtor_code = ""
        if doc_type == "purchase":
            creditor_code = (
                party.vendor_code
                or resolve_creditor_code(party.name, party.gst_regno, self._entity_memory)
            )
        else:
            debtor_code = (
                party.vendor_code
                or resolve_creditor_code(party.name, party.gst_regno, self._entity_memory)
            )
        net = _line_net_amount(line, inv)
        tax = _tax_amount(line, inv, self.clf)
        fx = inv.fx_rate if inv.fx_rate is not None else ""
        qty = line.quantity if line.quantity is not None else 1
        unit_price = round(net / qty, 2) if qty else net
        # Provenance fields: populated when available on the invoice metadata
        source_file_id: str = getattr(inv, "source_file_id", None) or ""
        ai_status: str = getattr(inv, "ai_status", None) or ""
        ai_note: str = getattr(inv, "ai_note", None) or ""
        return {
            "invoice_number": inv.invoice_number or "",
            "invoice_date": _fmt_date(inv.invoice_date),
            "due_date": _fmt_date(inv.due_date or inv.invoice_date),
            "vendor_name": inv.supplier.name or "",
            "customer_name": inv.customer.name or "",
            "entity_tax_id": party.gst_regno or "",
            "description": line.description,
            "sub_total": net,
            "tax_amount": tax,
            "total_amount": round(net + tax, 2),
            "account_code": line.account_code or "",
            "tax_code": tax_code,
            "creditor_code": creditor_code,
            "debtor_code": debtor_code,
            "currency": inv.currency,
            "currency_rate": fx,
            # New fields for real ERP column layouts
            "supplier_invoice_no": inv.invoice_number or "",
            "unit_price": unit_price,
            "qty": qty,
            "uom": "UNIT",
            "source_file_id": source_file_id,
            "ai_status": ai_status,
            "ai_note": ai_note,
        }

    def _profile_row(self, inv: NormalizedInvoice, line: InvoiceLine, doc_type: str) -> dict:
        context = self._row_context(inv, line, doc_type)
        field_map = self._field_map(doc_type)
        cols = self.sales_cols if doc_type == "sales" else self.purchase_cols
        row = {col: context.get(field_map.get(col, ""), "") for col in cols}
        # Apply per-doc-type constants last (override field-mapped values)
        for col, val in self._constants(doc_type).items():
            if col in row:
                row[col] = val
        return row

    def column_for_field(self, field_name: str, doc_type: str) -> str | None:
        """Return the actual ERP column name for a logical *field_name* in *doc_type*.

        Inverts the profile ``purchase_fields`` / ``sales_fields`` map (which is
        ``{column: context_key}``) to find the column whose context_key equals
        *field_name*.  Returns ``None`` when the field has no mapped column for
        this doc type (e.g. ``creditor_code`` on a sales sheet).
        """
        field_map = self._field_map(doc_type)
        for col, ctx_key in field_map.items():
            if ctx_key == field_name:
                return col
        return None

    def _purchase_row(self, inv, line):
        return self._profile_row(inv, line, "purchase")

    def _sales_row(self, inv, line):
        return self._profile_row(inv, line, "sales")

    def write_workbook(
        self,
        path: str | Path,
        purchases: Optional[list[NormalizedInvoice]] = None,
        sales: Optional[list[NormalizedInvoice]] = None,
    ) -> Path:
        """Write a workbook with per-doc-type sheet names from the profile.

        Overrides the base class to use ``_purchase_sheet`` / ``_sales_sheet``
        from the profile YAML instead of the hardcoded "Purchase"/"Sales" titles,
        so AutoCount gets "AP Invoice"/"AR Invoice" and SQL Account gets
        "SLPH_Invoice_Cash_Debit_Credit".
        """
        purchases, sales = purchases or [], sales or []
        wb = Workbook()
        for i, (title, cols, invs, doc) in enumerate([
            (self._purchase_sheet, self.purchase_cols, purchases, "purchase"),
            (self._sales_sheet, self.sales_cols, sales, "sales"),
        ]):
            sheet = wb.active if i == 0 else wb.create_sheet(title)
            sheet.title = title
            sheet.append(cols)
            for row in self.rows(invs, doc):
                sheet.append([row.get(c, "") for c in cols])
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path


class AutoCountExporter(ProfileLedgerExporter):
    def __init__(self, classifier: Optional[TaxClassifier] = None):
        super().__init__(_load_erp_profile("autocount.yaml"), classifier)


class SqlAccountExporter(ProfileLedgerExporter):
    def __init__(self, classifier: Optional[TaxClassifier] = None):
        super().__init__(_load_erp_profile("sql_account.yaml"), classifier)


EXPORTERS = {
    "qbs": QbsLedgerExporter,
    "xero": XeroLedgerExporter,
    "autocount": AutoCountExporter,
    "sql_account": SqlAccountExporter,
}

# Maps every known display/alias form (lowercased, stripped) to the canonical exporter key.
_SOFTWARE_ALIASES: dict[str, str] = {
    "qbs": "qbs",
    "qbs ledger": "qbs",
    "qbsledger": "qbs",
    "xero": "xero",
    "xero ledger": "xero",
    "xeroledger": "xero",
    "autocount": "autocount",
    "sql account": "sql_account",
    "sql_account": "sql_account",
    "sqlaccount": "sql_account",
}


def normalize_software_key(value: Optional[str]) -> Optional[str]:
    """Return the canonical exporter key for *value*, or None.

    Recognised keys: ``"qbs"``, ``"xero"``, ``"autocount"``, ``"sql_account"``.
    Accepts any casing and surrounding whitespace.  Callers decide the fallback
    when None is returned (genuinely unrecognised software name).
    """
    return _SOFTWARE_ALIASES.get((value or "").strip().lower())


def get_exporter(system: str, classifier: Optional[TaxClassifier] = None) -> LedgerExporter:
    key = normalize_software_key(system)
    if key is None:
        raise ValueError(f"unknown export system '{system}'; have {list(EXPORTERS)}")
    return EXPORTERS[key](classifier)


# Map exporter column headers to readable snake_case names for review notes.
_FIELD_LABELS = {
    "*ContactName": "contact_name", "*InvoiceNumber": "invoice_number",
    "*InvoiceDate": "invoice_date", "*DueDate": "due_date", "*Quantity": "quantity",
    "*UnitAmount": "unit_amount", "*AccountCode": "account_code", "*TaxType": "tax_type",
    "*Description": "description",
    "Vendor Name": "vendor_name", "Customer Name": "customer_name",
    "Invoice Number": "invoice_number", "Invoice Date": "invoice_date",
    "Sub Total": "sub_total", "Total Amount": "total", "Amount": "amount", "Total": "total",
    "Account Code / COA": "account_code",
    "Tax Code": "tax_code",
    "Creditor Code": "creditor_code",
    "Debtor Code": "debtor_code",
    "Acc No": "account_code",
    "Account Code": "account_code",
}


def _field_label(col: str) -> str:
    return _FIELD_LABELS.get(col, col.lstrip("*"))


def _is_empty(value) -> bool:
    """A required cell is missing when it is None or a blank/whitespace string."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def validate_required_fields(
    inv: NormalizedInvoice, exporter: LedgerExporter, doc_type: str
) -> list[str]:
    """Return readable labels for export-required fields that are missing/empty.

    Builds the rows exactly as they would be exported and checks the software's
    required columns. Invoice-level fields (empty on every line) are reported once;
    line-level fields are reported with a 1-based line number. Empty list = complete.
    """
    required = exporter.required_fields(doc_type)
    if not required:
        return []
    rows = exporter.rows([inv], doc_type)
    if not rows:
        # No lines were extracted — flag the doc so it surfaces for review rather
        # than being silently written as an empty shell or silently dropped.
        return ["no line items extracted"]
    n = len(rows)
    missing: list[str] = []
    for col in required:
        empty_lines = [i for i, row in enumerate(rows, start=1) if _is_empty(row.get(col, ""))]
        if not empty_lines:
            continue
        label = _field_label(col)
        if len(empty_lines) == n:
            # invoice-level (missing on every line) — report once, no line number
            missing.append(label)
        else:
            for i in empty_lines:
                missing.append(f"{label} (line {i})")
    return missing


def collect_export_unmapped_summary(
    batches: list[dict],
    exporter: LedgerExporter,
) -> dict:
    """Collect export rows with blank required tax/creditor columns."""
    details: list[dict] = []
    for batch in batches:
        sheet = batch.get("sheet") or "Purchase"
        row_doc_type = "sales" if sheet == "Sales" else "purchase"
        required = exporter.required_fields(row_doc_type)
        # Resolve the actual invoice_number column name for this exporter/doc_type.
        inv_col: str | None = None
        if isinstance(exporter, ProfileLedgerExporter):
            inv_col = exporter.column_for_field("invoice_number", row_doc_type)
        for idx, row in enumerate(batch.get("rows") or [], start=1):
            missing = [col for col in required if _is_empty(row.get(col, ""))]
            if missing:
                # Try profile-derived column first, then legacy fallbacks.
                invoice_number = (
                    (row.get(inv_col) if inv_col else None)
                    or row.get("Doc No")
                    or row.get("Invoice Number")
                    or row.get("*InvoiceNumber")
                    or row.get("DocNo")
                    or row.get("DOCNO(20)")
                    or ""
                )
                details.append(
                    {
                        "sheet": sheet,
                        "row": idx,
                        "missing": missing,
                        "invoice_number": invoice_number,
                    }
                )
    return {"count": len(details), "details": details}


def format_unmapped_export_note(summary: dict | None) -> str:
    """Human-readable note for approval / delivery cards."""
    if not summary:
        return ""
    count = int(summary.get("count") or 0)
    if count <= 0:
        return ""
    noun = "row" if count == 1 else "rows"
    return f"{count} {noun} need creditor or tax codes before ERP import"


def compute_doc_flag_breakdown(
    batches: list[dict],
    exporter: LedgerExporter,
) -> dict:
    """Per-document reconcile status + flag-reason counts (WS-1.4).

    Returns a dict:
        {
          "reconciles": bool,            # True iff no flags raised
          "n_total": int,                # total rows across all batches
          "reasons": {
            "blank_account": int,        # rows with empty account_code column
            "missing_tax": int,          # rows with empty tax_code column
            "missing_creditor": int,     # rows with empty creditor_code/debtor_code column
            "missing_invoice_number": int,
          }
        }

    A doc "reconciles" iff none of the three primary flag reasons fired
    (blank account / missing tax / missing creditor). TaxType is NOT in the
    formal ``required_*`` list for profile ERPs (it can be blank and the line
    still imports), but a blank tax code is a "tax-unresolved" signal the
    user wants surfaced — so we count it as a flag even though the row
    wouldn't fail the strict required-field check. A doc with all flags
    zero renders as ✓ reconciled on the delivery card.

    Used by the batch delivery card to surface per-doc visibility (✓/✗ + reason
    breakdown) — previously the unmapped count was rendered but the per-reason
    breakdown was discarded. Now a multi-file dropper sees exactly which rows
    need attention and why.
    """
    if not isinstance(exporter, ProfileLedgerExporter):
        # QBS / Xero: required-field check via the legacy literal columns.
        # (These exporters don't expose column_for_field for all required fields,
        # but the required_fields() list still drives the unmapped summary.)
        legacy = {"blank_account": 0, "missing_tax": 0, "missing_creditor": 0,
                  "missing_invoice_number": 0}
        n_total = 0
        for batch in batches:
            sheet = batch.get("sheet") or "Purchase"
            row_doc_type = "sales" if sheet == "Sales" else "purchase"
            required = exporter.required_fields(row_doc_type)
            for row in batch.get("rows") or []:
                n_total += 1
                for col in required:
                    if _is_empty(row.get(col, "")):
                        # Map legacy column names to per-reason buckets.
                        if col in ("Account Code / COA", "Account Code", "*AccountCode", "AccNo", "Acc No", "_ACCOUNT(10)"):
                            legacy["blank_account"] += 1
        return {
            "reconciles": sum(legacy.values()) == 0,
            "n_total": n_total,
            "reasons": legacy,  # not fully decomposed for legacy exporters
        }

    reasons = {"blank_account": 0, "missing_tax": 0, "missing_creditor": 0,
               "missing_invoice_number": 0}
    n_total = 0

    for batch in batches:
        sheet = batch.get("sheet") or "Purchase"
        row_doc_type = "sales" if sheet == "Sales" else "purchase"
        account_col = exporter.column_for_field("account_code", row_doc_type)
        tax_col = exporter.column_for_field("tax_code", row_doc_type)
        creditor_col = exporter.column_for_field("creditor_code", row_doc_type)
        debtor_col = exporter.column_for_field("debtor_code", row_doc_type)
        inv_col = exporter.column_for_field("invoice_number", row_doc_type)

        for row in batch.get("rows") or []:
            n_total += 1
            # Decompose per-reason counts. A single row can contribute to
            # multiple reasons (e.g. blank account + blank tax).
            if account_col and _is_empty(row.get(account_col, "")):
                reasons["blank_account"] += 1
            if tax_col and _is_empty(row.get(tax_col, "")):
                reasons["missing_tax"] += 1
            if (creditor_col and _is_empty(row.get(creditor_col, ""))) or \
               (debtor_col and _is_empty(row.get(debtor_col, ""))):
                reasons["missing_creditor"] += 1
            if inv_col and _is_empty(row.get(inv_col, "")):
                reasons["missing_invoice_number"] += 1

    # ✓ reconciled iff none of the three primary flag reasons fired.
    reconciles = (
        reasons["blank_account"] == 0
        and reasons["missing_tax"] == 0
        and reasons["missing_creditor"] == 0
    )
    return {
        "reconciles": reconciles,
        "n_total": n_total,
        "reasons": reasons,
    }


def format_flag_breakdown_note(breakdown: dict | None) -> str:
    """Human-readable flag-reason breakdown for a single doc (WS-1.4).

    Returns a short string like:
        "✓ reconciled"     (when reconciles=True)
        "✗ 2 blank accounts · 1 missing tax code"  (when reconciles=False)
    """
    if not breakdown:
        return ""
    if breakdown.get("reconciles"):
        return "✓ reconciled"
    reasons = breakdown.get("reasons") or {}
    parts: list[str] = []
    if reasons.get("blank_account"):
        n = reasons["blank_account"]
        parts.append(f"{n} blank account{'s' if n != 1 else ''}")
    if reasons.get("missing_tax"):
        n = reasons["missing_tax"]
        parts.append(f"{n} missing tax code{'s' if n != 1 else ''}")
    if reasons.get("missing_creditor"):
        n = reasons["missing_creditor"]
        parts.append(f"{n} missing creditor/debtor code{'s' if n != 1 else ''}")
    if reasons.get("missing_invoice_number"):
        n = reasons["missing_invoice_number"]
        parts.append(f"{n} missing invoice number{'s' if n != 1 else ''}")
    if not parts:
        return "✗ not reconciled"
    return "✗ " + " · ".join(parts)


def format_extraction_doc_count_note(doc_count: int, page_count: int) -> str:
    """G3 delivery-card line: extracted N documents from M pages (WS-2.4).

    Returns ``""`` when counts are missing or invalid so older sessions without
    extraction metadata stay silent instead of showing a broken line.
    """
    if doc_count < 1 or page_count < 1:
        return ""
    doc_word = "document" if doc_count == 1 else "documents"
    page_word = "page" if page_count == 1 else "pages"
    return f"Extracted {doc_count} {doc_word} from {page_count} {page_word}"


def collect_import_readiness(
    batches: list[dict],
    exporter: LedgerExporter,
    *,
    unmapped: dict | None = None,
) -> dict:
    """Collect distinct ERP codes referenced across all export rows.

    Only meaningful for :class:`ProfileLedgerExporter` (code-keyed ERPs like
    AutoCount and SQL Account).  For QBS / Xero exporters returns an empty dict.

    Returns a dict with keys:
      - ``software``: display name of the ERP (e.g. "AutoCount")
      - ``tax_codes``: sorted list of distinct non-empty tax code values
      - ``party_codes``: sorted list of distinct non-empty creditor/debtor codes
      - ``account_codes``: sorted list of distinct non-empty GL account codes
      - ``unmapped``: the result of ``collect_export_unmapped_summary`` (reused)
    """
    if not isinstance(exporter, ProfileLedgerExporter):
        return {}

    tax_codes: set[str] = set()
    party_codes: set[str] = set()
    account_codes: set[str] = set()

    for batch in batches:
        sheet = batch.get("sheet") or "Purchase"
        row_doc_type = "sales" if sheet == "Sales" else "purchase"

        tax_col = exporter.column_for_field("tax_code", row_doc_type)
        creditor_col = exporter.column_for_field("creditor_code", row_doc_type)
        debtor_col = exporter.column_for_field("debtor_code", row_doc_type)
        account_col = exporter.column_for_field("account_code", row_doc_type)

        for row in batch.get("rows") or []:
            if tax_col:
                v = (row.get(tax_col) or "").strip()
                if v:
                    tax_codes.add(v)
            if creditor_col:
                v = (row.get(creditor_col) or "").strip()
                if v:
                    party_codes.add(v)
            if debtor_col:
                v = (row.get(debtor_col) or "").strip()
                if v:
                    party_codes.add(v)
            if account_col:
                v = (row.get(account_col) or "").strip()
                if v:
                    account_codes.add(v)

    return {
        "software": exporter.software_name,
        "tax_codes": sorted(tax_codes),
        "party_codes": sorted(party_codes),
        "account_codes": sorted(account_codes),
        "unmapped": unmapped or {},
    }


def format_import_readiness_note(readiness: dict | None) -> str:
    """Concise human-readable checklist for the delivery card.

    Returns ``""`` when *readiness* is empty or not applicable (QBS/Xero).

    Example output:
        "AutoCount import — needs these codes in your company: tax SV-6, SV-8
         · creditors 400-G0001, 400-T0001 · accounts 610-0000. If your company
         uses different tax-code names, upload your tax-code list and we'll
         match it. ⚠️ 2 rows need a creditor code first (see above)."
    """
    if not readiness:
        return ""
    software = (readiness.get("software") or "").strip()
    if not software:
        return ""

    _MAX_CODES = 8

    def _fmt_list(codes: list[str]) -> str:
        if len(codes) <= _MAX_CODES:
            return ", ".join(codes)
        shown = ", ".join(codes[:_MAX_CODES])
        extra = len(codes) - _MAX_CODES
        return f"{shown} …+{extra} more"

    parts: list[str] = []
    tax_codes = readiness.get("tax_codes") or []
    party_codes = readiness.get("party_codes") or []
    account_codes = readiness.get("account_codes") or []

    if tax_codes:
        parts.append(f"tax {_fmt_list(tax_codes)}")
    if party_codes:
        parts.append(f"creditors/debtors {_fmt_list(party_codes)}")
    if account_codes:
        parts.append(f"accounts {_fmt_list(account_codes)}")

    if not parts:
        return ""

    codes_str = " · ".join(parts)
    note = (
        f"{software} import — needs these codes to exist in your company: "
        f"{codes_str}. "
        "If your company uses different tax-code names, upload your tax-code "
        "list and we'll match it."
    )
    unmapped = readiness.get("unmapped") or {}
    unmapped_count = int(unmapped.get("count") or 0)
    if unmapped_count > 0:
        noun = "row needs" if unmapped_count == 1 else "rows need"
        note = f"{note} ⚠️ {unmapped_count} {noun} a creditor code first (see above)."
    return note


_SHEET_INVALID = str.maketrans({c: None for c in "[]:*?/\\"})


def _sheet_title(name: str) -> str:
    return (name or "Bank").translate(_SHEET_INVALID).strip()[:31] or "Bank"


def _last4_digits(*sources: str | None) -> str:
    """Return the last 4 numeric digits found across the given strings (left-padded).

    Used to build a deterministic tab name like ``DBS - 5545 - CNH`` from
    a free-form ``bank_name`` + structured ``account_number``. We fall back
    across sources so an LLM that already packed the digits into ``bank_name``
    still produces a stable title.
    """
    digits = "".join(c for c in "".join(s or "" for s in sources) if c.isdigit())
    return digits[-4:].rjust(4, "0") if digits else "0000"


def _bank_label(bank_name: str) -> str:
    """Extract the bank-only label from a possibly-prefixed ``bank_name``.

    The LLM extraction prompt sometimes returns ``"OCBC - 5001"`` or
    ``"DBS Bank Ltd - 5545"``. We keep just the bank portion (text before the
    first ``-``) so the structured ``account_number`` + ``currency`` are the
    sole source of the last-4 + currency suffix appended by ``bank_sheet_title``.
    """
    raw = (bank_name or "").strip()
    if " - " in raw:
        return raw.split(" - ", 1)[0].strip()
    if "-" in raw and not any(c.isdigit() for c in raw.split("-", 1)[0]):
        return raw.split("-", 1)[0].strip()
    return raw or "Bank"


def bank_sheet_title(
    *,
    bank_name: str,
    account_number: str | None,
    currency: str,
) -> str:
    """Build ``"<Bank> - XXXX - CCY"`` for a bank statement's Excel tab.

    Deterministic (code-owned, not LLM-owned) so multi-currency statements of
    the same account split into distinct tabs instead of being silently merged
    by ``SlackLedgerStore._merge_bank_statement``. Result is sanitized via
    :func:`_sheet_title` (invalid Excel chars stripped, capped at 31 chars).
    """
    label = _bank_label(bank_name)
    last4 = _last4_digits(account_number, bank_name)
    ccy = (currency or "SGD").upper()
    return _sheet_title(f"{label} - {last4} - {ccy}")


def _parse_ddmmyyyy(value) -> Optional[datetime]:
    """Parse a ``DD/MM/YYYY`` date string into a ``datetime`` (None if unparseable).

    Robust to ``date``/``datetime`` inputs and a few separator variants; never
    raises — an unparseable value yields ``None`` so callers can keep stable order
    instead of crashing on a malformed row.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except (ValueError, TypeError):
            continue
    return None


class BankStatementExporter:
    """Write an accountant-grade, continuous bank-statement workbook.

    One sheet per account; every txn row preserved. Unlike a per-statement-block
    layout, each account sheet is ONE continuous chain across the whole FY:

    - Statements (months) are sorted ascending by date regardless of upload order,
      and transactions within a month are sorted ascending by date too.
    - The running ``Balance`` (col E) chains top-to-bottom across the ENTIRE sheet:
      every txn row = ``=<prev_balance_cell> + Deposit - Withdrawal``, crossing
      month boundaries (it does NOT re-seed each month).
    - Each month keeps one ``BALANCE B/F`` marker row as a visual separator. Its
      running ``Balance`` *carries forward* from the prior row (``=<prev_E>``),
      while ``Stated Balance`` (col F) holds that month's stated opening B/F. The
      FIRST B/F of the FY seeds from its own stated opening (``=F<row>``).
    - A cross-month continuity ``Check`` (col G) on each B/F row flags ``GAP`` when
      the carried-forward balance ≠ the stated B/F (a missing/duplicate month);
      txn rows keep the per-row ``CHECK`` against their own stated balance.
    - A single per-account ``TOTALS`` row at the bottom sums all withdrawals/deposits.

    The sheet is REBUILT from a normalized list of month-blocks on every append, so
    the result is identical no matter what order statements arrived in.
    """

    BANK_COLS = [
        "Date", "Description", "Withdrawal", "Deposit",
        "Balance", "Currency", "Math_Check",
    ]

    #: Description markers used to detect row roles.
    OPENING_MARKER = "BALANCE B/F"
    TOTALS_MARKER = "TOTALS"

    def bank_rows(self, stmt: BankStatement) -> list[dict]:
        """Build the data rows for one statement (values only; formulas added later).

        The opening row seeds ``Stated Balance`` from the opening balance; each txn
        row carries its withdrawal/deposit + the extracted ``Stated Balance``. A
        trailing ``TOTALS`` row closes the block. These value rows are the unit a
        caller hands to the store as a batch; the store normalizes + re-sorts them
        into the continuous chain (formula cells are filled on rebuild).
        """
        rows: list[dict] = []
        rows.append({
            "Description": self.OPENING_MARKER,
            "Balance": _num(stmt.opening_balance),
            "Currency": stmt.currency,
        })
        for txn in stmt.transactions:
            rows.append({
                "Date": _fmt_date(txn.date),
                "Description": txn.description,
                "Withdrawal": _num(txn.withdrawal),
                "Deposit": _num(txn.deposit),
                "Balance": _num(txn.balance),
                "Currency": stmt.currency,
            })
        rows.append({
            "Description": self.TOTALS_MARKER,
            "Currency": stmt.currency,
        })
        return rows

    # ------------------------------------------------------------------ #
    # Continuous-chain rebuild
    # ------------------------------------------------------------------ #

    @classmethod
    def rows_to_blocks(cls, rows: list[dict]) -> list[dict]:
        """Split a flat list of value-row dicts (a statement) into month-blocks.

        Each ``BALANCE B/F`` marker starts a new block carrying its stated opening
        balance; subsequent non-marker rows are that block's transactions. ``TOTALS``
        markers are dropped (the per-account total is regenerated on rebuild). A
        leading run of transactions with no preceding B/F forms a headless block
        (``stated_bf=None``) so nothing is lost.
        """
        blocks: list[dict] = []
        current: Optional[dict] = None
        for row in rows:
            desc = row.get("Description")
            if desc == cls.OPENING_MARKER:
                current = {
                    "stated_bf": row.get("Balance"),
                    "currency": row.get("Currency"),
                    "transactions": [],
                }
                blocks.append(current)
            elif desc == cls.TOTALS_MARKER:
                continue
            else:
                if current is None:
                    current = {"stated_bf": None, "currency": row.get("Currency"), "transactions": []}
                    blocks.append(current)
                current["transactions"].append(dict(row))
        return blocks

    @classmethod
    def _block_sort_key(cls, block: dict):
        """Sort key for a month-block: earliest parseable txn date.

        Blocks whose dates are all unparseable sort last but keep stable relative
        order (Python's sort is stable), so a malformed month never crashes the build.
        """
        dates = [
            d for d in (_parse_ddmmyyyy(t.get("Date")) for t in block["transactions"])
            if d is not None
        ]
        return (0, min(dates)) if dates else (1, datetime.max)

    @classmethod
    def _block_signature(cls, block: dict):
        """Identity of a statement block for dedup: opening balance + its txns.

        Two blocks with the same opening balance and the same ordered list of
        ``(Date, Description, Withdrawal, Deposit)`` transactions are the SAME
        statement re-uploaded (Balance is excluded — it is recomputed). Used to
        collapse statements that were appended more than once (e.g. during the
        doc_key format transition that duplicated months in the Sample Bank Client sheet).
        """
        def _norm(v):
            if v is None:
                return ""
            s = str(v).strip()
            if not s:
                return ""
            # Normalise numerics so 200 / 200.0 / "200.00" compare equal across
            # the fresh-block vs read-back-from-xlsx round trip.
            try:
                return f"{float(s):.2f}"
            except (TypeError, ValueError):
                return s

        txns = tuple(
            (_norm(t.get("Date")), _norm(t.get("Description")),
             _norm(t.get("Withdrawal")), _norm(t.get("Deposit")))
            for t in block.get("transactions", [])
        )
        return (_norm(block.get("stated_bf")), _norm(block.get("currency")), txns)

    @classmethod
    def dedupe_blocks(cls, blocks: list[dict]) -> list[dict]:
        """Drop blocks that are byte-for-byte duplicate statements (keep the first).

        Safety net for re-merges: an identical statement appended N times yields N
        identical blocks; this keeps one. Distinct months never collide (their txn
        dates differ), and duplicate transactions WITHIN one statement are kept
        (they live in the same block). Preserves input order.
        """
        seen: set = set()
        unique: list[dict] = []
        for block in blocks:
            sig = cls._block_signature(block)
            # A truly empty block (no bf, no txns) carries no data — keep at most one.
            if sig in seen:
                continue
            seen.add(sig)
            unique.append(block)
        return unique

    @classmethod
    def sort_blocks(cls, blocks: list[dict]) -> list[dict]:
        """Return month-blocks sorted ascending by date, txns sorted within each.

        Months are ordered by their earliest transaction date (upload-order
        independent); within a month, transactions are ordered ascending by date.
        Unparseable dates keep their stable relative position (never crash).
        """
        ordered = sorted(blocks, key=cls._block_sort_key)
        for block in ordered:
            block["transactions"].sort(
                key=lambda t: (0, _parse_ddmmyyyy(t.get("Date")))
                if _parse_ddmmyyyy(t.get("Date")) is not None
                else (1, datetime.max)
            )
        return ordered

    @classmethod
    def rebuild_account_sheet(
        cls,
        sheet,
        blocks: list[dict],
        cols: list[str],
        *,
        key_col: Optional[str] = None,
    ) -> None:
        """Wipe a sheet's body and rewrite it as one continuous chain of month-blocks.

        ``blocks`` is the normalized, already-sorted list from :meth:`sort_blocks`
        (each ``{"stated_bf", "currency", "transactions": [row_dict]}``). For each
        block we emit a ``BALANCE B/F`` row then its txn rows; a single ``TOTALS``
        row closes the whole account. The header row (row 1) is preserved; an
        optional ``key_col`` (the hidden dedupe column) is carried per-row from each
        row dict's ``key_col`` value so dedupe survives the rebuild.
        """
        # Clear everything below the header.
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)

        currency = ""
        for block in blocks:
            if block.get("currency"):
                currency = block["currency"]
                break

        def emit(row: dict) -> None:
            values = [row.get(c, "") for c in cols]
            if key_col is not None:
                values.append(row.get(key_col, ""))
            sheet.append(values)

        for block in blocks:
            emit({
                "Description": cls.OPENING_MARKER,
                "Balance": block.get("stated_bf"),
                "Currency": block.get("currency") or currency,
                key_col: block.get("bf_key", "") if key_col else "",
            })
            for txn in block["transactions"]:
                emit(txn)

        # Single per-account TOTALS row at the bottom.
        emit({"Description": cls.TOTALS_MARKER, "Currency": currency})

        cls.apply_bank_formulas(sheet, cols)

    @classmethod
    def apply_bank_formulas(cls, sheet, cols: Optional[list[str]] = None) -> None:
        """Write Math_Check formulas and TOTALS SUM over the sheet.

        Balance holds the actual bank-stated value on every row — no formula.
        Math_Check (col G) validates each row arithmetically. EVERY cell
        reference is wrapped in ``N()`` so a non-numeric cell (a stray currency
        code, a label, a blank) coerces to 0 instead of poisoning the whole
        formula with ``#VALUE!`` — the robustness pattern from the reference
        workbook. (See QA 2026-06-14: the un-coerced formula produced ``#VALUE!``
        whenever a 'SGD' string leaked into the Balance column.)

        - ``BALANCE B/F`` row: first B/F of the FY gets ``✅`` (no prior row);
          later months get ``=IF(ROUND(N(E_bf)-N(E_prev),2)=0,"✅","GAP")`` — a
          ``GAP`` means the stated opening doesn't match the prior closing balance.
        - txn row: ``=IF(ROUND(N(E)-(N(E_prev)+N(Deposit)-N(Withdrawal)),2)=0,
          "✅","❌ Exp: "&<expected>)`` — the expected running balance is shown on
          a mismatch so the bookkeeper sees the discrepancy at a glance.
        - ``TOTALS`` row: ``=SUM(...)`` over the Withdrawal/Deposit txn range.
        """
        header = [c.value for c in sheet[1]] if sheet.max_row >= 1 else []
        if not header:
            return
        idx = {name: i for i, name in enumerate(header)}
        if "Balance" not in idx or "Description" not in idx:
            return

        def col(name: str) -> Optional[str]:
            i = idx.get(name)
            return get_column_letter(i + 1) if i is not None else None

        c_desc = idx["Description"]
        bal_col = col("Balance")
        check_col = col("Math_Check")
        wd_col = col("Withdrawal")
        dep_col = col("Deposit")

        prev_balance_row: Optional[int] = None
        seen_first_bf = False
        first_txn_row: Optional[int] = None
        last_txn_row: Optional[int] = None
        totals_row: Optional[int] = None

        for r in range(2, sheet.max_row + 1):
            desc = sheet.cell(row=r, column=c_desc + 1).value
            if desc == cls.OPENING_MARKER:
                if check_col and bal_col:
                    if not seen_first_bf or prev_balance_row is None:
                        # First B/F of the FY — no prior row to compare against.
                        sheet[f"{check_col}{r}"] = "✅"
                    else:
                        sheet[f"{check_col}{r}"] = (
                            f'=IF(ROUND(N({bal_col}{r})-N({bal_col}{prev_balance_row}),2)=0,"✅","GAP")'
                        )
                prev_balance_row = r
                seen_first_bf = True
            elif desc == cls.TOTALS_MARKER:
                totals_row = r
            else:
                if first_txn_row is None:
                    first_txn_row = r
                last_txn_row = r
                if check_col and bal_col and prev_balance_row is not None:
                    # Expected running balance = prev balance + deposit − withdrawal,
                    # every term N()-coerced so text cells become 0 (no #VALUE!).
                    expected = f"N({bal_col}{prev_balance_row})"
                    if dep_col:
                        expected += f"+N({dep_col}{r})"
                    if wd_col:
                        expected += f"-N({wd_col}{r})"
                    sheet[f"{check_col}{r}"] = (
                        f'=IF(ROUND(N({bal_col}{r})-({expected}),2)=0,"✅",'
                        f'"❌ Exp: "&ROUND({expected},2))'
                    )
                prev_balance_row = r

        # TOTALS row: SUM over all txn rows for Withdrawal and Deposit.
        if totals_row is not None and first_txn_row is not None and last_txn_row is not None:
            if wd_col:
                sheet[f"{wd_col}{totals_row}"] = f"=SUM({wd_col}{first_txn_row}:{wd_col}{last_txn_row})"
            if dep_col:
                sheet[f"{dep_col}{totals_row}"] = f"=SUM({dep_col}{first_txn_row}:{dep_col}{last_txn_row})"

    def write_bank_workbook(self, path: str | Path, statements: list[BankStatement]) -> Path:
        """Write a workbook with one continuous-chain sheet per account.

        Statements sharing a bank/account sheet are merged into a single continuous
        ledger (sorted by date, one chain), so a year of monthly statements for an
        account lands as one sorted, cross-month-reconciling sheet. Multi-currency
        statements of the same account split into distinct sheets (one per currency).
        """
        wb = Workbook()
        # Group statements by their sheet title, preserving first-seen order.
        grouped: dict[str, list[BankStatement]] = {}
        for stmt in statements:
            grouped.setdefault(
                bank_sheet_title(
                    bank_name=stmt.bank_name,
                    account_number=stmt.account_number,
                    currency=stmt.currency or "SGD",
                ),
                [],
            ).append(stmt)

        for i, (title, stmts) in enumerate(grouped.items()):
            sheet = wb.active if i == 0 else wb.create_sheet()
            sheet.title = title
            sheet.append(self.BANK_COLS)
            blocks: list[dict] = []
            for stmt in stmts:
                blocks.extend(self.rows_to_blocks(self.bank_rows(stmt)))
            self.rebuild_account_sheet(sheet, self.sort_blocks(blocks), self.BANK_COLS)
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path


def get_bank_exporter() -> BankStatementExporter:
    return BankStatementExporter()
