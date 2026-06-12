"""End-to-end document processing pipeline.

Deterministic core called by the Slack layer:
  classify → resolve direction → extract → normalize → tax classify →
  categorize (COA account codes) → route → consolidate into FY workbooks.

Design principle: every LLM-calling step is injected via a keyword-only
parameter that defaults to the real function. Tests pass fake callables; the
real Slack layer passes nothing (uses the defaults). No Gemini call is ever
made inside this module itself.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Callable, Optional

from openpyxl import Workbook

from .classify.document_classifier import (
    ClassificationResult,
    classify_file,
    resolve_direction,
)
from .export.categorizer import categorize_invoice
from .export.client_context import (
    ClientContext,
    category_mapping_from_state,
    coa_from_state,
    entity_memory_from_state,
)
from .export.exporters import (
    BankStatementExporter,
    _sheet_title,
    get_bank_exporter,
    get_exporter,
    validate_required_fields,
)
from .export.models import BankStatement, BankTransaction, NormalizedInvoice
from .export.routing import DocRoute, route_document
from .extract.bank_statement_extractor import (
    ExtractedBankStatement,
    extract_bank_file,
    to_bank_statements,
)
from .extract.invoice_extractor import (
    ExtractedInvoice,
    extract_file,
    reconcile,
    to_normalized,
)

# --------------------------------------------------------------------------- #
# Result dataclasses
# --------------------------------------------------------------------------- #

@dataclass
class ProcessedDoc:
    path: str
    doc_type: str                           # "invoice" | "receipt" | "bank_statement"
    direction: Optional[str]                # "purchase" | "sales" | None (bank)
    normalized: Optional[NormalizedInvoice] # invoice / receipt
    bank: Optional[ExtractedBankStatement]  # bank_statement
    route: DocRoute
    reconciled: bool
    note: str                               # human-readable status / reconcile message / error


@dataclass
class BatchResult:
    workbooks: dict[str, bytes]     # filename -> xlsx bytes
    docs: list[ProcessedDoc]
    errors: list[str]


# --------------------------------------------------------------------------- #
# Internal helpers
# --------------------------------------------------------------------------- #

def _parse_iso(s: Optional[str]) -> Optional[date]:
    """Parse an ISO date string (YYYY-MM-DD) tolerantly; return None on failure."""
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _bank_representative_date(ex: ExtractedBankStatement) -> date:
    """Pick the best representative date from a bank statement for FY routing.

    Priority:
    1. Latest transaction date across all accounts (most reliable).
    2. Today (fallback when all txn dates are absent or unparseable).
    """
    best: Optional[date] = None
    for acct in ex.accounts:
        for txn in acct.transactions:
            d = _parse_iso(txn.date)
            if d is not None:
                if best is None or d > best:
                    best = d
    return best if best is not None else date.today()


def _effective_fye_month(client: ClientContext) -> tuple[int, bool]:
    """Return (fye_month, defaulted). Defaults to 12 (calendar year) when None."""
    if client.fye_month is not None:
        return client.fye_month, False
    return 12, True


def _build_ledger_workbook(
    exporter,
    purchases: list[NormalizedInvoice],
    sales: list[NormalizedInvoice],
) -> bytes:
    """Build an in-memory Ledger workbook and return its bytes."""
    wb = Workbook()
    for i, (title, cols, invs, doc_type_str) in enumerate([
        ("Purchase", exporter.purchase_cols, purchases, "purchase"),
        ("Sales", exporter.sales_cols, sales, "sales"),
    ]):
        sheet = wb.active if i == 0 else wb.create_sheet(title)
        sheet.title = title
        sheet.append(cols)
        for row in exporter.rows(invs, doc_type_str):
            sheet.append([row.get(c, "") for c in cols])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_bank_workbook(
    bank_exporter: BankStatementExporter,
    statements: list[BankStatement],
) -> bytes:
    """Build an in-memory BankStatement workbook and return its bytes."""
    wb = Workbook()
    used: dict[str, int] = {}
    for i, stmt in enumerate(statements):
        title = _sheet_title(stmt.bank_name)
        if title in used:
            used[title] += 1
            suffix = f" ({used[title]})"
            title = title[: 31 - len(suffix)] + suffix
        else:
            used[title] = 0
        sheet = wb.active if i == 0 else wb.create_sheet()
        sheet.title = title
        sheet.append(bank_exporter.BANK_COLS)
        for row in bank_exporter.bank_rows(stmt):
            sheet.append([row.get(c, "") for c in bank_exporter.BANK_COLS])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Single-document processing
# --------------------------------------------------------------------------- #

def process_document(
    path: str | Path,
    client: ClientContext,
    *,
    classify_fn: Callable = classify_file,
    direction_fn: Callable = resolve_direction,
    extract_fn: Callable = extract_file,
    bank_fn: Callable = extract_bank_file,
    categorize_fn: Callable = categorize_invoice,
) -> ProcessedDoc:
    """Classify, extract, normalize, categorize, and route a single document.

    All LLM-calling steps are injected via keyword-only parameters so that tests
    can pass pure-Python stubs and never touch Gemini / the network.

    Returns a ProcessedDoc; never raises — any per-doc exception is caught and
    recorded in ``note`` so a batch keeps going.
    """
    path = Path(path)
    fye_month, fye_defaulted = _effective_fye_month(client)
    client_id = client.client_id or "unknown"

    try:
        # ------------------------------------------------------------------ #
        # Step 1 — classify
        # ------------------------------------------------------------------ #
        cls: ClassificationResult = classify_fn(str(path))
        doc_type = (cls.doc_type or "other").strip().lower()

        # ------------------------------------------------------------------ #
        # Step 2a — bank statement branch
        # ------------------------------------------------------------------ #
        if doc_type == "bank_statement":
            result = bank_fn(str(path))
            # bank_fn may return (ExtractedBankStatement, mode_str) or just
            # ExtractedBankStatement — handle both for testability.
            if isinstance(result, tuple):
                ex_bank, _mode = result
            else:
                ex_bank = result

            rep_date = _bank_representative_date(ex_bank)
            route = route_document(
                doc_type="bank_statement",
                direction=None,
                doc_date=rep_date,
                fye_month=fye_month,
                client_id=client_id,
                filename=path.name,
            )
            note = "ok" + (" (fye_month defaulted to 12)" if fye_defaulted else "")
            return ProcessedDoc(
                path=str(path),
                doc_type=doc_type,
                direction=None,
                normalized=None,
                bank=ex_bank,
                route=route,
                reconciled=True,
                note=note,
            )

        # ------------------------------------------------------------------ #
        # Step 2b — invoice / receipt branch
        # ------------------------------------------------------------------ #
        direction: Optional[str] = direction_fn(cls, client_name=client.client_name)
        effective_direction = direction if direction in ("purchase", "sales") else "purchase"

        ex: ExtractedInvoice = extract_fn(str(path))
        normalized: NormalizedInvoice = to_normalized(
            ex,
            direction=effective_direction,
            our_gst_registered=client.tax_registered,
        )

        reconciled, rec_note = reconcile(ex)

        # ------------------------------------------------------------------ #
        # Step 3 — categorize (fill account codes per line)
        # ------------------------------------------------------------------ #
        state = client.to_state()
        categorize_fn(
            normalized,
            coa=coa_from_state(state),
            category_mapping=category_mapping_from_state(state),
            entity_memory=entity_memory_from_state(state),
        )

        # ------------------------------------------------------------------ #
        # Step 3b — validate export-required fields for the client's software.
        # Flag (don't drop): the row is still written so no data is lost, but a
        # missing required field marks the doc for review.
        # ------------------------------------------------------------------ #
        exporter = get_exporter(client.accounting_software)
        missing = validate_required_fields(normalized, exporter, effective_direction)
        if missing:
            reconciled = False
            normalized.reconciled = False
            review_note = "needs review: missing " + ", ".join(missing)
            normalized.reconcile_note = (
                f"{normalized.reconcile_note}; {review_note}"
                if normalized.reconcile_note
                else review_note
            )
            rec_note = f"{rec_note}; {review_note}" if rec_note else review_note

        # ------------------------------------------------------------------ #
        # Step 4 — route
        # ------------------------------------------------------------------ #
        doc_date = normalized.invoice_date or date.today()
        route = route_document(
            doc_type=doc_type,
            direction=direction,
            doc_date=doc_date,
            fye_month=fye_month,
            client_id=client_id,
            filename=path.name,
        )

        note_parts = [rec_note]
        if fye_defaulted:
            note_parts.append("fye_month defaulted to 12")
        note = "; ".join(note_parts)

        return ProcessedDoc(
            path=str(path),
            doc_type=doc_type,
            direction=direction,
            normalized=normalized,
            bank=None,
            route=route,
            reconciled=reconciled,
            note=note,
        )

    except Exception as exc:  # noqa: BLE001
        # Build a stub route so the dataclass is always complete.
        _fye = fye_month
        _doc_date = date.today()
        stub_route = route_document(
            doc_type="invoice",
            direction="purchase",
            doc_date=_doc_date,
            fye_month=_fye,
            client_id=client_id,
            filename=Path(path).name,
        )
        return ProcessedDoc(
            path=str(path),
            doc_type="unknown",
            direction=None,
            normalized=None,
            bank=None,
            route=stub_route,
            reconciled=False,
            note=f"ERROR: {exc}",
        )


# --------------------------------------------------------------------------- #
# Batch processing + workbook consolidation
# --------------------------------------------------------------------------- #

def process_batch(
    paths: list[str | Path],
    client: ClientContext,
    **inject,
) -> BatchResult:
    """Process a list of documents and consolidate into FY workbooks.

    ``inject`` is forwarded to each ``process_document`` call as keyword
    arguments (classify_fn, direction_fn, extract_fn, bank_fn, categorize_fn).

    Returns a BatchResult with:
    - ``workbooks``: dict of filename -> xlsx bytes (in-memory, not written to disk).
    - ``docs``: list of ProcessedDoc (one per path, including errors).
    - ``errors``: list of error strings for docs that failed.
    """
    docs: list[ProcessedDoc] = []
    errors: list[str] = []

    for p in paths:
        doc = process_document(p, client, **inject)
        docs.append(doc)
        if doc.note.startswith("ERROR"):
            errors.append(f"{p}: {doc.note}")

    # ------------------------------------------------------------------ #
    # Group invoice/receipt docs by FY workbook
    # ------------------------------------------------------------------ #
    # ledger_groups[workbook_filename] = {"Purchase": [...], "Sales": [...]}
    ledger_groups: dict[str, dict[str, list[NormalizedInvoice]]] = {}
    # bank_groups[workbook_filename] = list[ExtractedBankStatement]
    bank_groups: dict[str, list[ExtractedBankStatement]] = {}

    for doc in docs:
        if doc.note.startswith("ERROR") or (doc.normalized is None and doc.bank is None):
            continue
        wb_name = doc.route.workbook
        if doc.doc_type == "bank_statement" and doc.bank is not None:
            bank_groups.setdefault(wb_name, []).append(doc.bank)
        elif doc.normalized is not None:
            sheet = doc.route.sheet or "Purchase"
            ledger_groups.setdefault(wb_name, {"Purchase": [], "Sales": []})
            ledger_groups[wb_name].setdefault(sheet, []).append(doc.normalized)

    workbooks: dict[str, bytes] = {}

    # ------------------------------------------------------------------ #
    # Build Ledger workbooks
    # ------------------------------------------------------------------ #
    exporter = get_exporter(client.accounting_software)
    for wb_name, sheets in ledger_groups.items():
        purchases = sheets.get("Purchase", [])
        sales = sheets.get("Sales", [])
        workbooks[wb_name] = _build_ledger_workbook(exporter, purchases, sales)

    # ------------------------------------------------------------------ #
    # Build BankStatement workbooks
    # ------------------------------------------------------------------ #
    bank_exporter = get_bank_exporter()
    for wb_name, ex_banks in bank_groups.items():
        # Convert each ExtractedBankStatement -> list[BankStatement] and flatten
        all_stmts: list[BankStatement] = []
        for ex_bank in ex_banks:
            all_stmts.extend(to_bank_statements(ex_bank))
        if all_stmts:
            workbooks[wb_name] = _build_bank_workbook(bank_exporter, all_stmts)

    return BatchResult(workbooks=workbooks, docs=docs, errors=errors)
