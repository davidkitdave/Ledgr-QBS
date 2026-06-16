"""COA ingest core — pure file readers and store orchestration.

No Slack API calls here; all Slack interaction is injected via ``say_fn``.
"""

from __future__ import annotations

import csv
import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------------- #
# Outcome dataclass
# --------------------------------------------------------------------------- #

@dataclass
class CoaIngestOutcome:
    client_id: Optional[str]
    n_accounts: int
    status: str   # "active" | "no_profile" | "empty" | "validation_failed" | "error"
    note: str


# --------------------------------------------------------------------------- #
# Column-name aliases (case-insensitive normalisation)
# --------------------------------------------------------------------------- #

_COL_ALIASES: dict[str, str] = {
    "account code": "code",
    "account_code": "code",
    "code": "code",
    "description": "description",
    "account type": "account_type",
    "account_type": "account_type",
    "financial statement": "financial_statement",
    "financial_statement": "financial_statement",
    "nature": "nature",
    "ai search keywords": "keywords",
    "keywords": "keywords",
}


def _normalise_header(raw: str) -> Optional[str]:
    return _COL_ALIASES.get(raw.strip().lower())


def _row_to_dict(header_keys: list[Optional[str]], values: tuple) -> Optional[dict]:
    """Map a row of values to a COA dict using normalised header keys.

    Returns None if both code and description are empty (skip blank rows).
    """
    row: dict = {}
    for key, val in zip(header_keys, values, strict=False):
        if key is None:
            continue
        row[key] = str(val).strip() if val is not None else ""
    code = row.get("code", "").strip()
    desc = row.get("description", "").strip()
    if not code and not desc:
        return None
    # Ensure all expected keys are present (default to empty string)
    for k in ("code", "description", "account_type", "financial_statement", "nature", "keywords"):
        row.setdefault(k, "")
    return row


def _parse_sheet_rows(ws_rows: list[tuple]) -> list[dict]:
    """Parse a list of (header_row, *data_rows) tuples into COA dicts."""
    if not ws_rows:
        return []
    header_keys = [_normalise_header(str(h)) if h is not None else None for h in ws_rows[0]]
    results = []
    for raw_row in ws_rows[1:]:
        # skip fully-empty rows
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw_row):
            continue
        d = _row_to_dict(header_keys, raw_row)
        if d is not None:
            results.append(d)
    return results


# --------------------------------------------------------------------------- #
# File readers (pure — no Slack)
# --------------------------------------------------------------------------- #

def coa_rows_from_file(path: str) -> list[dict]:
    """Parse a COA spreadsheet or CSV into a list of row dicts.

    For .xlsx/.xls:
      1. Attempt ``load_client_setup`` (looks for a sheet named "COA").
      2. If that yields no accounts, fall back to reading the first worksheet
         directly with header detection.

    For .csv:
      Parse with ``csv.DictReader``, mapping spec headers (case-insensitive).
    """
    ext = Path(path).suffix.lower()

    if ext in (".xlsx", ".xls"):
        # Primary path: use load_client_setup which handles the "COA" sheet
        try:
            from invoice_processing.export.client_context import load_client_setup
            ctx = load_client_setup(path)
            if ctx.coa:
                return [
                    {
                        "code": acc.code or "",
                        "description": acc.description or "",
                        "account_type": acc.account_type or "",
                        "financial_statement": acc.financial_statement or "",
                        "nature": acc.nature or "",
                        "keywords": acc.keywords or "",
                    }
                    for acc in ctx.coa
                ]
        except Exception:
            logger.warning("COA load_client_setup failed for %s; trying fallback", path, exc_info=True)

        # Fallback: read the first worksheet directly
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = wb.worksheets[0]
                rows = list(ws.iter_rows(values_only=True))
            finally:
                wb.close()
            return _parse_sheet_rows(rows)
        except Exception:
            logger.warning("COA xlsx fallback parse failed for %s", path, exc_info=True)
            return []

    elif ext == ".csv":
        results: list[dict] = []
        try:
            with open(path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.DictReader(fh)
                for raw in reader:
                    row: dict = {}
                    for raw_key, val in raw.items():
                        norm = _normalise_header(raw_key)
                        if norm:
                            row[norm] = (val or "").strip()
                    for k in ("code", "description", "account_type",
                              "financial_statement", "nature", "keywords"):
                        row.setdefault(k, "")
                    code = row.get("code", "").strip()
                    desc = row.get("description", "").strip()
                    if not code and not desc:
                        continue
                    results.append(row)
        except Exception:
            logger.warning("COA csv parse failed for %s", path, exc_info=True)
            return []
        return results

    return []


def standard_coa_rows() -> list[dict]:
    """Return the built-in standard SG SME COA as a list of row dicts.

    Eval/dev only — not exposed in production onboarding.
    """
    data_path = Path(__file__).parent / "data" / "standard_sg_sme_coa.json"
    with open(data_path, encoding="utf-8") as fh:
        return json.load(fh)


# --------------------------------------------------------------------------- #
# Orchestration (injected store + say_fn)
# --------------------------------------------------------------------------- #

def ingest_coa(
    *,
    channel_id: str,
    store,
    rows: list[dict],
    say_fn: Callable,
) -> CoaIngestOutcome:
    """Persist COA rows for the client registered on ``channel_id``.

    Args:
        channel_id: Slack channel to resolve the client.
        store:      ProfileStore-compatible object (get_by_channel / save_coa / set_status).
        rows:       Parsed COA row dicts (from ``coa_rows_from_file`` or ``standard_coa_rows``).
        say_fn:     Callable(**kwargs) that posts a message to the channel.

    Returns:
        CoaIngestOutcome describing what happened.
    """
    from app.blocks import coa_saved_blocks, coa_validation_failed_blocks, needs_setup_blocks
    from app.coa_validate import validate_coa

    ctx = store.get_by_channel(channel_id)

    if ctx is None:
        say_fn(blocks=needs_setup_blocks())
        return CoaIngestOutcome(
            client_id=None,
            n_accounts=0,
            status="no_profile",
            note="No client profile found for this channel.",
        )

    if not rows:
        say_fn(text="I couldn't read any accounts from that file. Please check the format and try again.")
        return CoaIngestOutcome(
            client_id=ctx.client_id,
            n_accounts=0,
            status="empty",
            note="No accounts parsed from the provided rows.",
        )

    validation = validate_coa(rows)
    if not validation.ok:
        say_fn(blocks=coa_validation_failed_blocks(validation.errors))
        return CoaIngestOutcome(
            client_id=ctx.client_id,
            n_accounts=0,
            status="validation_failed",
            note="; ".join(validation.errors),
        )

    store.save_coa(ctx.client_id, rows)
    store.set_status(ctx.client_id, "active")

    say_fn(blocks=coa_saved_blocks(len(rows)))
    return CoaIngestOutcome(
        client_id=ctx.client_id,
        n_accounts=len(rows),
        status="active",
        note=f"Saved {len(rows)} accounts; client is now active.",
    )
