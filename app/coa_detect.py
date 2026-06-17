"""Lightweight spreadsheet classification for the COA confirmation UX.

Decides whether a dropped spreadsheet is most likely a Chart of Accounts upload
(so the user should be asked to confirm COA use) or a ledger/transaction export
(so it should fall through to the document pipeline).

The detector is deliberately conservative: it never silently mutates data.
Callers in :mod:`accounting_agents.slack_runner` ask the user to confirm
before any persist. See ADR-0006 (Path A: parse, validate, echo parse-back
summary, firm confirms).
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Iterable, Optional

from app.coa_validate import CoaValidationResult, validate_coa

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------------- #
# Result types
# --------------------------------------------------------------------------- #


class SpreadsheetKind(str, Enum):
    """Coarse classification for a dropped spreadsheet."""

    COA_CANDIDATE = "coa_candidate"
    LEDGER_CANDIDATE = "ledger_candidate"
    UNKNOWN = "unknown_spreadsheet"


@dataclass
class CoaPreview:
    """Parse + validate outcome for a candidate COA file (no persistence)."""

    rows: list[dict]
    validation: CoaValidationResult
    source: str  # "xlsx" | "csv"

    @property
    def n_accounts(self) -> int:
        return len(self.rows)

    @property
    def n_income(self) -> int:
        return sum(1 for r in self.rows if _is_income(r))

    @property
    def n_expense(self) -> int:
        return sum(1 for r in self.rows if _is_expense(r))

    def sample(self, n: int = 5) -> list[dict]:
        return self.rows[: max(0, n)]


# --------------------------------------------------------------------------- #
# Heuristics
# --------------------------------------------------------------------------- #

# Headers that strongly suggest a chart of accounts (case-insensitive, after
# the same normalisation ``app.coa_ingest._normalise_header`` applies).
_COA_HEADER_KEYS = {"description", "account_type", "code", "financial_statement", "nature"}

# Headers that strongly suggest a ledger/transaction export. Heuristic only —
# we only act on this when COA signals are absent.
_LEDGER_HEADER_KEYS = {
    "contact",
    "invoice_date",
    "invoice_number",
    "unit_amount",
    "total",
    "description",  # alone is weak; combined with the others = ledger
}

_COA_FILENAME_HINTS = (
    "client setup",
    "chart of accounts",
    "chart_of_accounts",
)
_LEDGER_FILENAME_HINTS = ("ledger_fy", "ledger fy", "ledger_")


def _is_income(row: dict) -> bool:
    at = (row.get("account_type") or "").strip().lower()
    return at in {
        "revenue",
        "income",
        "sales",
        "other income",
        "otherincome",
    } or any(k in at for k in ("revenue", "income", "sales"))


def _is_expense(row: dict) -> bool:
    at = (row.get("account_type") or "").strip().lower()
    return at in {
        "expense",
        "expenses",
        "cost of sales",
        "cos",
        "direct costs",
        "overhead",
        "overheads",
        "cost",
    } or any(k in at for k in ("expense", "cost", "overhead"))


def _normalise_headers(header_row: Iterable) -> set[str]:
    """Normalise raw header text into the keys used by the detector heuristics.

    COA headers (description / account_type / code / …) follow the same alias
    table :mod:`app.coa_ingest` uses. Non-COA headers are lowercased and have
    punctuation stripped so ``Invoice #`` and ``Invoice Date`` survive as
    ``invoice_number`` / ``invoice_date`` for the ledger-shape heuristic.
    """
    import re

    aliases = {
        "account code": "code",
        "account_code": "code",
        "code": "code",
        "description": "description",
        "account type": "account_type",
        "account_type": "account_type",
        "financial statement": "financial_statement",
        "financial_statement": "financial_statement",
        "nature": "nature",
        "ai search keywords": "keywords",
        "keywords": "keywords",
    }
    normalised: set[str] = set()
    for h in header_row:
        raw = str(h or "").strip()
        if not raw:
            continue
        key = aliases.get(raw.lower(), raw.lower())
        # Punctuation-stripped form used for ledger-shape matching.
        key = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
        if key:
            normalised.add(key)
    return normalised


def _first_sheet_header(path: str) -> Optional[list]:
    """Return the first-row header values for the first worksheet, if any."""
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                if not wb.worksheets:
                    return None
                ws = wb.worksheets[0]
                header = next(ws.iter_rows(values_only=True, max_row=1), None)
                return list(header) if header else None
            finally:
                wb.close()
        except Exception:  # noqa: BLE001 - heuristic only
            logger.debug("first-sheet header read failed for %s", path, exc_info=True)
            return None
    if ext == ".csv":
        import csv

        try:
            with open(path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.reader(fh)
                return next(reader, None)
        except Exception:  # noqa: BLE001
            logger.debug("csv header read failed for %s", path, exc_info=True)
            return None
    return None


def _sheet_names(path: str) -> list[str]:
    ext = Path(path).suffix.lower()
    if ext not in (".xlsx", ".xls"):
        return []
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:  # noqa: BLE001
        logger.debug("sheet names read failed for %s", path, exc_info=True)
        return []


def _filename_hint(path: str, hints: tuple[str, ...]) -> bool:
    name = Path(path).name.lower()
    return any(h in name for h in hints)


def _is_coa_signals(headers: set[str], sheet_names: list[str], path: str) -> bool:
    """True when the spreadsheet looks like a chart of accounts.

    Sheet name ``COA`` (QBS Client Setup) and at-least-two header hits
    (description + account_type, etc.) are STRONG signals. Filename hints
    alone are NOT enough — a ``Ledger_FY2025.xlsx`` with the string "coa"
    somewhere in the path would otherwise be mis-routed.
    """
    if "COA" in {s.strip() for s in sheet_names}:
        return True
    hits = headers & _COA_HEADER_KEYS
    if len(hits) >= 2:
        return True
    if "description" in hits and "account_type" in hits:
        return True
    return False


def _is_ledger_signals(headers: set[str], path: str) -> bool:
    """True when the spreadsheet looks like a ledger/transaction export.

    Requires at least two ledger-shape headers AND a ledger-flavoured filename
    hint, OR a very strong header match (Contact + Invoice Date + Total).
    """
    strong = {"contact", "invoice_date", "total"}
    if strong.issubset(headers):
        return True
    hits = headers & _LEDGER_HEADER_KEYS
    if len(hits) >= 3 and _filename_hint(path, _LEDGER_FILENAME_HINTS):
        return True
    return False


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #


def classify_spreadsheet(path: str) -> SpreadsheetKind:
    """Classify a downloaded spreadsheet file as COA / ledger / unknown.

    Pure — does not write to disk or call any network. Returns
    :class:`SpreadsheetKind.UNKNOWN` on any parse error so callers fall back
    safely to a disambiguation card.
    """
    if not Path(path).exists():
        return SpreadsheetKind.UNKNOWN
    ext = Path(path).suffix.lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        return SpreadsheetKind.UNKNOWN

    header_row = _first_sheet_header(path)
    if not header_row:
        return SpreadsheetKind.UNKNOWN
    headers = _normalise_headers(header_row)
    sheets = _sheet_names(path) if ext in (".xlsx", ".xls") else []

    if _is_coa_signals(headers, sheets, path):
        return SpreadsheetKind.COA_CANDIDATE
    if _is_ledger_signals(headers, path):
        return SpreadsheetKind.LEDGER_CANDIDATE
    return SpreadsheetKind.UNKNOWN


def preview_coa(path: str) -> Optional[CoaPreview]:
    """Parse + validate a candidate COA file without persisting anything.

    Returns ``None`` if the file cannot be read at all (caller should ask the
    user to re-upload). The returned preview carries the validation result
    so the UI can disable the confirm button when there are hard errors.
    """
    from app.coa_ingest import coa_rows_from_file

    ext = Path(path).suffix.lower().lstrip(".") or "xlsx"
    try:
        rows = coa_rows_from_file(path)
    except Exception:  # noqa: BLE001
        logger.exception("preview_coa: parse failed for %s", path)
        return None
    validation = validate_coa(rows)
    return CoaPreview(rows=rows, validation=validation, source=ext)
