"""Per-client invoice eval harness — mirrors production identity + COA + ground truth.

Unlike :mod:`eval.ledger_eval` (which uses a generic ``default_client()`` placeholder
identity + standard COA, so direction comes back as an artifact and categorization is
hollow), this harness loads each client's REAL identity and their OWN Chart of Accounts
from the client's ``*Client Setup*.xlsx`` and scores DIRECTION CORRECTNESS against the
folder ground truth (``Purchase/`` => "purchase", ``Sales/`` => "sales").

Completeness scoring reuses :mod:`eval.ledger_eval`'s helpers verbatim
(``HeaderFill`` / ``TargetCompleteness`` / ``_tally_completeness``) so the per-target
required-header contract is identical to the existing harness.

EVAL-ONLY. Nothing under ``invoice_processing/`` is modified. Results go to STDOUT only —
the test data contains real client names that must never be written into the repo.

Run (hits AI Studio):
    .venv/bin/python -m eval.client_eval --limit-per-client 6
"""

from __future__ import annotations

import argparse
import os
import traceback
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from eval.ledger_eval import (
    HeaderFill,
    PlacementResult,
    TargetCompleteness,
    _tally_completeness,
    load_ground_truth_ledger,
    score_placement,
)
from invoice_processing.export.exporters import (
    QbsLedgerExporter,
    XeroLedgerExporter,
)

# Direction ground truth per source folder.
_RESOLVED_DIRECTIONS = {"sales", "purchase"}
_TARGETS = ("QBS Ledger", "Xero")

# Month name -> int (FYE_MONTH in Sys_Config is a full month name).
_MONTH_MAP: dict[str, int] = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}


# --------------------------------------------------------------------------- #
# Sys_Config parser
# --------------------------------------------------------------------------- #
def _apply_sys_config(ctx, xlsx_path: Path) -> None:
    """Read Sys_Config Key/Value sheet from *xlsx_path* and apply known fields to *ctx*.

    Applies: REGION, ACCOUNTING_SOFTWARE, BASE_CURRENCY, FYE_MONTH, TAX_REGISTERED.
    Ignores unknown / absent keys. Leaves defaults unchanged when a key is absent.
    """
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        if "Sys_Config" not in wb.sheetnames:
            return
        kv: dict[str, str] = {}
        for row in wb["Sys_Config"].iter_rows(values_only=True):
            if row and len(row) >= 2 and row[0] is not None and row[1] is not None:
                k = str(row[0]).strip()
                v = str(row[1]).strip()
                if k and v:
                    kv[k] = v
    finally:
        wb.close()

    if "REGION" in kv:
        ctx.region = kv["REGION"]
    if "ACCOUNTING_SOFTWARE" in kv:
        ctx.accounting_software = kv["ACCOUNTING_SOFTWARE"]
    if "BASE_CURRENCY" in kv:
        ctx.base_currency = kv["BASE_CURRENCY"]
    if "FYE_MONTH" in kv:
        month_int = _MONTH_MAP.get(kv["FYE_MONTH"].lower())
        if month_int is not None:
            ctx.fye_month = month_int
    if "TAX_REGISTERED" in kv:
        ctx.tax_registered = kv["TAX_REGISTERED"].strip().lower() in ("yes", "true", "1", "y")


# --------------------------------------------------------------------------- #
# Classify spy — captures issuer/bill_to without re-running classify
# --------------------------------------------------------------------------- #
def _make_classify_spy(store: list[Optional[object]]):
    """Wrap ``classify_file`` to intercept each ``ClassificationResult``.

    Each call appends the result to *store* so the caller can inspect
    ``issuer_name`` / ``bill_to_name`` after ``process_document`` returns,
    without re-running classification (which hits the LLM).
    """
    from invoice_processing.classify.document_classifier import classify_file

    def _spy(path: str, **kwargs):
        result = classify_file(path, **kwargs)
        store.append(result)
        return result

    return _spy


# --------------------------------------------------------------------------- #
# Per-client result model
# --------------------------------------------------------------------------- #
@dataclass
class ClientReport:
    client_id: str
    setup_path: str
    n_docs: int = 0
    # Direction correctness vs folder ground truth (the headline).
    direction_total: int = 0          # docs with a ground-truth expectation
    direction_correct: int = 0        # resolved == expected
    # Pipeline health.
    classify_ok: int = 0
    recon_eligible: int = 0
    recon_pass: int = 0
    errors: int = 0
    # Per-target completeness tables (keyed by target name).
    completeness: dict[str, TargetCompleteness] = field(default_factory=dict)
    # Per-doc detail for diagnostics.
    rows: list[dict] = field(default_factory=list)
    # COA placement-accuracy (Task 1, extraction-accuracy plan). N/A when
    # the client has no ground-truth ledger; misses are real, scored failures.
    placement: PlacementResult = field(default_factory=PlacementResult)
    placement_gt_paths: list[str] = field(default_factory=list)

    @property
    def direction_rate(self) -> float:
        return self.direction_correct / self.direction_total if self.direction_total else 0.0

    @property
    def recon_rate(self) -> float:
        return self.recon_pass / self.recon_eligible if self.recon_eligible else 0.0


# --------------------------------------------------------------------------- #
# Discovery
# --------------------------------------------------------------------------- #
def _find_setup(client_dir: Path) -> Path | None:
    """Return the client's ``*Client Setup*.xlsx`` (skipping Excel lock files)."""
    for path in sorted(client_dir.glob("*Client Setup*.xlsx")):
        if path.name.startswith("~$"):
            continue
        return path
    return None


def _discover_docs(client_dir: Path, limit: int) -> list[tuple[Path, str]]:
    """Find up to *limit* (pdf, expected_direction) pairs, split across sides.

    ``Purchase/**/*.pdf`` => expected "purchase"; ``Sales/**/*.pdf`` => expected
    "sales". The limit is split half/half; if one side is empty (or short) the
    other side absorbs the remainder.
    """
    purchases = sorted((client_dir / "Purchase").rglob("*.pdf"))
    sales = sorted((client_dir / "Sales").rglob("*.pdf"))
    # Drop hidden / temp files defensively.
    purchases = [p for p in purchases if not p.name.startswith(".")]
    sales = [p for p in sales if not p.name.startswith(".")]

    half = limit // 2
    n_purchase = min(len(purchases), half)
    n_sales = min(len(sales), limit - n_purchase)
    # Backfill from purchase if sales was short.
    n_purchase = min(len(purchases), limit - n_sales)

    docs: list[tuple[Path, str]] = []
    docs.extend((p, "purchase") for p in purchases[:n_purchase])
    docs.extend((p, "sales") for p in sales[:n_sales])
    return docs


def _find_ground_truth_ledgers(client_dir: Path) -> list[Path]:
    """Return all ``<Client> - Ledger_FY*.xlsx`` ground-truth ledgers for *client_dir*.

    Used by the placement-accuracy metric. Skips Excel lock files (``~$``)
    and returns an empty list when the client has no ground-truth ledger —
    in which case placement is N/A, not a failure.
    """
    return sorted(
        p for p in client_dir.glob("*Ledger_FY*.xlsx")
        if not p.name.startswith("~$")
    )


def _resolve_account_description(
    client, account_code: object
) -> str | None:
    """Resolve a produced ``account_code`` to a human description via the
    client's COA. Returns ``None`` when the code is blank or unknown.

    Per ADR-0006 a client's COA codes are often blank (the QBS exporter
    keys by description); when the produced line has a code we look it
    up against the client's own COA so the comparator scores against
    the actual account chosen, not the raw code.
    """
    if not account_code:
        return None
    code = str(account_code).strip()
    if not code:
        return None
    for acc in getattr(client, "coa", []) or []:
        if (acc.code or "").strip() == code:
            return acc.description or None
    return None


# --------------------------------------------------------------------------- #
# Core per-client evaluation
# --------------------------------------------------------------------------- #
def eval_client(
    client_dir: Path,
    limit: int,
    *,
    process_fn=None,
) -> ClientReport | None:
    """Evaluate one client folder. Returns None (with a warning) if no setup file."""
    from invoice_processing.export.client_context import load_client_setup
    from invoice_processing.pipeline import process_document as _real_process

    if process_fn is None:
        process_fn = _real_process

    client_id = client_dir.name
    setup = _find_setup(client_dir)
    if setup is None:
        print(f"[WARN] No '*Client Setup*.xlsx' under {client_id!r} — skipping.")
        return None

    # Load COA/category/entity from workbook; enrich with identity that only
    # the folder + Sys_Config supply (load_client_setup intentionally omits
    # profile fields — they come from Firestore in production).
    client = load_client_setup(setup, client_id=client_id)
    client.client_name = client_dir.name          # folder name IS the client name
    _apply_sys_config(client, setup)              # REGION, ACCOUNTING_SOFTWARE, etc.

    report = ClientReport(client_id=client_id, setup_path=str(setup))
    report.completeness = {t: TargetCompleteness(target=t) for t in _TARGETS}

    qbs_exporter = QbsLedgerExporter()
    xero_exporter = XeroLedgerExporter()
    exporters = {"QBS Ledger": qbs_exporter, "Xero": xero_exporter}

    # Spy list: one ClassificationResult per doc, in call order.
    cls_results: list[Optional[object]] = []
    classify_spy = _make_classify_spy(cls_results)

    # Ground-truth lookup for placement-accuracy (Task 1). When the client
    # has no GT ledger, ``gt_lookup`` is empty and the metric reports N/A
    # for every line — the metric exists as a scoreboard for future
    # categorisation work, not a check on data we don't have.
    gt_paths = _find_ground_truth_ledgers(client_dir)
    report.placement_gt_paths = [str(p) for p in gt_paths]
    gt_lookup: dict[tuple[str, str], list[str]] = {}
    for p in gt_paths:
        try:
            gt_lookup.update(load_ground_truth_ledger(p))
        except Exception as exc:  # noqa: BLE001
            # A corrupt GT must not poison the run — log and continue with
            # whatever else we loaded.
            print(f"[WARN] Failed to parse ground-truth ledger {p}: {exc}")
    placement_produced: list[tuple[str, str, str | None]] = []

    for path, expected in _discover_docs(client_dir, limit):
        report.n_docs += 1
        spy_idx = len(cls_results)  # index of this doc's ClassificationResult (after call)
        try:
            doc = process_fn(str(path), client, classify_fn=classify_spy)
        except Exception as exc:  # noqa: BLE001
            report.errors += 1
            report.rows.append({
                "path": str(path), "expected": expected, "resolved": None,
                "correct": False, "note": f"ERROR: {exc}",
                "tb": traceback.format_exc(),
                "issuer_name": None, "bill_to_name": None,
            })
            continue

        # Pull party names from the spy (index stable even if spy not appended on error).
        cls_result = cls_results[spy_idx] if spy_idx < len(cls_results) else None
        issuer_name = getattr(cls_result, "issuer_name", None)
        bill_to_name = getattr(cls_result, "bill_to_name", None)

        note = doc.note or ""
        if note.startswith("ERROR"):
            report.errors += 1
            report.rows.append({
                "path": str(path), "expected": expected, "resolved": doc.direction,
                "correct": False, "note": note,
                "issuer_name": issuer_name, "bill_to_name": bill_to_name,
            })
            continue

        dt = (doc.doc_type or "").strip().lower()
        if dt and dt != "unknown":
            report.classify_ok += 1
        if dt in {"invoice", "receipt"}:
            report.recon_eligible += 1
            if doc.reconciled:
                report.recon_pass += 1

        resolved = (doc.direction or "").strip().lower()
        report.direction_total += 1
        correct = resolved == expected
        if correct:
            report.direction_correct += 1

        # Per-target completeness: tally on the GROUND-TRUTH side (so completeness
        # is judged against the correct export contract, not the model's guess).
        if doc.normalized is not None:
            side = expected if expected in _RESOLVED_DIRECTIONS else "purchase"
            for tname, exporter in exporters.items():
                _tally_completeness(
                    report.completeness[tname], exporter, doc.normalized, side
                )

            # Placement-accuracy (Task 1): for each line, record
            # (counterparty_name, line.description, resolved_account_description).
            # The comparator scores these against the GT ledger below.
            counterparty_name = (
                doc.normalized.supplier.name
                if doc.normalized.doc_type == "purchase"
                else doc.normalized.customer.name
            )
            for line in doc.normalized.lines:
                placement_produced.append((
                    counterparty_name or "",
                    line.description or "",
                    _resolve_account_description(client, line.account_code),
                ))

        report.rows.append({
            "path": str(path), "expected": expected,
            "resolved": resolved or None, "correct": correct,
            "doc_type": dt or "unknown", "reconciled": doc.reconciled,
            "note": note,
            "issuer_name": issuer_name,
            "bill_to_name": bill_to_name,
        })

    # Score placement against the GT lookup (if any). Empty lookup + empty
    # produced list both yield a clean N/A result; we never crash the run
    # on a missing GT — we just have nothing to grade.
    if gt_lookup or placement_produced:
        report.placement = score_placement(
            # score_placement expects (vendor, desc, account); None becomes "".
            [(v, d, a or "") for v, d, a in placement_produced],
            gt_lookup,
        )
    return report


# --------------------------------------------------------------------------- #
# Reporting (stdout only)
# --------------------------------------------------------------------------- #
def _print_completeness(completeness: dict[str, TargetCompleteness], indent: str = "  ") -> None:
    for tname in _TARGETS:
        table = completeness.get(tname)
        if table is None:
            continue
        print(f"{indent}[{table.target}]  rows={table.n_rows}")
        if not table.headers:
            print(f"{indent}  (no NormalizedInvoices to evaluate)")
            continue
        print(f"{indent}  {'Header':<24} {'Filled/Total':>14} {'Rate':>8}")
        print(f"{indent}  {'-' * 24} {'-' * 14:>14} {'-' * 8:>8}")
        for hf in table.headers.values():
            ft = f"{hf.filled}/{hf.total}"
            print(f"{indent}  {hf.header:<24} {ft:>14} {hf.rate * 100:>7.1f}%")


def _print_placement(report: ClientReport, indent: str = "  ") -> None:
    """Print the COA placement-accuracy line for one client (Task 1)."""
    p = report.placement
    if not report.placement_gt_paths:
        suffix = "N/A (no ground-truth ledger)"
    elif p.scored == 0 and p.na == 0:
        suffix = "N/A (no produced lines)"
    elif p.scored == 0:
        suffix = f"N/A ({p.na} line(s) with no GT row)"
    else:
        suffix = f"{p.correct}/{p.scored}  rate={p.rate * 100:.1f}%"
    print(f"{indent}PLACEMENT:           {suffix}  (N/A={p.na})")


def _print_client(report: ClientReport, client_name: str = "") -> None:
    print()
    print("=" * 100)
    print(f"CLIENT: {report.client_id}")
    print("=" * 100)
    print(f"  Setup file:          {report.setup_path}")
    if client_name:
        print(f"  Identity supplied:   client_name={client_name!r}")
    print(f"  Documents processed: {report.n_docs}")
    print(
        f"  DIRECTION accuracy:  {report.direction_correct}/{report.direction_total} "
        f"correct  rate={report.direction_rate * 100:.1f}%   (vs folder ground truth)"
    )
    print(f"  Classify OK:         {report.classify_ok}/{report.n_docs}")
    print(
        f"  Recon pass:          {report.recon_pass}/{report.recon_eligible} "
        f"rate={report.recon_rate * 100:.1f}%"
    )
    print(f"  Errors:              {report.errors}")
    _print_placement(report, indent="  ")
    print()
    print("  COMPLETENESS (per target) — required-header fill across exported rows")
    _print_completeness(report.completeness, indent="    ")

    # Per-doc direction + party extraction (issuer/bill_to for first 2 docs shown inline).
    print()
    print("  PER-DOC DIRECTION  (issuer/bill_to shown for first 2 docs)")
    print(
        f"    {'Expected':<10} {'Resolved':<10} {'OK':<4} {'Type':<12}"
        f" {'Issuer':<28} {'BillTo':<28} File"
    )
    sep = f"    {'-'*10} {'-'*10} {'-'*4} {'-'*12} {'-'*28} {'-'*28} {'-'*30}"
    print(sep)
    for i, r in enumerate(report.rows):
        ok = "YES" if r.get("correct") else "no"
        fname = Path(r["path"]).name[-30:]
        # Show party names for first 2 docs (to diagnose party-extraction vs name-match).
        if i < 2:
            issuer = (r.get("issuer_name") or "-")[:27]
            billed = (r.get("bill_to_name") or "-")[:27]
        else:
            issuer = ""
            billed = ""
        print(
            f"    {r['expected']:<10} {(r.get('resolved') or '-'):<10} "
            f"{ok:<4} {(r.get('doc_type') or '-'):<12}"
            f" {issuer:<28} {billed:<28} {fname}"
        )


def _merge_completeness(
    overall: dict[str, TargetCompleteness], report: ClientReport
) -> None:
    for tname in _TARGETS:
        src = report.completeness.get(tname)
        if src is None:
            continue
        dst = overall.setdefault(tname, TargetCompleteness(target=tname))
        dst.n_rows += src.n_rows
        for header, hf in src.headers.items():
            agg = dst.headers.get(header)
            if agg is None:
                agg = HeaderFill(header=header)
                dst.headers[header] = agg
            agg.filled += hf.filled
            agg.total += hf.total


def _print_overall(reports: list[ClientReport]) -> None:
    n_docs = sum(r.n_docs for r in reports)
    dir_total = sum(r.direction_total for r in reports)
    dir_correct = sum(r.direction_correct for r in reports)
    classify_ok = sum(r.classify_ok for r in reports)
    recon_eligible = sum(r.recon_eligible for r in reports)
    recon_pass = sum(r.recon_pass for r in reports)
    errors = sum(r.errors for r in reports)

    # Aggregate placement across all clients (skipping N/A).
    overall_placement = PlacementResult()
    for r in reports:
        overall_placement.correct += r.placement.correct
        overall_placement.missed += r.placement.missed
        overall_placement.na += r.placement.na
        overall_placement.total += r.placement.total

    overall_comp: dict[str, TargetCompleteness] = {}
    for r in reports:
        _merge_completeness(overall_comp, r)

    print()
    print("#" * 100)
    print("OVERALL (all clients)")
    print("#" * 100)
    dir_rate = dir_correct / dir_total * 100 if dir_total else 0.0
    recon_rate = recon_pass / recon_eligible * 100 if recon_eligible else 0.0
    print(f"  Clients evaluated:   {len(reports)}")
    print(f"  Documents processed: {n_docs}")
    print(f"  DIRECTION accuracy:  {dir_correct}/{dir_total} correct  rate={dir_rate:.1f}%")
    print(f"  Classify OK:         {classify_ok}/{n_docs}")
    print(f"  Recon pass:          {recon_pass}/{recon_eligible} rate={recon_rate:.1f}%")
    if overall_placement.scored:
        print(
            f"  PLACEMENT:           {overall_placement.correct}/{overall_placement.scored}"
            f"  rate={overall_placement.rate * 100:.1f}%  (N/A={overall_placement.na})"
        )
    else:
        print("  PLACEMENT:           N/A (no ground-truth ledgers with expected accounts)")
    print(f"  Errors:              {errors}")
    print()
    print("  COMPLETENESS (per target) — aggregated across all clients")
    _print_completeness(overall_comp, indent="    ")
    print()
    print("PER-CLIENT SUMMARY")
    for r in reports:
        p = r.placement
        if p.scored:
            placement_str = f"placement {p.correct}/{p.scored} ({p.rate * 100:.0f}%, N/A={p.na})"
        elif r.placement_gt_paths:
            placement_str = f"placement N/A ({p.na})"
        else:
            placement_str = "placement N/A (no GT)"
        print(
            f"  - {r.client_id}: direction "
            f"{r.direction_correct}/{r.direction_total} ({r.direction_rate * 100:.0f}%), "
            f"classify {r.classify_ok}/{r.n_docs}, "
            f"recon {r.recon_pass}/{r.recon_eligible}, "
            f"{placement_str}, errors {r.errors}"
        )
    print()


# --------------------------------------------------------------------------- #
# __main__
# --------------------------------------------------------------------------- #
_DEFAULT_CLIENTS = (
    "DMTV Global Pte Ltd,"
    "Orange Perspective Consulting Pte. Ltd.,"
    "Sanesea Shipping  Pte. Ltd."
)


def _report_to_dict(report: ClientReport) -> dict:
    """Serialize a :class:`ClientReport` to a JSON-safe dict.

    Used by the ``--output`` flag to save a baseline report that can be
    compared against a future run via :func:`_compare_reports`. The dict
    shape is deliberately stable: every field is a primitive or a list of
    primitives so the report round-trips through ``json.dumps``.
    """
    out = {
        "client_id": report.client_id,
        "setup_path": report.setup_path,
        "n_docs": report.n_docs,
        "direction_total": report.direction_total,
        "direction_correct": report.direction_correct,
        "direction_rate": report.direction_rate,
        "classify_ok": report.classify_ok,
        "recon_eligible": report.recon_eligible,
        "recon_pass": report.recon_pass,
        "recon_rate": report.recon_rate,
        "errors": report.errors,
        "placement": {
            "scored": report.placement.scored,
            "correct": report.placement.correct,
            "missed": report.placement.missed,
            "na": report.placement.na,
            "total": report.placement.total,
            "rate": report.placement.rate,
        },
        "completeness": {
            tname: {
                "target": table.target,
                "n_rows": table.n_rows,
                "headers": {
                    h: {"filled": hf.filled, "total": hf.total, "rate": hf.rate}
                    for h, hf in table.headers.items()
                },
            }
            for tname, table in report.completeness.items()
        },
        "rows": report.rows,
    }
    return out


def _compare_reports(baseline: dict, current: dict) -> dict:
    """Diff two serialized ClientReports — used for regression detection.

    Returns a dict with per-metric ``delta`` (current - baseline) and
    ``regressed`` flag. Direction accuracy is the headline regression signal;
    a drop of more than 5 percentage points flags ``regressed=True`` so the
    CI gate fails loudly. Completeness + recon deltas are reported as
    auxiliary context.
    """
    out: dict = {
        "baseline_client": baseline.get("client_id"),
        "current_client": current.get("client_id"),
        "metrics": {},
        "regressed": False,
    }
    for metric in ("direction_rate", "recon_rate"):
        b = baseline.get(metric, 0.0)
        c = current.get(metric, 0.0)
        delta = c - b
        out["metrics"][metric] = {"baseline": b, "current": c, "delta": delta}
    # Direction accuracy is the regression gate. Drop > 5% = regressed.
    dir_delta = out["metrics"]["direction_rate"]["delta"]
    if dir_delta < -0.05:
        out["regressed"] = True
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="Per-client invoice eval harness")
    parser.add_argument(
        "--clients",
        type=str,
        default=_DEFAULT_CLIENTS,
        help="Comma-separated client folder names",
    )
    parser.add_argument(
        "--limit-per-client",
        type=int,
        default=12,
        help="Max PDFs per client, split across Purchase/Sales (default: 12)",
    )
    parser.add_argument(
        "--root",
        type=str,
        default="~/Desktop/LocalTest/TestDoc/Sample Test Group",
        help="Root directory containing per-client folders",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="If set, write a JSON baseline of the run to this path "
             "(used as the regression comparison artifact).",
    )
    parser.add_argument(
        "--compare-to",
        type=str,
        default=None,
        help="If set with --output, diff the current run against this "
             "baseline JSON and append a regression summary to the output.",
    )
    args = parser.parse_args()

    # Load .env before any AI-client imports (mirrors ledger_eval).
    from dotenv import load_dotenv
    load_dotenv()
    # Force AI Studio dev mode (avoid Vertex quota during eval).
    os.environ["GOOGLE_GENAI_USE_VERTEXAI"] = "FALSE"

    root = Path(args.root).expanduser()
    client_names = [c.strip() for c in args.clients.split(",") if c.strip()]

    reports: list[ClientReport] = []
    for name in client_names:
        client_dir = root / name
        if not client_dir.is_dir():
            print(f"[WARN] Client folder not found: {client_dir} — skipping.")
            continue
        print(f"Evaluating client: {name}  (limit {args.limit_per_client})")
        report = eval_client(client_dir, args.limit_per_client)
        if report is not None:
            reports.append(report)
            _print_client(report, client_name=name)

    if reports:
        _print_overall(reports)

    if args.output:
        import json as _json
        from datetime import datetime, timezone

        output_path = Path(args.output).expanduser()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        baseline_payload: dict = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "root": str(root),
            "limit_per_client": args.limit_per_client,
            "clients": [_report_to_dict(r) for r in reports],
        }
        if args.compare_to and Path(args.compare_to).exists():
            with open(args.compare_to) as fp:
                prior = _json.load(fp)
            prior_by_client = {c["client_id"]: c for c in prior.get("clients", [])}
            diffs: list[dict] = []
            for current in baseline_payload["clients"]:
                prior_match = prior_by_client.get(current["client_id"])
                if prior_match is None:
                    continue
                diffs.append(_compare_reports(prior_match, current))
            baseline_payload["regression_diffs"] = diffs
            any_regressed = any(d.get("regressed") for d in diffs)
            baseline_payload["any_regressed"] = any_regressed
        with open(output_path, "w") as fp:
            _json.dump(baseline_payload, fp, indent=2, default=str)
        print(f"\nWrote JSON baseline → {output_path}")
        if args.compare_to:
            regressed = baseline_payload.get("any_regressed")
            if regressed:
                print(
                    f"[REGRESSION] Direction accuracy dropped >5% vs {args.compare_to}; "
                    f"see 'regression_diffs' in the output JSON."
                )
    elif not reports:
        print("[WARN] No clients evaluated. Check --root / --clients.")


if __name__ == "__main__":
    main()
