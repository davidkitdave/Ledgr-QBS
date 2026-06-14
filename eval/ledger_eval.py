"""Ledger document-processing evaluation harness.

Measures pipeline health and accuracy-proxy metrics across a sample of
invoice / receipt PDFs:
  - classification rate (doc_type resolved, not "unknown" / error)
  - reconciliation pass-rate (among invoice/receipt docs)
  - COA categorization fill-rate (lines with account_code populated)
  - direction resolution rate (sales-vs-purchase resolved, not "unknown")
  - per-target completeness: per-required-header fill for each export
    target (QBS Ledger + Xero), aligned to ADR-0005's completeness contract
  - COA placement accuracy: did each produced line land under the correct
    COA description (per ADR-0006, account codes are blank by design — the
    QBS exporter keys by *description*). See ``load_ground_truth_ledger``
    and ``score_placement``.

Run:
    uv run python -m eval.ledger_eval [--limit 6] [--root /path/to/TestDoc]

The ``run_eval`` function is hermetically testable — inject a stub
``process_fn`` to avoid any Gemini / network calls.
"""

from __future__ import annotations

import argparse
import os
import re
import traceback
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook

from invoice_processing.export.exporters import (
    QbsLedgerExporter,
    XeroLedgerExporter,
    _is_empty,
)

# ---------------------------------------------------------------------------
# Public data models
# ---------------------------------------------------------------------------

@dataclass
class DocEval:
    path: str
    doc_type: str
    direction: Optional[str]
    reconciled: bool
    n_lines: int
    n_lines_with_account: int       # account_code filled
    tax_treatments: list[str]
    note: str
    error: Optional[str]


@dataclass
class HeaderFill:
    """Fill tally for one required header in one export target."""
    header: str
    filled: int = 0
    total: int = 0

    @property
    def rate(self) -> float:
        return self.filled / self.total if self.total else 0.0


@dataclass
class TargetCompleteness:
    """Per-required-header fill table for one export target (QBS or Xero).

    ``headers`` is keyed by required-header name (preserving exporter order);
    each :class:`HeaderFill` tallies non-empty cells across every exported row
    of every processed NormalizedInvoice. ``n_rows`` is the total rows tallied.
    """
    target: str                     # "QBS Ledger" | "Xero"
    headers: dict[str, HeaderFill] = field(default_factory=dict)
    n_rows: int = 0


@dataclass
class EvalReport:
    n_docs: int
    classify_ok: int                # doc_type resolved (not "unknown" / error)
    recon_pass: int                 # reconciled True among invoices / receipts
    recon_rate: float
    categorized_lines: int
    total_lines: int
    categorization_fill_rate: float
    errors: int
    # Direction resolution (among doc_type == "invoice")
    direction_eligible: int = 0     # invoices considered for direction
    direction_resolved: int = 0     # direction in {sales, purchase}
    direction_rate: float = 0.0
    # Per-target completeness tables, keyed by target name ("QBS Ledger"/"Xero")
    completeness: dict[str, TargetCompleteness] = field(default_factory=dict)
    docs: list[DocEval] = field(default_factory=list)


@dataclass
class PlacementResult:
    """Tally of COA placement-accuracy comparisons for one client.

    Per the extraction-accuracy plan (Task 1) and ADR-0006, QBS clients key
    accounts by *description* (codes are often blank by design). A produced
    line is scored against the (vendor, description) ground-truth row that
    matches; ``correct`` increments when the engine's chosen account
    description is in the row's expected list, ``missed`` when it isn't,
    and ``na`` when the produced line has no ground-truth row to grade
    (or the ground-truth row has no expected account). The headline
    ``rate`` is ``correct / scored``; ``scored`` excludes ``na`` so an
    empty GT does not poison the score.
    """
    correct: int = 0
    missed: int = 0
    na: int = 0
    # Cached: total lines the comparator was asked to score (correct + missed + na).
    total: int = 0

    @property
    def scored(self) -> int:
        return self.correct + self.missed

    @property
    def rate(self) -> float:
        return self.correct / self.scored if self.scored else 0.0


# ---------------------------------------------------------------------------
# Core evaluation logic (injectable / hermetically testable)
# ---------------------------------------------------------------------------

_INVOICE_RECEIPT_TYPES = {"invoice", "receipt"}
_UNKNOWN_TYPES = {"unknown", ""}
_RESOLVED_DIRECTIONS = {"sales", "purchase"}


def _tally_completeness(
    table: TargetCompleteness,
    exporter,
    inv,
    side: str,
) -> None:
    """Build *inv*'s export rows for *exporter* and tally each required header.

    *side* is the exporter doc_type ("purchase" / "sales"). Uses the exporter's
    OWN ``required_fields`` + ``rows`` so the header contract is never re-derived
    here. A required header is "filled" on a row when its cell is non-empty
    (per the exporter's ``_is_empty``). Aggregates in place into *table*.
    """
    required = exporter.required_fields(side)
    rows = exporter.rows([inv], side)
    for col in required:
        hf = table.headers.get(col)
        if hf is None:
            hf = HeaderFill(header=col)
            table.headers[col] = hf
    for row in rows:
        table.n_rows += 1
        for col in required:
            hf = table.headers[col]
            hf.total += 1
            if not _is_empty(row.get(col, "")):
                hf.filled += 1


def run_eval(
    paths: list[str | Path],
    client,
    *,
    process_fn: Callable = None,
) -> EvalReport:
    """Run ``process_fn(path, client)`` for each path and aggregate metrics.

    A document that raises an exception becomes a :class:`DocEval` with
    ``error`` set and is counted in ``errors``; it never crashes the run.

    Args:
        paths:      Iterable of PDF file paths.
        client:     A :class:`~invoice_processing.export.client_context.ClientContext`.
        process_fn: Callable ``(path, client) -> ProcessedDoc``.  Defaults to
                    the real ``invoice_processing.pipeline.process_document``.

    Returns:
        An :class:`EvalReport` with aggregate metrics and per-doc detail.
    """
    if process_fn is None:
        from invoice_processing.pipeline import process_document
        process_fn = process_document

    doc_evals: list[DocEval] = []

    classify_ok = 0
    recon_pass = 0
    recon_eligible = 0      # only invoice / receipt docs count for recon
    categorized_lines = 0
    total_lines = 0
    errors = 0

    direction_eligible = 0
    direction_resolved = 0

    qbs_exporter = QbsLedgerExporter()
    xero_exporter = XeroLedgerExporter()
    qbs_table = TargetCompleteness(target="QBS Ledger")
    xero_table = TargetCompleteness(target="Xero")

    for p in paths:
        path_str = str(p)
        try:
            doc = process_fn(path_str, client)

            # Detect errors recorded inside the ProcessedDoc (pipeline catches internally)
            doc_error: Optional[str] = None
            if doc.note and doc.note.startswith("ERROR"):
                doc_error = doc.note
                errors += 1
            else:
                # Classification OK: doc_type is resolved and not "unknown"
                dt = (doc.doc_type or "").strip().lower()
                if dt and dt not in _UNKNOWN_TYPES:
                    classify_ok += 1

                # Reconciliation: only invoice / receipt docs
                if dt in _INVOICE_RECEIPT_TYPES:
                    recon_eligible += 1
                    if doc.reconciled:
                        recon_pass += 1

                # Direction resolution: among invoices (doc_type == "invoice")
                dir_norm = (doc.direction or "").strip().lower()
                if dt == "invoice":
                    direction_eligible += 1
                    if dir_norm in _RESOLVED_DIRECTIONS:
                        direction_resolved += 1

                # Categorization lines
                if doc.normalized is not None:
                    for line in doc.normalized.lines:
                        total_lines += 1
                        if line.account_code:
                            categorized_lines += 1

                # Per-target completeness: evaluate BOTH exporters for the
                # appropriate side. Unknown direction is evaluated as purchase
                # (but does NOT get direction credit above).
                if doc.normalized is not None:
                    side = dir_norm if dir_norm in _RESOLVED_DIRECTIONS else "purchase"
                    _tally_completeness(qbs_table, qbs_exporter, doc.normalized, side)
                    _tally_completeness(xero_table, xero_exporter, doc.normalized, side)

            # Per-doc tax treatments
            tax_treatments: list[str] = []
            n_lines = 0
            n_lines_with_account = 0
            if doc.normalized is not None:
                for line in doc.normalized.lines:
                    n_lines += 1
                    if line.account_code:
                        n_lines_with_account += 1
                    if line.tax_treatment:
                        tax_treatments.append(line.tax_treatment)

            doc_evals.append(DocEval(
                path=path_str,
                doc_type=doc.doc_type or "unknown",
                direction=doc.direction,
                reconciled=doc.reconciled,
                n_lines=n_lines,
                n_lines_with_account=n_lines_with_account,
                tax_treatments=tax_treatments,
                note=doc.note or "",
                error=doc_error,
            ))

        except Exception as exc:  # noqa: BLE001
            errors += 1
            tb = traceback.format_exc()
            doc_evals.append(DocEval(
                path=path_str,
                doc_type="unknown",
                direction=None,
                reconciled=False,
                n_lines=0,
                n_lines_with_account=0,
                tax_treatments=[],
                note=f"ERROR: {exc}",
                error=f"{exc}\n{tb}",
            ))

    n_docs = len(doc_evals)

    # Recon rate: among recon-eligible docs; 0.0 when none eligible
    recon_rate = recon_pass / recon_eligible if recon_eligible else 0.0

    # Fill rate: lines with account_code / total lines; 0.0 when no lines
    fill_rate = categorized_lines / total_lines if total_lines else 0.0

    # Direction rate: resolved / eligible (invoices); 0.0 when none eligible
    direction_rate = direction_resolved / direction_eligible if direction_eligible else 0.0

    return EvalReport(
        n_docs=n_docs,
        classify_ok=classify_ok,
        recon_pass=recon_pass,
        recon_rate=recon_rate,
        categorized_lines=categorized_lines,
        total_lines=total_lines,
        categorization_fill_rate=fill_rate,
        errors=errors,
        direction_eligible=direction_eligible,
        direction_resolved=direction_resolved,
        direction_rate=direction_rate,
        completeness={"QBS Ledger": qbs_table, "Xero": xero_table},
        docs=doc_evals,
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_BANK_KEYWORDS = ("bank", "statement", "bankstatement", "bsmt")


def _looks_like_bank(path: Path) -> bool:
    """Heuristic: skip obvious bank-statement PDFs."""
    name_lower = path.stem.lower()
    return any(kw in name_lower for kw in _BANK_KEYWORDS)


def discover_samples(
    root: str | Path = "~/Desktop/LocalTest/TestDoc",
    limit: int = 6,
) -> list[str]:
    """Glob a small set of invoice / receipt PDFs under *root*.

    Searches ``GST SR:ZR/`` and ``MYDoc/`` subdirectories first; falls back
    to any PDF found under *root*. Skips obvious bank-statement files.
    Caps at *limit* paths.

    Args:
        root:  Root directory to search (tilde-expanded).
        limit: Maximum number of paths to return.

    Returns:
        A list of absolute path strings (≤ *limit* entries).
    """
    root = Path(root).expanduser().resolve()
    found: list[Path] = []

    # Priority dirs: invoice / receipt areas
    priority_patterns = [
        root / "**" / "GST SR*" / "*.pdf",
        root / "**" / "GST ZR*" / "*.pdf",
        root / "**" / "MYDoc" / "**" / "*.pdf",
        root / "**" / "Invoice*" / "*.pdf",
        root / "**" / "Receipt*" / "*.pdf",
    ]

    seen: set[str] = set()
    for pattern in priority_patterns:
        for p in sorted(Path(root).glob(str(pattern.relative_to(root)))):
            if str(p) not in seen and not _looks_like_bank(p):
                seen.add(str(p))
                found.append(p)
            if len(found) >= limit:
                break
        if len(found) >= limit:
            break

    # Fallback: any PDF under root (excluding bank statements)
    if len(found) < limit:
        for p in sorted(root.rglob("*.pdf")):
            if str(p) not in seen and not _looks_like_bank(p):
                seen.add(str(p))
                found.append(p)
            if len(found) >= limit:
                break

    return [str(p) for p in found[:limit]]


def default_client():
    """Return a ready-to-use :class:`ClientContext` seeded with the standard COA.

    Uses:
        - ``fye_month=3`` (March year-end)
        - ``accounting_software="QBS Ledger"``
        - ``tax_registered=True``
        - ``coa`` seeded from :func:`~app.coa_ingest.standard_coa_rows`
    """
    from app.coa_ingest import standard_coa_rows
    from invoice_processing.export.client_context import ClientContext, CoaAccount

    rows = standard_coa_rows()
    coa = [
        CoaAccount(
            code=r.get("code") or None,
            description=r.get("description") or "",
            account_type=r.get("account_type") or None,
            financial_statement=r.get("financial_statement") or None,
            nature=r.get("nature") or None,
            keywords=r.get("keywords") or None,
        )
        for r in rows
    ]
    return ClientContext(
        client_id="eval-default",
        client_name="Eval Default Client",
        fye_month=3,
        accounting_software="QBS Ledger",
        tax_registered=True,
        coa=coa,
    )


# ---------------------------------------------------------------------------
# COA placement-accuracy (Task 1, extraction-accuracy plan)
# ---------------------------------------------------------------------------

# Token-set ratio at or above this threshold counts as a fuzzy match.
# 0.85 is deliberately generous: real-world ground-truth descriptions and
# produced descriptions share words even when a token is added/removed
# (e.g. "Consulting" vs "Consulting services" → 0.67 on raw ratio but the
# token-set version rises to ~0.82, so we use 0.85 as a soft floor).
_PLACEMENT_FUZZY_THRESHOLD = 0.85


def _normalise_text(value: object) -> str:
    """Lowercase, strip, collapse internal whitespace, drop common punctuation.

    Used for matching vendor / customer names and descriptions from the
    ground-truth ledger and the produced output. ``None`` becomes ``""`` so
    a missing field never accidentally matches.
    """
    if value is None:
        return ""
    s = str(value).lower().strip()
    # Collapse any non-alphanumeric run to a single space.
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _token_set_ratio(a: str, b: str) -> float:
    """Jaccard-like token overlap (intersection / union), 0.0 – 1.0.

    Robust to word-order and dropped/added filler words: "Consulting
    services" and "Services, consulting" score 1.0, "Consulting" and
    "Consulting services" score ~0.67. We bias against SequenceMatcher
    on raw text because letter ordering varies between extractor output
    and ledger hand-entry. Lowercased internally so "Service" matches
    "service" regardless of how the caller passed them in.
    """
    ta = {t for t in a.lower().split() if t}
    tb = {t for t in b.lower().split() if t}
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def _fuzzy_contains(needle: str, haystack: list[str]) -> bool:
    """True if *needle* matches any of *haystack* by exact-or-fuzzy match.

    Both sides are normalised at compare time so the caller can pass raw
    text on either side without worrying about case or punctuation.
    """
    if not needle:
        return False
    needle_n = _normalise_text(needle)
    for h in haystack:
        h_n = _normalise_text(h)
        if not h_n:
            continue
        if needle_n == h_n:
            return True
        if _token_set_ratio(needle_n, h_n) >= _PLACEMENT_FUZZY_THRESHOLD:
            return True
    return False


def _match_key(needle_key: tuple[str, str], gt_keys: set[tuple[str, str]]) -> bool:
    """True if a fuzzy match for *needle_key* exists in *gt_keys*."""
    nv, nd = needle_key
    for gv, gd in gt_keys:
        if not nv or not gv or not nd or not gd:
            continue
        if nv == gv and nd == gd:
            return True
        # Only fuzzy-match descriptions when vendor is exact; vendors are
        # legal names and we don't want to silently match "Acme Corp" to
        # "Acme Corporation".
        if nv == gv and _token_set_ratio(nd, gd) >= _PLACEMENT_FUZZY_THRESHOLD:
            return True
    return False


def load_ground_truth_ledger(path: Path) -> dict[tuple[str, str], list[str]]:
    """Parse a Cast Unity ground-truth ledger and return a placement lookup.

    Reads the ``Sales`` and ``Purchase`` sheets of ``<Client> - Ledger_FY*.xlsx``,
    normalises vendor/customer + description to a tuple key, and maps it to
    the list of expected account descriptions (taken from the
    ``Account Code / COA`` column, which real ledgers leave blank for
    "no code assigned" — see ADR-0006). A row whose account column is
    blank is still keyed (mapping to an empty list) so callers can
    distinguish "no ground truth" (key absent) from "ground truth has no
    expected account" (key present, empty list).

    Returns an empty dict when the workbook has neither Sales nor Purchase.
    """
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheets = [s for s in ("Sales", "Purchase") if s in wb.sheetnames]
        if not sheets:
            return {}

        # Per-sheet header layout (from real Cast Unity files, observed 2026-06):
        #   Sales:   (date, number, customer, description, ..., total, account, ...)
        #   Purchase:(number, date, vendor, [tax id], description, ..., total, account, ...)
        # We find the columns by header name so we don't break if the order shifts.
        lookup: dict[tuple[str, str], list[str]] = {}
        for sn in sheets:
            ws = wb[sn]
            rows = ws.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                continue
            # Normalise header strings: "Customer Name" / "Vendor Name" — case-insensitive contains.
            headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
            try:
                party_idx = next(
                    i for i, h in enumerate(headers)
                    if "customer name" in h or "vendor name" in h
                )
            except StopIteration:
                # Sheet has no party column (e.g. legacy layout) — skip it.
                continue
            try:
                desc_idx = next(i for i, h in enumerate(headers) if h == "description")
            except StopIteration:
                continue
            # Account column may be absent; default to None index → always blank.
            account_idx: int | None = next(
                (i for i, h in enumerate(headers) if "account code" in h or "coa" in h),
                None,
            )
            for row in rows:
                if not row or len(row) <= max(party_idx, desc_idx):
                    continue
                party = row[party_idx]
                desc = row[desc_idx]
                party_n = _normalise_text(party)
                desc_n = _normalise_text(desc)
                if not party_n or not desc_n:
                    continue
                expected: list[str] = []
                if account_idx is not None and account_idx < len(row):
                    cell = row[account_idx]
                    if cell is not None and str(cell).strip():
                        # Keep the expected description human-readable; the
                        # comparator normalises both sides at match time.
                        expected.append(str(cell).strip())
                lookup[(party_n, desc_n)] = expected
    finally:
        wb.close()
    return lookup


def score_placement(
    produced: list[tuple[str, str, str]],
    gt: dict[tuple[str, str], list[str]],
) -> PlacementResult:
    """Score a list of produced (vendor, description, account_description) lines
    against a ground-truth lookup built by :func:`load_ground_truth_ledger`.

    A produced line is N/A when no ground-truth row matches its
    (vendor, description) key, OR the matched row's expected-account list
    is empty (the ground-truth didn't assign an account). Otherwise the
    line is correct when the engine's chosen account description matches
    any expected entry under the same fuzzy rule used by the loader.
    """
    result = PlacementResult()
    gt_keys = set(gt.keys())

    for vendor, desc, account in produced:
        result.total += 1
        vendor_n = _normalise_text(vendor)
        desc_n = _normalise_text(desc)
        account_n = _normalise_text(account)

        # Find the matching GT key (fuzzy on description, exact on vendor).
        matched_key: tuple[str, str] | None = None
        for gv, gd in gt_keys:
            if gv != vendor_n:
                continue
            if gd == desc_n:
                matched_key = (gv, gd)
                break
            if _token_set_ratio(desc_n, gd) >= _PLACEMENT_FUZZY_THRESHOLD:
                matched_key = (gv, gd)
                break

        if matched_key is None:
            result.na += 1
            continue
        expected = gt[matched_key]
        if not expected:
            result.na += 1
            continue
        if _fuzzy_contains(account_n, expected):
            result.correct += 1
        else:
            result.missed += 1
    return result


# ---------------------------------------------------------------------------
# Pretty-print report
# ---------------------------------------------------------------------------

def _print_report(report: EvalReport) -> None:
    print()
    print("=" * 100)
    print("LEDGER EVAL REPORT")
    print("=" * 100)
    print(
        f"{'Path':<55} {'Type':<14} {'Dir':<10} {'Rec':>4} "
        f"{'Lines':>6} {'w/Acct':>7} {'Treatments':<20} {'Note'}"
    )
    print("-" * 100)
    for d in report.docs:
        rec_s = "YES" if d.reconciled else "no"
        treats = ",".join(sorted(set(d.tax_treatments))) or "-"
        note_s = (d.error or d.note or "")[:40]
        print(
            f"{d.path[-55:]:<55} {d.doc_type:<14} {(d.direction or '-'):<10} "
            f"{rec_s:>4} {d.n_lines:>6} {d.n_lines_with_account:>7} "
            f"{treats:<20} {note_s}"
        )
    print("=" * 100)
    print()
    print("AGGREGATE METRICS")
    print(f"  Documents processed:          {report.n_docs}")
    print(f"  Classify OK (not unknown):    {report.classify_ok} / {report.n_docs}")
    print(f"  Recon pass (invoice/receipt): {report.recon_pass}  rate={report.recon_rate*100:.1f}%")
    print(f"  Lines with account_code:      {report.categorized_lines} / {report.total_lines}"
          f"  fill={report.categorization_fill_rate*100:.1f}%")
    print(f"  Errors:                       {report.errors}")
    print()

    # ------------------------------------------------------------------ #
    # DIRECTION (sales-vs-purchase resolution among invoices)
    # ------------------------------------------------------------------ #
    print("DIRECTION")
    print(
        f"  direction_resolved {report.direction_resolved}/{report.direction_eligible} "
        f"rate={report.direction_rate*100:.1f}%"
        + ("  (no invoice docs)" if report.direction_eligible == 0 else "")
    )
    print()

    # ------------------------------------------------------------------ #
    # COMPLETENESS (per target) — per-required-header fill table
    # ------------------------------------------------------------------ #
    print("COMPLETENESS (per target)")
    print("  Per-required-header fill across all exported rows (filled/total, rate%).")
    for target_name in ("QBS Ledger", "Xero"):
        table = report.completeness.get(target_name)
        if table is None:
            continue
        print()
        print(f"  [{table.target}]  rows={table.n_rows}")
        if not table.headers:
            print("    (no NormalizedInvoices to evaluate)")
            continue
        print(f"    {'Header':<22} {'Filled/Total':>14} {'Rate':>8}")
        print(f"    {'-'*22} {'-'*14:>14} {'-'*8:>8}")
        for hf in table.headers.values():
            ft = f"{hf.filled}/{hf.total}"
            print(f"    {hf.header:<22} {ft:>14} {hf.rate*100:>7.1f}%")
    print()

    # Verdict
    if report.n_docs == 0:
        print("VERDICT: No documents processed.")
    elif report.errors == report.n_docs:
        print("VERDICT: ALL documents errored — check pipeline configuration.")
    else:
        classify_rate = report.classify_ok / report.n_docs if report.n_docs else 0
        verdicts = []
        if classify_rate >= 0.80:
            verdicts.append(f"classify OK ({classify_rate*100:.0f}%)")
        else:
            verdicts.append(f"classify BELOW 80% ({classify_rate*100:.0f}%)")
        if report.recon_rate >= 0.80:
            verdicts.append(f"recon OK ({report.recon_rate*100:.0f}%)")
        elif report.recon_pass == 0 and report.recon_rate == 0.0:
            verdicts.append("recon n/a (no invoice/receipt docs)")
        else:
            verdicts.append(f"recon BELOW 80% ({report.recon_rate*100:.0f}%)")
        print("VERDICT: " + " | ".join(verdicts))
    print()


# ---------------------------------------------------------------------------
# __main__ runner
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ledger document eval harness")
    parser.add_argument(
        "--limit",
        type=int,
        default=6,
        help="Max number of PDFs to evaluate (default: 6)",
    )
    parser.add_argument(
        "--root",
        type=str,
        default="~/Desktop/LocalTest/TestDoc",
        help="Root directory to discover sample PDFs",
    )
    args = parser.parse_args()

    # Load .env before any AI-client imports (mirrors bank_eval.py)
    from dotenv import load_dotenv
    load_dotenv()

    # Force AI Studio dev mode (avoid Vertex quota during eval)
    os.environ["GOOGLE_GENAI_USE_VERTEXAI"] = "FALSE"

    client = default_client()
    paths = discover_samples(root=args.root, limit=args.limit)

    if not paths:
        print(f"[WARN] No sample PDFs found under: {args.root}")
        print("       Pass --root to specify a directory containing invoice/receipt PDFs.")
    else:
        print(f"Evaluating {len(paths)} document(s) from: {args.root}")
        for p in paths:
            print(f"  {p}")

    report = run_eval(paths, client)
    _print_report(report)
