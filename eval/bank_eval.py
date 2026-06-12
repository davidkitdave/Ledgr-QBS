"""Bank-statement extraction evaluation harness.

Measures running-balance reconciliation pass-rate and opening/closing balance
ground-truth match across 8 test clients.

Run:  uv run python eval/bank_eval.py
"""

from __future__ import annotations

import glob
import os
import traceback
from pathlib import Path
from typing import Optional

# Load .env before any AI-client imports
from dotenv import load_dotenv

load_dotenv()

import openpyxl

from invoice_processing.extract.bank_statement_extractor import (
    extract_bank_file,
    to_bank_statements,
)

# ---------------------------------------------------------------------------
# Ground-truth helpers
# ---------------------------------------------------------------------------

BASE_DIR = Path("/Users/davidkitdave/Desktop/LocalTest/TestDoc/Cast Unity")

CLIENTS = [
    "AAAC GALLERY PRIVATELIMITED",
    "Akar Enterprises Pte. Ltd.",
    "Bfit Life Pte Ltd",
    "DMTV Global Pte Ltd",
    "Envistore Pte. Ltd.",
    "Orange Perspective Consulting Pte. Ltd.",
    "TC Studio Pte. Ltd.",
    "Ur Doctors Pte. Ltd.",
]

TOL_ABS = 0.01
TOL_REL = 0.001


def _within_tol(value: float, reference: float) -> bool:
    return abs(value - reference) <= max(TOL_ABS, TOL_REL * abs(reference))


def client_balance_set(client_dir: Path) -> set[float]:
    """Load every *BankStatement_FY*.xlsx, collect all Balance-column values."""
    balances: set[float] = set()
    pattern = str(client_dir / "*BankStatement_FY*.xlsx")
    for xlsx_path in glob.glob(pattern):
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                if sheet_name.strip().lower() == "sys_config":
                    continue
                ws = wb[sheet_name]
                # Find header row: scan first 10 rows for a cell == "Balance"
                balance_col_idx: Optional[int] = None
                header_row_idx: Optional[int] = None
                rows = list(ws.iter_rows(max_row=15, values_only=True))
                for ridx, row in enumerate(rows):
                    for cidx, cell in enumerate(row):
                        if cell is not None and str(cell).strip().lower() == "balance":
                            balance_col_idx = cidx
                            header_row_idx = ridx
                            break
                    if balance_col_idx is not None:
                        break

                if balance_col_idx is None:
                    continue

                # Collect all numeric values below the header row
                all_rows = list(ws.iter_rows(values_only=True))
                for row in all_rows[header_row_idx + 1 :]:
                    if balance_col_idx < len(row):
                        val = row[balance_col_idx]
                        try:
                            f = float(val)
                            balances.add(round(f, 2))
                        except (TypeError, ValueError):
                            pass
            wb.close()
        except Exception as exc:
            print(f"  [WARN] Could not read {xlsx_path}: {exc}")
    return balances


# ---------------------------------------------------------------------------
# Per-statement grading
# ---------------------------------------------------------------------------


def grade_statement(pdf_path: Path, balance_set: set[float]) -> dict:
    """Extract and grade one PDF; aggregate across all accounts it contains."""
    ex, mode_used = extract_bank_file(pdf_path, mode="auto")
    stmts = to_bank_statements(ex, mode_used=mode_used, source_file_id=str(pdf_path))

    total_rows = rows_true = rows_false = rows_none = 0
    accounts = len(stmts)
    bank_names: list[str] = []
    opening_match = False
    closing_match = False
    reconcile_notes: list[str] = []

    for stmt in stmts:
        bank_names.append(stmt.bank_name or "?")
        if stmt.reconcile_note:
            reconcile_notes.append(f"{stmt.bank_name}: {stmt.reconcile_note}")

        for txn in stmt.transactions:
            total_rows += 1
            if txn.math_ok is True:
                rows_true += 1
            elif txn.math_ok is False:
                rows_false += 1
            else:
                rows_none += 1

        # Opening balance ground-truth match
        if stmt.opening_balance is not None and balance_set:
            if any(_within_tol(stmt.opening_balance, b) for b in balance_set):
                opening_match = True

        # Closing balance ground-truth match
        if stmt.closing_balance is not None and balance_set:
            if any(_within_tol(stmt.closing_balance, b) for b in balance_set):
                closing_match = True

    checkable = rows_true + rows_false
    empty = checkable == 0

    if empty:
        # No checkable transaction rows. An account with no activity where
        # opening == closing is a perfectly correct extraction (e.g. a newly
        # opened account: STARTING BALANCE 0.00, no transactions). Use the
        # first account's opening and the last account's closing balance.
        opening = stmts[0].opening_balance if stmts else None
        closing = stmts[-1].closing_balance if stmts else None

        if (
            opening is not None
            and closing is not None
            and abs(opening - closing) <= max(TOL_ABS, TOL_REL * abs(opening))
        ):
            pass_rate = 1.0  # empty-but-consistent
        else:
            pass_rate = 0.0  # no rows but opening != closing => genuinely missed rows
    else:
        pass_rate = rows_true / checkable

    return {
        "pdf": pdf_path.name,
        "mode": mode_used,
        "accounts": accounts,
        "total_rows": total_rows,
        "rows_true": rows_true,
        "rows_false": rows_false,
        "rows_none": rows_none,
        "empty": empty,
        "pass_rate": pass_rate,
        "opening_match": opening_match,
        "closing_match": closing_match,
        "bank_names": ", ".join(bank_names),
        "reconcile_notes": " | ".join(reconcile_notes),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    print("Bank-Statement Extraction Evaluation")
    print("=" * 100)

    results: list[dict] = []

    for client in CLIENTS:
        client_dir = BASE_DIR / client
        if not client_dir.exists():
            print(f"[WARN] Client dir not found: {client_dir}")
            continue

        # First 2 PDFs (sorted) under BankStatement/**/*.pdf
        bs_dir = client_dir / "BankStatement"
        pdfs = sorted(bs_dir.rglob("*.pdf")) if bs_dir.exists() else []
        if not pdfs:
            print(f"[WARN] No PDFs found for {client}")
            continue

        # Ground-truth balance set (computed once per client)
        balance_set = client_balance_set(client_dir)

        for pdf_path in pdfs[:2]:
            print(
                f"\nGrading: [{client}]  pdf={pdf_path.name}  "
                f"gt_balances={len(balance_set)}"
            )

            try:
                result = grade_statement(pdf_path, balance_set)
                result["client"] = client
                result["error"] = None
                results.append(result)
            except Exception as exc:
                tb = traceback.format_exc()
                print(f"  ERROR: {exc}\n{tb}")
                results.append({
                    "client": client,
                    "pdf": pdf_path.name,
                    "mode": "ERROR",
                    "accounts": 0,
                    "total_rows": 0,
                    "rows_true": 0,
                    "rows_false": 0,
                    "rows_none": 0,
                    "empty": False,
                    "pass_rate": 0.0,
                    "opening_match": False,
                    "closing_match": False,
                    "bank_names": "",
                    "reconcile_notes": str(exc),
                    "error": str(exc),
                })

    # ---------------------------------------------------------------------------
    # Print table
    # ---------------------------------------------------------------------------
    print()
    print("=" * 140)
    print(
        f"{'Client':<42} {'PDF':<38} {'Mode':<8} {'Accts':>5} "
        f"{'T/F/N':>12} {'Empty':>6} {'Pass%':>7} {'Open?':>6} {'Close?':>7}"
    )
    print("-" * 140)

    graded = [r for r in results if r.get("error") is None]
    errored = [r for r in results if r.get("error") is not None]

    for r in results:
        tfn = f"{r['rows_true']}/{r['rows_false']}/{r['rows_none']}"
        pass_pct = f"{r['pass_rate']*100:.1f}%"
        empty_str = "YES" if r.get("empty") else "no"
        open_m = "YES" if r["opening_match"] else "no"
        close_m = "YES" if r["closing_match"] else "no"
        mode_str = r["mode"] if not r.get("error") else "ERROR"
        print(
            f"{r['client']:<42} {r['pdf']:<38} {mode_str:<8} {r['accounts']:>5} "
            f"{tfn:>12} {empty_str:>6} {pass_pct:>7} {open_m:>6} {close_m:>7}"
        )
        if r.get("reconcile_notes"):
            # Wrap long notes
            note = r["reconcile_notes"]
            print(f"  {'':42} reconcile_notes: {note}")

    print("=" * 140)

    # Overall summary
    if graded:
        # ---- Overall: ALL graded statements (empties scored per Fix 1) ----
        mean_pass = sum(r["pass_rate"] for r in graded) / len(graded)
        n_90 = sum(1 for r in graded if r["pass_rate"] >= 0.90)
        mean_open = sum(1 for r in graded if r["opening_match"]) / len(graded)
        mean_close = sum(1 for r in graded if r["closing_match"]) / len(graded)

        # ---- Non-empty breakdown: only statements with >=1 transaction ----
        non_empty = [r for r in graded if not r.get("empty")]
        empties = [r for r in graded if r.get("empty")]
        empty_consistent = [r for r in empties if r["pass_rate"] >= 0.90]
        ne_mean_pass = (
            sum(r["pass_rate"] for r in non_empty) / len(non_empty)
            if non_empty
            else 0.0
        )

        print()
        print("OVERALL SUMMARY (ALL graded statements; empties scored per Fix 1)")
        print(f"  Statements graded:          {len(graded)} / {len(results)}")
        print(f"  Errors:                     {len(errored)}")
        print(f"  Empty (consistent):         {len(empty_consistent)} / {len(empties)} empty")
        print(f"  Mean pass-rate (OVERALL):   {mean_pass*100:.1f}%  (target >=90%)")
        print(f"  Statements with rate>=0.90: {n_90} / {len(graded)}")
        print(f"  Opening-balance match rate: {mean_open*100:.1f}%")
        print(f"  Closing-balance match rate: {mean_close*100:.1f}%")
        print()
        print("NON-EMPTY BREAKDOWN (statements with >=1 transaction — real extraction quality)")
        print(f"  Non-empty statements:       {len(non_empty)}")
        if non_empty:
            ne_90 = sum(1 for r in non_empty if r["pass_rate"] >= 0.90)
            print(f"  Mean pass-rate (NON-EMPTY): {ne_mean_pass*100:.1f}%")
            print(f"  Non-empty with rate>=0.90:  {ne_90} / {len(non_empty)}")
        else:
            print("  Mean pass-rate (NON-EMPTY): n/a (no non-empty statements)")
        print()

        if mean_pass >= 0.90:
            print("VERDICT: PASS — overall reconciliation pass-rate meets the >=0.90 target.")
        else:
            print(f"VERDICT: BELOW TARGET — mean pass-rate {mean_pass*100:.1f}% < 90%.")

        # Flag low-scoring statements (excluding correctly-empty ones)
        low = [r for r in graded if r["pass_rate"] < 0.90 and not r.get("empty")]
        if low:
            print()
            print("Low-scoring statements (pass_rate < 0.90, excluding correctly-empty):")
            for r in low:
                print(f"  [{r['client']}]  pdf={r['pdf']}  pass={r['pass_rate']*100:.1f}%")
                if r.get("reconcile_notes"):
                    print(f"    reconcile_notes: {r['reconcile_notes']}")
    else:
        print("No statements were successfully graded.")

    if errored:
        print()
        print("Errored statements:")
        for r in errored:
            print(f"  [{r['client']}]  {r['pdf']}  -> {r['error']}")


if __name__ == "__main__":
    main()
