import io
import json
import openpyxl
from google.adk.tools import LongRunningFunctionTool, ToolContext
from google.genai import types, Client
from pydantic import BaseModel
from google.cloud import firestore
from . import config

class FormattedRow(BaseModel):
    row: list[str]

class FormattedData(BaseModel):
    headers: list[str]
    rows: list[FormattedRow]

def _dynamic_format_mapping(universal_data: dict | list, target_headers: str) -> tuple[list[str], list[list[str]]]:
    """Dynamically maps universal schema data to target CSV headers using Gemini."""
    client = Client()
    prompt = f"""
    You are an expert data mapper.
    Map the following Universal Schema JSON data:
    {json.dumps(universal_data)}
    
    Into these exact CSV columns:
    {target_headers}
    
    If a column has no mapping, leave it as an empty string. Output the headers and the mapped rows exactly.
    """
    response = client.models.generate_content(
        model=config.MODEL_LITE,
        contents=prompt,
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=FormattedData,
            temperature=0.0
        )
    )
    result = response.parsed
    return result.headers, [r.row for r in result.rows]

def _request_human_approval(journal_entry_details: str) -> dict:
    """Request human accountant approval for a proposed journal entry.
    
    This will pause execution and notify the human via Slack (stubbed here).
    
    Args:
        journal_entry_details: The formatted details of the journal entry to approve.
    """
    # In a real scenario, this integrates with the slack package.
    return {
        "status": "pending",
        "message": f"Please review and approve this journal entry: {journal_entry_details}"
    }

request_human_approval = LongRunningFunctionTool(func=_request_human_approval)

def categorize_transaction(description: str, amount: float) -> str:
    """Categorizes a bank transaction based on its description.
    
    Args:
        description: The text description of the bank transaction.
        amount: The monetary amount of the transaction.
    """
    desc = description.lower()
    if "aws" in desc or "google cloud" in desc:
        return "Software & Hosting"
    elif "wework" in desc or "rent" in desc:
        return "Rent Expense"
    elif "gusto" in desc or "payroll" in desc:
        return "Payroll Expense"
    return "Uncategorized Expense"

def create_journal_entry(debit_account: str, credit_account: str, amount: float, description: str) -> str:
    """Finalizes and posts the journal entry to the ERP (stubbed).
    
    Args:
        debit_account: The General Ledger account to debit.
        credit_account: The General Ledger account to credit.
        amount: The total amount for the entry.
        description: The memo for the entry.
    """
    return f"Successfully posted JE: Debit {debit_account} {amount}, Credit {credit_account} {amount} ({description})"

async def generate_invoice_excel(
    universal_data: dict | list,
    target_headers: str,
    financial_year: str,
    client_id: str,
    tool_context: ToolContext
) -> dict:
    """Generates an Excel file by dynamically mapping universal schema data to the client's accounting software headers.
    
    Args:
        universal_data: The extracted JSON data in the Universal Schema.
        target_headers: A string representing the client's requested CSV headers (e.g. "Date, Vendor, Amount").
        financial_year: The financial year folder to save under (e.g., "FY2026").
        client_id: The client identifier.
    """
    mapped_headers, mapped_rows = _dynamic_format_mapping(universal_data, target_headers)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Data"
    
    ws.append(mapped_headers)
    for row in mapped_rows:
        ws.append(row)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    # Save as an artifact in a structured path
    blob = types.Blob(mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", data=file_stream.read())
    part = types.Part(inline_data=blob)
    
    filename = f"invoice_export_{financial_year}_{client_id}.xlsx"
    
    version = await tool_context.save_artifact(filename, part)
    
    return {
        "status": "success", 
        "message": f"Excel file {filename} generated and saved successfully.",
        "artifact_version": version
    }

async def generate_bank_statement_excel(
    accounts_data: list[dict],
    target_headers: str,
    financial_year: str,
    client_id: str,
    tool_context: ToolContext
) -> dict:
    """Generates a multi-tab Excel file grouped by account UUID, mapped to the client's target headers.
    
    Args:
        accounts_data: List of dicts, each containing 'account_uuid' and 'transactions'.
        target_headers: The exact CSV header row expected by the accounting software.
        financial_year: The financial year folder.
        client_id: The client identifier.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    for account in accounts_data:
        uuid = account.get("account_uuid", "Unknown")
        ws = wb.create_sheet(title=f"Account_{uuid}")
        
        mapped_headers, mapped_rows = _dynamic_format_mapping(account.get("transactions", []), target_headers)
        
        ws.append(mapped_headers)
        for row in mapped_rows:
            ws.append(row)
            
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    blob = types.Blob(mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", data=file_stream.read())
    part = types.Part(inline_data=blob)
    
    filename = f"statement_export_{financial_year}_{client_id}.xlsx"
    
    version = await tool_context.save_artifact(filename, part)
    
    return {
        "status": "success", 
        "message": f"Multi-tab Bank Statement Excel {filename} generated successfully.",
        "artifact_version": version
    }

async def fetch_client_profile(client_id: str, tool_context: ToolContext) -> dict:
    """Fetches the client's onboarding profile (FYE, GST status) from Firestore and caches it in the session state.
    
    Args:
        client_id: The unique identifier for the client (e.g., from their Slack channel).
        
    Returns:
        dict: The client profile data.
    """
    try:
        # Initialize Firestore client
        # It automatically uses the GOOGLE_CLOUD_PROJECT from the environment
        db = firestore.Client()
        
        # We assume there is a 'clients' collection where document ID is the client_id
        doc_ref = db.collection('clients').document(client_id)
        doc = doc_ref.get()
        
        if doc.exists:
            profile_data = doc.to_dict()
            
            # Save the profile data to the ADK user-persistent state
            # This ensures we don't have to query Firestore again for this user
            tool_context.state["user:fye_month"] = profile_data.get("fye_month", "Unknown")
            tool_context.state["user:gst_registered"] = profile_data.get("gst_registered", False)
            tool_context.state["user:client_name"] = profile_data.get("client_name", "Unknown Client")
            
            # Load the required accounting software headers from the profile (with fallbacks for testing)
            tool_context.state["user:invoice_target_headers"] = profile_data.get("invoice_target_headers", "Date, Vendor Name, Total Amount, GL Account")
            tool_context.state["user:bank_target_headers"] = profile_data.get("bank_target_headers", "Date, Description, Amount, AccountCode")
            
            return {
                "status": "success",
                "message": f"Successfully loaded profile for {client_id}.",
                "profile": profile_data
            }
        else:
            return {
                "status": "error",
                "message": f"Client ID {client_id} not found in Firestore 'clients' collection."
            }
    except Exception as e:
        return {
            "status": "error",
            "message": f"Failed to fetch profile from Firestore: {str(e)}"
        }
