"""``LocalLedgerStore`` — disk-backed FY ledger workbook for ADK playground testing.

Drop-in replacement for ``SlackLedgerStore`` that reads/writes Excel workbooks on
the local filesystem instead of Slack.  The workbook format is IDENTICAL to the Slack
version (same ``openpyxl`` + ``exporters.py`` code), so you can open the output in
Excel and compare it directly against the production workbook.

Layout::

    {output_dir}/
      {client_id}/
        FY{fy}_invoice.xlsx      ← invoice ledger (Purchase + Sales sheets)
        FY{fy}_bank.xlsx         ← bank statement ledger
        .seen_doc_keys_{fy}.json ← dedup state (mirrors Firestore pointer)

Concurrency: single-process only (no locks).  Fine for playground / eval.
"""

from __future__ import annotations

import io
import json
import logging
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

from invoice_processing.export.exporters import (
    BankStatementExporter,
    get_exporter,
)

logger = logging.getLogger(__name__)

_DEFAULT_OUTPUT_DIR = "./playground_output"


class LocalLedgerStore:
    """Disk-backed ledger store with the same public API as ``SlackLedgerStore``.

    Args:
        output_dir: Root directory for all client workbooks.
    """

    def __init__(self, output_dir: str = _DEFAULT_OUTPUT_DIR) -> None:
        self._root = Path(output_dir)

    # ------------------------------------------------------------------ #
    # Path helpers
    # ------------------------------------------------------------------ #

    def _client_dir(self, client_id: str) -> Path:
        d = self._root / client_id
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _workbook_path(self, client_id: str, fy: str, kind: str = "invoice") -> Path:
        return self._client_dir(client_id) / f"FY{fy}_{kind}.xlsx"

    def _seen_path(self, client_id: str, fy: str) -> Path:
        return self._client_dir(client_id) / f".seen_doc_keys_{fy}.json"

    # ------------------------------------------------------------------ #
    # Dedup state (mirrors Firestore seen_doc_keys)
    # ------------------------------------------------------------------ #

    def _load_seen(self, client_id: str, fy: str) -> set:
        p = self._seen_path(client_id, fy)
        if p.exists():
            return set(json.loads(p.read_text()))
        return set()

    def _save_seen(self, client_id: str, fy: str, seen: set) -> None:
        p = self._seen_path(client_id, fy)
        p.write_text(json.dumps(sorted(seen)))

    # ------------------------------------------------------------------ #
    # Workbook helpers (same logic as SlackLedgerStore)
    # ------------------------------------------------------------------ #

    @staticmethod
    def _exporter_for(software: str):
        return get_exporter(software or "qbs")

    def _fresh_invoice_workbook(self, software: str) -> Workbook:
        exporter = self._exporter_for(software)
        wb = Workbook()
        for i, (title, cols) in enumerate(
            (("Purchase", exporter.purchase_cols), ("Sales", exporter.sales_cols))
        ):
            sheet = wb.active if i == 0 else wb.create_sheet(title)
            sheet.title = title
            sheet.append(list(cols))
        return wb

    def _fresh_bank_workbook(self) -> Workbook:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Bank"
        sheet.append(list(BankStatementExporter.BANK_COLS))
        return wb

    @staticmethod
    def _get_or_create_sheet(wb: Workbook, sheet_name: str, kind: str):
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        if kind == "bank" and "Bank" in wb.sheetnames and wb["Bank"].max_row <= 1:
            sheet = wb["Bank"]
            sheet.title = sheet_name
            return sheet
        sheet = wb.create_sheet(sheet_name)
        if kind == "bank":
            sheet.append(list(BankStatementExporter.BANK_COLS))
        return sheet

    @staticmethod
    def _append_rows_to_sheet(sheet, cols: list[str], rows: list[dict]) -> int:
        from accounting_agents.ledger_store import _invoice_date_sort_key

        sorted_rows = sorted(rows, key=_invoice_date_sort_key)
        for row in sorted_rows:
            sheet.append([row.get(c, "") for c in cols])
        return len(sorted_rows)

    def _merge_bank_statement(self, sheet, cols: list[str], rows: list[dict]) -> int:
        """Merge bank rows into the sheet (delegates to SlackLedgerStore internals)."""
        # Import the class method we need — reuse the same logic.
        from accounting_agents.ledger_store import SlackLedgerStore

        SlackLedgerStore._ensure_bank_header(sheet, cols)
        existing_blocks = SlackLedgerStore._read_bank_blocks(sheet, cols)
        new_blocks = BankStatementExporter.rows_to_blocks(rows)

        before = len(existing_blocks) + len(new_blocks)
        deduped_blocks = BankStatementExporter.dedupe_blocks(existing_blocks + new_blocks)
        existing_sig = {
            BankStatementExporter._block_signature(b) for b in existing_blocks
        }
        added = 0
        for block in deduped_blocks:
            if BankStatementExporter._block_signature(block) not in existing_sig:
                added += len(block["transactions"]) + 1
        if len(deduped_blocks) < before:
            logger.info(
                "bank merge: collapsed %d duplicate block(s) on %s",
                before - len(deduped_blocks), sheet.title,
            )
        all_blocks = BankStatementExporter.sort_blocks(deduped_blocks)
        BankStatementExporter.rebuild_account_sheet(sheet, all_blocks, cols)
        return added

    # ------------------------------------------------------------------ #
    # Public API (mirrors SlackLedgerStore)
    # ------------------------------------------------------------------ #

    def append_rows(
        self,
        *,
        client_id: str,
        fy: str,
        batches: list[dict],
        software: str = "qbs",
        kind: str = "invoice",
        client_name: str = "",
        # Ignored params kept for interface compatibility.
        slack_client: Any = None,
        channel_id: str = "",
        replace: bool = False,
    ) -> dict:
        """Append rows to the local FY workbook (same interface as SlackLedgerStore)."""
        wb_path = self._workbook_path(client_id, fy, kind)
        seen = self._load_seen(client_id, fy)

        if wb_path.exists():
            wb = load_workbook(wb_path)
        elif kind == "bank":
            wb = self._fresh_bank_workbook()
        else:
            wb = self._fresh_invoice_workbook(software)

        if kind == "bank":
            cols = list(BankStatementExporter.BANK_COLS)
        else:
            exporter = self._exporter_for(software)
            sheet_cols = {
                "Purchase": list(exporter.purchase_cols),
                "Sales": list(exporter.sales_cols),
            }

        appended = 0
        deduped = 0

        for batch in batches:
            sheet_name = batch["sheet"]
            doc_key = str(batch["doc_key"])
            rows = batch.get("rows") or []

            if doc_key in seen:
                deduped += 1
                continue

            sheet = self._get_or_create_sheet(wb, sheet_name, kind)
            if kind == "bank":
                n = self._merge_bank_statement(sheet, cols, rows)
            else:
                cols_for_sheet = sheet_cols.get(sheet_name, list(exporter.purchase_cols))
                n = self._append_rows_to_sheet(sheet, cols_for_sheet, rows)

            if n == 0:
                deduped += 1
            else:
                appended += n
                seen.add(doc_key)

        # Save.
        wb.save(wb_path)
        self._save_seen(client_id, fy, seen)

        return {
            "workbook_path": str(wb_path),
            "appended": appended,
            "deduped": deduped,
            "filename": wb_path.name,
        }

    def read_rows(self, client_id: str, fy: str | None = None) -> list[dict]:
        """Read all rows from the local FY workbook.

        If ``fy`` is None, reads from the latest FY found.
        """
        if fy is None:
            fy = self.latest_fy(client_id)
        if not fy:
            return []

        # Try invoice first, then bank.
        for kind in ("invoice", "bank"):
            wb_path = self._workbook_path(client_id, fy, kind)
            if wb_path.exists():
                wb = load_workbook(wb_path)
                rows: list[dict] = []
                for sheet in wb.worksheets:
                    if sheet.max_row < 1:
                        continue
                    headers = [c.value for c in sheet[1]]
                    for ws_row_num, row in enumerate(
                        sheet.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        row_dict: dict = {"_sheet": sheet.title, "_row": ws_row_num}
                        for idx, header in enumerate(headers):
                            if header is not None:
                                row_dict[header] = row[idx] if idx < len(row) else None
                        if any(v is not None for k, v in row_dict.items() if k not in ("_sheet", "_row")):
                            rows.append(row_dict)
                return rows
        return []

    def latest_fy(self, client_id: str) -> str | None:
        """Return the highest FY label that has a local workbook."""
        client_dir = self._root / client_id
        if not client_dir.exists():
            return None
        fys = set()
        for p in client_dir.glob("FY*_*.xlsx"):
            # Parse "FY2026_invoice.xlsx" → "2026"
            name = p.stem  # "FY2026_invoice"
            fy_part = name.split("_")[0]  # "FY2026"
            if fy_part.startswith("FY"):
                fys.add(fy_part[2:])
        return max(fys) if fys else None

    def fy_pointers(self, client_id: str) -> list[dict]:
        """Return summary info for all FYs (for chat-lane state seeding)."""
        client_dir = self._root / client_id
        if not client_dir.exists():
            return []
        seen_fys: dict[str, dict] = {}
        for p in client_dir.glob("FY*_*.xlsx"):
            name = p.stem
            parts = name.split("_", 1)
            fy = parts[0][2:] if parts[0].startswith("FY") else parts[0]
            kind = parts[1] if len(parts) > 1 else "unknown"
            if fy not in seen_fys:
                rows = self.read_rows(client_id, fy)
                seen_fys[fy] = {
                    "fy": fy,
                    "kind": kind,
                    "row_count": len(rows),
                    "has_data": len(rows) > 0,
                    "workbook_path": str(p),
                }
        return sorted(seen_fys.values(), key=lambda s: s["fy"], reverse=True)
