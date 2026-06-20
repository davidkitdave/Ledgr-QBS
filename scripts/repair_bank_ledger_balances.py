"""One-time repair script: rebuild a BankStatement workbook into the uniform
static-value style used by the current pipeline.

Usage
-----
    python scripts/repair_bank_ledger_balances.py /path/to/BankStatement_FY2025.xlsx

The input workbook is read-only.  The repaired workbook is written next to it
with a ``_repaired`` suffix, e.g. ``BankStatement_FY2025_repaired.xlsx``.  The
script is idempotent: running it twice on the same file (or on a file already in
the correct format) produces the same output.

What it fixes
-------------
Commit 6ca4e48 migrated the bank Balance column from an Excel-formula chain
(``=E_prev + Deposit − Withdrawal``) to static numeric values.  Workbooks that
span the migration (e.g. Sample Bank Client's ``BankStatement_FY2025``) have Jan–Mar stored as
formula strings and Apr+ as statics.  When openpyxl reads such a workbook with
the default ``data_only=False`` the formula cells come back as strings
(``"=E2+D3-C3"``); if loaded with ``data_only=True`` they come back as ``None``
(because no cached value was ever written).  Either way the running balance
breaks.

This script:
1. Reads every account sheet (skipping only blank sheets).
2. Detects and migrates the old 8-col header (``Stated Balance`` → ``Balance``,
   ``Check`` → ``Math_Check``) if present.
3. Recomputes the running Balance deterministically from
   ``stated_bf + Σ(deposit − withdrawal)`` — formula strings and None values are
   treated as missing and replaced with the recomputed value.
4. Rebuilds each sheet as one continuous, date-sorted static chain (the same
   layout ``rebuild_account_sheet`` always produces).
5. Writes the repaired workbook to ``<original_stem>_repaired.xlsx``.

The original file is never modified.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

# Make the project root importable when run as a script.
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from openpyxl import load_workbook  # noqa: E402

from accounting_agents.ledger_store import SlackLedgerStore  # noqa: E402
from invoice_processing.export.exporters import BankStatementExporter  # noqa: E402


# Legacy column name → current canonical name.
_LEGACY_COL_MAP: dict[str, str] = {
    "Stated Balance": "Balance",
    "Check": "Math_Check",
}

_COLS = list(BankStatementExporter.BANK_COLS)


def _migrate_header_row(ws) -> None:
    """Rename legacy column headers in row 1 to current canonical names."""
    if ws.max_row < 1:
        return
    for cell in ws[1]:
        if cell.value in _LEGACY_COL_MAP:
            cell.value = _LEGACY_COL_MAP[cell.value]


def _repair_sheet(ws) -> None:
    """Repair a single account sheet in-place."""
    if ws.max_row < 1:
        return

    # Step 1: migrate legacy header names.
    _migrate_header_row(ws)

    # Step 2: read existing data into month-blocks (recomputes Balance internally).
    blocks = SlackLedgerStore._read_bank_blocks(ws, _COLS)

    if not blocks:
        return

    # Step 3: sort blocks by date and rebuild the sheet as a clean static chain.
    sorted_blocks = BankStatementExporter.sort_blocks(blocks)
    BankStatementExporter.rebuild_account_sheet(ws, sorted_blocks, _COLS)


def repair(input_path: Path) -> Path:
    """Repair *input_path* and write the result next to it with ``_repaired`` suffix.

    Returns the path of the written output file.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    output_path = input_path.with_stem(input_path.stem + "_repaired")

    # Load with data_only=False so formula strings are preserved as strings
    # (the recompute logic in _read_bank_blocks handles them explicitly).
    wb = load_workbook(input_path, data_only=False)

    for ws in wb.worksheets:
        if ws.max_row <= 1:
            # Skip header-only or blank sheets.
            continue
        _repair_sheet(ws)

    wb.save(output_path)
    return output_path


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Rebuild a BankStatement workbook into the uniform static-balance style.",
    )
    parser.add_argument(
        "workbook",
        type=Path,
        help="Path to the BankStatement_FY<year>.xlsx to repair.",
    )
    args = parser.parse_args(argv)
    input_path: Path = args.workbook.expanduser().resolve()

    print(f"Reading:  {input_path}")
    output_path = repair(input_path)
    print(f"Repaired: {output_path}")
    print("Done.  The original file was not modified.")


if __name__ == "__main__":
    main()
