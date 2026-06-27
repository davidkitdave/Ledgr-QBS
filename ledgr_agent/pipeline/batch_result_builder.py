"""Map light-path per-file results into the shared BatchResult contract."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from ledgr_agent.export.erp_projection import normalize_system_key, project
from ledgr_agent.export.light_workbook import build_workbook_bytes
from ledgr_agent.review.grouping import merge_soft_warnings, partition_and_group_reasons
from ledgr_agent.schemas.batch_result import BatchResult, BatchStatus
from ledgr_agent.schemas.credit import CreditSummary
from ledgr_agent.schemas.review import ReviewRequest, SoftWarning


@dataclass
class LightFileResult:
    path: str
    kind: str
    status: str = "success"
    message: str | None = None
    documents: list[dict[str, Any]] = field(default_factory=list)
    bank_workbook: dict[str, Any] | None = None
    extraction_meta: dict[str, Any] = field(default_factory=dict)


def _review_from_note(
    note: str | None,
    file_name: str,
) -> tuple[list[ReviewRequest], list[SoftWarning]]:
    if not note:
        return [], []
    if note.lower().startswith("needs review:"):
        note = note[len("needs review:") :]
    note = note.strip()
    if not note:
        return [], []
    reasons = [r.strip() for r in note.split(";") if r.strip()]
    return partition_and_group_reasons(reasons, file_name=file_name)


def _determine_batch_status(
    *,
    blocked_reason: str | None,
    processed_docs: list[Any],
    missing_files: list[str],
) -> BatchStatus:
    if blocked_reason:
        return "blocked"
    if not processed_docs and missing_files:
        return "error"
    if not processed_docs and not missing_files:
        return "error"
    error_count = sum(1 for doc in processed_docs if doc.note.startswith("ERROR"))
    review_count = sum(
        1
        for doc in processed_docs
        if not doc.reconciled and not doc.note.startswith("ERROR")
    )
    if processed_docs and error_count == len(processed_docs):
        return "error"
    if review_count and error_count:
        return "partial"
    if review_count:
        return "needs_review"
    if error_count or missing_files:
        return "partial"
    return "success"


def _erp_export_summaries(workbooks: dict[str, bytes]) -> list[dict[str, object]]:
    from io import BytesIO

    from openpyxl import load_workbook

    summaries: list[dict[str, object]] = []
    for file_name, data in sorted(workbooks.items()):
        summary: dict[str, object] = {
            "file_name": file_name,
            "byte_size": len(data),
            "format": "xlsx",
        }
        try:
            workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
            sheets = []
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                headers = [str(value) if value is not None else "" for value in rows[0]] if rows else []
                sheets.append(
                    {
                        "sheet_name": sheet.title,
                        "headers": headers,
                        "row_count": max(len(rows) - 1, 0),
                    }
                )
            workbook.close()
            if sheets:
                summary["sheets"] = sheets
        except Exception:
            pass
        summaries.append(summary)
    return summaries


def _target_system(client: object) -> str:
    software = getattr(client, "accounting_software", None) or "qbs"
    key = normalize_system_key(str(software))
    if key in {"qbs", "xero", "autocount", "sql_account"}:
        return key
    return "qbs"


def _doc_posted(doc: dict[str, Any], *, path: str, doc_type: str) -> bool:
    if doc.get("status") == "error":
        return False
    if not doc.get("lines"):
        return False
    notes = str(doc.get("notes") or "")
    if notes.lower().startswith("error"):
        return False
    return True


def build_batch_result(
    file_results: list[LightFileResult],
    *,
    client: object,
    source_files: list[str],
    missing_files: list[str],
    blocked_reason: str | None = None,
    llm_call_count: int = 0,
    models_used: list[str] | None = None,
    strong_model_used: bool = False,
    documents_skipped_before_llm: int = 0,
    elapsed_ms: int | None = None,
    credits: CreditSummary | None = None,
) -> BatchResult:
    """Convert light fan-out/fan-in results into :class:`BatchResult`."""
    system = _target_system(client)
    per_file: list[dict[str, object]] = []
    posted_documents: list[dict[str, object]] = []
    skipped_documents: list[dict[str, object]] = []
    review_requests = []
    soft_warnings = []
    export_rows: list[dict[str, object]] = []
    workbook_sheets: list[dict[str, Any]] = []
    processed_docs_for_status: list[LightFileResult] = []

    for result in file_results:
        file_name = Path(result.path).name
        if result.status == "error":
            per_file.append(
                {
                    "path": result.path,
                    "file_name": file_name,
                    "doc_type": result.kind,
                    "direction": None,
                    "reconciled": False,
                    "workbook": None,
                    "sheet": None,
                    "note": f"ERROR: {result.message or 'read failed'}",
                }
            )
            skipped_documents.append(per_file[-1])
            continue

        processed_docs_for_status.append(result)
        if result.kind == "bank" and result.bank_workbook:
            sheets = result.bank_workbook.get("sheets") or []
            workbook_sheets.extend(sheets)
            reconciled_all = all(s.get("reconciled") for s in sheets) if sheets else False
            per_file.append(
                {
                    "path": result.path,
                    "file_name": file_name,
                    "doc_type": "bank_statement",
                    "direction": None,
                    "reconciled": reconciled_all,
                    "workbook": "batch_export.xlsx",
                    "sheet": sheets[0].get("title") if sheets else None,
                    "note": "ok",
                }
            )
            if reconciled_all:
                posted_documents.append(
                    {
                        "path": result.path,
                        "file_name": file_name,
                        "doc_type": "bank_statement",
                        "direction": None,
                        "reconciled": True,
                        "workbook": "batch_export.xlsx",
                        "sheet": sheets[0].get("title") if sheets else None,
                        "note": "ok",
                        "account_count": len(sheets),
                    }
                )
            else:
                skipped_documents.append(per_file[-1])
            for sheet in sheets:
                title = sheet.get("title") or "Bank"
                for row in sheet.get("rows") or []:
                    if not isinstance(row, dict):
                        continue
                    export_rows.append(
                        {
                            "workbook": "batch_export.xlsx",
                            "sheet": title,
                            "source_file": file_name,
                            **row,
                        }
                    )
            continue

        doc_count = len(result.documents)
        per_file.append(
            {
                "path": result.path,
                "file_name": file_name,
                "doc_type": "invoice" if doc_count == 1 else "mixed",
                "direction": None,
                "reconciled": doc_count > 0,
                "workbook": "batch_export.xlsx" if doc_count else None,
                "sheet": None,
                "note": "ok" if doc_count else "ERROR: no documents extracted",
                "document_count": doc_count,
            }
        )
        if not doc_count:
            skipped_documents.append(per_file[-1])
            continue

        for index, doc in enumerate(result.documents, start=1):
            doc_type = str(doc.get("doc_type") or "purchase")
            document_kind = str(doc.get("document_kind") or "invoice")
            posted = _doc_posted(doc, path=result.path, doc_type=document_kind)
            summary = {
                "path": result.path,
                "file_name": file_name,
                "doc_type": document_kind,
                "direction": doc_type,
                "reconciled": posted,
                "workbook": "batch_export.xlsx",
                "sheet": "Sales" if doc_type == "sales" else "Purchase",
                "note": doc.get("notes") or ("ok" if posted else "needs review: missing lines"),
                "invoice_number": doc.get("invoice_number"),
                "invoice_date": doc.get("invoice_date"),
                "total": doc.get("grand_total"),
            }
            if posted:
                posted_documents.append(summary)
            else:
                skipped_documents.append(summary)
                hard, soft = _review_from_note(str(summary["note"]), file_name)
                review_requests.extend(hard)
                soft_warnings.extend(soft)

            if not doc.get("lines"):
                continue
            try:
                projection = project(doc, systems=[system])
            except Exception:
                continue
            sheet_result = projection.get("results", {}).get(system) or {}
            sheet_name = sheet_result.get("sheet") or summary["sheet"]
            for row in sheet_result.get("rows") or []:
                export_rows.append(
                    {
                        "workbook": "batch_export.xlsx",
                        "sheet": sheet_name,
                        "source_file": file_name,
                        "source_doc_index": index,
                        "invoice_number": doc.get("invoice_number"),
                        **row,
                    }
                )
            workbook_sheets.append(
                {
                    "title": sheet_name,
                    "columns": sheet_result.get("columns") or [],
                    "rows": sheet_result.get("rows") or [],
                }
            )

    for missing in missing_files:
        skipped_documents.append(
            {
                "path": missing,
                "file_name": Path(missing).name,
                "note": "ERROR: file not found",
            }
        )

    workbooks: dict[str, bytes] = {}
    if workbook_sheets:
        workbooks["batch_export.xlsx"] = build_workbook_bytes(workbook_sheets)

    status: BatchStatus = _determine_batch_status(
        blocked_reason=blocked_reason,
        processed_docs=_processed_doc_shims(processed_docs_for_status),
        missing_files=missing_files,
    )

    validation_summary: dict[str, object] = {}
    if blocked_reason:
        validation_summary["block_reason"] = blocked_reason
    if missing_files:
        validation_summary["missing_files"] = list(missing_files)

    client_id = getattr(client, "client_id", None) or "unknown"
    firm_id = getattr(client, "firm_id", None)

    return BatchResult(
        status=status,
        client_id=str(client_id),
        firm_id=firm_id,
        source_files=list(source_files),
        per_file=per_file,
        posted_documents=posted_documents,
        skipped_documents=skipped_documents,
        review_requests=review_requests,
        soft_warnings=merge_soft_warnings(soft_warnings),
        erp_exports=_erp_export_summaries(workbooks),
        export_rows=export_rows,
        credits=credits or CreditSummary(credit_status="not_checked"),
        models_used=list(models_used or []),
        validation_summary=validation_summary,
        llm_call_count=llm_call_count,
        strong_model_used=strong_model_used,
        elapsed_ms=elapsed_ms,
        documents_requested=len(source_files),
        documents_processed=len(posted_documents),
        documents_skipped_before_llm=documents_skipped_before_llm,
    )


def _processed_doc_shims(results: list[LightFileResult]) -> list[Any]:
    """Adapt light results for :func:`determine_batch_status`."""

    class _Shim:
        def __init__(self, note: str, reconciled: bool) -> None:
            self.note = note
            self.reconciled = reconciled

    shims: list[Any] = []
    for result in results:
        if result.status == "error":
            shims.append(_Shim(f"ERROR: {result.message}", False))
            continue
        if result.kind == "bank":
            sheets = (result.bank_workbook or {}).get("sheets") or []
            reconciled = bool(sheets) and all(s.get("reconciled") for s in sheets)
            shims.append(_Shim("ok" if reconciled else "needs review", reconciled))
            continue
        if result.documents:
            shims.append(_Shim("ok", True))
        else:
            shims.append(_Shim("ERROR: no documents extracted", False))
    return shims
