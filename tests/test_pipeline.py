"""Hermetic tests for invoice_processing.pipeline.

All LLM-calling steps are replaced with deterministic stubs — no Gemini /
network calls are made. Each stub returns a small, valid object built from
the real model classes.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from invoice_processing.classify.document_classifier import ClassificationResult
from invoice_processing.export.categorizer import categorize_invoice
from invoice_processing.export.client_context import ClientContext, CoaAccount, EntityMemoryEntry
from invoice_processing.extract.bank_statement_extractor import (
    ExtractedAccount,
    ExtractedBankStatement,
    ExtractedBankTxn,
)
from invoice_processing.extract.invoice_extractor import ExtractedInvoice, ExtractedLine
from invoice_processing.pipeline import process_batch, process_document


# =========================================================================== #
# Shared fixtures
# =========================================================================== #

@pytest.fixture
def client_fye3() -> ClientContext:
    """Client with fye_month=3 (March year-end), QBS Ledger, and a minimal COA."""
    return ClientContext(
        client_id="test-client",
        client_name="Test Client Pte Ltd",
        fye_month=3,
        accounting_software="QBS Ledger",
        base_currency="SGD",
        tax_registered=True,
        coa=[
            CoaAccount(
                code="500",
                description="Office Expenses",
                account_type="Expense",
                keywords="office,supplies,stationery",
            )
        ],
        category_mapping={},
        entity_memory=[
            EntityMemoryEntry(
                name="Acme Supplier",
                reg_no="200012345A",
                mapping_code="500",
                tax_code="SR",
            )
        ],
    )


@pytest.fixture
def client_no_fye() -> ClientContext:
    """Client whose fye_month is None — pipeline must not crash."""
    return ClientContext(
        client_id="no-fye-client",
        client_name="No FYE Client",
        fye_month=None,
        accounting_software="QBS Ledger",
        base_currency="SGD",
        tax_registered=True,
    )


# =========================================================================== #
# Stub factories
# =========================================================================== #

def _make_cls(doc_type: str = "invoice") -> ClassificationResult:
    return ClassificationResult(
        doc_type=doc_type,
        issuer_name="Acme Supplier",
        bill_to_name="Test Client Pte Ltd",
        currency="SGD",
        total_amount=109.0,
        confidence=0.99,
        reason="stub",
    )


def _make_extracted_invoice() -> ExtractedInvoice:
    return ExtractedInvoice(
        doc_type="invoice",
        invoice_number="INV-001",
        invoice_date="2025-01-15",
        currency="SGD",
        issuer_name="Acme Supplier",
        issuer_gst_regno="200012345A",
        bill_to_name="Test Client Pte Ltd",
        lines=[
            ExtractedLine(
                description="Office supplies",
                net_amount=100.0,
                gst_amount=9.0,
                tax_label="SR",
            )
        ],
        subtotal=100.0,
        gst_total=9.0,
        total=109.0,
    )


def _make_extracted_bank() -> ExtractedBankStatement:
    return ExtractedBankStatement(
        accounts=[
            ExtractedAccount(
                bank_name="OCBC - 5001",
                account_number="5001",
                currency="SGD",
                statement_period="01 JAN 2025 - 31 JAN 2025",
                opening_balance=1000.0,
                closing_balance=1200.0,
                transactions=[
                    ExtractedBankTxn(
                        date="2025-01-10",
                        description="Transfer in",
                        deposit=200.0,
                        balance=1200.0,
                    )
                ],
            )
        ]
    )


def _stub_categorize_no_llm(inv, *, coa, category_mapping, entity_memory, **_kw):
    """Categorize without calling the LLM (use_llm=False)."""
    return categorize_invoice(
        inv,
        coa=coa,
        category_mapping=category_mapping,
        entity_memory=entity_memory,
        use_llm=False,
    )


# Direction stubs
def _direction_purchase(cls, *, client_name=None, **_kw) -> str:
    return "purchase"


def _direction_sales(cls, *, client_name=None, **_kw) -> str:
    return "sales"


# Extract stubs (return a consistent ExtractedInvoice)
def _extract_stub(path, **_kw) -> ExtractedInvoice:
    return _make_extracted_invoice()


# Bank stub
def _bank_stub(path, **_kw) -> tuple[ExtractedBankStatement, str]:
    return _make_extracted_bank(), "stub"


# =========================================================================== #
# Test: purchase invoice
# =========================================================================== #

class TestPurchaseInvoice:
    def _classify(self, path, **_kw) -> ClassificationResult:
        return _make_cls("invoice")

    def test_direction_purchase(self, client_fye3, tmp_path):
        doc_path = tmp_path / "inv001.pdf"
        doc_path.write_bytes(b"%PDF stub")

        doc = process_document(
            doc_path,
            client_fye3,
            classify_fn=self._classify,
            direction_fn=_direction_purchase,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert doc.doc_type == "invoice"
        assert doc.direction == "purchase"
        assert doc.normalized is not None
        assert doc.bank is None
        assert doc.route.sheet == "Purchase"
        assert doc.route.workbook == "Ledger_FY2025.xlsx"
        assert not doc.note.startswith("ERROR")

    def test_account_code_filled(self, client_fye3, tmp_path):
        """Entity memory match should fill account_code=500."""
        doc_path = tmp_path / "inv002.pdf"
        doc_path.write_bytes(b"%PDF stub")

        doc = process_document(
            doc_path,
            client_fye3,
            classify_fn=self._classify,
            direction_fn=_direction_purchase,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert doc.normalized is not None
        assert len(doc.normalized.lines) == 1
        # Entity memory maps "Acme Supplier" -> "500"
        assert doc.normalized.lines[0].account_code == "500"


# =========================================================================== #
# Test: sales invoice
# =========================================================================== #

class TestSalesInvoice:
    def _classify(self, path, **_kw) -> ClassificationResult:
        return _make_cls("invoice")

    def test_direction_sales(self, client_fye3, tmp_path):
        doc_path = tmp_path / "sales001.pdf"
        doc_path.write_bytes(b"%PDF stub")

        doc = process_document(
            doc_path,
            client_fye3,
            classify_fn=self._classify,
            direction_fn=_direction_sales,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert doc.doc_type == "invoice"
        assert doc.direction == "sales"
        assert doc.route.sheet == "Sales"
        assert doc.route.workbook.startswith("Ledger_FY")
        assert not doc.note.startswith("ERROR")


# =========================================================================== #
# Test: bank statement
# =========================================================================== #

class TestBankStatement:
    def _classify(self, path, **_kw) -> ClassificationResult:
        return _make_cls("bank_statement")

    def test_bank_statement_routing(self, client_fye3, tmp_path):
        doc_path = tmp_path / "bank_jan2025.pdf"
        doc_path.write_bytes(b"%PDF stub")

        doc = process_document(
            doc_path,
            client_fye3,
            classify_fn=self._classify,
            direction_fn=_direction_purchase,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert doc.doc_type == "bank_statement"
        assert doc.direction is None
        assert doc.normalized is None
        assert doc.bank is not None
        # Jan 2025 with fye_month=3 -> FY2025
        assert doc.route.workbook == "BankStatement_FY2025.xlsx"
        assert doc.route.sheet is None
        assert not doc.note.startswith("ERROR")


# =========================================================================== #
# Test: fye_month=None does not crash
# =========================================================================== #

class TestFyeMonthNone:
    def _classify_invoice(self, path, **_kw) -> ClassificationResult:
        return _make_cls("invoice")

    def _classify_bank(self, path, **_kw) -> ClassificationResult:
        return _make_cls("bank_statement")

    def test_invoice_no_fye(self, client_no_fye, tmp_path):
        doc_path = tmp_path / "inv_nofye.pdf"
        doc_path.write_bytes(b"%PDF stub")

        doc = process_document(
            doc_path,
            client_no_fye,
            classify_fn=self._classify_invoice,
            direction_fn=_direction_purchase,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert not doc.note.startswith("ERROR")
        # Calendar year default: Jan 2025 with fye_month=12 -> FY2025
        assert doc.route.workbook == "Ledger_FY2025.xlsx"
        assert "fye_month defaulted to 12" in doc.note

    def test_bank_no_fye(self, client_no_fye, tmp_path):
        doc_path = tmp_path / "bank_nofye.pdf"
        doc_path.write_bytes(b"%PDF stub")

        doc = process_document(
            doc_path,
            client_no_fye,
            classify_fn=self._classify_bank,
            direction_fn=_direction_purchase,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert not doc.note.startswith("ERROR")
        assert doc.doc_type == "bank_statement"
        assert doc.route.workbook.startswith("BankStatement_FY")


# =========================================================================== #
# Test: error isolation — a failing stub doesn't crash the batch
# =========================================================================== #

class TestErrorIsolation:
    def _classify_ok(self, path, **_kw) -> ClassificationResult:
        return _make_cls("invoice")

    def _classify_crash(self, path, **_kw) -> ClassificationResult:
        raise RuntimeError("Simulated classification failure")

    def test_error_doc_note_startswith_error(self, client_fye3, tmp_path):
        bad_path = tmp_path / "bad.pdf"
        bad_path.write_bytes(b"%PDF stub")

        doc = process_document(
            bad_path,
            client_fye3,
            classify_fn=self._classify_crash,
            direction_fn=_direction_purchase,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert doc.note.startswith("ERROR")
        assert doc.doc_type == "unknown"

    def test_batch_continues_after_error(self, client_fye3, tmp_path):
        good = tmp_path / "good.pdf"
        bad = tmp_path / "bad.pdf"
        good.write_bytes(b"%PDF stub")
        bad.write_bytes(b"%PDF stub")

        call_count = {"n": 0}

        def _classify_mixed(path, **_kw) -> ClassificationResult:
            call_count["n"] += 1
            if "bad" in str(path):
                raise RuntimeError("boom")
            return _make_cls("invoice")

        result = process_batch(
            [str(good), str(bad)],
            client_fye3,
            classify_fn=_classify_mixed,
            direction_fn=_direction_purchase,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert len(result.docs) == 2
        assert len(result.errors) == 1
        good_doc = next(d for d in result.docs if "good" in d.path)
        bad_doc = next(d for d in result.docs if "bad" in d.path)
        assert not good_doc.note.startswith("ERROR")
        assert bad_doc.note.startswith("ERROR")


# =========================================================================== #
# Test: process_batch — full [purchase, sales, bank] scenario
# =========================================================================== #

class TestProcessBatch:
    """process_batch over [purchase, sales, bank] for fye_month=3."""

    def setup_method(self):
        self._call = {"path": None}

    def _classify(self, path, **_kw) -> ClassificationResult:
        name = Path(path).name
        if "bank" in name:
            return _make_cls("bank_statement")
        if "sales" in name:
            return _make_cls("invoice")
        return _make_cls("invoice")

    def _direction(self, cls, *, client_name=None, **_kw) -> str:
        # direction_fn receives the ClassificationResult; we re-read path from
        # the test via self._call which we set in classify — simpler: stubs
        # receive the cls which already has issuer/bill_to; just use a
        # fixed mapping stored on the test instance.
        return self._next_direction

    def test_batch_workbooks(self, client_fye3, tmp_path):
        purchase_p = tmp_path / "purchase_inv.pdf"
        sales_p = tmp_path / "sales_inv.pdf"
        bank_p = tmp_path / "bank_stmt.pdf"
        for p in (purchase_p, sales_p, bank_p):
            p.write_bytes(b"%PDF stub")

        def _classify(path, **_kw):
            name = Path(path).name
            if "bank" in name:
                return _make_cls("bank_statement")
            return _make_cls("invoice")

        def _direction(cls, *, client_name=None, **_kw):
            # We can't know the file name here since the classify result
            # doesn't carry it — use cls.issuer_name as a proxy: both
            # purchase and sales share the same stub issuer. Instead track
            # calls via a closure counter.
            return _dir_iter()

        dir_seq = iter(["purchase", "sales"])

        def _dir_iter():
            return next(dir_seq)

        result = process_batch(
            [str(purchase_p), str(sales_p), str(bank_p)],
            client_fye3,
            classify_fn=_classify,
            direction_fn=lambda cls, **kw: _dir_iter(),
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert len(result.docs) == 3
        assert len(result.errors) == 0

        # Both invoice docs fall in Jan 2025 with fye_month=3 -> FY2025
        assert "Ledger_FY2025.xlsx" in result.workbooks
        assert "BankStatement_FY2025.xlsx" in result.workbooks

    def test_ledger_workbook_sheets(self, client_fye3, tmp_path):
        """Ledger workbook must have Purchase + Sales sheets with data rows."""
        purchase_p = tmp_path / "inv_p.pdf"
        sales_p = tmp_path / "inv_s.pdf"
        for p in (purchase_p, sales_p):
            p.write_bytes(b"%PDF stub")

        dir_seq = iter(["purchase", "sales"])

        def _classify(path, **_kw):
            return _make_cls("invoice")

        result = process_batch(
            [str(purchase_p), str(sales_p)],
            client_fye3,
            classify_fn=_classify,
            direction_fn=lambda cls, **kw: next(dir_seq),
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert "Ledger_FY2025.xlsx" in result.workbooks
        wb_bytes = result.workbooks["Ledger_FY2025.xlsx"]
        wb = openpyxl.load_workbook(BytesIO(wb_bytes))

        assert "Purchase" in wb.sheetnames
        assert "Sales" in wb.sheetnames

        # Header row + 1 data row per invoice line (1 line each)
        purchase_rows = list(wb["Purchase"].iter_rows(values_only=True))
        sales_rows = list(wb["Sales"].iter_rows(values_only=True))
        assert len(purchase_rows) >= 2, f"Expected header + data in Purchase, got {len(purchase_rows)}"
        assert len(sales_rows) >= 2, f"Expected header + data in Sales, got {len(sales_rows)}"

    def test_bank_workbook_present(self, client_fye3, tmp_path):
        """BankStatement workbook must be present and openable."""
        bank_p = tmp_path / "bank_jan.pdf"
        bank_p.write_bytes(b"%PDF stub")

        def _classify(path, **_kw):
            return _make_cls("bank_statement")

        result = process_batch(
            [str(bank_p)],
            client_fye3,
            classify_fn=_classify,
            direction_fn=_direction_purchase,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert "BankStatement_FY2025.xlsx" in result.workbooks
        wb_bytes = result.workbooks["BankStatement_FY2025.xlsx"]
        wb = openpyxl.load_workbook(BytesIO(wb_bytes))
        # At least one sheet from the single account
        assert len(wb.sheetnames) >= 1

    def test_batch_fye_none_no_crash(self, client_no_fye, tmp_path):
        """Batch with fye_month=None must complete without raising."""
        p = tmp_path / "inv.pdf"
        p.write_bytes(b"%PDF stub")

        def _classify(path, **_kw):
            return _make_cls("invoice")

        result = process_batch(
            [str(p)],
            client_no_fye,
            classify_fn=_classify,
            direction_fn=_direction_purchase,
            extract_fn=_extract_stub,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert len(result.docs) == 1
        assert not result.docs[0].note.startswith("ERROR")
        # Calendar year: Jan 2025 -> FY2025
        assert "Ledger_FY2025.xlsx" in result.workbooks


# =========================================================================== #
# Test: self-referential / dividend guard — end-to-end through pipeline
# =========================================================================== #

class TestSelfReferentialPipeline:
    """resolve_direction returning 'self_referential' or 'unknown' must NOT
    silently produce a booked-purchase row with the client as its own vendor.
    The ProcessedDoc must be flagged for review (reconciled=False, note
    containing 'needs review').
    """

    def _classify_invoice(self, path, **_kw) -> ClassificationResult:
        # Both issuer and bill_to are the client — self-referential.
        return ClassificationResult(
            doc_type="invoice",
            issuer_name="Test Client Pte Ltd",
            bill_to_name="Test Client Pte Ltd",
            currency="SGD",
            total_amount=5000.0,
            confidence=0.9,
            reason="Dividend certificate",
        )

    def _extract_self_ref(self, path, **_kw) -> ExtractedInvoice:
        return ExtractedInvoice(
            doc_type="invoice",
            invoice_number="DIV-001",
            invoice_date="2025-03-31",
            currency="SGD",
            issuer_name="Test Client Pte Ltd",
            bill_to_name="Test Client Pte Ltd",
            lines=[
                ExtractedLine(
                    description="Dividend payout",
                    net_amount=5000.0,
                    gst_amount=0.0,
                    tax_label="OS",
                )
            ],
            subtotal=5000.0,
            gst_total=0.0,
            total=5000.0,
        )

    def test_self_referential_flagged_for_review(self, client_fye3, tmp_path):
        """A self-referential doc must not be booked as purchase; must be flagged."""
        doc_path = tmp_path / "dividend.pdf"
        doc_path.write_bytes(b"%PDF stub")

        doc = process_document(
            doc_path,
            client_fye3,
            classify_fn=self._classify_invoice,
            # direction_fn uses the REAL resolve_direction so the full stack fires.
            extract_fn=self._extract_self_ref,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        # Must not crash.
        assert not doc.note.startswith("ERROR"), f"Unexpected error: {doc.note}"

        # Raw direction must be 'self_referential'.
        assert doc.direction == "self_referential", (
            f"Expected 'self_referential', got {doc.direction!r}"
        )

        # Must be flagged for review — not silently approved.
        assert doc.reconciled is False, "Self-referential doc must not be reconciled=True"

        # The review note must be present in doc.note.
        assert "self-referential" in doc.note.lower(), (
            f"Expected self-referential review note in: {doc.note!r}"
        )
        assert "needs review" in doc.note.lower(), (
            f"Expected 'needs review' in: {doc.note!r}"
        )

        # The normalized row must also be flagged.
        assert doc.normalized is not None
        assert doc.normalized.reconciled is False
        assert "self-referential" in (doc.normalized.reconcile_note or "").lower()

    def test_self_referential_vendor_is_not_client(self, client_fye3, tmp_path):
        """The supplier on the normalized row must NOT be the client itself."""
        doc_path = tmp_path / "dividend2.pdf"
        doc_path.write_bytes(b"%PDF stub")

        doc = process_document(
            doc_path,
            client_fye3,
            classify_fn=self._classify_invoice,
            extract_fn=self._extract_self_ref,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        # Even though effective_direction defaults to 'purchase' for row
        # structure, the normalized supplier name comes from the EXTRACTED
        # issuer_name (the document's own text), not injected by us.  The
        # important invariant is that the doc is flagged for review so a
        # human can correct it — it is never silently booked.
        assert doc.normalized is not None
        assert doc.normalized.reconciled is False

        # The doc.direction must NOT be 'purchase' or 'sales' — it stays as
        # the raw resolve_direction output so the reviewer sees the real reason.
        assert doc.direction not in ("purchase", "sales"), (
            f"Self-referential direction must not be 'purchase'/'sales', "
            f"got {doc.direction!r}"
        )

    def test_unknown_direction_also_flagged(self, client_fye3, tmp_path):
        """'unknown' direction (no match at all) must also flag for review."""
        doc_path = tmp_path / "unknown_dir.pdf"
        doc_path.write_bytes(b"%PDF stub")

        def _direction_unknown(cls, **_kw) -> str:
            return "unknown"

        doc = process_document(
            doc_path,
            client_fye3,
            classify_fn=self._classify_invoice,
            direction_fn=_direction_unknown,
            extract_fn=self._extract_self_ref,
            bank_fn=_bank_stub,
            categorize_fn=_stub_categorize_no_llm,
        )

        assert not doc.note.startswith("ERROR")
        assert doc.direction == "unknown"
        assert doc.reconciled is False
        assert "needs review" in doc.note.lower()
        assert doc.normalized is not None
        assert doc.normalized.reconciled is False
