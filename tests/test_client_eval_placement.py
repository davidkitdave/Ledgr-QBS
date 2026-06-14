"""TDD tests for COA placement-accuracy metric (Task 1 of extraction-accuracy plan).

Per ADR-0006 a client's COA codes are often blank by design — the QBS exporter
keys by *description*. So the placement accuracy is judged by **description
match**, not code presence.

What we measure: for each produced line, did the engine pick an account whose
description matches the ground-truth expected description for that
(vendor, description) pair? Clients with no ground-truth ledger are N/A (not
failures).

These tests pin:
- `load_ground_truth_ledger` parses the Sales/Purchase sheets of a
  ``<Client> - Ledger_FY*.xlsx`` and returns a normalised lookup.
- `score_placement` returns a `PlacementResult` with per-line 1/0/N/A
  classification using a case-insensitive, whitespace-collapsed, fuzzy
  (token-set ratio >= 0.85) match.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from eval.ledger_eval import (
    PlacementResult,
    load_ground_truth_ledger,
    score_placement,
)


# --------------------------------------------------------------------------- #
# Fixtures: build a small in-memory ground-truth ledger
# --------------------------------------------------------------------------- #

def _write_gt_workbook(path: Path) -> None:
    """Build a minimal ground-truth ledger: 2 sales + 2 purchase rows.

    The "Account Code / COA" column is intentionally LEFT BLANK in real
    ground-truth (per ADR-0006 — codes are blank by design, descriptions key
    the match). For the first iteration we additionally populate it with a
    human-curated expected description so we can score a *known* mapping.
    """
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Sales sheet
    ws_s = wb.create_sheet("Sales")
    ws_s.append([
        "Invoice Date", "Invoice Number", "Customer Name", "Description",
        "Source Amount", "Currency", "Currency Rate", "Amount", "Tax Amount",
        "Total", "Account Code / COA", "Processing Date", "Source File ID",
    ])
    ws_s.append([
        "01/06/2025", "INV-1", "Acme Corp", "Consulting services",
        1000.0, "SGD", 1.0, 1000.0, 0.0, 1000.0,
        "Service Revenue", "2025-06-01 10:00:00", "acme-inv1.pdf",
    ])
    ws_s.append([
        "15/07/2025", "INV-2", "Beta  Pte  Ltd", "Retainer Fee",
        2000.0, "SGD", 1.0, 2000.0, 0.0, 2000.0,
        "Retainer Revenue", "2025-07-15 10:00:00", "beta-inv2.pdf",
    ])

    # Purchase sheet
    ws_p = wb.create_sheet("Purchase")
    ws_p.append([
        "Invoice Number", "Invoice Date", "Vendor Name", "Entity Tax ID",
        "Description", "Source Amount", "Currency", "Currency Rate",
        "Sub Total", "Tax Amount", "Total Amount", "Account Code / COA",
        "Processing Date", "Source File ID",
    ])
    ws_p.append([
        "PO-1", "02/06/2025", "Office Depot", "201234567A",
        "Stationery", 50.0, "SGD", 1.0, 50.0, 0.0, 50.0,
        "Office Supplies", "2025-06-02 10:00:00", "od-po1.pdf",
    ])
    ws_p.append([
        "PO-2", "20/07/2025", "Grab  Singapore", None,
        "Transport for client meeting", 35.0, "SGD", 1.0, 35.0, 0.0, 35.0,
        "Travel & Transport", "2025-07-20 10:00:00", "grab-po2.pdf",
    ])

    wb.save(str(path))
    wb.close()


@pytest.fixture
def gt_path(tmp_path: Path) -> Path:
    p = tmp_path / "TestClient - Ledger_FY2025.xlsx"
    _write_gt_workbook(p)
    return p


# --------------------------------------------------------------------------- #
# Tests
# --------------------------------------------------------------------------- #

def test_load_ground_truth_ledger_returns_normalised_lookup(gt_path: Path):
    """GT loader returns a dict keyed by (vendor_norm, description_norm)."""
    gt = load_ground_truth_ledger(gt_path)
    assert ("acme corp", "consulting services") in gt
    assert gt[("acme corp", "consulting services")] == ["Service Revenue"]
    # Vendor / description are whitespace-collapsed AND case-insensitive
    assert ("beta pte ltd", "retainer fee") in gt
    assert ("office depot", "stationery") in gt
    assert ("grab singapore", "transport for client meeting") in gt
    # The four ground-truth rows all produced entries
    assert len(gt) == 4


def test_load_ground_truth_ledger_returns_empty_when_no_sheets(tmp_path: Path):
    """A workbook with no Sales/Purchase sheets yields an empty dict (N/A)."""
    p = tmp_path / "EmptyLedger.xlsx"
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    wb.create_sheet("Sys_Config").append(["K", "V"])
    wb.save(str(p))
    wb.close()
    assert load_ground_truth_ledger(p) == {}


def test_load_ground_truth_ledger_skips_blank_account_column(tmp_path: Path):
    """Rows whose Account Code / COA cell is blank are still keyed in the lookup
    but map to an empty expected list — they are not failures, they are N/A."""
    p = tmp_path / "PartialLedger.xlsx"
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet("Purchase")
    ws.append([
        "Invoice Number", "Invoice Date", "Vendor Name", "Description",
        "Total Amount", "Account Code / COA",
    ])
    ws.append(["A", "01/01/2025", "Vendor X", "Thing", 10.0, None])
    ws.append(["B", "02/01/2025", "Vendor Y", "Other", 20.0, "Cost of Sales"])
    wb.save(str(p))
    wb.close()
    gt = load_ground_truth_ledger(p)
    assert gt[("vendor x", "thing")] == []            # blank column → empty expected
    assert gt[("vendor y", "other")] == ["Cost of Sales"]


def test_score_placement_counts_matches(gt_path: Path):
    """All-correct produced lines score 1.0; N/A expected rows are excluded."""
    gt = load_ground_truth_ledger(gt_path)
    # Two produced lines, both under the correct account description.
    produced = [
        # (vendor, description, account_description)
        ("Acme Corp",  "Consulting services",  "Service Revenue"),
        ("Beta  Pte  Ltd", "Retainer Fee",  "Retainer Revenue"),
    ]
    res: PlacementResult = score_placement(produced, gt)
    assert res.scored == 2
    assert res.correct == 2
    assert res.missed == 0
    assert res.na == 0
    assert res.rate == 1.0


def test_score_placement_counts_misses(gt_path: Path):
    """A line whose chosen account description does not match the expected
    counts as a miss (correct=0, missed=1)."""
    gt = load_ground_truth_ledger(gt_path)
    produced = [
        ("Acme Corp", "Consulting services", "Sundry Income"),  # wrong
        ("Beta  Pte  Ltd", "Retainer Fee",  "Retainer Revenue"),  # right
    ]
    res = score_placement(produced, gt)
    assert res.scored == 2
    assert res.correct == 1
    assert res.missed == 1
    assert res.rate == 0.5


def test_score_placement_treats_unexpected_vendor_as_n_a(gt_path: Path):
    """A produced (vendor, description) not in the GT is N/A, not a miss."""
    gt = load_ground_truth_ledger(gt_path)
    produced = [
        ("Acme Corp", "Consulting services", "Service Revenue"),  # scored
        ("Unknown Vendor", "Anything", "Whatever"),                # N/A
    ]
    res = score_placement(produced, gt)
    assert res.scored == 1
    assert res.na == 1
    assert res.correct == 1
    assert res.missed == 0


def test_score_placement_treats_blank_gt_account_as_n_a(tmp_path: Path):
    """A produced line whose GT row has no expected account description is N/A,
    regardless of what the engine picked — we cannot grade what isn't there."""
    p = tmp_path / "NoCodeLedger.xlsx"
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet("Purchase")
    ws.append(["Vendor Name", "Description", "Total Amount", "Account Code / COA"])
    ws.append(["Vendor X", "Thing", 10.0, None])  # no expected account
    wb.save(str(p))
    wb.close()
    gt = load_ground_truth_ledger(p)
    produced = [("Vendor X", "Thing", "Anything Goes")]
    res = score_placement(produced, gt)
    assert res.scored == 0
    assert res.na == 1
    assert res.rate == 0.0


def test_score_placement_fuzzy_description_match(gt_path: Path):
    """Fuzzy match: produced description with a small token-set diff (>= 0.85
    token-set ratio) counts as a match, paired with a fuzzy account too."""
    gt = load_ground_truth_ledger(gt_path)
    # "consulting services" vs "Consulting  Services" — same tokens, differ only in whitespace/casing
    produced = [
        ("ACME CORP",  "Consulting  Services",  "Service Revenue"),
    ]
    res = score_placement(produced, gt)
    assert res.scored == 1
    assert res.correct == 1


def test_score_placement_no_ground_truth_yields_zero_scored(gt_path: Path):
    """When *gt* is empty, every line is N/A — rate is 0.0 (not a failure)."""
    res = score_placement([("V", "D", "Account")], {})
    assert res.scored == 0
    assert res.na == 1
    assert res.rate == 0.0


def test_placement_result_defaults():
    """PlacementResult exposes the four counters the eval prints."""
    r = PlacementResult()
    assert r.scored == 0
    assert r.correct == 0
    assert r.missed == 0
    assert r.na == 0
    assert r.rate == 0.0
