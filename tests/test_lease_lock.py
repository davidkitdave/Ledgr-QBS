"""Hermetic tests for ``FirestoreLeaseLock`` (WS5b — cross-instance ledger lock).

NO real Firestore is imported and NO real sleeping occurs: the lease takes an
injected ``firestore_ns`` (``tests._fake_firestore.FakeFirestoreNs``), an injected
``sleep`` (records calls instead of blocking), an injected ``now`` (a controllable
monotonic for the wait-deadline) and an injected ``rng`` (deterministic backoff).
The fake "server clock" backing snapshot ``read_time``/``update_time`` is advanced
explicitly so staleness/takeover can be exercised without wall-clock time.

Each test pins one WS5b failure mode:
  1. lease blocks a second acquirer until release
  2. stale lease takeover (clock advanced past TTL)
  3. release is holder-scoped (late release by a reclaimed holder is a no-op)
  4. acquire timeout raises LOUD and writes nothing
  5. Firestore error during acquire fails CLOSED (propagates, no critical section)
  6. two instances sharing one client serialize a concurrent append (no lost rows)
  7. idempotent retry after a crash-before-release does not double-append
"""

from __future__ import annotations

import io
import random
import threading

import pytest
from openpyxl import load_workbook

from accounting_agents.lease_lock import (
    DEFAULT_LEASE_SECONDS,
    FirestoreLeaseLock,
    LeaseAcquireTimeout,
)
from accounting_agents.ledger_store import SlackLedgerStore
from tests._fake_firestore import FakeFirestore, FakeFirestoreNs
from tests.test_ledger_store import FakeSlackClient, _row


def _lease(
    db: FakeFirestore,
    *,
    instance_id: str = "rev-A",
    now=None,
    sleep=None,
    rng_seed: int = 0,
) -> FirestoreLeaseLock:
    """Build a fully-injected lease (no real firestore / sleep / clock)."""
    sleeps: list[float] = []

    def _record_sleep(seconds: float) -> None:
        sleeps.append(seconds)

    lock = FirestoreLeaseLock(
        db,
        instance_id=instance_id,
        firestore_ns=FakeFirestoreNs(),
        sleep=sleep or _record_sleep,
        now=now or (lambda: 0.0),
        rng=random.Random(rng_seed),
    )
    lock._recorded_sleeps = sleeps  # type: ignore[attr-defined]
    return lock


# --------------------------------------------------------------------------- #
# 1. lease blocks a second acquirer until release
# --------------------------------------------------------------------------- #
def test_lease_blocks_second_acquirer():
    db = FakeFirestore()
    a = _lease(db, instance_id="A")
    b = _lease(db, instance_id="B")

    token_a = a.acquire("c1", "2026")
    assert token_a.startswith("A:")

    # B cannot win while A holds a live lease; cap B's wait so it raises fast.
    b._max_wait_seconds = 0.0  # deadline already passed after the first failed try
    with pytest.raises(LeaseAcquireTimeout):
        b.acquire("c1", "2026")

    # After A releases, B wins.
    a.release("c1", "2026", token_a)
    token_b = b.acquire("c1", "2026")
    assert token_b.startswith("B:")
    assert token_b != token_a


# --------------------------------------------------------------------------- #
# 2. stale lease takeover (server clock advanced past TTL)
# --------------------------------------------------------------------------- #
def test_stale_lease_takeover():
    db = FakeFirestore()
    a = _lease(db, instance_id="A")
    b = _lease(db, instance_id="B")

    token_a = a.acquire("c1", "2026")

    # A "dies" without releasing. Advance the fake SERVER clock past the TTL so
    # the staleness predicate (read_time - update_time > lease_seconds) trips.
    db.clock.advance(DEFAULT_LEASE_SECONDS + 1)

    token_b = b.acquire("c1", "2026")  # reclaims the stale lease
    assert token_b.startswith("B:")
    assert token_b != token_a


# --------------------------------------------------------------------------- #
# 3. release is holder-scoped
# --------------------------------------------------------------------------- #
def test_release_is_holder_scoped():
    db = FakeFirestore()
    a = _lease(db, instance_id="A")
    b = _lease(db, instance_id="B")

    token_a = a.acquire("c1", "2026")
    db.clock.advance(DEFAULT_LEASE_SECONDS + 1)
    token_b = b.acquire("c1", "2026")  # B reclaims the stale lease

    # A revives and calls release with its OLD token — must NOT delete B's lease.
    a.release("c1", "2026", token_a)

    # B's lease survives: a fresh acquirer C is still blocked.
    c = _lease(db, instance_id="C")
    c._max_wait_seconds = 0.0
    with pytest.raises(LeaseAcquireTimeout):
        c.acquire("c1", "2026")

    # And B can still cleanly release its own lease.
    b.release("c1", "2026", token_b)
    assert c.acquire("c1", "2026").startswith("C:")


# --------------------------------------------------------------------------- #
# 4. acquire timeout raises LOUD and writes nothing
# --------------------------------------------------------------------------- #
def test_acquire_timeout_raises_loud():
    db = FakeFirestore()
    holder = _lease(db, instance_id="HOLD")
    holder.acquire("c1", "2026")  # permanently held, kept fresh (clock never advances)

    # A controllable monotonic that ticks forward each call so the deadline elapses.
    ticks = iter([0.0, 0.0, 50.0, 130.0, 130.0, 130.0])

    waiter = FirestoreLeaseLock(
        db,
        instance_id="WAIT",
        firestore_ns=FakeFirestoreNs(),
        sleep=lambda s: None,
        now=lambda: next(ticks),
        rng=random.Random(0),
        max_wait_seconds=120,
    )

    store_before = dict(db._store)
    with pytest.raises(LeaseAcquireTimeout):
        waiter.acquire("c1", "2026")

    # No new lease doc / no mutation landed for the waiter.
    assert db._store == store_before


# --------------------------------------------------------------------------- #
# 5. Firestore error during acquire fails CLOSED
# --------------------------------------------------------------------------- #
def test_firestore_unavailable_fails_closed():
    db = FakeFirestore()
    db.transaction_raises = True  # every txn.set raises → simulate firestore down
    lease = _lease(db, instance_id="A")

    entered_critical_section = {"hit": False}
    with pytest.raises(RuntimeError) as exc:
        token = lease.acquire("c1", "2026")
        entered_critical_section["hit"] = True  # unreachable
        lease.release("c1", "2026", token)

    assert "unavailable" in str(exc.value)
    assert entered_critical_section["hit"] is False
    # Fail-closed: nothing written.
    assert db._store == {}


# --------------------------------------------------------------------------- #
# 6. two instances, one client, serialize a concurrent append (no lost rows)
# --------------------------------------------------------------------------- #
def test_two_instances_one_client_serialize():
    slack = FakeSlackClient()
    db = FakeFirestore()  # ONE shared Firestore → shared lease docs

    # Two distinct Cloud Run instances, each with its OWN in-process lock + lease,
    # sharing the same Firestore and Slack file store.
    store_a = SlackLedgerStore(
        db, opener=slack.opener(), lease=_lease(db, instance_id="rev-A")
    )
    store_b = SlackLedgerStore(
        db, opener=slack.opener(), lease=_lease(db, instance_id="rev-B")
    )

    barrier = threading.Barrier(2)
    errors: list[BaseException] = []

    def _append(store: SlackLedgerStore, doc_key: str, desc: str) -> None:
        try:
            barrier.wait()
            store.append_rows(
                client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
                software="qbs", kind="invoice",
                batches=[{"sheet": "Purchase", "doc_key": doc_key, "rows": [_row(desc)]}],
            )
        except BaseException as exc:  # noqa: BLE001 - surface to assertion
            errors.append(exc)

    t1 = threading.Thread(target=_append, args=(store_a, "A:Purchase:INV-1", "alpha"))
    t2 = threading.Thread(target=_append, args=(store_b, "B:Purchase:INV-2", "beta"))
    t1.start()
    t2.start()
    t1.join()
    t2.join()

    assert not errors, f"concurrent append raised: {errors}"

    # Both doc_keys serialized in → both rows present, none lost.
    ptr = store_a.get_pointer("c1", "2026")
    data = slack.files[ptr["slack_file_id"]]
    wb = load_workbook(io.BytesIO(data))
    descs = [c for row in wb["Purchase"].iter_rows(min_row=2, values_only=True) for c in row]
    assert "alpha" in descs and "beta" in descs
    seen = set(ptr["seen_doc_keys"])
    assert {"A:Purchase:INV-1", "B:Purchase:INV-2"} <= seen


# --------------------------------------------------------------------------- #
# 7. idempotent retry after a crash-before-release
# --------------------------------------------------------------------------- #
def test_idempotent_retry_after_crash():
    slack = FakeSlackClient()
    db = FakeFirestore()
    lease = _lease(db, instance_id="rev-A")
    store = SlackLedgerStore(db, opener=slack.opener(), lease=lease)

    # First append uploads + sets the pointer (seen_doc_keys records the doc_key).
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )

    # Simulate the holder having "crashed" before releasing: clear the lease doc
    # by hand (as the staleness reclaim would), then RETRY the SAME doc_key.
    db._store.pop(("clients", "c1", "ledger_locks", "2026"), None)

    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )

    # Dedupe via seen_doc_keys: the retry appends nothing.
    assert result2["appended"] == 0
    assert result2["deduped"] == 1
    ptr = store.get_pointer("c1", "2026")
    data = slack.files[ptr["slack_file_id"]]
    rows = list(load_workbook(io.BytesIO(data))["Purchase"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1  # exactly one row — no double-append
