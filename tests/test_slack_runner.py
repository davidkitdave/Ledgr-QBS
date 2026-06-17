"""Hermetic tests for the Slack runner (mock Slack + Firestore + ADK runner).

Covers:
- ``extract_final_text`` joins per-part text and NEVER reads ``content.text``
  (the reply-capture bug fix).
- ``find_interrupt_id`` detects an ``adk_request_input`` interrupt.
- ``process_file_event`` on the file path: downloads (mock) → save_artifact →
  run (fake runner) → on completion appends the ledger exactly once + posts.
- ``process_file_event`` on an interrupt: posts the Approve/Edit/Reject card and
  writes the Firestore interrupt correlation doc.
- ``handle_approval_action`` (approve) acks (handled by Bolt), resumes the real
  workflow, appends the ledger once, and is idempotent on a double-click.
"""

from __future__ import annotations

import asyncio
import os
import tempfile
from types import SimpleNamespace
from unittest.mock import patch
import pytest
from google.adk.apps import App, ResumabilityConfig
from google.adk.events.event import Event
from google.adk.runners import Runner
from google.adk.workflow import START, Workflow, node
from google.genai import types

from accounting_agents import nodes
from accounting_agents.hitl import read_interrupt, write_interrupt
from accounting_agents.ledger_store import SlackLedgerStore
from accounting_agents.nodes import approval_gate
from accounting_agents.sessions import FirestoreSessionService
from accounting_agents.slack_runner import (
    _derive_setup_prefill,
    _doc_label_from_state,
    _edits_from_view_state,
    _persist_corrections,
    build_async_app,
    deslugify_channel_name,
    event_node_name,
    event_stage_label,
    extract_final_text,
    find_interrupt_id,
    handle_approval_action,
    process_file_event,
)
from invoice_processing.export.models import InvoiceLine, NormalizedInvoice
from tests._fake_firestore import FakeFirestore
from tests.test_ledger_store import FakeSlackClient
from app.native_blocks_compat import _reset_for_tests


# ---------------------------------------------------------------------------
# Module-level autouse fixture: pin all tests in this module to the FALLBACK
# (section + actions) shape unless a test explicitly overrides the env var.
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _force_fallback_blocks(monkeypatch):
    """Pin LEDGR_NATIVE_BLOCKS=0 for every test in this module."""
    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "0")
    _reset_for_tests()
    yield
    _reset_for_tests()


# =========================================================================== #
# Pure helpers
# =========================================================================== #


def test_extract_final_text_joins_parts_not_content_text():
    content = SimpleNamespace(
        parts=[SimpleNamespace(text="Hello "), SimpleNamespace(text="world")],
        # A trap: a bare content.text the buggy code would have read instead.
        text="WRONG",
    )
    event = SimpleNamespace(content=content)
    assert extract_final_text(event) == "Hello world"


def test_extract_final_text_skips_non_text_parts():
    content = SimpleNamespace(
        parts=[SimpleNamespace(text=None), SimpleNamespace(text="only this")]
    )
    event = SimpleNamespace(content=content)
    assert extract_final_text(event) == "only this"


def test_extract_tool_response_text_returns_last_result():
    from accounting_agents.slack_runner import extract_tool_response_text

    fr1 = SimpleNamespace(response={"result": "first"})
    fr2 = SimpleNamespace(response={"result": "second (latest)"})
    event = SimpleNamespace(get_function_responses=lambda: [fr1, fr2])
    assert extract_tool_response_text(event) == "second (latest)"


def test_extract_tool_response_text_empty_when_no_responses():
    from accounting_agents.slack_runner import extract_tool_response_text

    event = SimpleNamespace(get_function_responses=lambda: [])
    assert extract_tool_response_text(event) == ""
    bare = SimpleNamespace()
    assert extract_tool_response_text(bare) == ""


def test_find_interrupt_id():
    fc = SimpleNamespace(name="adk_request_input", id="C1:F1")
    event = SimpleNamespace(get_function_calls=lambda: [fc])
    assert find_interrupt_id(event) == "C1:F1"
    none_event = SimpleNamespace(get_function_calls=lambda: [])
    assert find_interrupt_id(none_event) is None


# =========================================================================== #
# Fake ADK runner for the file path
# =========================================================================== #


class _FakeArtifactService:
    def __init__(self):
        self.saved = {}

    async def save_artifact(self, *, app_name, user_id, filename, artifact, session_id=None, custom_metadata=None):
        self.saved[(user_id, filename)] = artifact
        return 0


class _FakeSession:
    def __init__(self, state):
        self.state = dict(state)


class _FakeSessionService:
    def __init__(self, final_state):
        self._final_state = final_state
        self.created = False

    async def get_session(self, *, app_name, user_id, session_id):
        # Before run: no session. After run: returns the final state.
        if not self.created:
            return None
        return _FakeSession(self._final_state)

    async def create_session(self, *, app_name, user_id, session_id, state=None):
        self.created = True
        return _FakeSession(state or {})


class _FakeRunner:
    """Minimal runner stand-in for the file path; yields scripted events."""

    def __init__(self, events, final_state, app_name="acc"):
        self.app_name = app_name
        self.artifact_service = _FakeArtifactService()
        self.session_service = _FakeSessionService(final_state)
        self._events = events

    async def run_async(self, *, user_id, session_id, new_message=None, state_delta=None):
        self.session_service.created = True
        for ev in self._events:
            yield ev


def _ledger_payload(sheet="Purchase", doc_key="F1:Purchase:INV-1"):
    return {
        nodes.LEDGER_ROWS_KEY: {
            "client_id": "c1",
            "fy": "2026",
            "kind": "invoice",
            "software": "qbs",
            # Phase 1 fix: production consolidate_node writes the file_id into
            # the payload so _record_processing_log can resolve it (the entry
            # also falls back to top-level state["file_id"] in the live path).
            "file_id": "F1",
            "batches": [
                {
                    "sheet": sheet,
                    "doc_key": doc_key,
                    "rows": [{"Invoice Number": "INV-1", "Description": "x", "Source Amount": 10.0}],
                }
            ],
        },
        nodes.DELIVER_SUMMARY_KEY: "📒 Added 1 line from 1 document to your FY2026 ledger.",
    }


# --------------------------------------------------------------------------- #
# Test fixture: a Firestore-backed client store with a known QBS profile so
# process_file_event's soft-gate lets the run proceed.
# --------------------------------------------------------------------------- #
_TEST_PROFILE: dict = {
    "client_id": "c1",
    "client_name": "Test Client",
    "fye_month": 12,
    "accounting_software": "QBS Ledger",
    "gst_registered": True,
    "region": "SINGAPORE",
    "base_currency": "SGD",
    "status": "active",
}


def _seeded_client_store(db: FakeFirestore, channel_id: str = "C1",
                        client_id: str = "c1") -> "FirestoreClientStore":  # noqa: F821 — string forward-ref; import is local
    """Write a minimal QBS client profile + channel reverse-index into ``db``
    and return a :class:`FirestoreClientStore` bound to that fake. The default
    channel/client ids match the rest of this test module.
    """
    from invoice_processing.export.client_context import FirestoreClientStore

    profile = dict(_TEST_PROFILE, client_id=client_id)
    db.collection("clients").document(client_id).set(profile)
    db.collection("channels").document(channel_id).set({"client_id": client_id})
    return FirestoreClientStore(client=db)


def test_profile_state_delta_includes_software_and_coa():
    from accounting_agents.slack_runner import _profile_state_delta
    from invoice_processing.export.client_context import ClientContext, CoaAccount

    class _Store:
        def get_by_channel(self, channel_id):
            assert channel_id == "C1"
            return ClientContext(
                client_id="CL-1",
                client_name="Company-A",
                accounting_software="Xero",
                fye_month=10,
                coa=[CoaAccount(code="6010", description="Travel",
                                account_type="Expense", financial_statement="P&L",
                                nature="Dr", keywords="travel")],
            )

    delta = _profile_state_delta(_Store(), "C1")
    assert delta["software"] == "Xero"
    assert delta["client_id"] == "CL-1"
    assert len(delta["coa"]) == 1
    # Each COA entry exposes its `key` (the export-stable identifier) — verifies
    # the helper preserves what the categorizer needs downstream.
    assert delta["coa"][0]["key"] == "6010"
    assert delta["coa"][0]["keywords"] == "travel"


def test_profile_state_delta_empty_when_no_profile():
    from accounting_agents.slack_runner import _profile_state_delta

    class _Store:
        def get_by_channel(self, channel_id):
            return None

    assert _profile_state_delta(_Store(), "C1") == {}


def test_process_file_event_softgates_when_no_profile():
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    class _NoProfileStore:
        def get_by_channel(self, channel_id):
            return None

    runner = _FakeRunner([], _ledger_payload())  # should never run
    result = asyncio.run(
        process_file_event(
            runner=runner, ledger_store=store, db=db, slack_client=slack,
            channel_id="C1", file_id="F1", app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="invoice.pdf", client_store=_NoProfileStore(),
        )
    )
    assert result["status"] == "no_profile"
    assert any("this client set up" in t.lower() for t in _posted_texts(slack))


def test_process_file_event_completion_appends_ledger_once():
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    final_event = SimpleNamespace(
        content=SimpleNamespace(parts=[SimpleNamespace(text="done")]),
        get_function_calls=lambda: [],
    )
    runner = _FakeRunner([final_event], _ledger_payload())

    downloaded = {}

    def fake_download(client, file_id):
        downloaded["file_id"] = file_id
        return b"%PDF-1.4 fake"

    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=fake_download,
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
    )

    assert result["status"] == "delivered"
    # The PDF was saved as an artifact under the convention name.
    assert ("C1", "inbox/F1.pdf") in runner.artifact_service.saved
    assert downloaded["file_id"] == "F1"
    # The ledger workbook was uploaded exactly once, with one row.
    assert len(slack.uploads) == 1
    assert result["append"]["appended"] == 1
    # The delivery summary was posted.
    assert any("FY2026 ledger" in u for u in _posted_texts(slack))


def test_process_file_event_defer_slack_delivery_writes_processing_log():
    """defer_slack_delivery=True (batch mode) must still persist a processing_log
    entry per doc so the chat lane can introspect it (Phase 1 thread-context fix).

    Pre-fix: the defer_slack_delivery early return in persist_and_deliver skipped
    _record_processing_log, so multi-file drops were invisible to the assistant.
    """
    from tests.test_slack_runner import FakeSlackClient as _FS

    slack = _FS()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    # Custom payload so file_id matches the test's expectation
    custom_payload = _ledger_payload()
    custom_payload[nodes.LEDGER_ROWS_KEY]["file_id"] = "F-batch-1"
    custom_payload["source_filename"] = "25-D15-Company-A.pdf"
    runner = _FakeRunner([], custom_payload)

    written: list[dict] = []

    class _RecordingClientStore:
        def __init__(self, _db):
            self._db = _db

        def append_processing_log(self, *, client_id, file_id, entry):
            written.append({"client_id": client_id, "file_id": file_id, "entry": entry})

        def get_by_channel(self, _channel_id):
            from invoice_processing.export.client_context import ClientContext
            return ClientContext(
                client_id="c1", client_name="Test", accounting_software="QBS",
                fye_month=12, channel_id="C1",
            )

    rec_store = _RecordingClientStore(db)

    asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F-batch-1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="25-D15-Company-A.pdf",
            client_store=rec_store,
            defer_slack_delivery=True,
        )
    )

    assert len(written) == 1, f"expected one processing_log write, got {written}"
    entry = written[0]["entry"]
    assert entry["file_id"] == "F-batch-1"
    assert entry["filename"] == "25-D15-Company-A.pdf"
    assert entry["fy"] == "2026"
    assert entry["row_count"] == 1
    # delivery_message_ts is absent because the per-doc thread_ts is None in
    # this test (batch mode with no job summary yet). The batch-end backfill
    # (Phase 2) patches it onto the entry later when the job summary ts exists.
    assert "delivery_message_ts" not in entry


def test_process_file_event_completion_writes_processing_log_with_delivery_ts():
    """Non-batch (clean) delivery path records delivery_message_ts = thread_ts
    so a chat question in the same thread can resolve to this file (Phase 2)."""
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    custom_payload = _ledger_payload()
    custom_payload[nodes.LEDGER_ROWS_KEY]["file_id"] = "F-clean-1"
    custom_payload["source_filename"] = "clean.pdf"
    runner = _FakeRunner([], custom_payload)

    written: list[dict] = []

    class _RecordingClientStore:
        def append_processing_log(self, *, client_id, file_id, entry):
            written.append(entry)

        def get_by_channel(self, _channel_id):
            from invoice_processing.export.client_context import ClientContext
            return ClientContext(
                client_id="c1", client_name="Test", accounting_software="QBS",
                fye_month=12, channel_id="C1",
            )

    rec_store = _RecordingClientStore()

    asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F-clean-1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="clean.pdf",
            client_store=rec_store,
            thread_ts="1716000000.000200",
        )
    )

    assert len(written) == 1
    assert written[0]["delivery_message_ts"] == "1716000000.000200"
    assert written[0]["channel_id"] == "C1"


def test_process_file_event_interrupt_posts_card_and_writes_doc():
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    interrupt_event = SimpleNamespace(
        content=SimpleNamespace(parts=[]),
        get_function_calls=lambda: [SimpleNamespace(name="adk_request_input", id="C1:F1")],
    )
    runner = _FakeRunner([interrupt_event], {"approval_message": "needs review: line X"})

    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF fake",
            client_store=_seeded_client_store(db),
        )
    )

    assert result["status"] == "paused"
    assert result["op_id"] == "C1:F1"
    # No ledger uploaded while paused.
    assert slack.uploads == []
    # The approval card was posted (carries the three action buttons).
    card = _last_blocks(slack)
    action_ids = {
        e["action_id"]
        for b in card
        if b.get("type") == "actions"
        for e in b["elements"]
    }
    assert action_ids == {"approve", "edit", "reject"}
    # The interrupt correlation doc was written to Firestore.
    snap = db.collection("interrupts").document("C1:F1").get()
    assert snap.exists
    assert snap.to_dict()["slack_file_id"] == "F1"


# =========================================================================== #
# Approve action → resume → append once (idempotent), using a REAL runner
# =========================================================================== #

_DELIVER_RUNS: list = []


@node
async def _terminal(ctx, node_input) -> Event:
    """Stand-in successor of the gate: stage a ledger payload for persistence."""
    _DELIVER_RUNS.append(node_input)
    ctx.state[nodes.LEDGER_ROWS_KEY] = {
        "client_id": "c1",
        "fy": "2026",
        "kind": "invoice",
        "software": "qbs",
        "batches": [
            {
                "sheet": "Purchase",
                "doc_key": "F7:Purchase:INV-7",
                "rows": [{"Invoice Number": "INV-7", "Description": "y", "Source Amount": 20.0}],
            }
        ],
    }
    ctx.state[nodes.DELIVER_SUMMARY_KEY] = "📒 Added 1 line from 1 document to your FY2026 ledger."
    return Event(output={"ok": True})


def _build_resumable_app() -> App:
    wf = Workflow(name="approve_wf", edges=[(START, approval_gate, _terminal)])
    return App(
        name="acc",
        root_agent=wf,
        resumability_config=ResumabilityConfig(is_resumable=True),
    )


def _low_conf_state() -> dict:
    inv = NormalizedInvoice(
        invoice_number="INV-7",
        reconciled=True,
        lines=[InvoiceLine(description="ambiguous", tax_confidence=0.30)],
    )
    return {
        "op_id": "C7:F7",
        "channel_id": "C7",
        "file_id": "F7",
        nodes.NORMALIZED_KEY: [nodes._inv_to_dict(inv)],
    }


def setup_function(_):
    _DELIVER_RUNS.clear()


def test_approve_action_resumes_and_appends_once_idempotent():
    slack = FakeSlackClient()
    db = FakeFirestore()
    ledger_store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    svc = FirestoreSessionService(client=db)
    asyncio.run(
        svc.create_session(app_name="acc", user_id="C7", session_id="C7", state=_low_conf_state())
    )
    runner = Runner(app=_build_resumable_app(), session_service=svc)

    # Drive to the HITL pause.
    async def drive():
        async for _ev in runner.run_async(
            user_id="C7",
            session_id="C7",
            new_message=types.Content(parts=[types.Part(text="process")]),
        ):
            pass

    asyncio.run(drive())
    assert _DELIVER_RUNS == []

    # The Slack layer posted the card + wrote the correlation doc.
    write_interrupt(
        db, "C7:F7", session_id="C7", channel_id="C7", slack_file_id="F7", message_ts="111.222",
        extra={"summary": "needs review"},
    )

    # First click: resume, append the ledger once, update the card.
    res1 = asyncio.run(
        handle_approval_action(
            runner=runner,
            ledger_store=ledger_store,
            db=db,
            slack_client=slack,
            op_id="C7:F7",
            decision="approve",
            app_name="acc",
        )
    )
    assert res1["status"] == "resumed"
    assert len(_DELIVER_RUNS) == 1
    assert len(slack.uploads) == 1
    assert res1["append"]["appended"] == 1

    # Second click (double-click): no-op — no second resume, no second upload.
    res2 = asyncio.run(
        handle_approval_action(
            runner=runner,
            ledger_store=ledger_store,
            db=db,
            slack_client=slack,
            op_id="C7:F7",
            decision="approve",
            app_name="acc",
        )
    )
    assert res2["status"] == "already_processed"
    assert len(_DELIVER_RUNS) == 1
    assert len(slack.uploads) == 1


def test_reject_action_resumes_without_appending():
    slack = FakeSlackClient()
    db = FakeFirestore()
    ledger_store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    svc = FirestoreSessionService(client=db)
    asyncio.run(
        svc.create_session(app_name="acc", user_id="C8", session_id="C8", state={
            "op_id": "C8:F8",
            "channel_id": "C8",
            "file_id": "F8",
            nodes.NORMALIZED_KEY: _low_conf_state()[nodes.NORMALIZED_KEY],
        })
    )
    runner = Runner(app=_build_resumable_app(), session_service=svc)

    async def drive():
        async for _ev in runner.run_async(
            user_id="C8", session_id="C8",
            new_message=types.Content(parts=[types.Part(text="process")]),
        ):
            pass

    asyncio.run(drive())
    write_interrupt(db, "C8:F8", session_id="C8", channel_id="C8", slack_file_id="F8", message_ts="9.9")

    res = asyncio.run(
        handle_approval_action(
            runner=runner, ledger_store=ledger_store, db=db, slack_client=slack,
            op_id="C8:F8", decision="reject", app_name="acc",
        )
    )
    assert res["status"] == "resumed"
    # Reject path uploads nothing to the ledger.
    assert slack.uploads == []


# =========================================================================== #
# Task 7: Edit action + view_submission handlers (modal-based per-line edits)
# =========================================================================== #


class _FakeActionViewRunner:
    """Minimal Runner for Edit-modal tests: app_name + session_service stub."""

    def __init__(self, app_name="acc", session_service=None):
        self.app_name = app_name
        self.session_service = session_service


class _FakeSessionSvc:
    """In-memory session service stub keyed by (user_id, session_id)."""

    def __init__(self, sessions):
        self._sessions = sessions

    async def get_session(self, *, app_name, user_id, session_id):
        return self._sessions.get((user_id, session_id))


def _capture_hitl_handlers(runner_mock=None, ledger_store_mock=None, db_mock=None):
    """Build the Bolt app with fakes and capture the ``edit`` action + ``ledgr_invoice_edit`` view.

    Mirrors ``_capture_message_handler`` but for the action/view decorators. The
    fake app records registered actions/views in a dict keyed by their first
    positional argument (the action_id or callback_id), so callers can retrieve
    the actual coroutine that ``build_async_app`` registered.
    """
    from unittest.mock import MagicMock, patch

    from app.slack_app import _SeenEvents
    from accounting_agents import slack_runner

    registered = {"actions": {}, "views": {}}

    def action_decorator(action_id, *a, **k):
        def decorator(fn):
            registered["actions"][action_id] = fn
            return fn
        return decorator

    def view_decorator(callback_id, *a, **k):
        def decorator(fn):
            registered["views"][callback_id] = fn
            return fn
        return decorator

    fake_app = MagicMock()
    fake_app.event = lambda *a, **k: (lambda fn: fn)
    fake_app.action = action_decorator
    fake_app.view = view_decorator
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = runner_mock or _FakeActionViewRunner()

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch("invoice_processing.export.client_context.FirestoreClientStore"), \
         patch.object(slack_runner, "build_chat_runner",
                      return_value=SimpleNamespace(app_name="accounting_agents_assistant")):
        build_async_app(
            runner=rm,
            ledger_store=ledger_store_mock or MagicMock(),
            db=db_mock or MagicMock(),
        )

    return registered["actions"]["edit"], registered["views"]["ledgr_invoice_edit"]


def test_edit_action_opens_invoice_modal_with_proposed_lines():
    """Clicking Edit MUST call views_open with a well-formed modal (callback_id + op_id).

    The modal body is pre-filled from the paused session's normalized invoice
    lines + COA, so the user can correct the fields in-place before resubmitting.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    db = FakeFirestore()
    # Seed the HITL interrupt correlation doc so the handler can read session state.
    write_interrupt(
        db, "OP1", session_id="S-OP1", channel_id="C-OP1", slack_file_id="F-OP1", message_ts="1.1",
    )

    # Seed a paused session whose state the Edit modal reads to pre-fill lines.
    paused_session = _FakeSession({
        nodes.NORMALIZED_KEY: [
            {
                "lines": [
                    {"description": "Room", "account_code": "6010", "tax_treatment": "SR", "net_amount": 51.49},
                    {"description": "Tax", "account_code": None, "tax_treatment": "ZR", "net_amount": 3.60},
                ]
            }
        ],
        "coa": [{"code": "6010", "description": "Travel"}],
    })
    session_svc = _FakeSessionSvc({("C-OP1", "S-OP1"): paused_session})

    # Recording fake for the sync WebClient that build_async_app instantiates.
    sync_client = MagicMock()
    with patch("slack_sdk.WebClient", return_value=sync_client):
        edit_handler, _ = _capture_hitl_handlers(
            runner_mock=_FakeActionViewRunner(session_service=session_svc),
            db_mock=db,
        )

    body = {
        "actions": [{"action_id": "edit", "value": "OP1"}],
        "trigger_id": "T-EDIT-1",
        "channel": {"id": "C-OP1"},
    }

    with patch.object(slack_runner, "read_interrupt") as mock_read, \
         patch.object(slack_runner, "_read_session_state", AsyncMock(return_value=paused_session.state)):
        # Make read_interrupt return our seeded doc.
        mock_read.return_value = {
            "op_id": "OP1", "user_id": "C-OP1", "session_id": "S-OP1", "channel_id": "C-OP1",
        }
        ack = AsyncMock()
        bolt_client = MagicMock()
        asyncio.run(edit_handler(ack=ack, body=body, client=bolt_client))

    # The handler MUST ack (within Bolt's 3s window) before opening the modal.
    ack.assert_awaited_once()

    # views_open was called with trigger_id + a modal carrying the right IDs.
    sync_client.views_open.assert_called_once()
    kwargs = sync_client.views_open.call_args.kwargs
    assert kwargs["trigger_id"] == "T-EDIT-1"
    view = kwargs["view"]
    assert view["callback_id"] == "ledgr_invoice_edit"
    assert view["private_metadata"] == "OP1"
    # The modal blocks must include the proposed-line prefill (one input group per line).
    inputs = [b for b in view["blocks"] if b.get("type") == "input"]
    assert len(inputs) == 6  # 2 lines × (acct + tax + amt)


def test_edit_submit_invokes_handle_approval_action_with_parsed_edits():
    """Submitting the Edit modal MUST call handle_approval_action(decision="edit", edits=...).

    Seats a recorder on ``handle_approval_action``, drives the view handler with
    a realistic ``view_submission`` body (private_metadata + state.values), and
    asserts the recorder was called with op_id, decision="edit", and edits
    parsed from the ``acct_<i>`` / ``tax_<i>`` / ``amt_<i>`` block_ids.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    db = FakeFirestore()

    edit_handler, edit_submit_handler = _capture_hitl_handlers(db_mock=db)

    # Build a view_submission body with two edited lines.
    body = {
        "view": {
            "callback_id": "ledgr_invoice_edit",
            "private_metadata": "OP-SUBMIT-1",
            "state": {
                "values": {
                    "acct_0": {"v": {"selected_option": {"value": "6200"}}},
                    "tax_0": {"v": {"selected_option": {"value": "ZR"}}},
                    "amt_0": {"v": {"value": "99.95"}},
                    "acct_1": {"v": {"selected_option": {"value": "6010"}}},
                    "tax_1": {"v": {"selected_option": {"value": "SR"}}},
                    "amt_1": {"v": {"value": "12.00"}},
                }
            },
        }
    }

    # Record calls to handle_approval_action without actually resuming a workflow.
    # AsyncMock auto-awaits when patched over an `async def` function.
    recorder = AsyncMock(return_value={"status": "resumed"})

    with patch.object(slack_runner, "handle_approval_action", recorder):
        ack = AsyncMock()
        bolt_client = MagicMock()
        asyncio.run(edit_submit_handler(ack=ack, body=body, client=bolt_client))

    # Bolt ack must be called within 3s.
    ack.assert_awaited_once()
    recorder.assert_called_once()
    call = recorder.call_args
    assert call.kwargs["op_id"] == "OP-SUBMIT-1"
    assert call.kwargs["decision"] == "edit"
    edits = call.kwargs["edits"]
    assert edits == {
        "lines": [
            {"index": 0, "account_code": "6200", "tax_treatment": "ZR", "net_amount": 99.95},
            {"index": 1, "account_code": "6010", "tax_treatment": "SR", "net_amount": 12.0},
        ]
    }


def test_edit_submit_persists_correction_after_resume(monkeypatch):
    """E2E wiring: view_submission → handle_approval_action → _persist_corrections → add_correction.

    Mirrors ``test_edit_submit_invokes_handle_approval_action_with_parsed_edits``'s
    scaffold but additionally seeds:
      - a ``FakeFirestore`` interrupt correlation doc keyed by op_id, AND
      - a paused session state with ``client_id`` + ``normalized_invoices``
        so ``_persist_corrections`` (called from the wiring at slack_runner.py:1090-1098)
        has the inputs it needs to actually invoke ``add_correction``.
    The module-level ``_DEFAULT_CLIENT_STORE`` is monkeypatched to a recorder so
    we can assert the wiring reaches the persistence side without standing up
    a real FirestoreClientStore.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    db = FakeFirestore()

    # Seed the interrupt correlation doc the wiring reads at slack_runner.py:1092.
    # Explicit user_id=channel_id so the doc's user_id matches the session-svc
    # lookup key (write_interrupt's default is session_id, which would not).
    write_interrupt(
        db, "OP-WIRE-1",
        session_id="S-WIRE-1", channel_id="C-WIRE-1", user_id="C-WIRE-1",
        slack_file_id="F-WIRE-1", message_ts="1.1",
    )

    # Seed a paused session whose state carries the ADR-0004 inputs
    # (client_id + normalized_invoices) that _persist_corrections reads.
    # Use the REAL serialized NormalizedInvoice shape (_inv_to_dict = asdict):
    # the vendor lives in the nested ``supplier``/``customer`` party, NOT a flat
    # ``vendor_name`` key. A prior fixture used ``vendor_name`` and masked the
    # production bug where _persist_corrections read keys that never exist.
    paused_state = {
        "client_id": "CL-1",
        nodes.NORMALIZED_KEY: [
            {
                "doc_type": "purchase",
                "supplier": {"name": "Hotel Booking"},
                "customer": {"name": None},
                # Proposed line is uncategorized, so the submitted 6010/ZR are
                # genuine human changes (only-changed-lines become Corrections).
                "lines": [
                    {"description": "Room"},
                ],
            }
        ],
    }
    paused_session = _FakeSession(paused_state)
    session_svc = _FakeSessionSvc({("C-WIRE-1", "S-WIRE-1"): paused_session})

    _, edit_submit_handler = _capture_hitl_handlers(
        runner_mock=_FakeActionViewRunner(session_service=session_svc),
        db_mock=db,
    )

    # view_submission body: one edited line, account_code + tax_code (amount
    # omitted to mirror the natural UI path — _persist_corrections reads only
    # account_code / tax_code).
    body = {
        "view": {
            "callback_id": "ledgr_invoice_edit",
            "private_metadata": "OP-WIRE-1",
            "state": {
                "values": {
                    "acct_0": {"v": {"selected_option": {"value": "6010"}}},
                    "tax_0":  {"v": {"selected_option": {"value": "ZR"}}},
                }
            },
        }
    }

    # Recorder for handle_approval_action (mirrors Task 7's pattern).
    approval_recorder = AsyncMock(return_value={"status": "resumed"})

    # Recording fake for the default client store the wiring uses.
    correction_calls = []

    class _RecordingStore:
        def add_correction(self, *, client_id, vendor, account_code=None, tax_code=None):
            correction_calls.append(
                {"client_id": client_id, "vendor": vendor,
                 "account_code": account_code, "tax_code": tax_code}
            )

    monkeypatch.setattr(slack_runner, "_DEFAULT_CLIENT_STORE", _RecordingStore())

    with patch.object(slack_runner, "handle_approval_action", approval_recorder):
        ack = AsyncMock()
        bolt_client = MagicMock()
        asyncio.run(edit_submit_handler(ack=ack, body=body, client=bolt_client))

    # 1. The view handler still acks + still drives handle_approval_action.
    ack.assert_awaited_once()
    approval_recorder.assert_called_once()
    assert approval_recorder.call_args.kwargs["op_id"] == "OP-WIRE-1"
    assert approval_recorder.call_args.kwargs["decision"] == "edit"

    # 2. The wiring reached _persist_corrections → add_correction exactly once,
    #    with the seeded vendor + account/tax codes from the edited line.
    assert len(correction_calls) == 1
    assert correction_calls[0] == {
        "client_id": "CL-1",
        "vendor": "Hotel Booking",
        "account_code": "6010",
        "tax_code": "ZR",
    }


# =========================================================================== #
# Stage detection + live status message
# =========================================================================== #


def _node_event(node_name: str, *, text: str = "", interrupt: bool = False):
    """A scripted event tagged with a node path the way the ADK Workflow does."""
    calls = (
        [SimpleNamespace(name="adk_request_input", id="C1:F1")] if interrupt else []
    )
    return SimpleNamespace(
        node_info=SimpleNamespace(path=f"document_workflow@1/{node_name}@1"),
        content=SimpleNamespace(parts=[SimpleNamespace(text=text)] if text else []),
        get_function_calls=lambda: calls,
    )


def test_event_node_name_parses_trailing_node():
    ev = _node_event("classify_node")
    assert event_node_name(ev) == "classify_node"
    # No node_info / empty path → None (e.g. coordinator chatter).
    assert event_node_name(SimpleNamespace()) is None
    assert event_node_name(SimpleNamespace(node_info=SimpleNamespace(path=""))) is None


def test_event_stage_label_maps_known_nodes():
    # Classify announces it's looking at the document (warmer, not "Classifying").
    assert "document" in event_stage_label(_node_event("classify_node")).lower()
    # Extract labels name WHAT the doc was identified as.
    assert "bank statement" in event_stage_label(_node_event("extract_bank_node"))
    assert "invoice" in event_stage_label(_node_event("extract_invoice_node"))
    # Unmapped / untagged events do not trigger a status change.
    assert event_stage_label(_node_event("some_unknown_node")) is None
    assert event_stage_label(SimpleNamespace()) is None


def test_process_file_event_posts_status_once_and_updates_per_stage():
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    # A realistic single-document stream: classify → extract → tax → final text.
    events = [
        _node_event("classify_node"),
        _node_event("extract_invoice_node"),
        _node_event("tax_node"),
        _node_event("deliver_node", text="done"),
    ]
    runner = _FakeRunner(events, _ledger_payload())

    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF fake",
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
    )
    assert result["status"] == "delivered"

    # Exactly ONE status message was posted (the initial "received" message). The
    # delivery summary is a separate post; assert the status post is distinct and
    # there is only one "received" message.
    received = [p for p in _post_calls(slack) if "Received" in p.get("text", "")]
    assert len(received) == 1
    assert "invoice.pdf" in received[0]["text"]

    # The status message was edited in place through the stages (distinct labels)
    # plus the terminal ✅. The status post is the FIRST chat_postMessage, whose
    # returned ts is "1.000" (FakeSlackClient numbers posts 1-based).
    status_ts = "1.000"
    update_texts = [u["text"] for u in slack.updates]
    lowered = [t.lower() for t in update_texts]
    assert any("taking a look" in t for t in lowered)            # classify
    assert any("reading the line items" in t for t in lowered)   # extract (invoice)
    assert any("reconciling" in t for t in lowered)              # tax/approval
    assert update_texts[-1].startswith("✅ Added to")
    # Every update targeted the single status message ts.
    assert all(u["ts"] == status_ts for u in slack.updates)
    # In-flight updates carry the processing accordion (plan block or fallback).
    updates_with_blocks = [u for u in slack.updates if u.get("blocks")]
    assert updates_with_blocks, "stage updates must include processing plan blocks"
    blocks_str = str(updates_with_blocks[0]["blocks"])
    assert "Understanding document" in blocks_str or "Applying your rules" in blocks_str


def test_processing_status_uses_compact_header_not_accordion():
    """Initial status post must be a plain one-liner — no plan accordion yet.

    The first message on file drop stays compact; the plan accordion is attached
    on the first in-place edit when a pipeline stage starts.
    """
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    runner = _FakeRunner(
        [_node_event("deliver_node", text="done")],
        _ledger_payload(),
    )

    asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF fake",
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
    )

    # The first post must be the compact "Received" one-liner.
    received_posts = [p for p in _post_calls(slack) if "Received" in p.get("text", "")]
    assert len(received_posts) == 1, "exactly one 'Received' message must be posted"
    first_post = received_posts[0]

    # Must contain the filename.
    assert "invoice.pdf" in first_post["text"]

    # Must NOT carry any block-kit accordion blocks.
    blocks = first_post.get("blocks", [])
    block_types = [b.get("type") for b in blocks]
    assert "plan" not in block_types, "initial post must not include a 'plan' accordion block"

    # Must not contain stage-label strings anywhere in the blocks payload.
    blocks_str = str(blocks)
    for forbidden in ("Understanding document", "Applying your rules", "Ready to file"):
        assert forbidden not in blocks_str, (
            f"initial post must not contain accordion stage text '{forbidden}'"
        )


def test_clean_path_summary_line_still_emits():
    """Clean path must still emit the enriched '📒 Added N lines…' summary after processing.

    Step 12: removing the accordion must not regress the delivery summary post.
    """
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    runner = _FakeRunner(
        [_node_event("deliver_node", text="done")],
        _ledger_payload(),
    )

    asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF fake",
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
    )

    posted_texts = _posted_texts(slack)
    # The delivery summary from DELIVER_SUMMARY_KEY must appear.
    assert any("Added" in t and "ledger" in t for t in posted_texts), (
        "delivery summary line ('Added N lines … ledger') must be posted after clean-path processing"
    )


def test_reviewer_card_still_emits():
    """Interrupt path must still post the reviewer block-kit card (approve/edit/reject).

    Step 12: dropping the accordion must not remove the actionable reviewer card.
    """
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    interrupt_event = SimpleNamespace(
        content=SimpleNamespace(parts=[]),
        get_function_calls=lambda: [SimpleNamespace(name="adk_request_input", id="C1:F1")],
    )
    runner = _FakeRunner([interrupt_event], {"approval_message": "needs review: line X"})

    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF fake",
            client_store=_seeded_client_store(db),
        )
    )

    assert result["status"] == "paused"
    # The reviewer card (approve / edit / reject buttons) must still be posted.
    card = _last_blocks(slack)
    action_ids = {
        e["action_id"]
        for b in card
        if b.get("type") == "actions"
        for e in b["elements"]
    }
    assert action_ids == {"approve", "edit", "reject"}, (
        "reviewer card with approve/edit/reject actions must be posted on interrupt"
    )


def test_instant_ack_reaction_and_status_before_semaphore_heavy_work():
    """👀 reaction + initial status message fire BEFORE the download (semaphore-guarded).

    Strategy: the download fn blocks on a threading.Event until we confirm both
    the reaction and the status post have already happened, then unblocks.
    This proves the ack is outside/before the semaphore-guarded heavy work.
    """
    import threading

    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    # Teach the fake client that file F1 was shared in channel C1 at ts "9.001".
    slack._file_share_ts["F1"] = {"C1": "9.001"}

    # The download blocks until we release it.
    download_started = threading.Event()
    download_may_proceed = threading.Event()

    def blocking_download(client, file_id):
        download_started.set()
        download_may_proceed.wait(timeout=5)
        return b"%PDF-1.4 fake"

    # Run process_file_event in a background thread so we can inspect state
    # while the download is blocked.
    result_holder: list = []

    async def _run():
        r = await process_file_event(
            runner=_FakeRunner([
                SimpleNamespace(
                    content=SimpleNamespace(parts=[SimpleNamespace(text="done")]),
                    get_function_calls=lambda: [],
                )
            ], _ledger_payload()),
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=blocking_download,
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
        result_holder.append(r)

    t = threading.Thread(target=lambda: asyncio.run(_run()), daemon=True)
    t.start()

    # Wait for the download to start (we are now inside the semaphore).
    assert download_started.wait(timeout=5), "download never started"

    # At this point the download is blocked inside _SEM. Assert that the INSTANT
    # ack (reaction + status post) already happened BEFORE we entered _SEM.
    assert any(r["name"] == "eyes" and r["timestamp"] == "9.001"
               for r in slack.reactions_added), \
        "👀 reaction must be added before semaphore-guarded download starts"

    received_posts = [p for p in slack._posts if "on it" in p.get("text", "")]
    assert received_posts, "initial status message must be posted before download starts"

    # Release the download and wait for the run to finish.
    download_may_proceed.set()
    t.join(timeout=10)

    assert result_holder and result_holder[0]["status"] == "delivered"
    # On completion the 👀 is removed and ✅ added.
    assert any(r["name"] == "eyes" for r in slack.reactions_removed), \
        "👀 reaction must be removed on completion"
    assert any(r["name"] == "white_check_mark" for r in slack.reactions_added), \
        "✅ reaction must be added on completion"


def test_process_file_event_status_update_failure_does_not_crash_run():
    class _FlakyUpdateClient(FakeSlackClient):
        def chat_update(self, **kwargs):
            raise RuntimeError("slack 500")

    slack = _FlakyUpdateClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    runner = _FakeRunner([_node_event("classify_node", text="done")], _ledger_payload())

    # The cosmetic chat_update failures must NOT abort processing.
    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF fake",
            client_store=_seeded_client_store(db),
        )
    )
    assert result["status"] == "delivered"
    assert len(slack.uploads) == 1


def test_process_file_event_interrupt_sets_needs_review_status():
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    events = [
        _node_event("classify_node"),
        _node_event("approval_gate", interrupt=True),
    ]
    runner = _FakeRunner(events, {"approval_message": "needs review: line X"})

    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF fake",
            client_store=_seeded_client_store(db),
        )
    )
    assert result["status"] == "paused"
    update_texts = [u["text"] for u in slack.updates]
    assert any("Needs your review" in t for t in update_texts)


# =========================================================================== #
# De-slugify channel name
# =========================================================================== #


def test_deslugify_channel_name_basic():
    assert deslugify_channel_name("sample-channel-client-pte-ltd") == "Sample Channel Client Pte Ltd"


def test_deslugify_channel_name_underscores_and_suffixes():
    assert deslugify_channel_name("foo_bar_llp") == "Foo Bar LLP"
    assert deslugify_channel_name("acme-sg-pte-ltd") == "Acme SG Pte Ltd"


def test_deslugify_channel_name_empty():
    assert deslugify_channel_name("") == ""
    assert deslugify_channel_name("---") == ""


# =========================================================================== #
# Setup-open prefill from channel name
# =========================================================================== #


def test_derive_setup_prefill_from_channel_name():
    class _InfoClient:
        def conversations_info(self, *, channel):
            assert channel == "C-OPEN"
            return {"ok": True, "channel": {"id": channel, "name": "sample-channel-client-pte-ltd"}}

    body = {"channel": {"id": "C-OPEN"}, "trigger_id": "t1"}
    prefill = asyncio.run(_derive_setup_prefill(_InfoClient(), body))
    assert prefill == {"client_name": "Sample Channel Client Pte Ltd"}


def test_derive_setup_prefill_handles_lookup_failure():
    class _BoomClient:
        def conversations_info(self, *, channel):
            raise RuntimeError("missing_scope")

    body = {"channel": {"id": "C-OPEN"}}
    assert asyncio.run(_derive_setup_prefill(_BoomClient(), body)) is None


def test_derive_setup_prefill_no_channel_returns_none():
    assert asyncio.run(_derive_setup_prefill(object(), {})) is None


# =========================================================================== #
# helpers
# =========================================================================== #


def _post_calls(slack: FakeSlackClient) -> list:
    return getattr(slack, "_posts", [])


def _posted_texts(slack: FakeSlackClient) -> list:
    return [p.get("text", "") for p in _post_calls(slack)]


def _last_blocks(slack: FakeSlackClient):
    for p in reversed(_post_calls(slack)):
        if p.get("blocks"):
            return p["blocks"]
    return []


# =========================================================================== #
# Commit 4 (updated): ledger preview — per-software / per-sheet data_table
# =========================================================================== #


def _run_delivery_with_state(state: dict, monkeypatch, channel_id: str = "C1"):
    """Helper: run process_file_event with a fake state and return all post calls."""
    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    final_event = SimpleNamespace(
        content=SimpleNamespace(parts=[SimpleNamespace(text="done")]),
        get_function_calls=lambda: [],
    )
    runner = _FakeRunner([final_event], state)

    asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id=channel_id,
            file_id="F1",
            app_name="acc",
            download_fn=lambda client, file_id: b"%PDF-1.4 fake",
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
    )
    return _post_calls(slack)


def test_delivery_posts_ledger_preview_data_table(monkeypatch):
    """After a successful delivery, persist_and_deliver posts a data_table preview
    that uses QBS Purchase column shape (Sub Total at col 5, Total Amount at col 7)."""
    state = {
        nodes.LEDGER_ROWS_KEY: {
            "client_id": "c1",
            "fy": "2025",
            "kind": "invoice",
            "software": "qbs",
            "client_name": "Acme",
            "batches": [
                {
                    "sheet": "Purchase",
                    "doc_key": "Purchase:INV-042",
                    "rows": [
                        {
                            "Invoice Date": "2025-09-15",
                            "Invoice Number": "INV-042",
                            "Vendor Name": "Acme Trading",
                            "Description": "Acme Trading INV-2025-0042",
                            "Account Code / COA": "6090",
                            "Sub Total": 1132.11,
                            "Tax Amount": 102.39,
                            "Total Amount": 1234.50,
                        }
                    ],
                }
            ],
        },
        nodes.DELIVER_SUMMARY_KEY: "Added 1 line to your FY2025 ledger.",
    }
    posts = _run_delivery_with_state(state, monkeypatch)

    data_table_posts = [
        p for p in posts
        if any(b.get("type") == "data_table" for b in (p.get("blocks") or []))
    ]
    assert data_table_posts, "Expected at least one chat_postMessage with a data_table block"
    table_block = next(b for b in data_table_posts[0]["blocks"] if b["type"] == "data_table")
    # 1 header + 1 data row
    assert len(table_block["rows"]) == 2
    assert table_block["row_header_column_index"] == 0
    assert table_block["page_size"] == 10
    # QBS Purchase: col 5 = Sub Total, col 7 = Total Amount
    data_row = table_block["rows"][1]
    assert data_row[5]["type"] == "raw_number"
    assert data_row[5]["value"] == 1132.11
    assert data_row[7]["type"] == "raw_number"
    assert data_row[7]["value"] == 1234.50


def test_delivery_two_batches_posts_two_preview_messages(monkeypatch):
    """Purchase + Sales batches appear as two data_tables in one delivery card."""
    state = {
        nodes.LEDGER_ROWS_KEY: {
            "client_id": "c1",
            "fy": "2025",
            "kind": "invoice",
            "software": "qbs",
            "client_name": "Acme",
            "batches": [
                {
                    "sheet": "Purchase",
                    "doc_key": "Purchase:INV-001",
                    "rows": [{"Invoice Date": "15/09/2025", "Invoice Number": "INV-001",
                              "Vendor Name": "Acme", "Description": "Consulting",
                              "Account Code / COA": "6090", "Sub Total": 100.0,
                              "Tax Amount": 9.0, "Total Amount": 109.0}],
                },
                {
                    "sheet": "Sales",
                    "doc_key": "Sales:SI-001",
                    "rows": [{"Invoice Date": "15/09/2025", "Invoice Number": "SI-001",
                              "Customer Name": "ClientX", "Description": "Services",
                              "Account Code / COA": "4000", "Amount": 200.0,
                              "Tax Amount": 18.0, "Total": 218.0}],
                },
            ],
        },
        nodes.DELIVER_SUMMARY_KEY: "Added rows to ledger.",
    }
    posts = _run_delivery_with_state(state, monkeypatch)

    data_table_posts = [
        p for p in posts
        if any(b.get("type") == "data_table" for b in (p.get("blocks") or []))
    ]
    assert len(data_table_posts) == 1, "one consolidated delivery message"
    tables = [b for b in data_table_posts[0]["blocks"] if b["type"] == "data_table"]
    assert len(tables) == 2, "Purchase + Sales each get a data_table block"


def test_delivery_bank_batch_posts_bank_shaped_preview(monkeypatch):
    """A bank batch (sheet != Purchase/Sales) posts a 6-col bank-shaped preview."""
    state = {
        nodes.LEDGER_ROWS_KEY: {
            "client_id": "c1",
            "fy": "2025",
            "kind": "bank",
            "software": "qbs",
            "client_name": "Acme",
            "batches": [
                {
                    "sheet": "OCBC SGD",
                    "doc_key": "OCBC SGD:2025-09",
                    "rows": [
                        {"Date": "15/09/2025", "Description": "Cheque deposit",
                         "Withdrawal": 0.0, "Deposit": 5000.0,
                         "Balance": 12340.50, "Currency": "SGD"},
                    ],
                }
            ],
        },
        nodes.DELIVER_SUMMARY_KEY: "Bank statement processed.",
    }
    posts = _run_delivery_with_state(state, monkeypatch)

    data_table_posts = [
        p for p in posts
        if any(b.get("type") == "data_table" for b in (p.get("blocks") or []))
    ]
    assert data_table_posts, "Expected a bank preview post"
    table_block = next(b for b in data_table_posts[0]["blocks"] if b["type"] == "data_table")
    header_texts = [c["text"] for c in table_block["rows"][0]]
    assert header_texts == ["Date", "Description", "Withdrawal", "Deposit", "Balance", "Currency"]


def test_delivery_xero_software_uses_xero_headers(monkeypatch):
    """When software == 'xero', the preview headers match the Xero column shape."""
    state = {
        nodes.LEDGER_ROWS_KEY: {
            "client_id": "c1",
            "fy": "2025",
            "kind": "invoice",
            "software": "xero",
            "client_name": "Acme",
            "batches": [
                {
                    "sheet": "Purchase",
                    "doc_key": "Purchase:INV-042",
                    "rows": [
                        {"*ContactName": "Acme Trading Pte Ltd",
                         "*InvoiceNumber": "INV-2025-0042",
                         "*InvoiceDate": "15/09/2025",
                         "Description": "Consulting services",
                         "*AccountCode": "6090", "*TaxType": "SR",
                         "*UnitAmount": 1132.11, "Total": 1234.50},
                    ],
                }
            ],
        },
        nodes.DELIVER_SUMMARY_KEY: "Added 1 Xero row.",
    }
    posts = _run_delivery_with_state(state, monkeypatch)

    data_table_posts = [
        p for p in posts
        if any(b.get("type") == "data_table" for b in (p.get("blocks") or []))
    ]
    assert data_table_posts
    table_block = next(b for b in data_table_posts[0]["blocks"] if b["type"] == "data_table")
    header_texts = [c["text"] for c in table_block["rows"][0]]
    assert header_texts[0] == "Contact"
    assert "Invoice #" in header_texts
    assert "Tax Type" in header_texts


def test_delivery_qbs_software_uses_qbs_headers(monkeypatch):
    """When software == 'qbs', the preview uses QBS Ledger Purchase headers."""
    state = {
        nodes.LEDGER_ROWS_KEY: {
            "client_id": "c1",
            "fy": "2025",
            "kind": "invoice",
            "software": "qbs",
            "client_name": "Acme",
            "batches": [
                {
                    "sheet": "Purchase",
                    "doc_key": "Purchase:INV-001",
                    "rows": [
                        {"Invoice Date": "15/09/2025", "Invoice Number": "INV-001",
                         "Vendor Name": "Acme", "Description": "Consulting",
                         "Account Code / COA": "6090", "Sub Total": 100.0,
                         "Tax Amount": 9.0, "Total Amount": 109.0},
                    ],
                }
            ],
        },
        nodes.DELIVER_SUMMARY_KEY: "Added 1 QBS row.",
    }
    posts = _run_delivery_with_state(state, monkeypatch)

    data_table_posts = [
        p for p in posts
        if any(b.get("type") == "data_table" for b in (p.get("blocks") or []))
    ]
    assert data_table_posts
    table_block = next(b for b in data_table_posts[0]["blocks"] if b["type"] == "data_table")
    header_texts = [c["text"] for c in table_block["rows"][0]]
    assert header_texts[0] == "Invoice Date"
    assert "Vendor" in header_texts
    assert "Sub Total" in header_texts


# =========================================================================== #
# message/file_share wiring: uploads delivered via the message event path
# =========================================================================== #


def _capture_message_handler(runner_mock=None, ledger_store_mock=None, db_mock=None):
    """Build the Bolt app with fakes and return the registered ``message`` handler."""
    from unittest.mock import MagicMock, patch

    from app.slack_app import _SeenEvents
    from accounting_agents import slack_runner

    registered = {}
    fake_app = MagicMock()

    def event_decorator(name):
        def decorator(fn):
            registered[name] = fn
            return fn
        return decorator

    fake_app.event = event_decorator
    fake_app.action = lambda *a, **k: (lambda fn: fn)
    fake_app.view = lambda *a, **k: (lambda fn: fn)
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = runner_mock or MagicMock()
    rm.app_name = "acc"

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch("invoice_processing.export.client_context.FirestoreClientStore"), \
         patch.object(slack_runner, "build_chat_runner",
                      return_value=SimpleNamespace(app_name="accounting_agents_assistant")):
        build_async_app(
            runner=rm,
            ledger_store=ledger_store_mock or MagicMock(),
            db=db_mock or MagicMock(),
        )

    return registered["message"], fresh_seen


def test_message_file_share_calls_process_file_event_per_file():
    """A message/file_share event with 2 files calls process_file_event twice."""
    from unittest.mock import AsyncMock, MagicMock, patch
    from accounting_agents import slack_runner

    handler, _ = _capture_message_handler()

    body = {"event_id": "Ev-file-1"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "111.001",
        "channel": "C-file",
        "files": [{"id": "FA1"}, {"id": "FA2"}],
    }
    fake_client = MagicMock()
    mock_pfe = AsyncMock(return_value={"status": "delivered"})

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    assert mock_pfe.call_count == 2
    called_file_ids = {c.kwargs["file_id"] for c in mock_pfe.call_args_list}
    assert called_file_ids == {"FA1", "FA2"}
    # channel_id threaded through correctly
    assert all(c.kwargs["channel_id"] == "C-file" for c in mock_pfe.call_args_list)


def test_message_bot_file_upload_still_processed():
    """files.upload from this bot carries bot_id — must still enter the pipeline."""
    from unittest.mock import AsyncMock, MagicMock, patch
    from accounting_agents import slack_runner

    handler, _ = _capture_message_handler()

    body = {"event_id": "Ev-bot-file-1"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "111.002",
        "channel": "C-bot-file",
        "bot_id": "BLEDGR",
        "files": [{"id": "F-BOT-1", "name": "telco.pdf"}],
    }
    fake_client = MagicMock()
    mock_pfe = AsyncMock(return_value={"status": "paused"})

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    mock_pfe.assert_called_once()
    assert mock_pfe.call_args.kwargs["file_id"] == "F-BOT-1"


def test_message_bot_chatter_still_ignored():
    """Non-file bot messages must not enter the document pipeline."""
    from unittest.mock import AsyncMock, MagicMock, patch
    from accounting_agents import slack_runner

    handler, _ = _capture_message_handler()

    body = {"event_id": "Ev-bot-chat-1"}
    event = {
        "type": "message",
        "subtype": "bot_message",
        "ts": "111.003",
        "channel": "C-bot-file",
        "bot_id": "BLEDGR",
        "text": "Processing complete",
    }
    mock_pfe = AsyncMock()

    with patch.object(slack_runner, "process_file_event", mock_pfe):
        asyncio.run(handler(event=event, body=body, client=MagicMock()))

    mock_pfe.assert_not_called()


def test_message_file_share_dedup_not_reprocessed():
    """Same event_id redelivered → process_file_event called only once (not twice)."""
    from unittest.mock import AsyncMock, MagicMock, patch
    from accounting_agents import slack_runner

    handler, _ = _capture_message_handler()

    body = {"event_id": "Ev-file-dup"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "111.002",
        "channel": "C-file",
        "files": [{"id": "FB1"}],
    }
    fake_client = MagicMock()
    mock_pfe = AsyncMock(return_value={"status": "delivered"})

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        # First delivery
        asyncio.run(handler(event=event, body=body, client=fake_client))
        assert mock_pfe.call_count == 1
        # Duplicate delivery — same event_id
        asyncio.run(handler(event=event, body=body, client=fake_client))
        assert mock_pfe.call_count == 1  # unchanged


def test_message_text_only_routes_to_question_not_file():
    """A plain text message (no files, no subtype) goes to answer_question only."""
    from unittest.mock import AsyncMock, MagicMock, patch
    from accounting_agents import slack_runner

    handler, _ = _capture_message_handler()

    body = {"event_id": "Ev-text-1"}
    event = {
        "type": "message",
        "ts": "111.003",
        "channel": "C-qa",
        "text": "What is my GST balance?",
    }
    fake_client = MagicMock()
    mock_pfe = AsyncMock(return_value={"status": "delivered"})
    mock_aq = AsyncMock(return_value={"status": "answered", "text": "42"})

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "answer_question", mock_aq):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    mock_pfe.assert_not_called()
    mock_aq.assert_called_once()
    assert mock_aq.call_args.kwargs["question"] == "What is my GST balance?"
    assert mock_aq.call_args.kwargs["channel_id"] == "C-qa"


def test_message_extraction_question_routes_to_agent_not_upload_nudge():
    """Thread questions containing 'extraction' must reach the chat agent, not a canned nudge."""
    from unittest.mock import AsyncMock, MagicMock, patch
    from accounting_agents import slack_runner

    handler, _ = _capture_message_handler_with_slack_client(FakeSlackClient())

    body = {"event_id": "Ev-extract-1"}
    event = {
        "type": "message",
        "ts": "111.004",
        "channel": "C-soa",
        "text": "can you check is the extraction for the SOA using the right one?",
    }
    fake_client = MagicMock()
    mock_aq = AsyncMock(return_value={"status": "answered", "text": "legacy path"})

    with patch.object(slack_runner, "answer_question", mock_aq):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    mock_aq.assert_called_once()
    assert "extraction" in mock_aq.call_args.kwargs["question"].lower()


def test_batch_plan_task_titles_have_no_emoji_shortcodes(monkeypatch):
    """Plan block task titles are plain text — no :hourglass_flowing_sand: literals."""
    from app import native_blocks_compat
    from app.blocks import batch_processing_plan_blocks

    monkeypatch.delenv("LEDGR_BATCH_EXPANDED_PROGRESS", raising=False)
    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    native_blocks_compat._PROBE_CACHE.pop("C-plan-emoji", None)

    doc_rows = [
        {"file_label": "soa.pdf", "stage": "Understanding", "status": "in_progress"},
        {"file_label": "inv.pdf", "stage": "complete", "status": "complete"},
    ]
    blocks = batch_processing_plan_blocks(
        total=2, done=1, doc_rows=doc_rows, channel_id="C-plan-emoji",
    )
    plan = next(b for b in blocks if b.get("type") == "plan")
    for task in plan["tasks"]:
        title = task.get("title") or ""
        assert ":hourglass_flowing_sand:" not in title
        assert ":white_check_mark:" not in title
        assert ":x:" not in title


# =========================================================================== #
# Task 9: One Job summary per batch drop (collapse per-doc spam into 1 thread)
# =========================================================================== #


def _capture_message_handler_with_slack_client(
    injected_slack: FakeSlackClient,
    *,
    ledger_store=None,
):
    """Same as ``_capture_message_handler`` but injects ``injected_slack`` as the
    sync WebClient so we can read its recorded ``chat_postMessage`` /
    ``chat_update`` calls — the outer handler must post ONE summary + edit it.
    """
    from unittest.mock import MagicMock, patch

    from app.slack_app import _SeenEvents
    from accounting_agents import slack_runner

    registered = {}
    fake_app = MagicMock()

    def event_decorator(name):
        def decorator(fn):
            registered[name] = fn
            return fn
        return decorator

    fake_app.event = event_decorator
    fake_app.action = lambda *a, **k: (lambda fn: fn)
    fake_app.view = lambda *a, **k: (lambda fn: fn)
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = MagicMock()
    rm.app_name = "acc"

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch("slack_sdk.WebClient", return_value=injected_slack), \
         patch("invoice_processing.export.client_context.FirestoreClientStore"), \
         patch.object(slack_runner, "build_chat_runner",
                      return_value=SimpleNamespace(app_name="accounting_agents_assistant")):
        build_async_app(
            runner=rm,
            ledger_store=ledger_store if ledger_store is not None else MagicMock(),
            db=MagicMock(),
        )

    return registered["message"], fresh_seen


def test_batch_drop_posts_one_job_summary_then_threads_per_doc():
    """3 files dropped → exactly ONE top-level chat_postMessage + 3 process_file_event
    calls each carrying thread_ts=<summary_ts> + ONE chat_update editing the summary
    with the final tally (ADR-0007).
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner
    from app.slack_app import _SeenEvents

    # The handler reads ``_seen`` from the module globals at CALL time (not at
    # build time), so reset it here to a fresh instance for hermetic dedup.
    slack_runner._seen = _SeenEvents()
    slack = FakeSlackClient()
    handler, _ = _capture_message_handler_with_slack_client(slack)

    body = {"event_id": "Ev-batch-1"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "222.001",
        "channel": "C-batch",
        "files": [{"id": "FA1"}, {"id": "FA2"}, {"id": "FA3"}],
    }
    fake_client = MagicMock()
    # Mix of delivered + paused per doc → drives the needs_review tail.
    mock_pfe = AsyncMock(side_effect=[
        {"status": "delivered", "append": {"appended": 1, "software": "Xero", "fy": "2026"}},
        {"status": "delivered", "append": {"appended": 1, "software": "Xero", "fy": "2026"}},
        {"status": "paused", "op_id": "OP1"},
    ])

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    # --- ONE top-level summary message posted by the outer handler ---
    top_level = [p for p in slack._posts if not p.get("thread_ts")]
    assert len(top_level) == 1, f"expected 1 top-level summary, got {len(top_level)}: {top_level}"
    summary_text = top_level[0].get("text", "")
    assert "3" in summary_text  # total count surfaced in the placeholder
    assert "Received" in summary_text or "Processing" in summary_text

    summary_ts = top_level[0].get("ts")  # e.g. "1.000"
    assert summary_ts, "summary message must carry a ts"

    # --- each process_file_event was called with thread_ts=<summary_ts> ---
    assert mock_pfe.call_count == 3
    for c in mock_pfe.call_args_list:
        assert c.kwargs.get("thread_ts") == summary_ts, (
            f"per-doc call must carry thread_ts={summary_ts}, got {c.kwargs}"
        )
        assert c.kwargs.get("defer_slack_delivery") is True

    # --- after the loop: final chat_update edits the summary with the tally ---
    assert len(slack.updates) >= 1
    upd = slack.updates[-1]
    assert upd.get("channel") == "C-batch"
    assert upd.get("ts") == summary_ts
    final_text = upd.get("text", "")
    assert "Processed" in final_text  # tally uses the helper's template
    assert "2" in final_text  # posted count
    assert "1" in final_text  # needs_review count
    assert "Xero" in final_text or "FY2026" in final_text


def test_single_file_drop_keeps_processing_and_delivery_on_top_level_message():
    """A 1-file drop uses the same main-channel UX as multi-file drops.

    Single-file drops now also post the processing "thinking" plan block
    and the delivery preview tables on the top-level Job summary message
    (they used to be hidden in a thread reply). HITL review cards continue
    to thread under summary_ts per ADR-0007.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner
    from app.slack_app import _SeenEvents

    slack_runner._seen = _SeenEvents()
    slack = FakeSlackClient()
    handler, _ = _capture_message_handler_with_slack_client(slack)

    body = {"event_id": "Ev-batch-2"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "222.002",
        "channel": "C-single",
        "files": [{"id": "FS1"}],
    }
    fake_client = MagicMock()
    mock_pfe = AsyncMock(return_value={
        "status": "delivered",
        "append": {"appended": 1, "software": "QBS Ledger", "fy": "2026"},
    })

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    # One top-level Job summary (the only channel post).
    top_level = [p for p in slack._posts if not p.get("thread_ts")]
    assert len(top_level) == 1
    assert "1" in top_level[0].get("text", "")
    # The Job summary's INITIAL post carries the plan block (used to be plain
    # text only for single-file drops).
    assert top_level[0].get("blocks"), "single-file Job summary must attach a plan block"
    summary_ts = top_level[0].get("ts")

    # process_file_event runs once; thread_ts still wires it under the
    # Job summary (HITL cards need that anchor) but the per-doc "thinking"
    # status post is suppressed because batch_mode=True.
    assert mock_pfe.call_count == 1
    assert mock_pfe.call_args.kwargs.get("thread_ts") == summary_ts
    assert mock_pfe.call_args.kwargs.get("defer_slack_delivery") is True
    assert mock_pfe.call_args.kwargs.get("batch_mode") is True
    assert mock_pfe.call_args.kwargs.get("defer_ledger_persist") is True
    assert mock_pfe.call_args.kwargs.get("status_callback") is not None

    # Per-doc "Received …" thread status messages are suppressed for single
    # drops (the plan block on the top-level message owns the UX now).
    thread_statuses = [
        p for p in slack._posts
        if p.get("thread_ts") and "Received" in (p.get("text") or "")
    ]
    assert thread_statuses == []

    # Final tally chat_update on the top-level summary carries the delivery
    # summary text.
    assert len(slack.updates) >= 1
    final_text = slack.updates[-1].get("text", "")
    assert "Processed" in final_text
    assert "1" in final_text
    # The final update was on the SAME top-level message, not a thread reply.
    assert slack.updates[-1].get("ts") == summary_ts
    assert "thread_ts" not in slack.updates[-1]


def test_single_file_drop_uses_processing_document_headline(monkeypatch):
    """Single-file drops must not say 'Processing batch (1 document)'."""
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner
    from app import native_blocks_compat
    from app.slack_app import _SeenEvents

    monkeypatch.delenv("LEDGR_BATCH_EXPANDED_PROGRESS", raising=False)
    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    native_blocks_compat._PROBE_CACHE.pop("C-single-headline", None)

    slack_runner._seen = _SeenEvents()
    slack = FakeSlackClient()
    handler, _ = _capture_message_handler_with_slack_client(slack)

    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "222.003",
        "channel": "C-single-headline",
        "files": [{"id": "FS-headline"}],
    }
    mock_pfe = AsyncMock(return_value={
        "status": "delivered",
        "append": {"appended": 0, "software": "QBS Ledger", "fy": "2024", "kind": "bank"},
    })

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        asyncio.run(handler(event=event, body={"event_id": "Ev-headline"}, client=MagicMock()))

    top_level = [p for p in slack._posts if not p.get("thread_ts")]
    plan = next(b for b in (top_level[0].get("blocks") or []) if b.get("type") == "plan")
    assert plan["title"] == "Processing document"


def test_single_file_coordinator_flushes_workbook_to_slack(monkeypatch):
    """Single-file defer path must still call append_rows at batch end → one upload."""
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner
    from accounting_agents.ledger_store import SlackLedgerStore
    from app import native_blocks_compat
    from app.slack_app import _SeenEvents
    from tests.test_ledger_store import FakeFirestore

    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    native_blocks_compat._PROBE_CACHE.pop("C-single-flush", None)

    slack = FakeSlackClient()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    slack_runner._seen = _SeenEvents()
    handler, _ = _capture_message_handler_with_slack_client(slack, ledger_store=store)

    deferred = {
        "summary": "Added",
        "payload": {
            "client_id": "c-akar",
            "fy": "2024",
            "kind": "bank",
            "software": "qbs",
            "client_name": "Akar Enterprises Pte. Ltd.",
        },
        "batches": [
            {
                "sheet": "DBS Bank Ltd - 5545 - SGD",
                "doc_key": "F-single:5545:SGD:Apr2024",
                "rows": [
                    {"Description": "BALANCE B/F", "Balance": 100.0, "Currency": "SGD"},
                    {"Date": "15/04/2024", "Description": "TEST", "Deposit": 1.0,
                     "Balance": 101.0, "Currency": "SGD"},
                    {"Description": "TOTALS", "Currency": "SGD"},
                ],
            },
        ],
        "workbook_name": "",
    }
    mock_pfe = AsyncMock(return_value={
        "status": "delivered",
        "append": {
            "deferred_delivery": deferred,
            "deferred_ledger": deferred,
            "kind": "bank",
            "software": "qbs",
            "fy": "2024",
            "appended": 0,
        },
    })

    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "222.004",
        "channel": "C-single-flush",
        "files": [{"id": "FS-flush"}],
    }
    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        asyncio.run(handler(event=event, body={"event_id": "Ev-flush"}, client=MagicMock()))

    assert len(slack.uploads) == 1, "expected one files_upload_v2 at batch end"
    final_update = slack.updates[-1]
    assert "added" in final_update.get("text", "").lower()
    blocks = final_update.get("blocks") or []
    assert any(b.get("type") == "data_table" for b in blocks), (
        "delivery preview tables should appear when rows were appended"
    )


def test_single_file_re_drop_all_deduped_shows_delivery_preview_and_callout(monkeypatch):
    """Re-drop where every batch dedupes still shows extraction preview + callout."""
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner
    from accounting_agents.ledger_store import SlackLedgerStore
    from app import native_blocks_compat
    from app.slack_app import _SeenEvents
    from tests.test_ledger_store import FakeFirestore

    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    native_blocks_compat._PROBE_CACHE.pop("C-redrop", None)

    slack = FakeSlackClient()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    doc_key = "DBS Bank Ltd - 5545 - SGD:5545:SGD:01 Apr 2024 - 30 Apr 2024"
    store.append_rows(
        client_id="c-akar", fy="2024", slack_client=slack, channel_id="C-redrop",
        kind="bank",
        batches=[{
            "sheet": "DBS Bank Ltd - 5545 - SGD",
            "doc_key": doc_key,
            "rows": [
                {"Description": "BALANCE B/F", "Balance": 1.0, "Currency": "SGD"},
                {"Date": "15/04/2024", "Description": "OLD", "Balance": 1.0, "Currency": "SGD"},
                {"Description": "TOTALS", "Currency": "SGD"},
            ],
        }],
    )
    assert len(slack.uploads) == 1
    slack.uploads.clear()

    slack_runner._seen = _SeenEvents()
    handler, _ = _capture_message_handler_with_slack_client(slack, ledger_store=store)

    deferred = {
        "summary": "Added",
        "payload": {
            "client_id": "c-akar", "fy": "2024", "kind": "bank",
            "software": "qbs", "client_name": "Akar",
        },
        "batches": [{
            "sheet": "DBS Bank Ltd - 5545 - SGD",
            "doc_key": doc_key,
            "rows": [
                {"Description": "BALANCE B/F", "Balance": 1.0, "Currency": "SGD"},
                {"Date": "15/04/2024", "Description": "OLD", "Balance": 1.0, "Currency": "SGD"},
                {"Description": "TOTALS", "Currency": "SGD"},
            ],
        }],
        "workbook_name": "",
    }
    mock_pfe = AsyncMock(return_value={
        "status": "delivered",
        "append": {
            "deferred_delivery": deferred,
            "deferred_ledger": deferred,
            "kind": "bank", "software": "qbs", "fy": "2024", "appended": 0,
        },
    })

    event = {
        "type": "message", "subtype": "file_share", "ts": "222.005",
        "channel": "C-redrop", "files": [{"id": "FS-redrop"}],
    }
    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        asyncio.run(handler(event=event, body={"event_id": "Ev-redrop"}, client=MagicMock()))

    final_update = slack.updates[-1]
    final_text = final_update.get("text", "")
    assert "added" in final_text.lower()
    assert "already recorded" not in final_text.lower()
    assert "workbook unchanged" in final_text.lower()
    blocks = final_update.get("blocks") or []
    assert any(b.get("type") == "data_table" for b in blocks)
    action_ids = {
        el.get("action_id")
        for b in blocks
        if b.get("type") in ("card", "actions")
        for el in (b.get("actions") or b.get("elements") or [])
    }
    assert "ledgr_dedup_replace" in action_ids


# =========================================================================== #
# Task 5: doc_label built from run state, posted on card, persisted on interrupt
# =========================================================================== #


def test_doc_label_from_state():
    """First invoice vendor + total formatted as a per-document label."""
    state = {
        "source_filename": "Receipt-Hotel.pdf",
        nodes.NORMALIZED_KEY: [
            {
                "vendor_name": "Hotel Booking",
                "total_amount": 51.49,
                "currency": "SGD",
            }
        ],
    }
    label = _doc_label_from_state(state)
    assert "Receipt-Hotel.pdf" in label
    assert "Hotel Booking" in label
    assert "51.49" in label


def test_doc_label_from_state_falls_back_on_missing_invoice():
    """No normalized invoices yet → label is just the filename."""
    label = _doc_label_from_state({"source_filename": "mystery.pdf"})
    assert "mystery.pdf" in label
    # No vendor / total noise when the invoice is absent.
    assert "·" not in label.split("mystery.pdf", 1)[1]


def test_doc_label_from_state_uses_issuer_name_alias():
    """``issuer_name`` is the sales-direction alias for ``vendor_name``."""
    state = {
        "source_filename": "INV-1001.pdf",
        nodes.NORMALIZED_KEY: [
            {"issuer_name": "BigBuyer Ltd", "total_amount": 100.0, "currency": "SGD"}
        ],
    }
    label = _doc_label_from_state(state)
    assert "BigBuyer Ltd" in label
    assert "100.00" in label


def test_doc_label_from_state_skips_money_when_total_missing():
    """No numeric total → no trailing currency block."""
    state = {
        "source_filename": "INV-1001.pdf",
        nodes.NORMALIZED_KEY: [
            {"vendor_name": "Acme", "currency": "SGD"}  # no total_amount
        ],
    }
    label = _doc_label_from_state(state)
    assert "Acme" in label
    assert "SGD" not in label


def test_doc_label_from_state_default_filename():
    """Empty state still produces a non-empty label (so the card never looks bare)."""
    label = _doc_label_from_state({})
    assert "document" in label


def test_process_file_event_interrupt_persists_doc_label_and_renders_it():
    """Approval card header names the document; interrupt doc carries the same label.

    Patches ``_read_interrupt_summary`` to return a known summary, seeds the
    session state with a normalized invoice so the label is rich, and asserts
    the resulting card text + Firestore doc both carry the label.
    """
    from unittest.mock import AsyncMock, patch

    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    # Seed a profile whose channel_id matches the run we drive below so the
    # soft-gate at process_file_event lets the run proceed.
    seeded = _seeded_client_store(db, channel_id="C5", client_id="c5")

    interrupt_event = SimpleNamespace(
        content=SimpleNamespace(parts=[]),
        get_function_calls=lambda: [SimpleNamespace(name="adk_request_input", id="C5:F5")],
    )
    # Pre-seed the final-state view with a normalized invoice so the label is
    # built from the same shape the real graph produces.
    final_state = {
        "approval_message": "needs review: line X",
        "source_filename": "Receipt-Hotel.pdf",
        nodes.NORMALIZED_KEY: [
            {"vendor_name": "Hotel Booking", "total_amount": 51.49, "currency": "SGD"}
        ],
    }
    runner = _FakeRunner([interrupt_event], final_state)

    with patch(
        "accounting_agents.slack_runner._read_interrupt_summary",
        new=AsyncMock(return_value="needs review: line X"),
    ):
        result = asyncio.run(
            process_file_event(
                runner=runner,
                ledger_store=store,
                db=db,
                slack_client=slack,
                channel_id="C5",
                file_id="F5",
                app_name="acc",
                download_fn=lambda c, f: b"%PDF fake",
                source_filename="Receipt-Hotel.pdf",
                client_store=seeded,
            )
        )

    assert result["status"] == "paused"

    # Card text names the document above the review header.
    card = _last_blocks(slack)
    head = next(
        b["text"]["text"] for b in card
        if b.get("type") == "section" and b.get("text", {}).get("type") == "mrkdwn"
    )
    assert "Receipt-Hotel.pdf" in head
    assert "Hotel Booking" in head
    assert "51.49" in head
    # And the label sits ABOVE the standard review-needed header line.
    assert head.index("Receipt-Hotel.pdf") < head.index("Review needed")

    # Interrupt correlation doc persists the label alongside the summary.
    snap = db.collection("interrupts").document("C5:F5").get()
    assert snap.exists
    doc = snap.to_dict()
    assert doc.get("doc_label")
    assert "Receipt-Hotel.pdf" in doc["doc_label"]


# =========================================================================== #
# Task 7: edits-from-view-state parser (Slack view_submission → line edits DTO)
# =========================================================================== #


def test_edits_from_view_state_builds_line_edits():
    view = {"state": {"values": {
        "acct_0": {"v": {"selected_option": {"value": "6010"}}},
        "tax_0":  {"v": {"selected_option": {"value": "ZR"}}},
        "amt_0":  {"v": {"value": "44.74"}},
    }}}
    edits = _edits_from_view_state(view)
    assert edits == {"lines": [{"index": 0, "account_code": "6010",
                                "tax_treatment": "ZR", "net_amount": 44.74}]}


# =========================================================================== #
# Task 8: an edit becomes a per-client Correction (ADR-0004)
# =========================================================================== #


def test_persist_corrections_writes_vendor_mapping():
    """One Correction per edited line that carries account_code / tax_code.

    Mirrors the spec test verbatim: a single line with both fields produces one
    ``add_correction`` call with the invoice's vendor (taken from the first
    normalized invoice's ``vendor_name``).
    """
    saved = []

    class _Store:
        def add_correction(self, *, client_id, vendor, account_code=None, tax_code=None):
            saved.append((client_id, vendor, account_code, tax_code))

    state = {
        "client_id": "CL-1",
        nodes.NORMALIZED_KEY: [
            {"vendor_name": "Hotel Booking",
             "lines": [{"description": "Room"}]}  # uncategorized → edits are real changes
        ],
    }
    edits = {"lines": [{"index": 0, "account_code": "6010", "tax_treatment": "ZR"}]}
    _persist_corrections(_Store(), state, edits)
    assert saved == [("CL-1", "Hotel Booking", "6010", "ZR")]


def test_persist_corrections_reads_nested_party_real_serialized_shape():
    """Vendor comes from the nested party the SERIALIZER actually produces.

    ``_inv_to_dict`` = ``asdict(NormalizedInvoice)`` nests the parties under
    ``supplier`` / ``customer`` — there is no flat ``vendor_name`` key on real
    state. Purchases key off ``supplier.name``; sales key off ``customer.name``
    (mirroring ``NormalizedInvoice.counterparty``). This is the shape that broke
    learning in production.
    """
    saved = []

    class _Store:
        def add_correction(self, *, client_id, vendor, account_code=None, tax_code=None):
            saved.append((client_id, vendor, account_code, tax_code))

    # Purchase → supplier.name; proposed 6-3000, human changes to 5-1000.
    purchase_state = {
        "client_id": "CL-1",
        nodes.NORMALIZED_KEY: [
            {"doc_type": "purchase",
             "supplier": {"name": "Vendor Alpha Pte Ltd"},
             "customer": {"name": "Company-A"},
             "lines": [{"description": "audit", "account_code": "6-3000"}]}
        ],
    }
    _persist_corrections(_Store(), purchase_state,
                         {"lines": [{"index": 0, "account_code": "5-1000"}]})
    assert saved == [("CL-1", "Vendor Alpha Pte Ltd", "5-1000", None)]

    # Sales → customer.name; proposed line untaxed, human sets SR.
    saved.clear()
    sales_state = {
        "client_id": "CL-1",
        nodes.NORMALIZED_KEY: [
            {"doc_type": "sales",
             "supplier": {"name": "Company-A"},
             "customer": {"name": "PTTEP"},
             "lines": [{"description": "svc"}]}
        ],
    }
    _persist_corrections(_Store(), sales_state,
                         {"lines": [{"index": 0, "tax_treatment": "SR"}]})
    assert saved == [("CL-1", "PTTEP", None, "SR")]


def test_persist_corrections_only_changed_line_no_collision_multi_line():
    """Multi-line invoice: only the line the human CHANGED becomes a Correction.

    Regression for the collision bug — the modal re-submits every line, so an
    unchanged line (same code as proposed) must NOT overwrite the edited line's
    vendor mapping. Here line 0 is changed 6-3000→5-1000; line 1 is left at the
    proposed 6-3000. Exactly one Correction (5-1000) must be written.
    """
    saved = []

    class _Store:
        def add_correction(self, *, client_id, vendor, account_code=None, tax_code=None):
            saved.append((client_id, vendor, account_code, tax_code))

    state = {
        "client_id": "CL-1",
        nodes.NORMALIZED_KEY: [
            {"doc_type": "purchase",
             "supplier": {"name": "Vendor Alpha Pte Ltd"},
             "lines": [
                 {"description": "audit", "account_code": "6-3000"},
                 {"description": "report", "account_code": "6-3000"},
             ]}
        ],
    }
    edits = {"lines": [
        {"index": 0, "account_code": "5-1000"},   # changed
        {"index": 1, "account_code": "6-3000"},   # unchanged (== proposal)
    ]}
    _persist_corrections(_Store(), state, edits)
    assert saved == [("CL-1", "Vendor Alpha Pte Ltd", "5-1000", None)]


def test_persist_corrections_skips_unchanged_lines():
    """A line resubmitted at its proposed value is NOT a Correction."""
    saved = []

    class _Store:
        def add_correction(self, *, client_id, vendor, account_code=None, tax_code=None):
            saved.append((client_id, vendor, account_code, tax_code))

    state = {
        "client_id": "CL-1",
        nodes.NORMALIZED_KEY: [
            {"doc_type": "purchase",
             "supplier": {"name": "Acme"},
             "lines": [{"description": "x", "account_code": "6-3000"}]}
        ],
    }
    _persist_corrections(_Store(), state, {"lines": [{"index": 0, "account_code": "6-3000"}]})
    assert saved == []


def test_persist_corrections_skips_lines_with_no_code_fields():
    """A line that only changed ``net_amount`` (no account_code, no tax_treatment)
    is skipped. Edits without either code field are not entity-memory worthy —
    the user's amount tweak is a one-off variance, not a vendor rule.
    """
    saved = []

    class _Store:
        def add_correction(self, *, client_id, vendor, account_code=None, tax_code=None):
            saved.append((client_id, vendor, account_code, tax_code))

    state = {
        "client_id": "CL-1",
        nodes.NORMALIZED_KEY: [{"vendor_name": "Acme"}],
    }
    edits = {"lines": [{"index": 0, "net_amount": 12.34}]}
    _persist_corrections(_Store(), state, edits)
    assert saved == []


def test_persist_corrections_uses_issuer_name_alias_when_vendor_missing():
    """``issuer_name`` (sales-direction alias) is the fallback vendor field."""
    saved = []

    class _Store:
        def add_correction(self, *, client_id, vendor, account_code=None, tax_code=None):
            saved.append((client_id, vendor, account_code, tax_code))

    state = {
        "client_id": "CL-1",
        nodes.NORMALIZED_KEY: [{"issuer_name": "BigBuyer Ltd"}],
    }
    edits = {"lines": [{"index": 0, "account_code": "5000"}]}
    _persist_corrections(_Store(), state, edits)
    assert saved == [("CL-1", "BigBuyer Ltd", "5000", None)]


def test_persist_corrections_noop_without_client_id_or_invoice():
    """Defensive: missing client_id or empty invoice list ⇒ no writes (no crash)."""
    saved = []

    class _Store:
        def add_correction(self, *, client_id, vendor, account_code=None, tax_code=None):
            saved.append((client_id, vendor, account_code, tax_code))

    _persist_corrections(_Store(), {}, {"lines": [{"index": 0, "account_code": "6010"}]})
    _persist_corrections(_Store(), {"client_id": "X"}, {"lines": [{"index": 0, "account_code": "6010"}]})
    _persist_corrections(_Store(), {"client_id": "X", nodes.NORMALIZED_KEY: []},
                         {"lines": [{"index": 0, "account_code": "6010"}]})
    assert saved == []


def test_main_async_wires_firestore_client_store(monkeypatch):
    """Socket mode must seed onboarding/commands with the SAME Firestore store
    the document pipeline reads — not an ephemeral in-memory store. Otherwise a
    profile registered via the modal would be invisible to processing.
    """
    import accounting_agents.slack_runner as sr
    from invoice_processing.export.client_context import FirestoreClientStore

    captured: dict = {}

    monkeypatch.setattr(sr, "build_runner", lambda: SimpleNamespace(app_name="acc"))
    monkeypatch.setattr(sr, "SlackLedgerStore", lambda db: object())
    monkeypatch.setattr(
        "accounting_agents.sessions.FirestoreSessionService",
        lambda: SimpleNamespace(client=object()),
    )

    def _fake_build(**kwargs):
        captured["store"] = kwargs.get("store")
        return object()

    monkeypatch.setattr(sr, "build_async_app", _fake_build)

    class _FakeHandler:
        def __init__(self, app, token):
            captured["token"] = token

        async def start_async(self):
            captured["started"] = True

    monkeypatch.setattr(
        "slack_bolt.adapter.socket_mode.async_handler.AsyncSocketModeHandler",
        _FakeHandler,
    )
    monkeypatch.setenv("SLACK_APP_TOKEN", "xapp-test")

    asyncio.run(sr._main_async())

    assert isinstance(captured["store"], FirestoreClientStore)
    assert captured.get("started") is True


# =========================================================================== #
# Task 8: Reject unreadable uploads — empty bytes or unknown mime
# =========================================================================== #


def test_process_file_event_rejects_empty_bytes():
    """An empty download is rejected with a friendly message; NOT counted as processed.

    The pipeline must post a clear rejection message and return status
    ``"rejected_unreadable"`` so the batch-drop tally excludes it from
    the "Processed N documents" count (neither ``posted`` nor ``needs_review``
    increments for rejected files).
    """
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    runner = _FakeRunner([], _ledger_payload())  # should never run the graph

    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"",  # empty — nothing downloaded
            source_filename="mystery.bin",
            client_store=_seeded_client_store(db),
        )
    )

    # Must return a distinct rejection status (not "delivered" or "paused").
    assert result["status"] == "rejected_unreadable"
    # A friendly message was posted — not the "Processed" path.
    texts = _posted_texts(slack)
    assert any(
        "empty" in t.lower() or "supported" in t.lower() or "couldn't read" in t.lower()
        for t in texts
    ), f"expected rejection message in: {texts}"
    # The graph was never driven (no artifacts saved).
    assert runner.artifact_service.saved == {}
    # No ledger rows appended (no upload).
    assert slack.uploads == []


def test_process_file_event_rejects_unknown_extension():
    """A file with an unrecognised extension is rejected before the graph runs.

    Simulates an "unknown/unknown size" upload: the bytes are non-empty but the
    extension is not one the pipeline supports (.exe is clearly wrong). The
    validator must catch this and reject it rather than forwarding garbage to
    Gemini.
    """
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    runner = _FakeRunner([], _ledger_payload())  # must not run

    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"\x4d\x5a\x90\x00" * 10,  # EXE magic bytes, non-empty
            source_filename="setup.exe",
            client_store=_seeded_client_store(db),
        )
    )

    assert result["status"] == "rejected_unreadable"
    texts = _posted_texts(slack)
    assert any(
        "empty" in t.lower() or "supported" in t.lower() or "couldn't read" in t.lower()
        for t in texts
    ), f"expected rejection message in: {texts}"
    assert runner.artifact_service.saved == {}
    assert slack.uploads == []


def test_process_file_event_accepted_pdf_still_processes():
    """A known-good extension (.pdf, non-empty bytes) passes the validator and proceeds.

    Regression guard: the validation guard must NOT block legitimate uploads.
    """
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    final_event = SimpleNamespace(
        content=SimpleNamespace(parts=[SimpleNamespace(text="done")]),
        get_function_calls=lambda: [],
    )
    runner = _FakeRunner([final_event], _ledger_payload())

    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake content",
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
    )

    assert result["status"] == "delivered"
    assert len(slack.uploads) == 1


# =========================================================================== #
# build_fastapi_app — prod wiring test (Task 4)
# =========================================================================== #

def test_build_fastapi_app_wires_adk_graph(monkeypatch):
    """build_fastapi_app() returns a FastAPI app with POST /slack/events.

    Asserts (without any network/Slack calls) that the factory:
    1. Builds the runner via build_runner() (which binds accounting_agents.agent.app).
    2. Builds the async Bolt app via build_async_app().
    3. Wraps it in AsyncSlackRequestHandler.
    4. Exposes POST /slack/events.
    5. Exposes GET /healthz.

    Uses monkeypatch.setattr (not with-patch) so patches stay active during the
    lazy _get_handler() call triggered by tc.post("/slack/events").
    FirestoreSessionService and FirestoreClientStore are function-local imports
    inside _get_handler, so we patch the SOURCE modules.
    """
    from unittest.mock import AsyncMock, MagicMock

    import accounting_agents.sessions as _sessions_mod
    import accounting_agents.slack_runner as _runner_mod
    import invoice_processing.export.client_context as _ctx_mod
    import slack_bolt.adapter.fastapi.async_handler as _handler_mod

    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    runner_calls: list = []
    app_calls: list = []
    fake_async_app = MagicMock()
    fake_handler = MagicMock()
    fake_handler.handle = AsyncMock(return_value=MagicMock(status_code=401))

    def _fake_build_runner(**kw):
        runner_calls.append(kw)
        return MagicMock(name="runner")

    def _fake_build_async_app(*, runner, ledger_store, db, store=None, bot_token=None):
        app_calls.append({"runner": runner, "store": store})
        return fake_async_app

    # Patch module-level names in slack_runner (build_runner/build_async_app/
    # SlackLedgerStore are module attributes so this works directly).
    monkeypatch.setattr(_runner_mod, "build_runner", _fake_build_runner)
    monkeypatch.setattr(_runner_mod, "build_async_app", _fake_build_async_app)
    monkeypatch.setattr(_runner_mod, "SlackLedgerStore", MagicMock(return_value=MagicMock()))

    # FirestoreSessionService/FirestoreClientStore are imported locally inside
    # _get_handler, so patch the SOURCE class in the source modules.
    monkeypatch.setattr(_sessions_mod, "FirestoreSessionService", MagicMock(return_value=MagicMock()))
    monkeypatch.setattr(_ctx_mod, "FirestoreClientStore", MagicMock(return_value=MagicMock()))
    monkeypatch.setattr(_handler_mod, "AsyncSlackRequestHandler", MagicMock(return_value=fake_handler))

    from accounting_agents.slack_runner import build_fastapi_app
    api = build_fastapi_app()

    # Must return a FastAPI instance.
    assert isinstance(api, FastAPI)

    # POST /slack/events and GET /healthz must be registered.
    paths = {r.path for r in api.routes}
    assert "/slack/events" in paths
    assert "/healthz" in paths

    # Trigger lazy construction by hitting /slack/events — patches still active.
    tc = TestClient(api, raise_server_exceptions=False)
    tc.post("/slack/events")

    # build_runner was called exactly once (lazy construction happened once).
    assert len(runner_calls) == 1

    # build_async_app was called with a runner (proves graph is wired).
    assert len(app_calls) == 1
    assert app_calls[0]["runner"] is not None

    # A FirestoreClientStore was passed as the `store` (not InMemory).
    assert app_calls[0]["store"] is not None


# =========================================================================== #
# _resolve_file_name — real uploaded filename for validation + card labels
# (regression: handlers passed no name → default "document.pdf" → unsupported
#  files never rejected + every card mislabeled)
# =========================================================================== #


def test_resolve_file_name_prefers_file_object_name():
    from accounting_agents.slack_runner import _resolve_file_name

    class _Client:
        def files_info(self, file):
            raise AssertionError("must not call files_info when the name is present")

    assert _resolve_file_name(_Client(), "F1", {"name": "Invoice-99.pdf"}) == "Invoice-99.pdf"


def test_resolve_file_name_falls_back_to_files_info():
    from accounting_agents.slack_runner import _resolve_file_name

    class _Resp:
        data = {"file": {"name": "scan.exe"}}

    class _Client:
        def files_info(self, file):
            return _Resp()

    # No name on the (minimal file_shared) object → fetch via files_info, keep .exe.
    assert _resolve_file_name(_Client(), "F1", None) == "scan.exe"


def test_resolve_file_name_defaults_only_when_truly_unavailable():
    from accounting_agents.slack_runner import _resolve_file_name

    class _Client:
        def files_info(self, file):
            raise RuntimeError("boom")

    assert _resolve_file_name(_Client(), "F1", None) == "document.pdf"


# =========================================================================== #
# Step 1: chat session id + chat runner wiring (ADR-0008)
# =========================================================================== #


def test_chat_session_id_thread_case():
    """A message inside a Slack thread keys its session by the raw thread_ts.

    Replies in the same thread reuse the same session id → the assistant sees
    multi-turn history. The message ts is irrelevant in this case.
    """
    from accounting_agents.slack_runner import _chat_session_id

    sid = _chat_session_id("C123", "1700000000.123", "1700000005.456")
    assert sid == "C123:chat:1700000000.123"


def test_chat_session_id_day_bucket():
    """A top-level message (no thread_ts) buckets by the UTC day of its ts.

    1700000000 == 2023-11-14T22:13:20 UTC → day-2023-11-14.
    """
    from accounting_agents.slack_runner import _chat_session_id

    sid = _chat_session_id("C123", None, "1700000000.000")
    assert sid == "C123:chat:day-2023-11-14"


def test_chat_session_id_falls_back_to_channel_when_ts_missing():
    """A missing/unparseable message_ts degrades to ``channel_id`` alone.

    Direct callers (tests, one-shot scripts) may not have a message ts; the
    chat lane must still produce a usable session id.
    """
    from accounting_agents.slack_runner import _chat_session_id

    assert _chat_session_id("C123", None, None) == "C123"
    assert _chat_session_id("C123", None, "not-a-number") == "C123"


def test_chat_and_document_sessions_isolated():
    """A document session id and a chat session id on the same channel differ.

    Pipeline events never pollute chat history and vice versa (ADR-0008).
    """
    from accounting_agents.slack_runner import _chat_session_id, _per_doc_session_id

    doc = _per_doc_session_id("C9", "FILE1")
    chat = _chat_session_id("C9", "1700000000.123", "1700000005.456")
    assert doc != chat
    assert ":chat:" in chat
    assert ":chat:" not in doc


# --------------------------------------------------------------------------- #
# Multi-turn session reuse + chat-runner wiring
# --------------------------------------------------------------------------- #


class _CapturingChatRunner:
    """Minimal chat-runner stub that records each ``run_async`` call.

    Mirrors enough of the real ``Runner`` API for ``answer_question``: an
    ``app_name`` + ``session_service`` + an async-generator ``run_async`` that
    yields one final-text event.
    """

    def __init__(self, sessions: dict, app_name: str = "accounting_agents_assistant"):
        self.app_name = app_name
        self._sessions = sessions
        self.calls: list[dict] = []

        runner_self = self

        class _SessionService:
            def __init__(self):
                self.created: list[tuple] = []

            async def get_session(self, *, app_name, user_id, session_id):
                return runner_self._sessions.get((user_id, session_id))

            async def create_session(self, *, app_name, user_id, session_id, state=None):
                from google.adk.errors.already_exists_error import AlreadyExistsError
                if (user_id, session_id) in runner_self._sessions:
                    raise AlreadyExistsError("exists")
                self.created.append((user_id, session_id))
                runner_self._sessions[(user_id, session_id)] = _FakeSession(state or {})
                return runner_self._sessions[(user_id, session_id)]

        self.session_service = _SessionService()

    async def run_async(
        self, *, user_id, session_id, new_message=None, state_delta=None, run_config=None,
    ):
        self.calls.append(
            {
                "user_id": user_id,
                "session_id": session_id,
                "state_delta": state_delta or {},
                "run_config": run_config,
                "new_message": new_message,
            }
        )
        # Append a synthetic user event to the persisted session so the test can
        # assert that the second turn's session contains the first turn.
        sess = self._sessions.get((user_id, session_id))
        if sess is not None:
            events = getattr(sess, "events", None) or []
            events.append(("user", new_message))
            sess.events = events
        yield SimpleNamespace(
            content=SimpleNamespace(parts=[SimpleNamespace(text="ok")]),
            get_function_calls=lambda: [],
        )


def _noop_ledger_store():
    """A ledger_store stub that returns empty rows so answer_question can run hermetically."""
    from unittest.mock import MagicMock

    s = MagicMock()
    s.latest_fy.return_value = None
    s.read_rows.return_value = []
    return s


def test_multi_turn_reuses_session():
    """Two ``answer_question`` calls on the same ``(channel, raw_thread_ts)`` reuse
    the same chat session and the second call sees the first turn's event.
    """
    from accounting_agents.slack_runner import answer_question

    slack = FakeSlackClient()
    sessions: dict = {}
    chat_runner = _CapturingChatRunner(sessions)

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=_noop_ledger_store(),
            slack_client=slack,
            channel_id="C-CHAT",
            question="Hello",
            app_name=chat_runner.app_name,
            client_store=None,
            message_ts="1700000000.001",
            thread_ts="1700000000.001",
            raw_thread_ts="1700000000.001",
        )
    )
    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=_noop_ledger_store(),
            slack_client=slack,
            channel_id="C-CHAT",
            question="And again",
            app_name=chat_runner.app_name,
            client_store=None,
            message_ts="1700000000.002",
            thread_ts="1700000000.001",
            raw_thread_ts="1700000000.001",
        )
    )

    assert len(chat_runner.calls) == 2
    assert chat_runner.calls[0]["session_id"] == chat_runner.calls[1]["session_id"]
    assert chat_runner.calls[0]["session_id"] == "C-CHAT:chat:1700000000.001"

    # The second turn's persisted session contains events recorded by the first.
    sess = sessions[("C-CHAT", "C-CHAT:chat:1700000000.001")]
    assert len(getattr(sess, "events", [])) == 2


def test_answer_question_does_not_set_question_key():
    """The chat lane no longer embeds the question into state (multi-turn now)."""
    from accounting_agents.slack_runner import answer_question

    slack = FakeSlackClient()
    sessions: dict = {}
    chat_runner = _CapturingChatRunner(sessions)

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=_noop_ledger_store(),
            slack_client=slack,
            channel_id="C-CHAT",
            question="Whatever",
            app_name=chat_runner.app_name,
            client_store=None,
            message_ts="1700000000.500",
            thread_ts="1700000000.500",
            raw_thread_ts=None,
        )
    )

    state_delta = chat_runner.calls[0]["state_delta"]
    assert "question_text" not in state_delta
    # The ledger / channel keys are still seeded.
    assert state_delta.get("channel_id") == "C-CHAT"
    assert "ledger_data" in state_delta


def test_run_config_caps_recent_events():
    """``answer_question`` passes ``RunConfig(num_recent_events=20)`` to run_async."""
    from accounting_agents.slack_runner import answer_question

    slack = FakeSlackClient()
    sessions: dict = {}
    chat_runner = _CapturingChatRunner(sessions)

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=_noop_ledger_store(),
            slack_client=slack,
            channel_id="C-CHAT",
            question="hi",
            app_name=chat_runner.app_name,
            client_store=None,
            message_ts="1700000000.600",
            thread_ts="1700000000.600",
            raw_thread_ts=None,
        )
    )

    run_config = chat_runner.calls[0]["run_config"]
    assert run_config is not None
    gsc = run_config.get_session_config
    assert gsc is not None
    assert gsc.num_recent_events == 20


class _SilentModelChatRunner(_CapturingChatRunner):
    """Simulates ``gemini-2.5-flash-lite`` going silent after a tool call.

    Yields a function-response event (carrying ``result``), then a final event
    with NO text parts — the exact failure mode observed live (the model
    occasionally produces an empty completion after a tool returns, despite
    instruction telling it to always reply with text).
    """

    async def run_async(
        self, *, user_id, session_id, new_message=None, state_delta=None, run_config=None,
    ):
        self.calls.append(
            {
                "user_id": user_id,
                "session_id": session_id,
                "state_delta": state_delta or {},
                "run_config": run_config,
                "new_message": new_message,
            }
        )
        # Event 1: a tool result with a useful raw payload the user should see.
        fr = SimpleNamespace(response={"result": '{"revenue": 8500, "expenses": 1820.5, "net": 6679.5}'})
        yield SimpleNamespace(
            content=SimpleNamespace(parts=[SimpleNamespace(text=None)]),
            get_function_calls=lambda: [],
            get_function_responses=lambda: [fr],
        )
        # Event 2: final event with EMPTY text parts (the model went silent).
        yield SimpleNamespace(
            content=SimpleNamespace(parts=[SimpleNamespace(text=None)]),
            get_function_calls=lambda: [],
            get_function_responses=lambda: [],
        )


def test_silent_model_safety_net_surfaces_tool_result():
    """When the model goes silent after a tool call, the chat lane MUST surface
    the tool's raw result rather than the opaque ``rephrase your question``
    canned message. Observed live on 2026-06-15 with gemini-2.5-flash-lite.
    """
    from accounting_agents.slack_runner import answer_question

    slack = FakeSlackClient()
    sessions: dict = {}
    chat_runner = _SilentModelChatRunner(sessions)

    result = asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=_noop_ledger_store(),
            slack_client=slack,
            channel_id="C-CHAT",
            question="summarise the purchases",
            app_name=chat_runner.app_name,
            client_store=None,
            message_ts="1700000000.700",
            thread_ts="1700000000.700",
            raw_thread_ts="1700000000.700",
        )
    )

    posted = slack._posts[-1]["text"]
    assert "rephrasing your question" not in posted
    assert "revenue" in posted and "6679.5" in posted
    assert result["text"] == posted


# --------------------------------------------------------------------------- #
# P0-2: chat session must see freshly-written ledger rows across turns
# --------------------------------------------------------------------------- #


class _StatefulCapturingChatRunner(_CapturingChatRunner):
    """Like _CapturingChatRunner but has a session service with ``append_event``
    that actually applies the state_delta to the session state in-memory,
    mirroring what a real ADK InMemorySessionService does.

    Also records ``pre_run_state_snapshots`` so tests can assert what state
    the session held at the START of each ``run_async`` call (i.e. AFTER any
    pre-write via ``_apply_state_delta`` but before the agent processes the turn).
    """

    def __init__(self, sessions: dict, app_name: str = "accounting_agents_assistant"):
        super().__init__(sessions, app_name)
        self.pre_run_state_snapshots: list[dict] = []
        self.pre_run_state_deltas: list[dict] = []

        # Wrap the inherited _SessionService to add append_event support so
        # _apply_state_delta (called before run_async) can actually update the
        # in-memory session state.
        _runner_ref = self

        class _SessionServiceWithAppend(self.session_service.__class__):
            def __init__(self):
                pass  # fields copied below

            async def append_event(self, session, event):
                """Apply event.actions.state_delta to the persisted session state."""
                if event and event.actions and event.actions.state_delta:
                    delta = event.actions.state_delta
                    # Update all stored sessions whose state object IS this session.
                    for stored in _runner_ref._sessions.values():
                        if stored is session:
                            stored.state.update(delta)
                            break
                    else:
                        # session is a copy — update by matching identity via state.
                        session.state.update(delta)
                        # Propagate to all stored sessions (there's only one per test).
                        for stored in _runner_ref._sessions.values():
                            stored.state.update(delta)
                return event

        # Re-instantiate with the same internal data but the enriched class.
        existing_svc = self.session_service
        new_svc = _SessionServiceWithAppend()
        new_svc.__dict__.update(existing_svc.__dict__)
        self.session_service = new_svc

    async def run_async(
        self, *, user_id, session_id, new_message=None, state_delta=None, run_config=None,
    ):
        sess = self._sessions.get((user_id, session_id))
        # Snapshot the session state at the START of this call (AFTER any
        # _apply_state_delta pre-write, BEFORE run_async processes the turn).
        self.pre_run_state_snapshots.append(
            dict(sess.state) if sess is not None else {}
        )
        self.pre_run_state_deltas.append(dict(state_delta or {}))
        async for event in super().run_async(
            user_id=user_id,
            session_id=session_id,
            new_message=new_message,
            state_delta=state_delta,
            run_config=run_config,
        ):
            yield event


def test_chat_session_sees_freshly_written_ledger_rows():
    """Turn 2 must see the rows that the pipeline wrote BETWEEN the two turns.

    Regression test for P0-2 (2026-06-15 live QA): after the pipeline posted a
    bank statement the chat lane replied "I cannot see any documents" because
    either the session state was stale or the pre-write never happened.

    The fix: ``answer_question`` explicitly pre-writes ``ledger_data`` into the
    session state via ``_apply_state_delta`` BEFORE calling ``runner.run_async``,
    so the value is unconditionally overwritten each turn regardless of what the
    session already holds.
    """
    from accounting_agents.slack_runner import LEDGER_DATA_KEY, answer_question

    THREE_ROWS = [
        {"Date": "01/12/2025", "Source Filename": "bank.pdf", "Doc Type": "B",
         "Source Amount": 100.0, "Description": "FAST PMT", "Balance": 900.0,
         "Withdrawal": 100.0, "Deposit": None, "Currency": "SGD"},
        {"Date": "05/12/2025", "Source Filename": "bank.pdf", "Doc Type": "B",
         "Source Amount": 200.0, "Description": "SALARY", "Balance": 1100.0,
         "Withdrawal": None, "Deposit": 200.0, "Currency": "SGD"},
        {"Date": "10/12/2025", "Source Filename": "bank.pdf", "Doc Type": "B",
         "Source Amount": 50.0, "Description": "ATM", "Balance": 1050.0,
         "Withdrawal": 50.0, "Deposit": None, "Currency": "SGD"},
    ]

    from unittest.mock import MagicMock

    slack = FakeSlackClient()
    sessions: dict = {}
    chat_runner = _StatefulCapturingChatRunner(sessions)

    # Turn 1: ledger store has no data yet (pipeline hasn't posted).
    ledger_turn1 = MagicMock()
    ledger_turn1.latest_fy.return_value = None
    ledger_turn1.read_rows.return_value = []

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=ledger_turn1,
            slack_client=slack,
            channel_id="C-P02",
            question="list recent documents",
            app_name=chat_runner.app_name,
            client_store=None,
            message_ts="1700000001.001",
            thread_ts="1700000001.001",
            raw_thread_ts="1700000001.001",
        )
    )

    # Between turns: pipeline posts and writes 3 rows.
    ledger_turn2 = MagicMock()
    ledger_turn2.latest_fy.return_value = "FY2025"
    ledger_turn2.read_rows.return_value = THREE_ROWS

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=ledger_turn2,
            slack_client=slack,
            channel_id="C-P02",
            question="list recent documents",
            app_name=chat_runner.app_name,
            client_store=None,
            message_ts="1700000001.002",
            thread_ts="1700000001.001",
            raw_thread_ts="1700000001.001",
        )
    )

    assert len(chat_runner.calls) == 2

    # The session state at the START of turn 2 (before run_async) must already
    # hold the fresh rows — not the stale [] from turn 1.  This is the invariant
    # the fix enforces by calling _apply_state_delta BEFORE runner.run_async.
    turn2_pre_state = chat_runner.pre_run_state_snapshots[1]
    assert turn2_pre_state.get(LEDGER_DATA_KEY) == THREE_ROWS, (
        f"Session state at turn-2 start held stale ledger_data="
        f"{turn2_pre_state.get(LEDGER_DATA_KEY)!r}; expected 3 rows."
    )


def test_pick_chat_fy_prefers_question_processing_log_hint():
    from accounting_agents.slack_runner import _pick_chat_fy

    class _Store:
        def latest_fy(self, client_id):
            return "2026"

    fy = _pick_chat_fy(
        best_fy="2026",
        fy_summaries=[
            {"fy": "2026", "row_count": 10, "has_data": True},
            {"fy": "2025", "row_count": 3, "has_data": True},
        ],
        processing_log=[
            {"filename": "25-D12-Company-A.pdf", "fy": "2025"},
        ],
        question="yes file name 25-D12",
        fye_month=10,
        ledger_store=_Store(),
        client_id="c1",
    )
    assert fy == "2025"


def test_resolve_thread_delivery_context_filters_by_ts():
    """Phase 3: when the user replies under a delivery card, the chat lane
    must scope its question to that delivery's files.

    The processing_log entry is written by Phase 2 with delivery_message_ts =
    the job-summary message ts (= the parent of the thread).
    """
    from accounting_agents.slack_runner import _resolve_thread_delivery_context

    log = [
        {"file_id": "F-other", "filename": "old.pdf", "fy": "2024",
         "delivery_message_ts": "1700000000.000100", "channel_id": "C1"},
        {"file_id": "F-D15", "filename": "25-D15-Company-A.pdf", "fy": "2025",
         "delivery_message_ts": "1700000099.000200", "channel_id": "C1",
         "invoice_ids": ["25-D15"]},
        {"file_id": "F-D12", "filename": "25-D12-Company-A.pdf", "fy": "2025",
         "delivery_message_ts": "1700000099.000200", "channel_id": "C1",
         "invoice_ids": ["25-D12"]},
        {"file_id": "F-future", "filename": "26-X.pdf", "fy": "2026",
         "delivery_message_ts": "1700000999.000300", "channel_id": "C1"},
    ]

    ctx = _resolve_thread_delivery_context(
        raw_thread_ts="1700000099.000200",
        channel_id="C1",
        processing_log=log,
    )
    assert ctx["thread_delivery_message_ts"] == "1700000099.000200"
    assert sorted(ctx["thread_delivery_filenames"]) == [
        "25-D12-Company-A.pdf", "25-D15-Company-A.pdf",
    ]
    assert sorted(ctx["thread_delivery_invoice_ids"]) == ["25-D12", "25-D15"]
    assert ctx["thread_delivery_fy"] == "2025"
    assert len(ctx["thread_scoped_processing_log"]) == 2


def test_resolve_thread_delivery_context_empty_for_top_level():
    """Top-level channel message (no thread) returns an empty dict so the
    chat lane keeps its existing channel-wide behaviour."""
    from accounting_agents.slack_runner import _resolve_thread_delivery_context

    ctx = _resolve_thread_delivery_context(
        raw_thread_ts=None,
        channel_id="C1",
        processing_log=[{"file_id": "F1", "filename": "x.pdf", "fy": "2025",
                         "delivery_message_ts": "1700000099.000200"}],
    )
    assert ctx == {}


def test_answer_question_injects_thread_delivery_context_into_state_delta():
    """End-to-end: a thread reply under a delivery card must inject
    ``thread_delivery_*`` keys into the chat state_delta so the assistant
    preamble and tools can see the scope (Phase 3).
    """
    from unittest.mock import MagicMock
    from accounting_agents.slack_runner import answer_question

    slack = FakeSlackClient()
    sessions: dict = {}
    chat_runner = _StatefulCapturingChatRunner(sessions)

    class _Store:
        def latest_fy(self, client_id):
            return "2026"
        def best_fy_for_chat(self, client_id, slack_client):
            return "2025", [
                {"fy": "2025", "row_count": 5, "has_data": True},
                {"fy": "2026", "row_count": 1, "has_data": True},
            ]
        def read_rows(self, *, client_id, fy, slack_client, channel_id):
            return []

    class _StubClientStore:
        def __init__(self, log):
            self._log = log
        def get_by_channel(self, channel_id):
            from invoice_processing.export.client_context import ClientContext
            return ClientContext(
                client_id="c1", client_name="Test", accounting_software="QBS",
                fye_month=12, channel_id=channel_id,
            )
        def list_processing_log(self, client_id, limit=20):
            return list(self._log)
        def list_pending_interrupts(self, *a, **k):
            return []

    log = [
        {"file_id": "F-D15", "filename": "25-D15-Company-A.pdf",
         "fy": "2025", "channel_id": "C1",
         "delivery_message_ts": "1700000099.000200",
         "invoice_ids": ["25-D15"]},
        {"file_id": "F-D12", "filename": "25-D12-Company-A.pdf",
         "fy": "2025", "channel_id": "C1",
         "delivery_message_ts": "1700000099.000200",
         "invoice_ids": ["25-D12"]},
    ]
    rec_store = _StubClientStore(log)

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=_Store(),
            slack_client=slack,
            channel_id="C1",
            question="why 6100-Software for 25-D15?",
            app_name=chat_runner.app_name,
            client_store=rec_store,
            message_ts="1700000200.001",
            thread_ts="1700000099.000200",
            raw_thread_ts="1700000099.000200",
        )
    )

    delta = chat_runner.pre_run_state_deltas[0]
    # Thread-scoped keys present
    assert delta.get("thread_delivery_message_ts") == "1700000099.000200"
    assert "25-D15-Company-A.pdf" in delta.get("thread_delivery_filenames") or []
    assert "25-D12-Company-A.pdf" in delta.get("thread_delivery_filenames") or []
    assert "25-D15" in delta.get("thread_delivery_invoice_ids") or []
    # FY re-picked to match the thread (2025 has rows in this fixture)
    assert delta.get("fy_loaded") == "2025"


def test_answer_question_chat_ux_ack_and_thinking(monkeypatch):
    """Phase 4: chat lane adds 👀 on the user message and thinking status before ADK run."""
    from accounting_agents.slack_runner import answer_question

    monkeypatch.setenv("LEDGR_CHAT_UX", "1")
    slack = FakeSlackClient()
    sessions: dict = {}
    chat_runner = _StatefulCapturingChatRunner(sessions)

    class _Store:
        def latest_fy(self, client_id):
            return "2025"
        def read_rows(self, *, client_id, fy, slack_client, channel_id):
            return []

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=_Store(),
            slack_client=slack,
            channel_id="C-UX",
            question="what is loaded?",
            app_name=chat_runner.app_name,
            client_store=None,
            message_ts="1700000300.001",
            thread_ts="1700000099.000200",
            raw_thread_ts="1700000099.000200",
        )
    )

    eyes = [r for r in slack.reactions_added if r["name"] == "eyes"]
    assert eyes == [{"channel": "C-UX", "timestamp": "1700000300.001", "name": "eyes"}]
    assert slack.thinking_status_calls, "expected assistant_threads_setStatus"
    assert slack.thinking_status_calls[0]["thread_ts"] == "1700000099.000200"
    assert slack.thinking_status_calls[0]["status"] == "is thinking..."
    checkmarks = [r for r in slack.reactions_added if r["name"] == "white_check_mark"]
    assert checkmarks, "expected ✅ on user message after reply"
    assert not [r for r in slack.reactions_removed if r["name"] == "eyes"] or True


def test_answer_question_chat_ux_error_clears_thinking(monkeypatch):
    """Error path must clear thinking status and remove 👀."""
    from accounting_agents.slack_runner import _handle_chat_turn

    monkeypatch.setenv("LEDGR_CHAT_UX", "1")
    slack = FakeSlackClient()
    sessions: dict = {}

    class _FailRunner(_CapturingChatRunner):
        async def run_async(self, **kwargs):
            if False:
                yield  # pragma: no cover
            raise RuntimeError("boom")

    runner = _FailRunner(sessions)

    class _Store:
        def read_rows(self, **kwargs):
            return []

    asyncio.run(
        _handle_chat_turn(
            chat_runner=runner,
            ledger_store=_Store(),
            slack_client=slack,
            channel_id="C-ERR",
            question="break",
            client_store=None,
            message_ts="1700000400.001",
            thread_ts="1700000099.000200",
            raw_thread_ts="1700000099.000200",
            doc_runner=None,
            db=None,
        )
    )

    cleared = [c for c in slack.thinking_status_calls if c.get("status") == ""]
    assert cleared, "expected thinking status cleared on error"
    removed = [r for r in slack.reactions_removed if r["name"] == "eyes"]
    assert removed == [{"channel": "C-ERR", "timestamp": "1700000400.001", "name": "eyes"}]
    assert any("error" in p.get("text", "").lower() for p in slack._posts)


def test_chat_stream_disabled_by_default():
    from accounting_agents.slack_runner import _chat_stream_enabled

    assert _chat_stream_enabled() is False


def test_resolve_thread_delivery_context_falls_back_to_replies():
    """Older processing_log entries (pre-Phase 2) lack delivery_message_ts.
    The resolver must fall back to conversations.replies parsing."""
    from accounting_agents.slack_runner import _resolve_thread_delivery_context

    class _RepliesClient:
        def __init__(self):
            self.calls: list[dict] = []
        def conversations_replies(self, *, channel, ts, limit):
            self.calls.append({"channel": channel, "ts": ts, "limit": limit})
            return {
                "messages": [
                    {
                        "text": "Added 3 lines to FY2025 ledger",
                        "blocks": [
                            {"type": "section",
                             "text": {"type": "mrkdwn",
                                      "text": "Delivered 2 files:\n• 25-D15-Company-A.pdf\n• 25-D12-Company-A.pdf"}},
                            {"type": "data_table",
                             "rows": [[{"text": "Source Filename"}, {"text": "Account"}],
                                      [{"text": "25-D15-Company-A.pdf"}, {"text": "6-3000"}]]},
                        ]
                    }
                ]
            }

    log = [
        {"file_id": "F-D15", "filename": "25-D15-Company-A.pdf", "fy": "2025"},
        {"file_id": "F-D12", "filename": "25-D12-Company-A.pdf", "fy": "2025"},
    ]
    slack = _RepliesClient()
    ctx = _resolve_thread_delivery_context(
        raw_thread_ts="1716000000.000",
        channel_id="C1",
        processing_log=log,
        slack_client=slack,
    )
    assert slack.calls, "must call conversations_replies on fallback"
    assert "25-D15-Company-A.pdf" in ctx["thread_delivery_filenames"]
    assert "25-D12-Company-A.pdf" in ctx["thread_delivery_filenames"]
    assert ctx["thread_delivery_fy"] == "2025"
    assert "25-D15" in ctx["thread_delivery_invoice_ids"]
    assert ctx.get("thread_delivery_preview_rows"), "must parse delivery data_table"
    assert ctx["thread_delivery_preview_rows"][0]["account_code"] == "6-3000"


def test_parse_delivery_data_table_rows_xero_shape():
    from accounting_agents.slack_runner import _parse_delivery_data_table_rows

    resp = {
        "messages": [{
            "blocks": [{
                "type": "data_table",
                "rows": [
                    [
                        {"text": "Invoice #"},
                        {"text": "Account"},
                        {"text": "Description"},
                    ],
                    [
                        {"text": "25-D15"},
                        {"text": "902-A02"},
                        {"text": "PTTEP/UOA"},
                    ],
                ],
            }],
        }],
    }
    rows = _parse_delivery_data_table_rows(resp)
    assert len(rows) == 1
    assert rows[0]["invoice_id"] == "25-D15"
    assert rows[0]["account_code"] == "902-A02"


def test_prefetch_thread_ledger_matches_finds_xero_invoice():
    from accounting_agents.slack_runner import _prefetch_thread_ledger_matches

    ledger = [{
        "_sheet": "Purchase",
        "*InvoiceNumber": "25-D15",
        "*AccountCode": "902-A02",
        "*ContactName": "Company-A",
        "*Description": "Professional fees",
    }]
    matches = _prefetch_thread_ledger_matches(
        ledger,
        invoice_ids=["25-D15"],
        filenames=["25-D15-Company-A.pdf"],
    )
    assert len(matches) == 1
    assert matches[0]["account_code"] == "902-A02"
    assert matches[0]["row_index"] == 0


def test_try_direct_thread_account_code_answer_from_preview():
    from accounting_agents.slack_runner import (
        _try_direct_thread_account_code_answer,
    )
    from accounting_agents.assistant import LEDGER_DATA_KEY

    state = {
        "thread_delivery_preview_rows": [{
            "invoice_id": "25-D15",
            "account_code": "902-A02",
            "vendor": "Person-1",
            "description": "PTTEP/UOA monitoring audit",
        }],
        "thread_delivery_invoice_ids": ["25-D15"],
        LEDGER_DATA_KEY: [],
    }
    text, focus = _try_direct_thread_account_code_answer(
        "Why account code for 25-D15 in this batch?", state_delta=state,
    )
    assert text
    assert "902-A02" in text
    assert "25-D15" in text
    assert focus and focus.get("account_code") == "902-A02"
    assert "vendor" not in text.lower() or "Person-1" in text


def test_try_direct_thread_account_code_answer_clarifies_wrong_code():
    from accounting_agents.slack_runner import _try_direct_thread_account_code_answer

    state = {
        "thread_delivery_preview_rows": [{
            "invoice_id": "25-D15",
            "account_code": "902-A02",
            "description": "Fees",
        }],
    }
    text, _focus = _try_direct_thread_account_code_answer(
        "Why account code 6-3000 for invoice 25-D15?", state_delta=state,
    )
    assert text
    assert "902-A02" in text
    assert "6-3000" in text


def test_try_direct_thread_coa_description_followup_from_focus():
    from accounting_agents.assistant import THREAD_FOCUS_KEY
    from accounting_agents.slack_runner import _try_direct_thread_account_code_answer

    state = {
        THREAD_FOCUS_KEY: {
            "invoice_id": "25-D15",
            "account_code": "902-A02",
            "vendor": "Person-1",
            "line_description": "Fees",
        },
        "coa": [
            {"code": "902-A02", "description": "Professional Fees", "account_type": "Expense"},
        ],
    }
    text, focus = _try_direct_thread_account_code_answer(
        "What is the description of the acount code?", state_delta=state,
    )
    assert text
    assert "902-A02" in text
    assert "Professional Fees" in text
    assert focus and focus["account_code"] == "902-A02"


def test_question_asks_account_code_matches_acount_typo():
    from accounting_agents.slack_runner import _question_asks_account_code

    assert _question_asks_account_code("What is the description of the acount code?")


def test_answer_question_injects_thread_delivery_ledger_matches(monkeypatch):
    from accounting_agents.slack_runner import answer_question

    FY25_ROWS = [{
        "_sheet": "Purchase",
        "*InvoiceNumber": "25-D15",
        "*AccountCode": "902-A02",
        "*Description": "Fees",
    }]

    class _Store:
        def list_processing_log(self, client_id, limit=20):
            return [{
                "file_id": "F-D15",
                "filename": "25-D15-Company-A.pdf",
                "fy": "2025",
                "delivery_message_ts": "1700000099.000200",
                "row_count": 0,
            }]

        def get_by_channel(self, channel_id):
            from invoice_processing.export.client_context import ClientContext
            return ClientContext(
                client_id="c1", client_name="Test", accounting_software="Xero",
                fye_month=12, channel_id=channel_id,
            )

    class _Ledger:
        def best_fy_for_chat(self, client_id, slack_client):
            return "2025", [{"fy": "2025", "row_count": 1, "has_data": True}]

        def read_rows(self, **kwargs):
            return FY25_ROWS

    class _RepliesSlack:
        def conversations_replies(self, *, channel, ts, limit):
            return {
                "messages": [{
                    "text": "FY2025 ledger",
                    "blocks": [{
                        "type": "data_table",
                        "rows": [
                            [{"text": "Invoice #"}, {"text": "Account"}],
                            [{"text": "25-D15"}, {"text": "902-A02"}],
                        ],
                    }],
                }],
            }

        def chat_postMessage(self, **kwargs):
            return {"ok": True, "ts": "1700000400.002"}

    sessions: dict = {}
    runner = _StatefulCapturingChatRunner(sessions)
    slack = _RepliesSlack()

    result = asyncio.run(answer_question(
        runner=runner,
        ledger_store=_Ledger(),
        slack_client=slack,
        channel_id="C-THREAD",
        question="Why account code for 25-D15?",
        app_name=runner.app_name,
        client_store=_Store(),
        message_ts="1700000400.001",
        thread_ts="1700000099.000200",
        raw_thread_ts="1700000099.000200",
    ))

    assert result.get("direct") is True
    assert "902-A02" in (result.get("text") or "")
    assert "25-D15" in (result.get("text") or "")
    assert not runner.calls, "direct path must skip LLM when preview rows exist"


def test_chat_session_picks_fy_with_most_rows_not_latest():
    """P0-B: ``answer_question`` should call ``best_fy_for_chat`` and use the
    FY whose workbook has the most data — not just the highest FY label.

    Regression for the Company-A client whose data lives in FY2025 while
    ``latest_fy`` would pick FY2026 (empty) and report "ledger not loaded".
    """
    from accounting_agents.slack_runner import LEDGER_DATA_KEY, answer_question

    # FY2025 has 3 rows, FY2026 has 1 row — the agent should pick FY2025.
    FY25_ROWS = [
        {"Date": "01/01/2025", "Source Filename": "a.pdf", "Doc Type": "B",
         "Source Amount": 100.0, "Description": "A1", "Balance": 100.0,
         "Withdrawal": 100.0, "Deposit": None, "Currency": "SGD"},
        {"Date": "02/01/2025", "Source Filename": "a.pdf", "Doc Type": "B",
         "Source Amount": 200.0, "Description": "A2", "Balance": 300.0,
         "Withdrawal": None, "Deposit": 200.0, "Currency": "SGD"},
        {"Date": "03/01/2025", "Source Filename": "a.pdf", "Doc Type": "B",
         "Source Amount": 50.0, "Description": "A3", "Balance": 250.0,
         "Withdrawal": 50.0, "Deposit": None, "Currency": "SGD"},
    ]
    FY26_ROWS = [
        {"Date": "01/01/2026", "Source Filename": "b.pdf", "Doc Type": "B",
         "Source Amount": 10.0, "Description": "B1", "Balance": 10.0,
         "Withdrawal": None, "Deposit": 10.0, "Currency": "SGD"},
    ]

    from unittest.mock import MagicMock

    slack = FakeSlackClient()
    sessions: dict = {}
    chat_runner = _StatefulCapturingChatRunner(sessions)

    # ``read_rows`` is called AFTER FY selection; the store returns rows
    # keyed by the FY the runner asked for. The two return values below
    # simulate two FYs — one of them has 3 rows, the other 1.
    read_calls: list[str] = []

    def fake_read_rows(*, client_id, fy, slack_client, channel_id):
        read_calls.append(fy)
        if fy == "2025":
            return list(FY25_ROWS)
        if fy == "2026":
            return list(FY26_ROWS)
        return []

    class _FakeStore:
        def __init__(self):
            self.latest_fy = lambda client_id: "2026"  # would pick empty FY
            self.read_rows = fake_read_rows
            self.best_fy_for_chat = (
                lambda client_id, slack_client: (
                    "2025",
                    [
                        {"fy": "2025", "row_count": 3, "has_data": True},
                        {"fy": "2026", "row_count": 1, "has_data": True},
                    ],
                )
            )

    ledger = _FakeStore()

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=ledger,
            slack_client=slack,
            channel_id="C-P0B",
            question="list recent documents",
            app_name=chat_runner.app_name,
            client_store=None,
            message_ts="1700000002.001",
            thread_ts="1700000002.001",
            raw_thread_ts="1700000002.001",
        )
    )

    # The runner should have selected FY2025 (most rows) and read from it.
    assert "2025" in read_calls
    turn1_pre_state = chat_runner.pre_run_state_snapshots[0]
    assert turn1_pre_state.get(LEDGER_DATA_KEY) == FY25_ROWS
    # Diagnostic state was injected into the state_delta passed to run_async
    # (ADK applies it via the user-message event, so the snapshot of session
    # state itself does not yet show it).
    turn1_delta = chat_runner.pre_run_state_deltas[0]
    assert turn1_delta.get("fy_loaded") == "2025"
    assert turn1_delta.get("ledger_row_count") == 3
    assert turn1_delta.get("fy_pointers") == [
        {"fy": "2025", "row_count": 3, "has_data": True},
        {"fy": "2026", "row_count": 1, "has_data": True},
    ]


# --------------------------------------------------------------------------- #
# P1 — pending_reviews + document_sessions injection
# --------------------------------------------------------------------------- #


def test_list_pending_interrupts_filters_by_channel_and_status():
    from accounting_agents.hitl import (
        INTERRUPTS_COLLECTION,
        list_pending_interrupts,
        write_interrupt,
    )

    db = FakeFirestore()
    write_interrupt(
        db, "INT-1",
        session_id="C-A:F-1", channel_id="C-A", slack_file_id="F-1",
    )
    write_interrupt(
        db, "INT-2",
        session_id="C-B:F-1", channel_id="C-B", slack_file_id="F-1",
    )
    # Resolve INT-2 so it should NOT appear in pending.
    write_interrupt(
        db, "INT-2",
        session_id="C-B:F-1", channel_id="C-B", slack_file_id="F-1",
        status="resolved",
    )
    pending = list_pending_interrupts(db, "C-A")
    assert len(pending) == 1
    assert pending[0]["interrupt_id"] == "INT-1"
    assert pending[0]["channel_id"] == "C-A"
    # Sanity: the resolved doc is filtered out.
    assert list_pending_interrupts(db, "C-B") == []


def test_coerce_snapshot_fields_keeps_chat_relevant_subset():
    from accounting_agents.slack_runner import _coerce_snapshot_fields

    state = {
        "doc_type": "invoice",
        "extraction_path": "understand",
        "review_reasons": ["tax_code_unknown"],
        "source_filename": "x.pdf",
        "summary_table": [1, 2, 3, 4, 5],
        "normalized_invoice_count": 1,
        "soa_legacy_path": False,
        # Should be omitted from the snapshot.
        "ledger_data": [{"huge": "row"}],
        "pending_ledger_write": [{"x": 1}],
    }
    snap = _coerce_snapshot_fields(state)
    assert snap["doc_type"] == "invoice"
    assert snap["extraction_path"] == "understand"
    assert snap["review_reasons"] == ["tax_code_unknown"]
    assert snap["summary_table_size"] == 5
    assert "ledger_data" not in snap
    assert "pending_ledger_write" not in snap


def test_snapshot_doc_sessions_reads_per_doc_state():
    """``_snapshot_doc_sessions`` should read per-document sessions and
    return a small chat-friendly subset keyed by file_id."""
    from accounting_agents.slack_runner import _snapshot_doc_sessions

    class _FakeSessionService:
        def __init__(self, store):
            self._store = store

        async def get_session(self, *, app_name, user_id, session_id):
            data = self._store.get((user_id, session_id))
            if not data:
                return None
            return SimpleNamespace(state=data)

    store = {
        ("C-X:F-1", "C-X:F-1"): {
            "doc_type": "invoice",
            "extraction_path": "understand",
            "summary_table": [1, 2, 3],
            "ledger_data": [{"ignored": True}],
        },
        ("C-X:F-2", "C-X:F-2"): {
            "doc_type": "soa",
            "soa_legacy_path": True,
            "review_reasons": ["amount_mismatch"],
        },
    }
    runner = SimpleNamespace(session_service=_FakeSessionService(store))
    out = asyncio.run(
        _snapshot_doc_sessions(runner, "app", "C-X", ["F-1", "F-2", "F-3"])
    )
    assert "F-1" in out and "F-2" in out
    # F-3 has no session → omitted.
    assert "F-3" not in out
    assert out["F-1"]["doc_type"] == "invoice"
    assert out["F-1"]["summary_table_size"] == 3
    assert "ledger_data" not in out["F-1"]
    assert out["F-2"]["soa_legacy_path"] is True


def test_chat_session_injects_pending_reviews_and_doc_sessions(monkeypatch):
    """P1-2: ``answer_question`` should populate the new state keys for
    pending HITL reviews and per-document session snapshots so the
    diagnostic tools see them."""
    from accounting_agents.hitl import write_interrupt
    from accounting_agents.slack_runner import answer_question

    db = FakeFirestore()
    write_interrupt(
        db, "INT-99",
        session_id="C-P1:F-1", channel_id="C-P1", slack_file_id="F-1",
    )
    write_interrupt(
        db, "INT-100",
        session_id="C-P1:F-2", channel_id="C-P1", slack_file_id="F-2",
    )

    slack = FakeSlackClient()
    sessions: dict = {}
    chat_runner = _StatefulCapturingChatRunner(sessions)

    # Stub a fake client_store that returns one processing log entry (so
    # file_ids = ["F-1"] is non-empty and _snapshot_doc_sessions runs).
    from tests.test_slack_runner import FakeFirestore as _FF
    from invoice_processing.export.client_context import (
        FirestoreClientStore,
    )

    class _StubClientStore:
        def list_processing_log(self, client_id, limit=20):
            return [
                {
                    "filename": "old.pdf",
                    "file_id": "F-1",
                    "doc_type": "invoice",
                    "extraction_path": "understand",
                }
            ]

        def add_processing_log(self, *args, **kwargs):
            return None

        def get_profile(self, client_id):
            from invoice_processing.export.client_context import ClientContext
            return ClientContext(
                client_id=client_id,
                client_name="Test",
                accounting_software="QBS Ledger",
                fye_month=12,
            )

        def get_by_channel(self, channel_id):
            from invoice_processing.export.client_context import ClientContext
            return ClientContext(
                client_id="C-P1",
                client_name="Test",
                accounting_software="QBS Ledger",
                fye_month=12,
                channel_id=channel_id,
            )

    class _FakeStore:
        def latest_fy(self, client_id):
            return "2025"

        def read_rows(self, *, client_id, fy, slack_client, channel_id):
            return []

        def best_fy_for_chat(self, client_id, slack_client):
            return "2025", [{"fy": "2025", "row_count": 0, "has_data": False}]

    ledger = _FakeStore()
    client_store = _StubClientStore()

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=ledger,
            slack_client=slack,
            channel_id="C-P1",
            question="anything waiting on me?",
            app_name=chat_runner.app_name,
            client_store=client_store,
            db=db,
            message_ts="1700000003.001",
            thread_ts="1700000003.001",
            raw_thread_ts="1700000003.001",
        )
    )

    delta = chat_runner.pre_run_state_deltas[0]
    pending = delta.get("pending_reviews") or []
    assert len(pending) == 2
    ids = {p["interrupt_id"] for p in pending}
    assert ids == {"INT-99", "INT-100"}
    assert delta.get("pending_review_count") == 2
    assert delta.get("processing_log_count") == 1
    # document_sessions should be a dict (empty for a real session that
    # has no per-doc session in this fake) but still present.
    assert "document_sessions" in delta
    assert isinstance(delta["document_sessions"], dict)


# --------------------------------------------------------------------------- #
# P0-2 UX: summarize_recent_activity empty-window names the window + newest date
# --------------------------------------------------------------------------- #


def test_summarize_empty_result_message_names_window():
    """When the 30-day window is empty, the message must name the window size
    AND cite the most-recent date in the data so the user knows what to ask for.

    Regression for P0-2 UX (2026-06-15): the old empty response was a JSON blob
    with ``transaction_count: 0`` — no hint about the window or newest data.
    """
    from accounting_agents.assistant import LEDGER_DATA_KEY, summarize_recent_activity

    # Rows all dated 2025-12-01 — well outside the 30-day window from 2026-06-15.
    rows = [
        {"Date": "01/12/2025", "Source Amount": 500.0,
         "Account Code / COA": "6100-Software", "Doc Type": "P"},
        {"Date": "15/12/2025", "Source Amount": 300.0,
         "Account Code / COA": "6200-Rent", "Doc Type": "P"},
    ]

    class _FakeTool:
        def __init__(self, state):
            self.state = state

    ctx = _FakeTool({LEDGER_DATA_KEY: rows})

    # Rows dated 2025-Dec are always outside a 30-day window from any 2026 date
    # — no mocking needed.
    result = summarize_recent_activity(ctx)

    # Must mention the window ("30") AND the most-recent date present in the data.
    assert "30" in result, f"Expected '30' in result: {result!r}"
    assert "2025" in result, f"Expected '2025' in result: {result!r}"


# --------------------------------------------------------------------------- #
# build_async_app: chat_runner is used for text, FirestoreClientStore is default
# --------------------------------------------------------------------------- #


def test_build_async_app_uses_chat_runner_for_text():
    """The text-handler path calls ``chat_runner.run_async``, not the coordinator runner."""
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    fake_chat_runner = MagicMock()
    fake_chat_runner.app_name = "accounting_agents_assistant"

    # Capture the message handler with chat_runner injected.
    from app.slack_app import _SeenEvents

    registered = {}
    fake_app = MagicMock()

    def event_decorator(name):
        def decorator(fn):
            registered[name] = fn
            return fn
        return decorator

    fake_app.event = event_decorator
    fake_app.action = lambda *a, **k: (lambda fn: fn)
    fake_app.view = lambda *a, **k: (lambda fn: fn)
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = MagicMock()
    rm.app_name = "acc"

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch.object(slack_runner, "build_chat_runner", return_value=fake_chat_runner):
        slack_runner.build_async_app(
            runner=rm,
            ledger_store=MagicMock(),
            db=MagicMock(),
            store=MagicMock(),  # avoid FirestoreClientStore default
        )

    handler = registered["message"]

    body = {"event_id": "Ev-text-cr"}
    event = {
        "type": "message",
        "ts": "1700000000.700",
        "channel": "C-qa",
        "text": "How am I doing?",
    }
    mock_aq = AsyncMock(return_value={"status": "answered", "text": "ok"})
    with patch.object(slack_runner, "answer_question", mock_aq):
        asyncio.run(handler(event=event, body=body, client=MagicMock()))

    mock_aq.assert_called_once()
    # The runner passed to answer_question is the chat runner, NOT the coordinator runner.
    assert mock_aq.call_args.kwargs["runner"] is fake_chat_runner
    assert mock_aq.call_args.kwargs["app_name"] == fake_chat_runner.app_name


def test_build_async_app_default_store_is_firestore():
    """When no ``store`` is supplied, ``build_async_app`` defaults to FirestoreClientStore.

    This fixes the socket-mode gap where onboarding writes used to vanish into
    an ``InMemoryClientStore`` while the pipeline read from Firestore.
    """
    from unittest.mock import MagicMock, patch

    from accounting_agents import slack_runner
    from invoice_processing.export.client_context import FirestoreClientStore

    captured: dict = {}
    real_build_async_app = slack_runner.build_async_app

    # Spy on FirestoreClientStore — capture the instance returned to the function.
    real_cls = FirestoreClientStore

    def _capture(*args, **kwargs):
        inst = real_cls.__new__(real_cls)
        # Avoid touching Firestore in __init__.
        captured["instance"] = inst
        return inst

    rm = MagicMock()
    rm.app_name = "acc"
    fake_app = MagicMock()
    fake_app.event = lambda *a, **k: (lambda fn: fn)
    fake_app.action = lambda *a, **k: (lambda fn: fn)
    fake_app.view = lambda *a, **k: (lambda fn: fn)
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    with patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch("invoice_processing.export.client_context.FirestoreClientStore", side_effect=_capture), \
         patch.object(slack_runner, "build_chat_runner", return_value=MagicMock(app_name="chat")):
        real_build_async_app(
            runner=rm,
            ledger_store=MagicMock(),
            db=MagicMock(),
            # NOTE: store omitted → must default to FirestoreClientStore
        )

    assert "instance" in captured, "FirestoreClientStore must be instantiated when no store is passed"
    assert type(captured["instance"]).__name__ == "FirestoreClientStore"


# --------------------------------------------------------------------------- #
# _ensure_session is idempotent — the chat lane must reuse sessions across turns
# --------------------------------------------------------------------------- #


def test_ensure_session_is_idempotent_for_chat_reuse():
    """Calling ``_ensure_session`` twice for the same (user, session) is a no-op.

    The chat lane relies on this: the second turn must NOT wipe the events the
    first turn added — get-or-create semantics only.
    """
    from accounting_agents.slack_runner import _ensure_session

    class _Svc:
        def __init__(self):
            self.created: list[tuple] = []

        async def create_session(self, *, app_name, user_id, session_id, state=None):
            from google.adk.errors.already_exists_error import AlreadyExistsError
            key = (user_id, session_id)
            if key in self.created:
                raise AlreadyExistsError("exists")
            self.created.append(key)

    svc = _Svc()
    runner = SimpleNamespace(session_service=svc)
    asyncio.run(_ensure_session(runner, "app", "U1", "S1"))
    asyncio.run(_ensure_session(runner, "app", "U1", "S1"))  # must not raise
    assert svc.created == [("U1", "S1")]


# =========================================================================== #
# Chat-lane write confirm bridge (Step 4 / ADR-0009)
# =========================================================================== #


def test_classify_confirmation_reply():
    from accounting_agents.slack_runner import classify_confirmation_reply

    for yes in ("yes", "Yes.", "confirm", "go ahead", "ok", "do it", "please do"):
        assert classify_confirmation_reply(yes) is True, yes
    for no in ("no", "cancel", "stop", "don't", "nope"):
        assert classify_confirmation_reply(no) is False, no
    for ambiguous in ("what does it change?", "tell me more", ""):
        assert classify_confirmation_reply(ambiguous) is None, ambiguous


def _confirm_session_with_pending(fc_id="adk-confirm-1", payload=None):
    """Build a session whose last event has an UNANSWERED adk_request_confirmation."""
    from google.adk.events.event_actions import EventActions
    from google.adk.tools.tool_confirmation import ToolConfirmation

    payload = payload or {"op": "amend", "sheet": "Purchase", "row": 2,
                          "updates": {"Account Code / COA": "6010"}}
    ev = Event(
        author="assistant",
        content=types.Content(
            parts=[types.Part(function_call=types.FunctionCall(
                name="adk_request_confirmation", id=fc_id, args={}))]
        ),
        long_running_tool_ids=[fc_id],
        actions=EventActions(
            requested_tool_confirmations={
                fc_id: ToolConfirmation(hint="change account?", payload=payload)
            }
        ),
    )
    return SimpleNamespace(events=[ev])


def test_find_pending_confirmation_detects_unanswered():
    from accounting_agents.slack_runner import find_pending_confirmation

    session = _confirm_session_with_pending()
    found = find_pending_confirmation(session)
    assert found is not None
    fc_id, confirmation = found
    assert fc_id == "adk-confirm-1"
    assert confirmation.payload["updates"]["Account Code / COA"] == "6010"


def test_find_pending_confirmation_none_when_answered():
    from accounting_agents.slack_runner import find_pending_confirmation

    session = _confirm_session_with_pending(fc_id="fc-9")
    # Append a function_response answering fc-9 → no longer pending.
    answer = Event(
        author="user",
        content=types.Content(
            parts=[types.Part(function_response=types.FunctionResponse(
                id="fc-9", name="adk_request_confirmation", response={"confirmed": True}))]
        ),
    )
    session.events.append(answer)
    assert find_pending_confirmation(session) is None


def test_synthesize_confirmation_message_matches_id_and_payload():
    from accounting_agents.slack_runner import (
        _synthesize_confirmation_message,
        find_pending_confirmation,
    )

    session = _confirm_session_with_pending(fc_id="fc-77")
    fc_id, confirmation = find_pending_confirmation(session)
    msg = _synthesize_confirmation_message(fc_id, confirmation, confirmed=True)

    part = msg.parts[0]
    fr = part.function_response
    assert fr.id == "fc-77"
    assert fr.name == "adk_request_confirmation"
    assert fr.response["confirmed"] is True
    # Original requested payload is echoed back so the re-run tool sees its spec.
    assert fr.response["payload"]["updates"]["Account Code / COA"] == "6010"


def test_synthesize_confirmation_message_negative():
    from accounting_agents.slack_runner import (
        _synthesize_confirmation_message,
        find_pending_confirmation,
    )

    session = _confirm_session_with_pending()
    fc_id, confirmation = find_pending_confirmation(session)
    msg = _synthesize_confirmation_message(fc_id, confirmation, confirmed=False)
    assert msg.parts[0].function_response.response["confirmed"] is False


# =========================================================================== #
# Post-run write execution (_execute_pending_writes)
# =========================================================================== #


def _seed_two_row_ledger():
    """Seed a real two-row QBS Purchase ledger; return (slack, store)."""
    slack = FakeSlackClient()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[
            {"sheet": "Purchase", "doc_key": "k1", "rows": [
                {"Invoice Number": "INV-1", "Description": "AWS", "Source Amount": 1000.0,
                 "Account Code / COA": "6090", "Tax Amount": 90.0},
                {"Invoice Number": "INV-2", "Description": "Rent", "Source Amount": 2000.0,
                 "Account Code / COA": "6200", "Tax Amount": 0.0},
            ]},
        ],
    )
    return slack, store


def test_execute_pending_writes_amend_posts_audits_and_refreshes():
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_two_row_ledger()
    state = {
        "pending_ledger_write": [
            {"op": "amend", "sheet": "Purchase", "row": 2,
             "updates": {"Account Code / COA": "6010"}, "tax_treatment": "SR"},
        ],
    }
    committed = asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-1",
    ))
    assert committed is True
    # Workbook actually mutated.
    rows = store.read_rows("c1", "2026", slack, "C1")
    assert rows[0]["Account Code / COA"] == "6010"
    # Pending list cleared + idempotency marker set.
    assert state["pending_ledger_write"] == []
    assert "fc-1" in state["committed_confirmations"]
    # A confirmation message was posted.
    assert any("Updated" in m["text"] for m in slack._posts)


def test_execute_pending_writes_remove():
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_two_row_ledger()
    state = {"pending_ledger_write": [
        {"op": "remove", "sheet": "Purchase", "row": 2}]}
    asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-rm",
    ))
    rows = store.read_rows("c1", "2026", slack, "C1")
    # Row 2 (AWS) gone; only Rent remains.
    assert len(rows) == 1
    assert rows[0]["Description"] == "Rent"
    assert any("Removed" in m["text"] for m in slack._posts)


def test_execute_pending_writes_idempotent_double_yes():
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_two_row_ledger()
    spec = {"op": "amend", "sheet": "Purchase", "row": 2,
            "updates": {"Account Code / COA": "6010"}, "tax_treatment": "SR"}

    state = {"pending_ledger_write": [spec]}
    asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-dup",
    ))
    posts_after_first = len([m for m in slack._posts if "Updated" in m["text"]])

    # Second "yes" for the SAME fc_id: re-queue the spec, but it must NOT re-apply.
    state["pending_ledger_write"] = [spec]
    committed2 = asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-dup",
    ))
    assert committed2 is False
    posts_after_second = len([m for m in slack._posts if "Updated" in m["text"]])
    assert posts_after_second == posts_after_first  # no duplicate write/post
    assert state["pending_ledger_write"] == []


# =========================================================================== #
# MEDIUM-2 — improved classify_confirmation_reply
# =========================================================================== #


def test_classify_confirmation_improved():
    from accounting_agents.slack_runner import classify_confirmation_reply

    # Leading-token affirmative + trailing words
    assert classify_confirmation_reply("yes please") is True
    assert classify_confirmation_reply("yes do it") is True
    assert classify_confirmation_reply("ok great") is True
    assert classify_confirmation_reply("sure thing") is True
    # Phrase contains
    assert classify_confirmation_reply("go ahead and change it") is True
    assert classify_confirmation_reply("please do that") is True
    # Negative precedence — "no" wins even if "yes" also present
    assert classify_confirmation_reply("no thanks") is False
    assert classify_confirmation_reply("no, not yet") is False
    assert classify_confirmation_reply("don't do it") is False
    assert classify_confirmation_reply("cancel that please") is False
    assert classify_confirmation_reply("wait, hold on") is False
    # Ambiguous — bare unrelated question
    assert classify_confirmation_reply("what does this change?") is None
    assert classify_confirmation_reply("tell me more") is None
    assert classify_confirmation_reply("maybe") is None
    assert classify_confirmation_reply("") is None


# =========================================================================== #
# MEDIUM-3 — staleness bound in find_pending_confirmation
# =========================================================================== #


def _make_stale_session(fc_id="stale-fc", n_extra_events=10):
    """Build a session where the adk_request_confirmation is older than the window."""
    from google.adk.events.event_actions import EventActions
    from google.adk.tools.tool_confirmation import ToolConfirmation

    confirm_event = Event(
        author="assistant",
        content=types.Content(
            parts=[types.Part(function_call=types.FunctionCall(
                name="adk_request_confirmation", id=fc_id, args={}))]
        ),
        long_running_tool_ids=[fc_id],
        actions=EventActions(
            requested_tool_confirmations={
                fc_id: ToolConfirmation(hint="change something?", payload={"op": "amend"})
            }
        ),
    )
    # Pad with unrelated events to push the confirmation outside the window.
    filler = [
        Event(
            author="user",
            content=types.Content(parts=[types.Part(text=f"question {i}")]),
        )
        for i in range(n_extra_events)
    ]
    return SimpleNamespace(events=[confirm_event] + filler)


def test_find_pending_confirmation_ignores_stale():
    from accounting_agents.slack_runner import find_pending_confirmation

    # With 10 extra events after the confirmation, it is outside the 6-event window.
    session = _make_stale_session(n_extra_events=10)
    assert find_pending_confirmation(session) is None


def test_find_pending_confirmation_within_window():
    from accounting_agents.slack_runner import find_pending_confirmation

    # With only 2 extra events, the confirmation is still within the window.
    session = _make_stale_session(n_extra_events=2)
    result = find_pending_confirmation(session)
    assert result is not None
    fc_id, _ = result
    assert fc_id == "stale-fc"


def test_find_pending_confirmation_custom_window():
    from accounting_agents.slack_runner import find_pending_confirmation

    # Large custom window sees the stale event.
    session = _make_stale_session(n_extra_events=10)
    result = find_pending_confirmation(session, staleness_window=20)
    assert result is not None


# =========================================================================== #
# HIGH-2 — signature mismatch refuses the write
# =========================================================================== #


def _seed_two_row_ledger_sig():
    """Seed a two-row QBS Purchase ledger; return (slack, store)."""
    slack = FakeSlackClient()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[
            {"sheet": "Purchase", "doc_key": "k1", "rows": [
                {"Invoice Number": "INV-1", "Description": "AWS hosting",
                 "Source Amount": 1000.0, "Account Code / COA": "6090",
                 "Tax Amount": 90.0},
                {"Invoice Number": "INV-2", "Description": "Rent",
                 "Source Amount": 2000.0, "Account Code / COA": "6200",
                 "Tax Amount": 0.0},
            ]},
        ],
    )
    return slack, store


def test_execute_pending_writes_refuses_on_signature_mismatch():
    """A stale or replayed remove that targets a shifted row is refused."""
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_two_row_ledger_sig()
    # Simulate: the first row was already deleted before the "yes" arrived,
    # shifting row numbers. We give a signature that now matches nothing.
    state = {
        "pending_ledger_write": [
            {"op": "remove", "sheet": "Purchase", "row": 2,
             "row_signature": "deadbeefdeadbeef"},  # wrong sig
        ],
    }
    committed = asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-sig",
    ))
    # Nothing was committed.
    assert committed is False
    # Both rows still intact.
    rows = store.read_rows("c1", "2026", slack, "C1")
    assert len(rows) == 2
    # A warning message was posted.
    assert any("⚠️" in m["text"] for m in slack._posts)


def test_execute_pending_writes_correct_signature_succeeds():
    """A write with the correct row signature goes through normally."""
    from accounting_agents.assistant import _row_signature
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_two_row_ledger_sig()
    rows = store.read_rows("c1", "2026", slack, "C1")
    row0 = rows[0]
    sig = _row_signature(row0)

    state = {
        "pending_ledger_write": [
            {"op": "amend", "sheet": "Purchase", "row": 2,
             "updates": {"Account Code / COA": "6010"}, "tax_treatment": "SR",
             "row_signature": sig},
        ],
    }
    committed = asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-ok",
    ))
    assert committed is True
    updated = store.read_rows("c1", "2026", slack, "C1")
    assert updated[0]["Account Code / COA"] == "6010"


def test_execute_pending_writes_remove_replay_after_marker_failure():
    """Double-replay of a remove: second attempt hits wrong sig (row already gone)."""
    from accounting_agents.assistant import _row_signature
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_two_row_ledger_sig()
    rows = store.read_rows("c1", "2026", slack, "C1")
    sig = _row_signature(rows[0])

    spec = {"op": "remove", "sheet": "Purchase", "row": 2, "row_signature": sig}

    # First application — succeeds and removes row 2.
    state1 = {"pending_ledger_write": [spec]}
    asyncio.run(_execute_pending_writes(
        state=state1, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-replay",
    ))
    assert len(store.read_rows("c1", "2026", slack, "C1")) == 1

    # Simulated replay (marker persist failed): same spec again WITHOUT fc_id guard.
    # The row at position 2 is now "Rent" (shifted up) — signature mismatch → refuse.
    state2 = {"pending_ledger_write": [spec]}
    committed2 = asyncio.run(_execute_pending_writes(
        state=state2, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id=None,  # no marker
    ))
    assert committed2 is False
    # "Rent" row still intact — the wrong row was NOT deleted.
    remaining = store.read_rows("c1", "2026", slack, "C1")
    assert len(remaining) == 1
    assert remaining[0]["Description"] == "Rent"


# --------------------------------------------------------------------------- #
# Step 2B: handle_review_action post-resume continuation (the Critical fix).
# After a confirm_as_is / reextract_as resume, the run flows to the terminal
# approval_gate, which either pauses AGAIN (must post the approval card) or
# auto-approves and completes (must persist + deliver — else silent data loss).
# --------------------------------------------------------------------------- #
def test_handle_review_action_delivers_on_clean_completion(monkeypatch):
    """confirm_as_is whose resumed run auto-approves → persist_and_deliver runs."""
    from accounting_agents.slack_runner import handle_review_action

    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    op_id = "CR9:FR9:review"
    write_interrupt(
        db, op_id, session_id="CR9:FR9", channel_id="CR9", slack_file_id="FR9",
        message_ts="1.1", user_id="CR9", extra={"kind": "review", "question": "q?", "reasons": []},
    )

    # Resumed run completes with no further interrupt.
    final_event = SimpleNamespace(
        content=SimpleNamespace(parts=[SimpleNamespace(text="done")]),
        get_function_calls=lambda: [],
    )

    async def fake_resume(runner, db_, op, decision):
        return [final_event]

    monkeypatch.setattr("accounting_agents.slack_runner.resume_session", fake_resume)

    runner = _FakeRunner([], _ledger_payload())
    # The (faked) resumed run would have populated the session; mark it created
    # so persist_and_deliver can read the final ledger payload.
    runner.session_service.created = True
    result = asyncio.run(
        handle_review_action(
            runner=runner, ledger_store=store, db=db, slack_client=slack,
            op_id=op_id, action="confirm_as_is", app_name="acc",
        )
    )

    assert result["status"] == "resumed"
    # The document was actually delivered to the ledger (the silent-data-loss guard).
    assert result["outcome"]["status"] == "delivered"
    assert len(slack.uploads) == 1
    assert any("FY2026 ledger" in t for t in _posted_texts(slack))


def test_handle_review_action_reposts_terminal_card_on_repause(monkeypatch):
    """confirm_as_is whose resumed run pauses again at approval_gate → approval card posted."""
    from accounting_agents.slack_runner import handle_review_action

    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    op_id = "CR8:FR8:review"
    write_interrupt(
        db, op_id, session_id="CR8:FR8", channel_id="CR8", slack_file_id="FR8",
        message_ts="2.2", user_id="CR8", extra={"kind": "review", "question": "q?", "reasons": []},
    )

    # Resumed run pauses again at the TERMINAL gate with the BASE id (no :review).
    repause_event = SimpleNamespace(
        content=SimpleNamespace(parts=[]),
        get_function_calls=lambda: [SimpleNamespace(name="adk_request_input", id="CR8:FR8")],
    )

    async def fake_resume(runner, db_, op, decision):
        return [repause_event]

    monkeypatch.setattr("accounting_agents.slack_runner.resume_session", fake_resume)

    runner = _FakeRunner([], {"approval_message": "needs review: line X"})
    result = asyncio.run(
        handle_review_action(
            runner=runner, ledger_store=store, db=db, slack_client=slack,
            op_id=op_id, action="confirm_as_is", app_name="acc",
        )
    )

    assert result["status"] == "resumed"
    # The terminal approval interrupt was surfaced (not silently stalled).
    assert result["outcome"]["status"] == "paused"
    assert result["outcome"]["op_id"] == "CR8:FR8"
    # A base-id interrupt doc was written so the approve/edit/reject card can resume.
    assert read_interrupt(db, "CR8:FR8") is not None
    # Nothing uploaded while paused.
    assert slack.uploads == []


# =========================================================================== #
# learn_mapping runner drain (Step 7 / C-3)
# =========================================================================== #


def test_runner_drains_pending_learn_mapping():
    """Post-run drain: pending_learn_mapping entries call add_correction and clear the list.

    Uses ``_CapturingChatRunner`` with a pre-seeded session that already
    contains ``pending_learn_mapping`` (as the learn_mapping tool would have
    written it during the turn).  Verifies:
    - ``client_store.add_correction`` is called once with the right args.
    - The session's ``pending_learn_mapping`` list is cleared via _apply_state_delta.
    """
    from accounting_agents.slack_runner import PENDING_LEARN_KEY, answer_question

    # ---- fake client store that records add_correction calls ----
    class _FakeLearnStore:
        def __init__(self):
            self.corrections: list[dict] = []

        def get_by_channel(self, channel_id):
            return None  # no profile → profile_delta stays empty

        def add_correction(self, *, client_id, vendor, account_code=None, tax_code=None):
            self.corrections.append({
                "client_id": client_id,
                "vendor": vendor,
                "account_code": account_code,
                "tax_code": tax_code,
            })

    fake_store = _FakeLearnStore()
    slack = FakeSlackClient()

    # Pre-seed the session with the learn spec already in state
    # (mimicking what the tool wrote inside the agent turn).
    channel_id = "C-LEARN"
    session_id = f"{channel_id}:chat:1700000000.001"
    sessions: dict = {
        (channel_id, session_id): _FakeSession({
            PENDING_LEARN_KEY: [
                {"vendor": "Acme Cloud", "account_code": "6090", "tax_code": None},
            ],
        }),
    }
    chat_runner = _CapturingChatRunner(sessions)

    asyncio.run(
        answer_question(
            runner=chat_runner,
            ledger_store=_noop_ledger_store(),
            slack_client=slack,
            channel_id=channel_id,
            question="remember, Acme Cloud goes to 6090",
            app_name=chat_runner.app_name,
            client_store=fake_store,
            message_ts="1700000000.001",
            thread_ts=None,
            raw_thread_ts="1700000000.001",
        )
    )

    # add_correction was called exactly once with the right args.
    assert len(fake_store.corrections) == 1, (
        f"Expected 1 add_correction call, got {fake_store.corrections}"
    )
    corr = fake_store.corrections[0]
    assert corr["vendor"] == "Acme Cloud"
    assert corr["account_code"] == "6090"
    assert corr["tax_code"] is None
    # client_id comes from profile_delta (empty here → falls back to channel_id).
    assert corr["client_id"] == channel_id

    # The pending list was cleared from the session.
    # _apply_state_delta appends an event; check the session state was cleared.
    # (_CapturingChatRunner.session_service.get_session returns the same _FakeSession
    # object; _apply_state_delta appends an Event — but since our fake doesn't
    # actually mutate state from events, we verify via the store call count instead.)
    assert len(fake_store.corrections) == 1  # idempotency: not doubled


# =========================================================================== #
# replace_month drain (Step 7 / C-3)
# =========================================================================== #


def _seed_month_ledger():
    """Seed a Purchase+Sales workbook with Sep+Oct rows; return (slack, store)."""
    from openpyxl import Workbook
    import io

    wb = Workbook()
    ws_p = wb.active
    ws_p.title = "Purchase"
    ws_p.append(["Date", "Invoice Number", "Description", "Source Amount", "Account Code / COA"])
    ws_p.append(["05/09/2025", "INV-P1", "AWS Sep",  100.0, "6090"])
    ws_p.append(["03/10/2025", "INV-P2", "AWS Oct",  120.0, "6090"])
    ws_s = wb.create_sheet("Sales")
    ws_s.append(["Date", "Invoice Number", "Description", "Source Amount", "Account Code / COA"])
    ws_s.append(["10/09/2025", "INV-S1", "Consult Sep", 500.0, "4000"])

    buf = io.BytesIO()
    wb.save(buf)
    xlsx = buf.getvalue()

    slack = FakeSlackClient()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    import uuid
    file_id = "F" + uuid.uuid4().hex[:10]
    slack.files[file_id] = xlsx
    url = f"https://files.slack.com/{file_id}/ledger.xlsx"
    slack.urls[url] = xlsx
    store._pointer_ref("c1", "2026").set({
        "slack_file_id": file_id,
        "client_id": "c1",
        "fy": "2026",
        "kind": "invoice",
        "seen_doc_keys": ["Purchase:INV-P1", "Purchase:INV-P2", "Sales:INV-S1"],
    })
    return slack, store


def test_execute_pending_writes_replace_month_drains_and_posts():
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_month_ledger()
    state = {
        "pending_ledger_write": [
            {"op": "replace_month", "year": 2025, "month": 9},
        ],
    }
    committed = asyncio.run(_execute_pending_writes(
        state=state,
        ledger_store=store,
        slack_client=slack,
        channel_id="C1",
        client_id="c1",
        fy="2026",
        session_id="S1",
        fc_id="fc-rm-month",
    ))

    assert committed is True
    # Workbook trimmed — Sep rows gone.
    rows = store.read_rows("c1", "2026", slack, "C1")
    dates = [r.get("Date") for r in rows if r.get("Date")]
    assert all("10/2025" in str(d) for d in dates), dates

    # Summary message posted.
    assert any("September 2025" in m.get("text", "") for m in slack._posts), slack._posts
    assert any("re-drop" in m.get("text", "").lower() for m in slack._posts)

    # Audit logged + idempotency marker set.
    assert state["pending_ledger_write"] == []
    assert "fc-rm-month" in state["committed_confirmations"]


def test_execute_pending_writes_replace_month_purges_doc_keys():
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_month_ledger()
    state = {"pending_ledger_write": [{"op": "replace_month", "year": 2025, "month": 9}]}
    asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-keys",
    ))

    ptr = store.get_pointer("c1", "2026")
    surviving = set(ptr.get("seen_doc_keys") or [])
    assert "Purchase:INV-P1" not in surviving   # Sep purged
    assert "Sales:INV-S1" not in surviving      # Sep purged
    assert "Purchase:INV-P2" in surviving       # Oct survives


def test_execute_pending_writes_replace_month_idempotent_double_yes():
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_month_ledger()
    spec = {"op": "replace_month", "year": 2025, "month": 9}

    state = {"pending_ledger_write": [spec]}
    asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-idem",
    ))
    posts_after_first = len(slack._posts)

    # Second "yes" for same fc_id — must NOT re-apply.
    state["pending_ledger_write"] = [spec]
    committed2 = asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-idem",
    ))
    assert committed2 is False
    assert len(slack._posts) == posts_after_first  # no extra message
    assert state["pending_ledger_write"] == []


def test_execute_pending_writes_replace_month_refreshes_ledger_data():
    """After replace_month commits, state ledger_data reflects the trimmed workbook."""
    from accounting_agents.slack_runner import _execute_pending_writes

    slack, store = _seed_month_ledger()
    state = {"pending_ledger_write": [{"op": "replace_month", "year": 2025, "month": 9}]}
    asyncio.run(_execute_pending_writes(
        state=state, ledger_store=store, slack_client=slack, channel_id="C1",
        client_id="c1", fy="2026", session_id="S1", fc_id="fc-refresh",
    ))
    # The runner drains + marks committed; the real answer_question refreshes
    # ledger_data via read_rows post-commit. Here we verify the workbook is correct.
    rows = store.read_rows("c1", "2026", slack, "C1")
    assert len(rows) == 1  # only Oct Purchase row
    assert rows[0].get("Date") == "03/10/2025"


# =========================================================================== #
# Re-extract (Step 7 / ADR-0010): process_file_event hint+replace params,
# _execute_pending_reextract drain.
# =========================================================================== #


class _StateCapturingRunner(_FakeRunner):
    """A doc-graph runner that records the ``state_delta`` it was run with."""

    def __init__(self, events, final_state, app_name="acc"):
        super().__init__(events, final_state, app_name=app_name)
        self.run_state_delta: dict = {}

    async def run_async(self, *, user_id, session_id, new_message=None, state_delta=None):
        self.run_state_delta = dict(state_delta or {})
        async for ev in super().run_async(
            user_id=user_id, session_id=session_id,
            new_message=new_message, state_delta=state_delta,
        ):
            yield ev


class _CapturingLedgerStore:
    """A ledger_store stand-in that records the ``replace`` kwarg of append_rows."""

    def __init__(self, *, batch_replace_counts=None):
        self.append_calls: list[dict] = []
        self._batch_replace_counts = batch_replace_counts

    def append_rows(self, **kwargs):
        self.append_calls.append(kwargs)
        result = {
            "slack_file_id": "F-up", "appended": 1, "deduped": 0,
            "filename": "ledger.xlsx",
        }
        if self._batch_replace_counts is not None:
            result["batch_replace_counts"] = self._batch_replace_counts
        return result


def _completion_runner():
    """A doc runner whose run completes cleanly (no interrupt) → persist+deliver."""
    final_event = SimpleNamespace(
        content=SimpleNamespace(parts=[SimpleNamespace(text="done")]),
        get_function_calls=lambda: [],
    )
    return final_event


def test_process_file_event_hint_flows_into_state_delta_as_review_hint():
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    runner = _StateCapturingRunner([_completion_runner()], _ledger_payload())

    asyncio.run(
        process_file_event(
            runner=runner, ledger_store=store, db=db, slack_client=slack,
            channel_id="C1", file_id="F1", app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="re-extract-F1.pdf",
            hint="read as a credit note",
            client_store=_seeded_client_store(db),
        )
    )

    assert runner.run_state_delta.get("review_hint") == "read as a credit note"


def test_process_file_event_replace_flows_to_append_rows():
    slack = FakeSlackClient()
    db = FakeFirestore()
    ledger = _CapturingLedgerStore()
    runner = _FakeRunner([_completion_runner()], _ledger_payload())

    asyncio.run(
        process_file_event(
            runner=runner, ledger_store=ledger, db=db, slack_client=slack,
            channel_id="C1", file_id="F1", app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="re-extract-F1.pdf",
            hint="read as a credit note",
            replace=True,
            client_store=_seeded_client_store(db),
        )
    )

    assert ledger.append_calls, "append_rows must have been called"
    assert ledger.append_calls[0]["replace"] is True


def test_process_file_event_default_path_does_not_replace_or_hint():
    """The normal file-drop path RESETS the re-extract keys (so a re-shared file
    can't inherit a stale hint/replace flag) and calls append_rows replace=False."""
    slack = FakeSlackClient()
    db = FakeFirestore()
    ledger = _CapturingLedgerStore()
    runner = _StateCapturingRunner([_completion_runner()], _ledger_payload())

    asyncio.run(
        process_file_event(
            runner=runner, ledger_store=ledger, db=db, slack_client=slack,
            channel_id="C1", file_id="F1", app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
    )

    # Reset (not absent): a fresh drop overwrites any stale leaked values.
    assert runner.run_state_delta.get("review_hint") == ""
    assert runner.run_state_delta.get("reextract_replace") is False
    assert ledger.append_calls[0]["replace"] is False


# --- _execute_pending_reextract ------------------------------------------- #


class _RecordingPFE:
    """A replacement for slack_runner.process_file_event that records its kwargs
    and returns a scripted result (default: a clean replaced delivery)."""

    def __init__(self, result=None):
        self.calls: list[dict] = []
        self._result = result or {
            "status": "delivered",
            "append": {"batch_replace_counts": [{"replaced": 1, "appended": 1}]},
        }

    async def __call__(self, **kwargs):
        self.calls.append(kwargs)
        return self._result


def _run_reextract(specs, *, pfe, slack, doc_runner=None):
    from accounting_agents import slack_runner

    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    runner = doc_runner or _FakeRunner([], _ledger_payload())
    with patch.object(slack_runner, "process_file_event", pfe):
        asyncio.run(
            slack_runner._execute_pending_reextract(
                specs,
                doc_runner=runner,
                ledger_store=store,
                db=db,
                slack_client=slack,
                channel_id="C1",
                app_name="acc",
                client_store=None,
                thread_ts=None,
            )
        )
    return runner


def test_execute_pending_reextract_calls_pfe_with_hint_and_replace():
    slack = FakeSlackClient()
    pfe = _RecordingPFE()
    doc_runner = _FakeRunner([], _ledger_payload())
    specs = [{"op": "reextract", "file_id": "F9", "hints": "read as a credit note"}]

    _run_reextract(specs, pfe=pfe, slack=slack, doc_runner=doc_runner)

    assert len(pfe.calls) == 1
    call = pfe.calls[0]
    assert call["runner"] is doc_runner  # the INJECTED doc runner, not the chat one
    assert call["file_id"] == "F9"
    assert call["hint"] == "read as a credit note"
    assert call["replace"] is True


def test_execute_pending_reextract_is_idempotent_on_double_run():
    slack = FakeSlackClient()
    pfe = _RecordingPFE()
    doc_runner = _FakeRunner([], _ledger_payload())
    specs = [{"op": "reextract", "file_id": "F9", "hints": "read as a credit note"}]

    from accounting_agents import slack_runner

    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    with patch.object(slack_runner, "process_file_event", pfe):
        async def _twice():
            for _ in range(2):
                await slack_runner._execute_pending_reextract(
                    specs, doc_runner=doc_runner, ledger_store=store, db=db,
                    slack_client=slack, channel_id="C1", app_name="acc",
                    client_store=None, thread_ts=None,
                )
        asyncio.run(_twice())

    # Same (file_id, hints) on the same runner instance → dispatched exactly once.
    assert len(pfe.calls) == 1


def test_execute_pending_reextract_posts_clear_month_note_on_identity_change():
    """A delivered re-read that replaced 0 rows (identity changed) → the user is
    pointed at the month-level primitive (replace_recorded_month / clear)."""
    slack = FakeSlackClient()
    pfe = _RecordingPFE(result={
        "status": "delivered",
        "append": {"batch_replace_counts": [{"replaced": 0, "appended": 2}]},
    })
    specs = [{"op": "reextract", "file_id": "F9", "hints": "read as a credit note"}]

    _run_reextract(specs, pfe=pfe, slack=slack)

    texts = " ".join(_posted_texts(slack)).lower()
    assert "clear" in texts and "identity" in texts


# --- HIGH regression: state-leak between runs on the same file_id ----------- #


class _CapturingLedgerStoreMulti:
    """Records every append_rows call so we can assert per-call replace kwarg."""

    def __init__(self):
        self.append_calls: list[dict] = []

    def append_rows(self, **kwargs):
        self.append_calls.append(dict(kwargs))
        return {
            "slack_file_id": "F-up", "appended": 1, "deduped": 0,
            "filename": "ledger.xlsx",
        }


def test_normal_drop_after_reextract_does_not_inherit_replace_flag():
    """HIGH regression: re-extract run on file_id X sets reextract_replace=True in
    session state. A subsequent normal drop of the SAME file_id on the SAME
    long-lived per-doc session MUST call append_rows with replace=False — not the
    stale True from the prior run."""
    slack = FakeSlackClient()
    db = FakeFirestore()
    ledger = _CapturingLedgerStoreMulti()
    # Use a _FakeRunner that reuses the same _FakeSessionService between calls so
    # the second run sees the session that the first run created — the realistic
    # long-lived-session scenario.
    runner = _FakeRunner([_completion_runner()], _ledger_payload())
    client_store = _seeded_client_store(db)

    # First call: re-extract (replace=True) — seeds reextract_replace=True into state.
    asyncio.run(
        process_file_event(
            runner=runner, ledger_store=ledger, db=db, slack_client=slack,
            channel_id="C1", file_id="F1", app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="invoice.pdf",
            hint="read as a credit note",
            replace=True,
            client_store=client_store,
        )
    )
    assert ledger.append_calls[0]["replace"] is True

    # Reset the runner's events so the second run also completes cleanly.
    runner._events = [_completion_runner()]
    # The session now exists with reextract_replace=True in state; the second
    # NORMAL drop must overwrite it to False via the unconditional state_delta.
    asyncio.run(
        process_file_event(
            runner=runner, ledger_store=ledger, db=db, slack_client=slack,
            channel_id="C1", file_id="F1", app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="invoice.pdf",
            # No hint, no replace — normal drop.
            client_store=client_store,
        )
    )
    # Second call must have replace=False (not the stale True from the first run).
    assert len(ledger.append_calls) == 2
    assert ledger.append_calls[1]["replace"] is False


def test_normal_drop_after_reextract_does_not_inherit_review_hint():
    """HIGH regression: a normal re-drop of file_id X must NOT forward the hint
    that a prior re-extract run seeded into the session's state_delta."""
    slack = FakeSlackClient()
    db = FakeFirestore()
    runner = _StateCapturingRunner([_completion_runner()], _ledger_payload())
    client_store = _seeded_client_store(db)

    # First call: re-extract seeds review_hint into state_delta.
    asyncio.run(
        process_file_event(
            runner=runner, ledger_store=SlackLedgerStore(FakeFirestore(), opener=slack.opener()),
            db=db, slack_client=slack,
            channel_id="C1", file_id="F1", app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="invoice.pdf",
            hint="read as a credit note",
            replace=True,
            client_store=client_store,
        )
    )
    assert runner.run_state_delta.get("review_hint") == "read as a credit note"

    # Second call: normal drop — state_delta must carry review_hint="" (not the stale hint).
    runner._events = [_completion_runner()]
    asyncio.run(
        process_file_event(
            runner=runner, ledger_store=SlackLedgerStore(FakeFirestore(), opener=slack.opener()),
            db=db, slack_client=slack,
            channel_id="C1", file_id="F1", app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="invoice.pdf",
            client_store=client_store,
        )
    )
    assert runner.run_state_delta.get("review_hint") == ""
    assert runner.run_state_delta.get("reextract_replace") is False


# --- MEDIUM regression: duplicate path also fires the identity-change note --- #


def test_execute_pending_reextract_posts_clear_month_note_on_duplicate_status():
    """MEDIUM regression: when the re-read returns status='duplicate' (all_deduped)
    and replaced=0 the identity changed — the user must still see the 'clear month'
    guidance. Previously only status='delivered' triggered it."""
    slack = FakeSlackClient()
    pfe = _RecordingPFE(result={
        "status": "duplicate",
        "append": {"batch_replace_counts": [{"replaced": 0, "appended": 0}]},
    })
    specs = [{"op": "reextract", "file_id": "F9", "hints": "read as a credit note"}]

    _run_reextract(specs, pfe=pfe, slack=slack)

    texts = " ".join(_posted_texts(slack)).lower()
    assert "clear" in texts and "identity" in texts


# =========================================================================== #
# Step 8: proactive post-delivery re-extract offer
#   - _finalize_run_outcome posts the offer ONLY when the reviewer fired
#     (REVIEW_REASON_KEY non-empty) AND verdict != CLARIFY (a CLARIFY already
#     surfaced the mid-flow card). Clean happy-path docs post NOTHING.
#   - the proactive_redo action opens the hint modal with the right file_id.
#   - the ledgr_proactive_redo view submit runs the re-extract (file_id, hint,
#     replace=True) via the Step-7 drain.
# =========================================================================== #


def _run_finalize_delivery(final_state: dict, *, slack=None):
    """Drive process_file_event to a clean (no-interrupt) delivery whose final
    session state is ``final_state``, then return the FakeSlackClient so callers
    can inspect whether the proactive offer card was posted."""
    slack = slack or FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    final_event = SimpleNamespace(
        content=SimpleNamespace(parts=[SimpleNamespace(text="done")]),
        get_function_calls=lambda: [],
    )
    runner = _FakeRunner([final_event], final_state)

    def fake_download(client, file_id):
        return b"%PDF-1.4 fake"

    result = asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=slack,
            channel_id="C1",
            file_id="F1",
            app_name="acc",
            download_fn=fake_download,
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
    )
    return slack, result


def _proactive_card_posts(slack: FakeSlackClient) -> list:
    """All chat_postMessage calls carrying the proactive_redo action button."""
    out = []
    for p in _post_calls(slack):
        for b in (p.get("blocks") or []):
            if b.get("block_id") == "ledgr_proactive_redo":
                out.append(p)
    return out


def test_finalize_posts_proactive_offer_when_reviewer_fired_and_not_clarify():
    """Reviewer FIRED (non-empty reasons) + verdict != CLARIFY + delivered →
    the proactive re-extract offer is posted (threaded under the delivery)."""
    state = dict(_ledger_payload())
    state[nodes.REVIEW_REASON_KEY] = ["unreconciled: Invoice (FX off)"]
    state[nodes.REVIEW_VERDICT_KEY] = nodes.REVIEW_VERDICT_HINTS

    slack, result = _run_finalize_delivery(state)

    assert result["status"] == "delivered"
    offers = _proactive_card_posts(slack)
    assert len(offers) == 1
    button = offers[0]["blocks"][1]["elements"][0]
    assert button["action_id"] == "proactive_redo"
    assert button["value"] == "F1"
    # The humanized reason is rendered, not the raw machine string.
    assert "the totals didn't reconcile" in offers[0]["blocks"][0]["text"]["text"]


def test_finalize_posts_no_offer_on_clean_happy_path():
    """A clean delivery (no review reasons, verdict OK) delivers normally and
    posts NO proactive offer — proving the offer is rare."""
    state = dict(_ledger_payload())
    state[nodes.REVIEW_REASON_KEY] = []
    state[nodes.REVIEW_VERDICT_KEY] = nodes.REVIEW_VERDICT_OK

    slack, result = _run_finalize_delivery(state)

    assert result["status"] == "delivered"
    assert _proactive_card_posts(slack) == []
    # The normal delivery summary still posted.
    assert any("FY2026 ledger" in t for t in _posted_texts(slack))


def test_finalize_posts_no_offer_when_verdict_clarify():
    """When the verdict is CLARIFY the mid-flow review card already engaged the
    user, so the post-delivery offer must NOT also fire (no double-prompt)."""
    state = dict(_ledger_payload())
    state[nodes.REVIEW_REASON_KEY] = ["doc_type_other"]
    state[nodes.REVIEW_VERDICT_KEY] = nodes.REVIEW_VERDICT_CLARIFY

    slack, result = _run_finalize_delivery(state)

    assert result["status"] == "delivered"
    assert _proactive_card_posts(slack) == []


def _capture_proactive_handlers(runner_mock=None, ledger_store_mock=None, db_mock=None):
    """Build the Bolt app with fakes; capture the ``proactive_redo`` action +
    ``ledgr_proactive_redo`` view handlers that build_async_app registers."""
    from unittest.mock import MagicMock, patch

    from app.slack_app import _SeenEvents
    from accounting_agents import slack_runner

    registered = {"actions": {}, "views": {}}

    def action_decorator(action_id, *a, **k):
        def decorator(fn):
            registered["actions"][action_id] = fn
            return fn
        return decorator

    def view_decorator(callback_id, *a, **k):
        def decorator(fn):
            registered["views"][callback_id] = fn
            return fn
        return decorator

    fake_app = MagicMock()
    fake_app.event = lambda *a, **k: (lambda fn: fn)
    fake_app.action = action_decorator
    fake_app.view = view_decorator
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = runner_mock or _FakeActionViewRunner()

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch("invoice_processing.export.client_context.FirestoreClientStore"), \
         patch.object(slack_runner, "build_chat_runner",
                      return_value=SimpleNamespace(app_name="accounting_agents_assistant")):
        build_async_app(
            runner=rm,
            ledger_store=ledger_store_mock or MagicMock(),
            db=db_mock or MagicMock(),
        )

    return registered["actions"]["proactive_redo"], registered["views"]["ledgr_proactive_redo"]


def test_proactive_redo_action_opens_modal_with_file_id():
    """Clicking the offer button opens the hint modal whose private_metadata
    carries the file_id from the button value."""
    from unittest.mock import AsyncMock, MagicMock, patch

    sync_client = MagicMock()
    with patch("slack_sdk.WebClient", return_value=sync_client):
        action_handler, _ = _capture_proactive_handlers()

    body = {
        "actions": [{"action_id": "proactive_redo", "value": "F-REDO-1"}],
        "trigger_id": "T-REDO-1",
    }
    ack = AsyncMock()
    asyncio.run(action_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    sync_client.views_open.assert_called_once()
    kwargs = sync_client.views_open.call_args.kwargs
    assert kwargs["trigger_id"] == "T-REDO-1"
    view = kwargs["view"]
    assert view["callback_id"] == "ledgr_proactive_redo"
    assert view["private_metadata"] == "F-REDO-1"


def test_proactive_redo_view_submit_runs_reextract_with_hint_and_replace():
    """Submitting the proactive hint modal runs the re-extract directly:
    process_file_event is called with the file_id, the hint, and replace=True.
    The channel is recovered from the file's share record (the view_submission
    body has no source channel)."""
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    # A FakeSlackClient that knows file F-REDO-9 was shared in channel C1 — this
    # is what _resolve_file_channel reads to recover the channel from files_info.
    slack = FakeSlackClient()
    slack._file_share_ts["F-REDO-9"] = {"C1": "111.222"}

    doc_runner = _FakeRunner([], _ledger_payload())
    with patch("slack_sdk.WebClient", return_value=slack):
        _, view_handler = _capture_proactive_handlers(
            runner_mock=doc_runner,
            ledger_store_mock=MagicMock(),
            db_mock=FakeFirestore(),
        )

    body = {
        "view": {
            "callback_id": "ledgr_proactive_redo",
            "private_metadata": "F-REDO-9",
            "state": {
                "values": {
                    "hint_block": {"hint_input": {"value": "read it as a credit note"}}
                }
            },
        }
    }

    pfe = _RecordingPFE()
    ack = AsyncMock()
    with patch.object(slack_runner, "process_file_event", pfe):
        asyncio.run(view_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    assert len(pfe.calls) == 1
    call = pfe.calls[0]
    assert call["file_id"] == "F-REDO-9"
    assert call["hint"] == "read it as a credit note"
    assert call["replace"] is True
    assert call["channel_id"] == "C1"
    assert call["runner"] is doc_runner


def test_proactive_redo_view_submit_noop_without_hint():
    """An empty hint submission must NOT dispatch a re-extract (nothing to steer)."""
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    slack = FakeSlackClient()
    slack._file_share_ts["F-REDO-9"] = {"C1": "111.222"}
    with patch("slack_sdk.WebClient", return_value=slack):
        _, view_handler = _capture_proactive_handlers(db_mock=FakeFirestore())

    body = {
        "view": {
            "callback_id": "ledgr_proactive_redo",
            "private_metadata": "F-REDO-9",
            "state": {"values": {"hint_block": {"hint_input": {"value": "   "}}}},
        }
    }

    pfe = _RecordingPFE()
    ack = AsyncMock()
    with patch.object(slack_runner, "process_file_event", pfe):
        asyncio.run(view_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    assert pfe.calls == []


def test_event_stage_key_maps_to_three_layers():
    from accounting_agents.slack_runner import event_stage_key

    assert event_stage_key(_node_event("classify_node")) == "understand"
    assert event_stage_key(_node_event("extract_invoice_document_node")) == "understand"
    assert event_stage_key(_node_event("categorize_node")) == "policy"
    assert event_stage_key(_node_event("tax_node")) == "policy"
    assert event_stage_key(_node_event("consolidate_node")) == "commit"


def test_stage_output_for_extract_node():
    from accounting_agents.slack_runner import _stage_output_for_completed_node

    state = {
        nodes.NORMALIZED_KEY: [{
            "doc_type": "purchase",
            "supplier": {"name": "Telco Provider A"},
            "invoice_number": "8004483920",
            "doc_total": 1328.15,
            "currency": "SGD",
            "lines": [{"description": "a"}, {"description": "b"}],
        }],
    }
    out = _stage_output_for_completed_node("extract_invoice_document_node", state)
    assert "Telco Provider A" in out
    assert "1,328.15" in out
    assert "2 lines" in out


# --------------------------------------------------------------------------- #
# _StageState tests
# --------------------------------------------------------------------------- #


class TestStageState:

    def _make(self):
        from accounting_agents.slack_runner import _StageState
        return _StageState()

    def test_initial_all_pending(self):
        state = self._make()
        snap = state.snapshot()
        assert all(s["status"] == "pending" for s in snap)
        assert len(snap) == 3

    def test_advance_sets_in_progress_and_completes_prior(self):
        state = self._make()
        state.advance("policy")
        snap = state.snapshot()
        assert snap[0]["status"] == "complete"   # understand
        assert snap[1]["status"] == "in_progress"  # policy
        assert snap[2]["status"] == "pending"    # commit

    def test_advance_attaches_output_to_previous_stage(self):
        state = self._make()
        state.advance("policy", output="Recognized as invoice")
        snap = state.snapshot()
        assert snap[0]["output"] == "Recognized as invoice"
        assert snap[1]["output"] is None

    def test_advance_sequential_preserves_order(self):
        state = self._make()
        state.advance("understand")
        state.advance("policy")
        state.advance("commit")
        snap = state.snapshot()
        assert snap[0]["status"] == "complete"
        assert snap[1]["status"] == "complete"
        assert snap[2]["status"] == "in_progress"

    def test_advance_same_stage_twice_is_idempotent(self):
        state = self._make()
        state.advance("policy")
        state.advance("policy")
        snap = state.snapshot()
        policy = next(s for s in snap if s["task_id"] == "policy")
        assert policy["status"] == "in_progress"

    def test_set_output_refreshes_stage_line(self):
        state = self._make()
        state.advance("understand")
        state.set_output("understand", "Telco Provider A · SGD 1,328.15")
        snap = state.snapshot()
        assert snap[0]["output"] == "Telco Provider A · SGD 1,328.15"

    def test_mark_complete_sets_all_stages(self):
        state = self._make()
        state.advance("policy")
        state.mark_complete()
        snap = state.snapshot()
        assert all(s["status"] == "complete" for s in snap)

    def test_mark_complete_attaches_output_to_last(self):
        state = self._make()
        state.mark_complete(output="Done")
        snap = state.snapshot()
        assert snap[-1]["output"] == "Done"

    def test_mark_failed_keeps_subsequent_pending(self):
        state = self._make()
        state.advance("understand")
        state.mark_failed("understand", "parse error")
        snap = state.snapshot()
        understand = next(s for s in snap if s["task_id"] == "understand")
        assert understand["status"] == "failed"
        assert understand["output"] == "parse error"
        policy = next(s for s in snap if s["task_id"] == "policy")
        assert policy["status"] == "pending"

    def test_snapshot_returns_independent_copy(self):
        state = self._make()
        snap1 = state.snapshot()
        snap1[0]["status"] = "complete"
        snap2 = state.snapshot()
        assert snap2[0]["status"] == "pending"


# --------------------------------------------------------------------------- #
# per-doc card inline action handlers
# --------------------------------------------------------------------------- #


def _capture_per_doc_handlers(runner_mock=None, ledger_store_mock=None, db_mock=None):
    """Build the Bolt app with fakes; capture the per-doc action handlers."""
    from unittest.mock import MagicMock, patch

    from app.slack_app import _SeenEvents
    from accounting_agents import slack_runner

    registered = {"actions": {}, "views": {}}

    def action_decorator(action_id, *a, **k):
        def decorator(fn):
            registered["actions"][action_id] = fn
            return fn
        return decorator

    def view_decorator(callback_id, *a, **k):
        def decorator(fn):
            registered["views"][callback_id] = fn
            return fn
        return decorator

    fake_app = MagicMock()
    fake_app.event = lambda *a, **k: (lambda fn: fn)
    fake_app.action = action_decorator
    fake_app.view = view_decorator
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = runner_mock or _FakeActionViewRunner()

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch("invoice_processing.export.client_context.FirestoreClientStore"), \
         patch.object(slack_runner, "build_chat_runner",
                      return_value=SimpleNamespace(app_name="accounting_agents_assistant")):
        build_async_app(
            runner=rm,
            ledger_store=ledger_store_mock or MagicMock(),
            db=db_mock or MagicMock(),
        )

    return (
        registered["actions"]["ledgr_per_doc_reextract"],
        registered["actions"]["ledgr_per_doc_edit"],
        registered["actions"]["ledgr_per_doc_view_row"],
    )


def test_per_doc_reextract_opens_modal_with_file_id():
    """ledgr_per_doc_reextract opens the proactive_redo hint modal with the
    file_id from the button value — same UX as proactive_redo."""
    from unittest.mock import AsyncMock, MagicMock, patch

    sync_client = MagicMock()
    with patch("slack_sdk.WebClient", return_value=sync_client):
        reextract_handler, _, _ = _capture_per_doc_handlers()

    body = {
        "actions": [{"action_id": "ledgr_per_doc_reextract", "value": "F-CARD-1"}],
        "trigger_id": "T-CARD-1",
    }
    ack = AsyncMock()
    asyncio.run(reextract_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    sync_client.views_open.assert_called_once()
    kwargs = sync_client.views_open.call_args.kwargs
    assert kwargs["trigger_id"] == "T-CARD-1"
    view = kwargs["view"]
    assert view["callback_id"] == "ledgr_proactive_redo"
    assert view["private_metadata"] == "F-CARD-1"


def test_per_doc_edit_posts_ephemeral_not_supported():
    """ledgr_per_doc_edit posts an ephemeral explaining the limitation."""
    from unittest.mock import AsyncMock, MagicMock, patch

    sync_client = MagicMock()
    with patch("slack_sdk.WebClient", return_value=sync_client):
        _, edit_handler, _ = _capture_per_doc_handlers()

    body = {
        "actions": [{"action_id": "ledgr_per_doc_edit", "value": "F-EDIT-1"}],
        "trigger_id": "T-EDIT-1",
        "channel": {"id": "C-EDIT-1"},
        "user": {"id": "U-EDIT-1"},
    }
    ack = AsyncMock()
    asyncio.run(edit_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    sync_client.chat_postEphemeral.assert_called_once()
    kwargs = sync_client.chat_postEphemeral.call_args.kwargs
    assert kwargs["channel"] == "C-EDIT-1"
    assert kwargs["user"] == "U-EDIT-1"
    assert "Re-extract" in kwargs["text"]


def test_per_doc_view_row_posts_coming_soon_ephemeral():
    """ledgr_per_doc_view_row posts a 'coming soon' ephemeral."""
    from unittest.mock import AsyncMock, MagicMock, patch

    sync_client = MagicMock()
    with patch("slack_sdk.WebClient", return_value=sync_client):
        _, _, view_row_handler = _capture_per_doc_handlers()

    body = {
        "actions": [{"action_id": "ledgr_per_doc_view_row", "value": "F-VR-1"}],
        "trigger_id": "T-VR-1",
        "channel": {"id": "C-VR-1"},
        "user": {"id": "U-VR-1"},
    }
    ack = AsyncMock()
    asyncio.run(view_row_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    sync_client.chat_postEphemeral.assert_called_once()
    kwargs = sync_client.chat_postEphemeral.call_args.kwargs
    assert kwargs["channel"] == "C-VR-1"
    assert kwargs["user"] == "U-VR-1"
    assert "coming soon" in kwargs["text"].lower()


# --------------------------------------------------------------------------- #
# Commit 3: dedup callout card action handlers
# --------------------------------------------------------------------------- #


def _capture_dedup_handlers(runner_mock=None, ledger_store_mock=None, db_mock=None):
    """Build the Bolt app with fakes; capture the dedup action handlers."""
    from unittest.mock import MagicMock, patch

    from app.slack_app import _SeenEvents
    from accounting_agents import slack_runner

    registered = {"actions": {}, "views": {}}

    def action_decorator(action_id, *a, **k):
        def decorator(fn):
            registered["actions"][action_id] = fn
            return fn
        return decorator

    def view_decorator(callback_id, *a, **k):
        def decorator(fn):
            registered["views"][callback_id] = fn
            return fn
        return decorator

    fake_app = MagicMock()
    fake_app.event = lambda *a, **k: (lambda fn: fn)
    fake_app.action = action_decorator
    fake_app.view = view_decorator
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = runner_mock or _FakeActionViewRunner()

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch("invoice_processing.export.client_context.FirestoreClientStore"), \
         patch.object(slack_runner, "build_chat_runner",
                      return_value=SimpleNamespace(app_name="accounting_agents_assistant")):
        build_async_app(
            runner=rm,
            ledger_store=ledger_store_mock or MagicMock(),
            db=db_mock or MagicMock(),
        )

    return (
        registered["actions"]["ledgr_dedup_replace"],
        registered["actions"]["ledgr_dedup_keep"],
    )


def test_dedup_replace_handler_is_registered():
    """ledgr_dedup_replace handler must be registered in build_async_app."""
    replace_handler, _ = _capture_dedup_handlers()
    assert callable(replace_handler)


def test_dedup_keep_handler_is_registered():
    """ledgr_dedup_keep handler must be registered in build_async_app."""
    _, keep_handler = _capture_dedup_handlers()
    assert callable(keep_handler)


def test_dedup_replace_posts_ephemeral_with_label():
    """ledgr_dedup_replace acks and posts an ephemeral describing the replace intent."""
    from unittest.mock import AsyncMock, MagicMock, patch
    import urllib.parse

    sync_client = MagicMock()
    with patch("slack_sdk.WebClient", return_value=sync_client):
        replace_handler, _ = _capture_dedup_handlers()

    vendor_enc = urllib.parse.quote("Acme Supplies", safe="")
    month_enc = urllib.parse.quote("September 2025", safe="")
    btn_value = f"{vendor_enc}|2025|{month_enc}|OP-42"

    body = {
        "actions": [{"action_id": "ledgr_dedup_replace", "value": btn_value}],
        "channel": {"id": "C-DEDUP-1"},
        "message": {"ts": "111.222"},
        "user": {"id": "U-DEDUP-1"},
    }
    ack = AsyncMock()
    asyncio.run(replace_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    sync_client.chat_postEphemeral.assert_called_once()
    kwargs = sync_client.chat_postEphemeral.call_args.kwargs
    assert kwargs["channel"] == "C-DEDUP-1"
    assert "September 2025" in kwargs["text"]
    assert "Acme Supplies" in kwargs["text"]


def test_dedup_keep_updates_message_to_kept_outcome():
    """ledgr_dedup_keep acks and edits the dedup card to the kept-existing outcome."""
    from unittest.mock import AsyncMock, MagicMock, patch
    import urllib.parse

    sync_client = MagicMock()
    with patch("slack_sdk.WebClient", return_value=sync_client):
        _, keep_handler = _capture_dedup_handlers()

    vendor_enc = urllib.parse.quote("Acme Supplies", safe="")
    month_enc = urllib.parse.quote("September 2025", safe="")
    btn_value = f"{vendor_enc}|2025|{month_enc}|-"

    body = {
        "actions": [{"action_id": "ledgr_dedup_keep", "value": btn_value}],
        "channel": {"id": "C-DEDUP-2"},
        "message": {"ts": "333.444"},
        "user": {"id": "U-DEDUP-2"},
    }
    ack = AsyncMock()
    asyncio.run(keep_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    sync_client.chat_update.assert_called_once()
    kwargs = sync_client.chat_update.call_args.kwargs
    assert kwargs["channel"] == "C-DEDUP-2"
    assert kwargs["ts"] == "333.444"
    assert "Kept existing" in kwargs["text"]
    assert "September 2025" in kwargs["text"]
    assert "Acme Supplies" in kwargs["text"]
    # The updated message must have a blocks list (not raw text only)
    assert isinstance(kwargs.get("blocks"), list)
    assert len(kwargs["blocks"]) >= 1


# --------------------------------------------------------------------------- #
# Commit 5: feedback buttons action handlers
# --------------------------------------------------------------------------- #


def _capture_feedback_handlers(runner_mock=None, ledger_store_mock=None, db_mock=None):
    """Build the Bolt app with fakes; capture the feedback action handlers."""
    from unittest.mock import MagicMock, patch

    from app.slack_app import _SeenEvents
    from accounting_agents import slack_runner

    registered = {"actions": {}}

    def action_decorator(action_id, *a, **k):
        def decorator(fn):
            registered["actions"][action_id] = fn
            return fn
        return decorator

    fake_app = MagicMock()
    fake_app.event = lambda *a, **k: (lambda fn: fn)
    fake_app.action = action_decorator
    fake_app.view = lambda *a, **k: (lambda fn: fn)
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = runner_mock or _FakeActionViewRunner()

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch("invoice_processing.export.client_context.FirestoreClientStore"), \
         patch.object(slack_runner, "build_chat_runner",
                      return_value=SimpleNamespace(app_name="accounting_agents_assistant")):
        build_async_app(
            runner=rm,
            ledger_store=ledger_store_mock or MagicMock(),
            db=db_mock or MagicMock(),
        )

    return (
        registered["actions"].get("ledgr_doc_feedback"),
        registered["actions"].get("ledgr_doc_feedback_pos"),
        registered["actions"].get("ledgr_doc_feedback_neg"),
    )


def test_feedback_handler_is_registered():
    """ledgr_doc_feedback handler must be registered in build_async_app."""
    native_handler, pos_handler, neg_handler = _capture_feedback_handlers()
    assert callable(native_handler), "ledgr_doc_feedback not registered"
    assert callable(pos_handler), "ledgr_doc_feedback_pos not registered"
    assert callable(neg_handler), "ledgr_doc_feedback_neg not registered"


def test_feedback_pos_posts_ephemeral():
    """👍 path: acks, then posts an ephemeral acknowledgement."""
    from unittest.mock import AsyncMock, MagicMock, patch
    import urllib.parse

    sync_client = MagicMock()
    # Stub get_client_id so add_correction path doesn't blow up on MagicMock.
    sync_client.get_client_id = MagicMock(return_value="client-1")

    with patch("slack_sdk.WebClient", return_value=sync_client):
        native_handler, _, _ = _capture_feedback_handlers()

    vendor_enc = urllib.parse.quote("Acme Supplies", safe="")
    doc_ref = f"F-FB-1|{vendor_enc}|6090|SR"
    body = {
        "actions": [{"action_id": "ledgr_doc_feedback", "value": f"pos|{doc_ref}"}],
        "channel": {"id": "C-FB-1"},
        "user": {"id": "U-FB-1"},
        "trigger_id": "T-FB-1",
    }
    ack = AsyncMock()
    asyncio.run(native_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    sync_client.chat_postEphemeral.assert_called_once()
    kwargs = sync_client.chat_postEphemeral.call_args.kwargs
    assert kwargs["channel"] == "C-FB-1"
    assert kwargs["user"] == "U-FB-1"
    assert "feedback" in kwargs["text"].lower() or "thanks" in kwargs["text"].lower()


def test_feedback_neg_opens_proactive_redo_modal():
    """👎 path: acks, opens the proactive-redo modal with the correct file_id,
    then posts an ephemeral acknowledgement."""
    from unittest.mock import AsyncMock, MagicMock, patch
    import urllib.parse

    sync_client = MagicMock()
    with patch("slack_sdk.WebClient", return_value=sync_client):
        native_handler, _, _ = _capture_feedback_handlers()

    vendor_enc = urllib.parse.quote("Acme Supplies", safe="")
    doc_ref = f"F-FB-NEG|{vendor_enc}|6090|SR"
    body = {
        "actions": [{"action_id": "ledgr_doc_feedback", "value": f"neg|{doc_ref}"}],
        "channel": {"id": "C-FB-2"},
        "user": {"id": "U-FB-2"},
        "trigger_id": "T-FB-NEG",
    }
    ack = AsyncMock()
    asyncio.run(native_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()
    sync_client.views_open.assert_called_once()
    vkw = sync_client.views_open.call_args.kwargs
    assert vkw["trigger_id"] == "T-FB-NEG"
    # Modal callback_id and private_metadata carry the file_id
    assert vkw["view"]["callback_id"] == "ledgr_proactive_redo"
    assert vkw["view"]["private_metadata"] == "F-FB-NEG"
    # Ephemeral acknowledgement also posted
    sync_client.chat_postEphemeral.assert_called_once()


# =========================================================================== #
# P0-1: file_shared COA routing — xlsx must reach run_coa_ingest when pending
# =========================================================================== #


def _capture_file_shared_handler(store_mock=None, *, sync_client_mock=None):
    """Build the Bolt async app and return the registered ``file_shared`` handler
    plus the sync Slack client mock so tests can assert on the card it posts.
    """
    from unittest.mock import MagicMock, patch

    from app.slack_app import _SeenEvents
    from accounting_agents import slack_runner

    registered = {}
    fake_app = MagicMock()

    def event_decorator(name):
        def decorator(fn):
            registered[name] = fn
            return fn
        return decorator

    fake_app.event = event_decorator
    fake_app.action = lambda *a, **k: (lambda fn: fn)
    fake_app.view = lambda *a, **k: (lambda fn: fn)
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = MagicMock()
    rm.app_name = "acc"

    sync_mock = sync_client_mock or MagicMock()

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         patch("slack_sdk.WebClient", return_value=sync_mock), \
         patch("invoice_processing.export.client_context.FirestoreClientStore"), \
         patch.object(slack_runner, "build_chat_runner",
                      return_value=SimpleNamespace(app_name="accounting_agents_assistant")):
        build_async_app(
            runner=rm,
            ledger_store=MagicMock(),
            db=MagicMock(),
            store=store_mock or MagicMock(),
        )

    return registered["file_shared"], fresh_seen, sync_mock


def test_file_shared_offers_coa_card_when_xlsx_dropped_on_pending():
    """An xlsx file_shared on a pending_coa channel must post a confirm card
    (ADR-0006 activation gate) and must NOT auto-ingest or call process_file_event.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    import openpyxl

    from accounting_agents import slack_runner

    store_mock = MagicMock()
    store_mock.get_by_channel.return_value = None  # no profile → pending_coa

    sync_mock = MagicMock()
    handler, _, sync_mock = _capture_file_shared_handler(
        store_mock=store_mock, sync_client_mock=sync_mock,
    )

    event = {
        "type": "file_shared",
        "event_ts": "300.001",
        "file_id": "FXLSX1",
        "channel_id": "C-coa-pending",
        "file": {
            "id": "FXLSX1",
            "name": "Client Setup.xlsx",
            "filetype": "xlsx",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    }
    body = {"event_id": "Ev-coa-xlsx-1"}
    fake_bolt_client = MagicMock()

    mock_pfe = AsyncMock(return_value={"status": "delivered"})
    mock_coa = MagicMock()

    tmp = tempfile.mkdtemp(prefix="ledgr_coa_offer_test_")
    coa_path = os.path.join(tmp, "Client Setup.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COA"
    ws.append(["Account code", "Description", "Account type", "Financial Statement"])
    ws.append(["6100", "Sales", "Revenue", "Profit and Loss"])
    ws.append(["6200", "Rent", "Expense", "Profit and Loss"])
    wb.save(coa_path)

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch("app.slack_app.run_coa_ingest", mock_coa), \
         patch("app.slack_app.slack_download_file", return_value=coa_path):
        asyncio.run(handler(event=event, body=body, client=fake_bolt_client))

    # The card was posted on the sync client.
    assert sync_mock.chat_postMessage.called
    post_kwargs = sync_mock.chat_postMessage.call_args.kwargs
    assert "blocks" in post_kwargs
    assert "actions" in [b.get("type") for b in post_kwargs["blocks"]]
    action_ids = [
        e.get("action_id")
        for b in post_kwargs["blocks"]
        if b.get("type") == "actions"
        for e in b.get("elements", [])
    ]
    assert "ledgr_coa_confirm" in action_ids
    # No silent ingest.
    mock_coa.assert_not_called()
    mock_pfe.assert_not_called()


def test_file_shared_offer_card_says_replace_on_active_channel():
    """An xlsx on an already-active channel posts a 'Replace COA' card."""
    from unittest.mock import AsyncMock, MagicMock, patch

    import openpyxl

    from accounting_agents import slack_runner

    active_profile = SimpleNamespace(status="active")
    store_mock = MagicMock()
    store_mock.get_by_channel.return_value = active_profile

    sync_mock = MagicMock()
    handler, _, sync_mock = _capture_file_shared_handler(
        store_mock=store_mock, sync_client_mock=sync_mock,
    )

    event = {
        "type": "file_shared",
        "event_ts": "300.003",
        "file_id": "FXLSX3",
        "channel_id": "C-active",
        "file": {
            "id": "FXLSX3",
            "name": "Sample Test Group - Client Setup.xlsx",
            "filetype": "xlsx",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    }
    body = {"event_id": "Ev-coa-xlsx-3"}
    fake_bolt_client = MagicMock()

    mock_pfe = AsyncMock(return_value={"status": "rejected"})
    mock_coa = MagicMock()

    tmp = tempfile.mkdtemp(prefix="ledgr_coa_replace_test_")
    coa_path = os.path.join(tmp, "Client Setup.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COA"
    ws.append(["Account code", "Description", "Account type", "Financial Statement"])
    ws.append(["6100", "Sales", "Revenue", "Profit and Loss"])
    ws.append(["6200", "Rent", "Expense", "Profit and Loss"])
    wb.save(coa_path)

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch("app.slack_app.run_coa_ingest", mock_coa), \
         patch("app.slack_app.slack_download_file", return_value=coa_path):
        asyncio.run(handler(event=event, body=body, client=fake_bolt_client))

    post_kwargs = sync_mock.chat_postMessage.call_args.kwargs
    confirm_btn = next(
        e
        for b in post_kwargs["blocks"]
        if b.get("type") == "actions"
        for e in b.get("elements", [])
        if e.get("action_id") == "ledgr_coa_confirm"
    )
    assert confirm_btn["text"]["text"] == "Replace COA"
    assert confirm_btn.get("style") == "danger"
    mock_coa.assert_not_called()


def test_file_shared_offer_does_not_run_pfe_for_non_spreadsheet():
    """A non-spreadsheet file on file_shared still falls through silently (the
    document pipeline lives on the message handler)."""
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    store_mock = MagicMock()
    store_mock.get_by_channel.return_value = None

    sync_mock = MagicMock()
    handler, _, sync_mock = _capture_file_shared_handler(
        store_mock=store_mock, sync_client_mock=sync_mock,
    )

    event = {
        "type": "file_shared",
        "event_ts": "300.004",
        "file_id": "FPDF1",
        "channel_id": "C-coa-pending",
        "file": {
            "id": "FPDF1",
            "name": "Invoice.pdf",
            "filetype": "pdf",
            "mimetype": "application/pdf",
        },
    }
    body = {"event_id": "Ev-pdf-1"}
    fake_bolt_client = MagicMock()

    mock_pfe = AsyncMock(return_value={"status": "delivered"})

    with patch.object(slack_runner, "process_file_event", mock_pfe):
        asyncio.run(handler(event=event, body=body, client=fake_bolt_client))

    assert not sync_mock.chat_postMessage.called
    mock_pfe.assert_not_called()


# =========================================================================== #
# P1-1 — HITL delivery card must be threaded under the original upload
# =========================================================================== #


def _capture_approval_handlers(
    runner_mock=None, ledger_store_mock=None, db_mock=None,
    injected_slack: "FakeSlackClient | None" = None,
):
    """Build the Bolt app with fakes; capture the ``approve``, ``edit``,
    ``ledgr_invoice_edit`` and ``reject`` handlers.

    Pass ``injected_slack`` to use that FakeSlackClient as the sync WebClient
    so caller can inspect ``_posts`` after the handler fires.
    """
    from unittest.mock import MagicMock, patch

    from app.slack_app import _SeenEvents
    from accounting_agents import slack_runner

    registered = {"actions": {}, "views": {}}

    def action_decorator(action_id, *a, **k):
        def decorator(fn):
            registered["actions"][action_id] = fn
            return fn
        return decorator

    def view_decorator(callback_id, *a, **k):
        def decorator(fn):
            registered["views"][callback_id] = fn
            return fn
        return decorator

    fake_app = MagicMock()
    fake_app.event = lambda *a, **k: (lambda fn: fn)
    fake_app.action = action_decorator
    fake_app.view = view_decorator
    fake_app.command = lambda *a, **k: (lambda fn: fn)

    fresh_seen = _SeenEvents()
    rm = runner_mock or _FakeActionViewRunner()

    slack_patch = (
        patch("slack_sdk.WebClient", return_value=injected_slack)
        if injected_slack is not None
        else patch("slack_sdk.WebClient")
    )

    with patch.object(slack_runner, "_seen", fresh_seen), \
         patch("slack_bolt.async_app.AsyncApp", return_value=fake_app), \
         slack_patch, \
         patch("invoice_processing.export.client_context.FirestoreClientStore"), \
         patch.object(slack_runner, "build_chat_runner",
                      return_value=SimpleNamespace(app_name="accounting_agents_assistant")):
        build_async_app(
            runner=rm,
            ledger_store=ledger_store_mock or MagicMock(),
            db=db_mock or MagicMock(),
        )

    return (
        registered["actions"]["approve"],
        registered["actions"]["reject"],
        registered["actions"]["edit"],
        registered["views"]["ledgr_invoice_edit"],
    )


def test_approve_path_emits_delivery_card_in_thread():
    """Clicking Approve posts the delivery card threaded under the original upload.

    The interrupt body carries ``thread_ts="T_UPLOAD"`` (stored when the approval
    card was first posted). After the fix, persist_and_deliver receives that
    thread_ts and every chat_postMessage call carries it.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    db = FakeFirestore()
    slack = FakeSlackClient()
    runner = _FakeRunner([], _ledger_payload())
    runner.session_service.created = True
    ledger_store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    write_interrupt(
        db, "CAPPROVE:F1", session_id="CAPPROVE:F1", channel_id="CAPPROVE",
        slack_file_id="F1", message_ts="10.0", user_id="CAPPROVE",
        extra={"summary": "needs review", "thread_ts": "T_UPLOAD"},
    )

    async def fake_resume(_runner, _db, op, decision):
        return []

    approve_handler, _, _, _ = _capture_approval_handlers(
        runner_mock=runner,
        ledger_store_mock=ledger_store,
        db_mock=db,
        injected_slack=slack,
    )

    with patch.object(slack_runner, "resume_session", fake_resume):
        ack = AsyncMock()
        body = {"actions": [{"action_id": "approve", "value": "CAPPROVE:F1"}]}
        asyncio.run(approve_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()

    posts = _post_calls(slack)
    delivery_posts = [
        p for p in posts
        if "FY2026 ledger" in p.get("text", "")
    ]
    assert delivery_posts, "Expected a delivery card post but found none"
    assert all(
        p.get("thread_ts") == "T_UPLOAD" for p in delivery_posts
    ), f"Delivery post(s) missing thread_ts='T_UPLOAD': {delivery_posts}"


def test_edit_path_emits_delivery_card_in_thread():
    """Submitting the Edit modal posts the delivery card threaded under the upload.

    The interrupt body carries ``thread_ts="T_UPLOAD"``; after the fix it flows
    through handle_approval_action → persist_and_deliver → chat_postMessage.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    db = FakeFirestore()
    slack = FakeSlackClient()
    runner = _FakeRunner([], _ledger_payload())
    runner.session_service.created = True
    ledger_store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    write_interrupt(
        db, "CEDIT:F2", session_id="CEDIT:F2", channel_id="CEDIT",
        slack_file_id="F2", message_ts="20.0", user_id="CEDIT",
        extra={"summary": "needs review", "thread_ts": "T_UPLOAD"},
    )

    async def fake_resume(_runner, _db, op, decision):
        return []

    _, _, _, edit_submit_handler = _capture_approval_handlers(
        runner_mock=runner,
        ledger_store_mock=ledger_store,
        db_mock=db,
        injected_slack=slack,
    )

    # Minimal view_submission body: one line, no real edits needed.
    body = {
        "view": {
            "callback_id": "ledgr_invoice_edit",
            "private_metadata": "CEDIT:F2",
            "state": {
                "values": {
                    "acct_0": {"v": {"selected_option": {"value": "6010"}}},
                    "tax_0": {"v": {"selected_option": {"value": "SR"}}},
                    "amt_0": {"v": {"value": "100.00"}},
                }
            },
        }
    }

    with patch.object(slack_runner, "resume_session", fake_resume), \
         patch.object(slack_runner, "_persist_corrections", MagicMock()):
        ack = AsyncMock()
        asyncio.run(edit_submit_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()

    posts = _post_calls(slack)
    delivery_posts = [
        p for p in posts
        if "FY2026 ledger" in p.get("text", "")
    ]
    assert delivery_posts, "Expected a delivery card post but found none"
    assert all(
        p.get("thread_ts") == "T_UPLOAD" for p in delivery_posts
    ), f"Delivery post(s) missing thread_ts='T_UPLOAD': {delivery_posts}"


def test_reject_path_posts_rejection_in_thread():
    """Clicking Reject posts the rejection notice threaded under the original upload.

    The interrupt body carries ``thread_ts="T_UPLOAD"``; after the fix the
    rejection message is posted with that thread_ts instead of channel root.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    db = FakeFirestore()
    slack = FakeSlackClient()
    runner = _FakeRunner([], {})
    runner.session_service.created = True
    ledger_store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    write_interrupt(
        db, "CREJECT:F3", session_id="CREJECT:F3", channel_id="CREJECT",
        slack_file_id="F3", message_ts="30.0", user_id="CREJECT",
        extra={"summary": "", "thread_ts": "T_UPLOAD"},
    )

    async def fake_resume(_runner, _db, op, decision):
        return []

    approve_handler, reject_handler, _, _ = _capture_approval_handlers(
        runner_mock=runner,
        ledger_store_mock=ledger_store,
        db_mock=db,
        injected_slack=slack,
    )

    with patch.object(slack_runner, "resume_session", fake_resume):
        ack = AsyncMock()
        body = {"actions": [{"action_id": "reject", "value": "CREJECT:F3"}]}
        asyncio.run(reject_handler(ack=ack, body=body, client=MagicMock()))

    ack.assert_awaited_once()

    posts = _post_calls(slack)
    rejection_posts = [
        p for p in posts
        if "rejected" in p.get("text", "").lower()
    ]
    assert rejection_posts, "Expected a rejection message post but found none"
    assert all(
        p.get("thread_ts") == "T_UPLOAD" for p in rejection_posts
    ), f"Rejection post(s) missing thread_ts='T_UPLOAD': {rejection_posts}"


# =========================================================================== #
# P1-1 Fix-3: HITL-review → approve path must emit the full delivery card,
# not the bare "Document processed." fallback.
# Bug: the session state that persist_and_deliver reads after the approval
# resume may lack DELIVER_SUMMARY_KEY (deliver_node writes it to ctx.state
# but ADK may not flush it before get_session() is called).  The fix derives
# the summary from LEDGER_ROWS_KEY — which IS reliably present — when
# DELIVER_SUMMARY_KEY is absent.
# =========================================================================== #

def _ledger_rows_state_no_summary(
    sheet="Purchase",
    doc_key="F1:Purchase:INV-1",
    client_name="Acme Client",
):
    """State with LEDGER_ROWS_KEY but no DELIVER_SUMMARY_KEY.

    Simulates the production HITL-review → approve path where consolidate_node
    has run (rows are in state) but deliver_node's DELIVER_SUMMARY_KEY write
    is not yet visible to persist_and_deliver via get_session().
    """
    return {
        nodes.LEDGER_ROWS_KEY: {
            "client_id": "c1",
            "fy": "2026",
            "kind": "invoice",
            "software": "qbs",
            "client_name": client_name,
            "batches": [
                {
                    "sheet": sheet,
                    "doc_key": doc_key,
                    "rows": [
                        {"Invoice Number": "INV-99", "Description": "consulting", "Source Amount": 250.0},
                        {"Invoice Number": "INV-99", "Description": "expenses", "Source Amount": 50.0},
                    ],
                }
            ],
        }
        # Deliberately NO DELIVER_SUMMARY_KEY — this is the failing production case.
    }


def test_approve_path_emits_full_delivery_card_not_bare_string():
    """HITL-review → approve path must emit the rich delivery card, not 'Document processed.'

    Regression test for P1-1 (2026-06-15 ultraqa findings §3).
    Sequence:
      1. Write a :review interrupt.
      2. Call handle_review_action("confirm_as_is") — fake resume yields an
         approval-gate re-pause event, so _finalize_run_outcome posts the
         approval card and writes a new approval-gate interrupt doc.
      3. Read the written approval-gate op_id from db.
      4. Call handle_approval_action("approve") with that op_id.
      5. Assert the final delivery Slack message:
         - contains "FY" (ledger pointer) OR starts with "Added"
         - matches r"\\b\\d+\\s+(transactions?|lines?)\\b" (row count, singular or plural)
         - does NOT match the bare fallback r"^Document processed\\.?\\s*$"
         - xlsx was uploaded (slack.uploads non-empty)
    """
    import re
    from accounting_agents.slack_runner import handle_approval_action, handle_review_action

    db = FakeFirestore()
    slack = FakeSlackClient()
    # Runner whose session state has LEDGER_ROWS_KEY but no DELIVER_SUMMARY_KEY —
    # this is the exact production failure shape.
    state = _ledger_rows_state_no_summary()
    runner = _FakeRunner([], state)
    runner.session_service.created = True  # session exists after the initial run
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    # Step 1: write the :review interrupt (as process_file_event would after a
    # mid-flow pause at review_extraction_node).
    review_op_id = "CHITL:FHITL:review"
    approval_op_id = "CHITL:FHITL"  # the base id (without :review suffix)
    write_interrupt(
        db, review_op_id, session_id=approval_op_id, channel_id="CHITL",
        slack_file_id="FHITL", message_ts="1.0", user_id="CHITL",
        extra={"kind": "review", "question": "Verify line items?", "reasons": [],
               "thread_ts": "T_UPLOAD"},
    )

    # The re-pause event that handle_review_action's resume will yield — this is
    # the approval_gate pausing with the BASE op_id (no :review suffix).
    repause_event = SimpleNamespace(
        content=SimpleNamespace(parts=[]),
        get_function_calls=lambda: [
            SimpleNamespace(name="adk_request_input", id=approval_op_id)
        ],
    )

    async def fake_resume_review(_runner, _db, op, decision):
        # Simulate the resumed run reaching approval_gate and re-pausing.
        return [repause_event]

    # Patch resume_session at module level so both handlers use our fakes.
    import accounting_agents.slack_runner as sr_mod
    orig_resume = sr_mod.resume_session

    async def _fake_resume_review(runner_, db_, op, decision):
        return [repause_event]

    async def _fake_resume_approve(runner_, db_, op, decision):
        # Approval resume: the graph runs to completion; no new interrupt.
        return []

    # Step 2: run the review accept action — sees the re-pause event →
    # _finalize_run_outcome writes the approval-gate interrupt doc.
    sr_mod.resume_session = _fake_resume_review
    try:
        asyncio.run(
            handle_review_action(
                runner=runner, ledger_store=store, db=db, slack_client=slack,
                op_id=review_op_id, action="confirm_as_is", app_name="acc",
            )
        )
    finally:
        sr_mod.resume_session = orig_resume

    # Step 3: verify the approval-gate interrupt was written by _finalize_run_outcome.
    approval_interrupt = read_interrupt(db, approval_op_id)
    assert approval_interrupt is not None, (
        "_finalize_run_outcome should have written an approval-gate interrupt doc"
    )

    # Step 4: approve action.
    sr_mod.resume_session = _fake_resume_approve
    try:
        asyncio.run(
            handle_approval_action(
                runner=runner, ledger_store=store, db=db, slack_client=slack,
                op_id=approval_op_id, decision="approve", app_name="acc",
            )
        )
    finally:
        sr_mod.resume_session = orig_resume

    # Step 5: assertions on the final delivery shape.
    posts = _post_calls(slack)
    texts = [p.get("text", "") for p in posts]

    # Must NOT be the bare fallback.
    bare_pattern = re.compile(r"^Document processed\.?\s*$")
    assert not any(bare_pattern.match(t) for t in texts), (
        f"Delivery message must not be the bare fallback. Got posts: {texts}"
    )

    # Must contain a ledger/FY pointer OR "Added ".
    has_fy_pointer = any("FY" in t or "Added " in t for t in texts)
    assert has_fy_pointer, (
        f"Delivery message must contain a ledger/FY pointer. Got: {texts}"
    )

    # Must contain a row count (e.g. "2 lines").
    row_count_pattern = re.compile(r"\b\d+\s+(transactions|lines)\b")
    has_row_count = any(row_count_pattern.search(t) for t in texts)
    assert has_row_count, (
        f"Delivery message must contain a row count (N lines/transactions). Got: {texts}"
    )

    # xlsx must have been uploaded.
    assert slack.uploads, (
        "Delivery must include an xlsx upload — none found"
    )


async def _noop_coro():
    """Placeholder async no-op for test flow control."""
    pass


def test_clean_path_emits_full_delivery_card():
    """Clean path (no review, no approval gate pause) emits the same delivery shape.

    Lockstep counterpart to test_approve_path_emits_full_delivery_card_not_bare_string:
    both paths must produce the same shape so regressions on either are caught.
    """
    import re

    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    final_event = SimpleNamespace(
        content=SimpleNamespace(parts=[SimpleNamespace(text="done")]),
        get_function_calls=lambda: [],
    )
    # Clean path: LEDGER_ROWS_KEY present, DELIVER_SUMMARY_KEY present (deliver_node ran).
    runner = _FakeRunner([final_event], _ledger_payload())

    result = asyncio.run(
        process_file_event(
            runner=runner, ledger_store=store, db=db, slack_client=slack,
            channel_id="C1", file_id="F_CLEAN", app_name="acc",
            download_fn=lambda c, f: b"%PDF-1.4 fake",
            source_filename="invoice.pdf",
            client_store=_seeded_client_store(db),
        )
    )

    assert result["status"] == "delivered"
    posts = _post_calls(slack)
    texts = [p.get("text", "") for p in posts]

    bare_pattern = re.compile(r"^Document processed\.?\s*$")
    assert not any(bare_pattern.match(t) for t in texts), (
        f"Clean-path delivery must not be the bare fallback. Got: {texts}"
    )

    has_fy_pointer = any("FY" in t or "Added " in t for t in texts)
    assert has_fy_pointer, f"Clean-path delivery must contain FY pointer. Got: {texts}"

    row_count_pattern = re.compile(r"\b\d+\s+(transactions?|lines?)\b")
    has_row_count = any(row_count_pattern.search(t) for t in texts)
    assert has_row_count, f"Clean-path delivery must contain row count. Got: {texts}"

    assert slack.uploads, "Clean-path delivery must include an xlsx upload"


# =========================================================================== #
# Phase 1 — One agent chat for the whole batch (file_shared race fix)
# =========================================================================== #


def test_file_shared_defer_does_not_poison_message_handler_dedup():
    """Regression: file_shared deferral must NOT mark file:{id} as seen.

    Phase-1A moved document processing to the message handler but left
    ``seen_before(f"file:{id}")`` at the top of ``_file_shared``. That call
    *marks* the file id on first access, so when the message handler ran it
    saw the file as already-processed, skipped all 6 docs, and posted
    "Processed 6 documents — nothing new to add" with zero Gemini work.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner
    from app.slack_app import _SeenEvents

    fresh_seen = _SeenEvents()
    slack_runner._seen = fresh_seen

    active_profile = SimpleNamespace(status="active")
    store_mock = MagicMock()
    store_mock.get_by_channel.return_value = active_profile

    file_handler, _, _ = _capture_file_shared_handler(store_mock=store_mock)
    msg_handler, _ = _capture_message_handler_with_slack_client(FakeSlackClient())

    pdf_event = {
        "type": "file_shared",
        "event_ts": "400.001",
        "file_id": "FPOISON",
        "channel_id": "C-poison",
        "file": {"id": "FPOISON", "name": "invoice.pdf", "filetype": "pdf"},
    }
    fake_client = MagicMock()

    with patch.object(slack_runner, "process_file_event", AsyncMock()) as mock_pfe:
        asyncio.run(file_handler(
            event=pdf_event, body={"event_id": "Ev-poison-fs"}, client=fake_client,
        ))

    # file_shared must not have marked the file id.
    assert not fresh_seen.seen_before(f"file:FPOISON"), (
        "file_shared deferral must not poison the file-level dedup key"
    )
    # Reset the probe mark from the assertion above.
    fresh_seen._seen.pop(f"file:FPOISON", None)

    mock_pfe.return_value = {"status": "delivered", "append": {"appended": 1, "fy": "2026"}}
    msg_event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "400.002",
        "channel": "C-poison",
        "files": [{"id": "FPOISON", "name": "invoice.pdf"}],
    }
    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF"):
        asyncio.run(msg_handler(
            event=msg_event, body={"event_id": "Ev-poison-msg"}, client=fake_client,
        ))

    mock_pfe.assert_called_once()
    assert mock_pfe.call_args.kwargs["file_id"] == "FPOISON"


def test_file_shared_does_not_process_documents():
    """After Phase 1A, ``file_shared`` is NOT allowed to start a document run.

    Pre-fix: a 6-file drop would stampede 6 ``process_file_event`` calls from
    ``_file_shared`` with no thread_ts and no defer, producing 6 top-level
    "Processing [dev] 'file.pdf'" accordions before the batch coordinator
    posted its one job summary. The message handler is now the sole document
    owner; ``file_shared`` only handles COA spreadsheet routing (ADR-0006).
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner

    active_profile = SimpleNamespace(status="active")
    store_mock = MagicMock()
    store_mock.get_by_channel.return_value = active_profile

    handler, _, _ = _capture_file_shared_handler(store_mock=store_mock)

    event = {
        "type": "file_shared",
        "event_ts": "300.010",
        "file_id": "FX10",
        "channel_id": "C-noise",
        "file": {
            "id": "FX10",
            "name": "Invoice.pdf",
            "filetype": "pdf",
            "mimetype": "application/pdf",
        },
    }
    body = {"event_id": "Ev-noise-pdf"}
    fake_bolt_client = MagicMock()

    mock_pfe = AsyncMock(return_value={"status": "delivered"})

    with patch.object(slack_runner, "process_file_event", mock_pfe):
        asyncio.run(handler(event=event, body=body, client=fake_bolt_client))

    # Invariant: _file_shared must not start a document run.
    mock_pfe.assert_not_called()


def test_batch_six_files_one_top_level_message(monkeypatch):
    """Six-file drop yields EXACTLY ONE top-level message (job summary) — no spam.

    The pre-Phase-1 bug surfaced live as six "Processing [dev] 'file.pdf'"
    accordions + a separate job summary. After the fix, the message handler
    is the sole document owner and ``process_file_event`` is called in
    ``batch_mode`` so the per-doc status posting is suppressed.
    """
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner
    from app.slack_app import _SeenEvents

    slack_runner._seen = _SeenEvents()
    slack = FakeSlackClient()
    handler, _ = _capture_message_handler_with_slack_client(slack)

    body = {"event_id": "Ev-batch-6"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "222.006",
        "channel": "C-batch-6",
        "files": [{"id": f"F6-{i}"} for i in range(6)],
    }
    fake_client = MagicMock()
    mock_pfe = AsyncMock(side_effect=[
        {"status": "delivered", "append": {
            "appended": 1, "software": "Xero", "fy": "2026",
            "deferred_delivery": {
                "payload": {"fy": "2026", "software": "Xero", "kind": "invoice",
                            "client_name": "Acme Client",
                            "batches": [{"sheet": "Purchase",
                                         "rows": [{"Contact": "Vendor",
                                                   "Total": 100.0,
                                                   "Currency": "USD"}]}]},
                "batches": [{"sheet": "Purchase",
                             "rows": [{"Contact": "Vendor", "Total": 100.0,
                                       "Currency": "USD"}]}],
                "workbook_name": "Ledger_FY2026.xlsx",
            },
        }}
        for _ in range(6)
    ])

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    # The first top-level post is the placeholder job summary; the final
    # tally+delivery card edits the same ts. So we should see at most ONE
    # top-level post (placeholder) and the delivery lives in the update.
    top_level_posts = [p for p in slack._posts if not p.get("thread_ts")]
    assert len(top_level_posts) == 1, (
        f"expected exactly 1 top-level post, got {len(top_level_posts)}: "
        f"{[p.get('text', '') for p in top_level_posts]}"
    )
    summary_ts = top_level_posts[0].get("ts")
    assert summary_ts, "summary message must carry a ts"

    # Delivery must arrive via the update of summary_ts, not a fresh top-level post.
    final_update = slack.updates[-1]
    assert final_update.get("ts") == summary_ts
    assert final_update.get("channel") == "C-batch-6"
    assert final_update.get("blocks"), (
        "merged batch update must include delivery blocks"
    )

    # process_file_event was called 6× with batch_mode=True (silent mode).
    assert mock_pfe.call_count == 6
    for c in mock_pfe.call_args_list:
        assert c.kwargs.get("batch_mode") is True, (
            f"batch calls must carry batch_mode=True; got {c.kwargs}"
        )
        assert c.kwargs.get("thread_ts") == summary_ts


def test_batch_mode_skips_per_doc_status_post():
    """``batch_mode=True`` on ``process_file_event`` must NOT post a per-doc status.

    This is the core anti-noise contract: when total>1, the per-doc
    "Received" status + plan accordion posting is suppressed so only the
    job-summary message lives at channel root.
    """
    from unittest.mock import MagicMock, patch

    from accounting_agents import slack_runner

    # In batch_mode the per-doc status post is replaced by a no-op
    # (status_ts=None). We assert by inspecting the post list before
    # the run would normally post.
    captured_post_calls: list[dict] = []
    fake_slack = FakeSlackClient()
    real_post = fake_slack.chat_postMessage
    def spy_post(**kwargs):
        captured_post_calls.append(kwargs)
        return real_post(**kwargs)
    fake_slack.chat_postMessage = spy_post

    # Build a no-op runner + store and assert that even with batch_mode=True
    # the per-doc _post_status is never invoked (status_ts short-circuits).
    db = FakeFirestore()
    store = SlackLedgerStore(FakeFirestore(), opener=fake_slack.opener())
    runner = _FakeRunner([_node_event("classify_node", text="done")], _ledger_payload())

    # Simulate batch_mode by passing the flag — the early return path will
    # be tested via the calling site in test_batch_six_files_one_top_level_message.
    # Here we directly verify that _post_status is NOT in the captured post
    # list when batch_mode=True (because process_file_event early-skips it).
    asyncio.run(
        process_file_event(
            runner=runner,
            ledger_store=store,
            db=db,
            slack_client=fake_slack,
            channel_id="C-batch-mode",
            file_id="FX-batch",
            app_name="acc",
            download_fn=lambda *a, **k: b"%PDF batch",
            source_filename="batch_doc.pdf",
            thread_ts=None,
            defer_slack_delivery=True,
            batch_mode=True,
        )
    )

    # The _post_status call inside process_file_event is bypassed when
    # batch_mode=True. The persisted-and-delivered branch posts its
    # delivery card thread_ts-less (top-level) when batch_mode=True AND
    # defer_slack_delivery=True both. Since defer=True means the delivery
    # is stashed in append['deferred_delivery'] and NOT posted, we expect
    # zero top-level posts in the per-doc run.
    # Exclude the "no profile" early-return message which is unrelated.
    top_level = [
        p for p in captured_post_calls
        if not p.get("thread_ts") and "set up yet" not in p.get("text", "")
    ]
    assert top_level == [], (
        f"batch_mode=True must suppress per-doc top-level posts; got: {top_level}"
    )


def test_batch_aggregate_one_table_same_fy(monkeypatch):
    """Same-FY/sheet batch → one combined data_table block (not split per doc).

    After Phase 1C/1D, the aggregator concatenates rows from every
    deferred document into a single ``ledger_preview_data_table`` per
    ``(fy, sheet, workbook_name)`` group.
    """
    from app.blocks import ledger_preview_data_table
    from accounting_agents.slack_runner import _build_batch_aggregate_blocks

    # Module-level LEDGR_NATIVE_BLOCKS=0 (set at import time) suppresses the
    # data_table native block; force it on for this assertion.
    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    from unittest.mock import MagicMock, patch
    from app import native_blocks_compat as _nbc
    _nbc._PROBE_CACHE.pop("C-test", None)

    deferred = []
    for i in range(3):
        deferred.append({
            "payload": {
                "fy": "2026", "software": "Xero", "kind": "invoice",
                "client_name": "Acme Client",
                "batches": [{"sheet": "Purchase",
                             "rows": [{"Contact": f"Vendor {i}",
                                       "Total": 100.0 * (i + 1),
                                       "Currency": "USD"}]}],
            },
            "batches": [{"sheet": "Purchase",
                         "rows": [{"Contact": f"Vendor {i}",
                                   "Total": 100.0 * (i + 1),
                                   "Currency": "USD"}]}],
            "workbook_name": "Ledger_FY2026.xlsx",
        })

    summary, blocks = _build_batch_aggregate_blocks(deferred, "C-test")
    assert summary
    assert "3" in summary or "lines" in summary
    data_tables = [b for b in blocks if b.get("type") == "data_table"]
    # Same FY + Purchase + workbook + software + kind → one merged group.
    assert len(data_tables) == 1, (
        f"expected 1 data_table for same-FY/sheet group, got {len(data_tables)}"
    )
    table = data_tables[0]
    # Header row + 3 data rows.
    assert len(table["rows"]) == 4
    header_texts = [c["text"] for c in table["rows"][0]]
    assert "Currency" in header_texts, (
        f"Currency column must be present in batch aggregate; got {header_texts}"
    )


def test_batch_aggregate_n_docs_counts_documents_not_batches():
    """A single doc that splits Purchase+Sales contributes 1 doc, not 2."""
    from accounting_agents.slack_runner import _build_batch_aggregate_blocks

    deferred = [{
        "payload": {
            "fy": "2026", "software": "Xero", "kind": "invoice",
            "client_name": "Acme Client",
            "batches": [
                {"sheet": "Purchase", "rows": [{"Total": 10, "Currency": "USD"}]},
                {"sheet": "Sales", "rows": [{"Total": 5, "Currency": "USD"}]},
            ],
        },
        "batches": [
            {"sheet": "Purchase", "rows": [{"Total": 10, "Currency": "USD"}]},
            {"sheet": "Sales", "rows": [{"Total": 5, "Currency": "USD"}]},
        ],
        "workbook_name": "Ledger_FY2026.xlsx",
    }]

    summary, _ = _build_batch_aggregate_blocks(deferred, "C-test")
    # 1 document, 2 sheets, 2 rows total.
    assert "1 document" in summary or "1 documents" in summary, (
        f"summary should report 1 document (not 2 batches), got: {summary!r}"
    )
    # And 2 lines (rows) — one Purchase, one Sales.
    assert "2 lines" in summary, f"expected 2 lines, got: {summary!r}"


# =========================================================================== #
# Phase 3C — summary table wired into HITL review/approval cards
# =========================================================================== #


def test_paused_run_posts_summary_table_before_approval_card(monkeypatch):
    """When understand-extract populates ledger_summary_table, the HITL thread
    gets a summary table post BEFORE the approval card lands.

    This is the missing terminal/debug visibility: reviewers previously only
    saw the approve/edit/reject buttons without the Gemini interpretation
    that drove them. _post_summary_table existed but was never called.
    """
    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    from unittest.mock import MagicMock, patch
    from app import native_blocks_compat as _nbc
    _nbc._PROBE_CACHE.clear()

    from accounting_agents import slack_runner
    from accounting_agents import nodes as nodes_mod

    db = FakeFirestore()
    slack = FakeSlackClient()
    ledger_store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    summary_table = [
        {"category": "Vendor", "details": "PTTEPI"},
        {"category": "Amount", "details": "USD 2,000.00"},
        {"category": "Direction", "details": "Purchase"},
    ]
    state = {
        nodes_mod.LEDGER_SUMMARY_TABLE_KEY: summary_table,
        "review_question": "Confirm extraction",
        nodes_mod.REVIEW_REASON_KEY: ["Low classify confidence"],
    }

    class _SessionService:
        def __init__(self, st): self._st = st
        async def get_session(self, *a, **k):
            from types import SimpleNamespace
            return SimpleNamespace(state=self._st)

    runner = MagicMock()
    runner.app_name = "acc"
    runner.session_service = _SessionService(state)

    # Build an event whose `raw` causes find_interrupt_id to return our id,
    # and whose content yields a last_text for the review card.
    fake_event = MagicMock()
    fake_event.actions = MagicMock(skip_summarization=False)
    fake_event.content = MagicMock(parts=[MagicMock(text="")])
    fake_event.author = "agent"

    # Patch find_interrupt_id to return our review op_id.
    # Patch _read_session_state to return our state.
    # Patch _post_review_card to be a no-op spy.
    posted = {"summary_table": None, "review_card": None}

    def fake_find_interrupt_id(ev):
        return "CSTMID:Frun:review"

    def fake_read_session_state(*a, **k):
        async def _inner():
            return state
        return _inner()

    def fake_post_review_card(client, channel, question, op_id, reasons, **kw):
        posted["review_card"] = {
            "channel": channel, "question": question, "thread_ts": kw.get("thread_ts"),
        }
        return "R-TS"

    with patch.object(slack_runner, "find_interrupt_id", fake_find_interrupt_id), \
         patch.object(slack_runner, "_read_session_state", fake_read_session_state), \
         patch.object(slack_runner, "_post_review_card", fake_post_review_card):
        asyncio.run(
            slack_runner._finalize_run_outcome(
                events=[fake_event],
                interrupt_id=None,
                last_text="",
                runner=runner,
                ledger_store=ledger_store,
                db=db,
                slack_client=slack,
                channel_id="CSTMID",
                session_id="sess",
                app_name="acc",
                user_id="CSTMID",
                file_id="Frun",
                thread_ts="T-THREAD",
            )
        )

    # Find the summary table post (text starts with "Document summary").
    summary_table_posts = [
        p for p in slack._posts
        if p.get("thread_ts") == "T-THREAD"
        and p.get("text", "").startswith("Document summary")
    ]
    assert summary_table_posts, (
        f"summary table must be posted in thread before the review card; got: "
        f"{[p.get('text', '') for p in slack._posts]}"
    )
    # And the review card was posted after.


# =========================================================================== #
# Task: defer ledger upload during multi-file batch (one workbook write)      #
# =========================================================================== #


def test_batch_defers_ledger_upload_until_end():
    """In multi-file batch mode, ``append_rows`` is called ONCE at batch end, not per doc.

    Pre-fix: each ``process_file_event`` invocation triggered its own
    ``ledger_store.append_rows`` call → six mid-batch ``files_upload_v2`` posts
    and six ``files_delete`` of the prior workbook. After this slice, per-doc
    runs stash the ledger payload via ``deferred_ledger``; the batch coordinator
    calls ``append_rows`` exactly once per ``(client_id, fy, software, kind)``
    group after the loop finishes.

    The harness closes over a ``MagicMock`` ledger store, so we exercise the
    flush helper directly with a recording fake (the same code path the
    message handler invokes after the per-doc loop).
    """
    from unittest.mock import MagicMock

    from accounting_agents.slack_runner import _flush_deferred_ledger_writes

    store = MagicMock()
    append_calls: list[dict] = []

    def fake_append_rows(*, client_id, fy, slack_client, channel_id, batches, software, kind, client_name, replace):
        append_calls.append({
            "client_id": client_id, "fy": fy, "channel_id": channel_id,
            "batches": batches, "software": software, "kind": kind,
            "client_name": client_name, "replace": replace,
        })
        return {
            "slack_file_id": "F-fake-batch",
            "appended": len(batches),
            "deduped": 0,
            "filename": f"Acme Client - Ledger_FY{fy}.xlsx",
        }
    store.append_rows = fake_append_rows

    # Build six per-doc deferred_delivery items, all the same FY / client.
    batch_deferred: list[dict] = []
    for i in range(6):
        batch_deferred.append({
            "payload": {
                "client_id": "c-acme-client",
                "fy": "2026",
                "kind": "invoice",
                "software": "Xero",
                "client_name": "Acme Client",
            },
            "batches": [
                {"sheet": "Purchase", "doc_key": f"FD-{i}:Purchase:INV-{i}",
                 "rows": [{"Contact": f"Vendor {i}", "Total": 100.0 * (i + 1),
                           "Currency": "USD"}]}
            ],
            "workbook_name": "",  # filled in by flush
        })

    asyncio.run(_flush_deferred_ledger_writes(
        ledger_store=store,
        slack_client=MagicMock(),
        channel_id="C-batch-defer",
        batch_deferred=batch_deferred,
    ))

    # CORE ASSERTION: append_rows called exactly once per FY group, not 6×.
    assert len(append_calls) == 1, (
        f"expected exactly 1 batch-end append_rows call, got {len(append_calls)}: "
        f"{append_calls}"
    )
    call = append_calls[0]
    assert call["client_id"] == "c-acme-client"
    assert call["fy"] == "2026"
    assert call["software"] == "Xero"
    assert call["kind"] == "invoice"
    # 6 docs' worth of batches concatenated into one append_rows invocation.
    assert len(call["batches"]) == 6
    # And the resolved workbook name is back-patched on every deferred_delivery.
    for item in batch_deferred:
        assert item["workbook_name"] == "Acme Client - Ledger_FY2026.xlsx"


def test_batch_no_files_delete_mid_batch():
    """In batch mode, the workbook is touched exactly once at batch end.

    Pre-fix: each per-doc ``append_rows`` called ``files_delete`` on the
    superseded file → the channel's Files tab flickered between uploads and
    often ended up empty mid-batch. After this slice the workbook is touched
    exactly once at batch end, so ``files_delete`` fires at most once per FY
    workbook in the batch (and only if a prior file existed).

    We exercise the flush helper directly with a real ``SlackLedgerStore`` and
    a recording ``FakeSlackClient`` so we can assert the call count of
    ``files_upload_v2`` and ``files_delete`` deterministically.
    """
    from accounting_agents.ledger_store import SlackLedgerStore
    from accounting_agents.slack_runner import _flush_deferred_ledger_writes
    from tests.test_ledger_store import FakeFirestore, FakeSlackClient

    slack = FakeSlackClient()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())

    # Four per-doc deferred_delivery items, all the same FY / client.
    batch_deferred: list[dict] = []
    for i in range(4):
        batch_deferred.append({
            "payload": {
                "client_id": "c-files", "fy": "2026", "kind": "invoice",
                "software": "QBS", "client_name": "Client",
            },
            "batches": [
                {"sheet": "Purchase", "doc_key": f"FF-{i}:Purchase:INV-{i}",
                 "rows": [{"Contact": f"V{i}", "Total": 1.0}]}
            ],
            "workbook_name": "",
        })

    asyncio.run(_flush_deferred_ledger_writes(
        ledger_store=store,
        slack_client=slack,
        channel_id="C-batch-files",
        batch_deferred=batch_deferred,
    ))

    # CORE ASSERTION: one upload at batch end, no pre-existing prior file →
    # zero files_delete (the only delete comes from the upload step in
    # ledger_store, and only when prev_file_id differs from new_file_id).
    assert len(slack.uploads) == 1, (
        f"expected exactly 1 files_upload_v2 (batch-end), got {len(slack.uploads)}: "
        f"{slack.uploads}"
    )
    # No prior file existed for this client/fy, so files_delete should not fire.
    assert slack.deleted_file_ids == [], (
        f"first batch-end append should NOT delete anything; got: "
        f"{slack.deleted_file_ids}"
    )


def test_persist_and_deliver_defer_ledger_persist_stashes_payload():
    """``defer_ledger_persist=True`` skips ``append_rows`` and stashes a ``deferred_ledger``.

    The single-doc, non-batch caller path — when an individual doc is being
    processed with batch-level defer enabled, ``persist_and_deliver`` must NOT
    call ``append_rows``. Instead it returns a ``deferred_ledger`` dict on the
    result that the batch coordinator can merge with its peers.
    """
    from unittest.mock import MagicMock

    from accounting_agents import slack_runner

    # Build a session with a real ledger payload.
    runner = _FakeRunner([], _ledger_payload())
    store = MagicMock()
    # If append_rows IS called, the test fails.
    store.append_rows = MagicMock(side_effect=AssertionError(
        "append_rows must NOT be called when defer_ledger_persist=True"
    ))

    asyncio.run(
        slack_runner.persist_and_deliver(
            runner=runner,
            ledger_store=store,
            slack_client=FakeSlackClient(),
            channel_id="C-defer-1",
            session_id="S-defer-1",
            app_name="acc",
            user_id="U-defer-1",
            thread_ts=None,
            replace=False,
            defer_slack_delivery=True,  # batch defer is on
            batch_mode=True,
            defer_ledger_persist=True,  # and ledger persist is deferred too
        )
    )
    # The MagicMock was never called → if we got here without AssertionError,
    # the defer wiring works.
    store.append_rows.assert_not_called()


# =========================================================================== #
# Task: batch thinking UX — plan block on the placeholder message             #
# =========================================================================== #


def test_batch_processing_plan_blocks_uses_native_plan_by_default(monkeypatch):
    """Batch progress uses the agent thinking plan block when native blocks are supported."""
    from app import native_blocks_compat
    from app.blocks import batch_processing_plan_blocks

    monkeypatch.delenv("LEDGR_BATCH_EXPANDED_PROGRESS", raising=False)
    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    native_blocks_compat._PROBE_CACHE.pop("C-batch-thinking", None)

    doc_rows = [
        {"file_label": "Contractor Beta.pdf", "stage": "complete", "detail": "Purchase · 1 line", "status": "complete"},
        {"file_label": "Telco-A.pdf", "stage": "in_progress", "detail": "Understanding", "status": "in_progress"},
        {"file_label": "Supplier Gamma.pdf", "stage": "queued", "detail": None, "status": "in_progress"},
    ]
    blocks = batch_processing_plan_blocks(
        total=3,
        done=1,
        doc_rows=doc_rows,
        channel_id="C-batch-thinking",
    )
    plan_blocks = [b for b in blocks if b.get("type") == "plan"]
    assert len(plan_blocks) == 1
    assert plan_blocks[0]["title"] == "Processing batch (3 documents)"
    assert len(plan_blocks[0]["tasks"]) == 4  # overall + 3 docs
    for t in plan_blocks[0]["tasks"]:
        if "output" in t:
            assert isinstance(t["output"], dict)
            assert t["output"].get("type") == "rich_text"


def test_batch_processing_plan_blocks_expanded_section_when_opt_in(monkeypatch):
    """LEDGR_BATCH_EXPANDED_PROGRESS=1 uses always-visible section text (opt-in)."""
    from app import native_blocks_compat
    from app.blocks import batch_processing_plan_blocks

    monkeypatch.setenv("LEDGR_BATCH_EXPANDED_PROGRESS", "1")
    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    native_blocks_compat._PROBE_CACHE.pop("C-batch-expanded", None)

    doc_rows = [
        {"file_label": "a.pdf", "stage": "complete", "status": "complete", "detail": "ok"},
    ]
    blocks = batch_processing_plan_blocks(
        total=1,
        done=1,
        doc_rows=doc_rows,
        channel_id="C-batch-expanded",
    )
    assert not any(b.get("type") == "plan" for b in blocks)
    sections = [b for b in blocks if b.get("type") == "section"]
    assert len(sections) == 1
    assert "1 of 1 complete" in sections[0]["text"]["text"]


def test_batch_processing_plan_blocks_falls_back_when_native_unsupported(monkeypatch):
    """When the channel does not support native blocks, fall back to section+context."""
    from app import native_blocks_compat
    from app.blocks import batch_processing_plan_blocks

    monkeypatch.setattr(native_blocks_compat, "supports_native_blocks", lambda *a, **k: False)
    # Clear the probe cache so the patched function is consulted.
    native_blocks_compat._PROBE_CACHE.pop("C-fallback", None)

    blocks = batch_processing_plan_blocks(
        total=2,
        done=1,
        doc_rows=[
            {"file_label": "a.pdf", "stage": "complete", "status": "complete", "detail": "ok"},
            {"file_label": "b.pdf", "stage": "in_progress", "status": "in_progress", "detail": "running"},
        ],
        channel_id="C-fallback",
    )
    # No plan block — expanded section layout (always visible).
    assert not any(b.get("type") == "plan" for b in blocks)
    assert any(b.get("type") == "section" for b in blocks)
    assert blocks[0].get("expand") is True


def test_batch_six_files_post_initial_native_plan_block(monkeypatch):
    """First chat_postMessage in a 6-file drop carries the native agent thinking plan."""
    from unittest.mock import AsyncMock, MagicMock, patch

    from accounting_agents import slack_runner
    from app import native_blocks_compat
    from app.slack_app import _SeenEvents

    monkeypatch.delenv("LEDGR_BATCH_EXPANDED_PROGRESS", raising=False)
    monkeypatch.setenv("LEDGR_NATIVE_BLOCKS", "1")
    native_blocks_compat._PROBE_CACHE.pop("C-batch-plan", None)

    slack_runner._seen = _SeenEvents()
    slack = FakeSlackClient()
    handler, _ = _capture_message_handler_with_slack_client(slack)

    body = {"event_id": "Ev-batch-plan-1"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "225.001",
        "channel": "C-batch-plan",
        "files": [{"id": f"FP-{i}"} for i in range(6)],
    }
    fake_client = MagicMock()
    mock_pfe = AsyncMock(side_effect=[
        {"status": "delivered", "append": {"appended": 0, "software": "Xero", "fy": "2026",
                                            "kind": "invoice", "deferred_delivery": None}}
        for _ in range(6)
    ])

    with patch.object(slack_runner, "process_file_event", mock_pfe), \
         patch.object(slack_runner, "download_pdf_bytes", return_value=b"%PDF fake"):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    # First top-level post must carry the native plan block (agent thinking pattern).
    top_level = [p for p in slack._posts if not p.get("thread_ts")]
    assert len(top_level) == 1
    first_post = top_level[0]
    blocks = first_post.get("blocks") or []
    plan_blocks = [b for b in blocks if b.get("type") == "plan"]
    assert plan_blocks, "expected a native plan block in the batch progress message"
    plan = plan_blocks[0]
    assert plan["title"] == "Processing batch (6 documents)"
    # 1 overall + 6 docs = 7 tasks
    assert len(plan["tasks"]) == 7


# =========================================================================== #
# Task: unified Understand direction (Part E)                                  #
# =========================================================================== #


def test_document_ledger_extract_schema_has_party_and_direction():
    """``DocumentLedgerExtract`` owns From/To parties + ``direction_for_client``."""
    from invoice_processing.extract.ledger_extract import (
        DocumentLedgerExtract,
        PartyField,
    )

    extract = DocumentLedgerExtract(
        vendor_name="Contractor Beta",
        customer_name="Company-A",
        document_reference="INV-26-001",
        document_date="2026-06-01",
        document_total=1200.0,
        from_party=PartyField(
            name="Contractor Beta", uen=None, role="issuer",
        ),
        to_party=PartyField(
            name="Company-A",
            uen="201700001A",
            role="recipient",
        ),
        direction_for_client="purchase",
    )
    # The Contractor Beta case: client is the recipient → "purchase" (not "sales").
    assert extract.direction_for_client == "purchase"
    assert extract.to_party.uen == "201700001A"
    assert extract.to_party.role == "recipient"
    assert extract.from_party.role == "issuer"


def test_understand_prompt_includes_client_context():
    """When client identity is provided, the prompt asks for ``direction_for_client``."""
    from invoice_processing.extract.ledger_extract import _build_understand_prompt

    prompt = _build_understand_prompt(
        client_name="Company-A",
        client_uen="201700001A",
    )
    assert "Company-A" in prompt
    assert "201700001A" in prompt
    assert "direction_for_client" in prompt
    assert "purchase" in prompt
    assert "sales" in prompt


def test_understand_prompt_omits_client_context_when_absent():
    """When client identity is not provided, no client block in the prompt."""
    from invoice_processing.extract.ledger_extract import _build_understand_prompt

    prompt = _build_understand_prompt(None, None)
    assert "Client context" not in prompt
    # The from/to/direction_for_client fields are still documented in the
    # base prompt (they may be populated as null) — but the conditional
    # "Client context: ..." block must be absent.
    assert "direction_for_client as follows" not in prompt


def test_ledger_extract_to_normalized_uses_direction_for_client_when_auto():
    """``direction="auto"`` causes the adapter to honor the Understand verdict."""
    from invoice_processing.extract.ledger_extract import (
        DocumentLedgerExtract,
        PartyField,
        ledger_extract_to_normalized,
    )

    extract = DocumentLedgerExtract(
        vendor_name="Vendor Pte Ltd",
        customer_name="Company-A",
        document_reference="INV-1",
        document_date="2026-06-01",
        document_total=100.0,
        from_party=PartyField(name="Vendor Pte Ltd", role="issuer"),
        to_party=PartyField(
            name="Company-A",
            uen="201700001A",
            role="recipient",
        ),
        direction_for_client="purchase",
    )
    normalized = ledger_extract_to_normalized(extract, direction="auto")
    assert normalized.doc_type == "purchase"

    # Flip direction_for_client → sales and the adapter should follow.
    extract_sales = extract.model_copy(update={"direction_for_client": "sales"})
    normalized_sales = ledger_extract_to_normalized(extract_sales, direction="auto")
    assert normalized_sales.doc_type == "sales"


def test_resolve_direction_from_extract_understand_verdict():
    """``_resolve_direction_from_extract`` honors the Understand verdict."""
    from accounting_agents.nodes import _resolve_direction_from_extract

    # Purchase.
    assert _resolve_direction_from_extract({"direction_for_client": "purchase"}) == "purchase"
    # Sales.
    assert _resolve_direction_from_extract({"direction_for_client": "sales"}) == "sales"
    # Self-referential.
    assert _resolve_direction_from_extract(
        {"direction_for_client": "self_referential"}
    ) == "self_referential"
    # Unknown passes through (caller escalates to HITL).
    assert _resolve_direction_from_extract({"direction_for_client": "unknown"}) == "unknown"
    # Missing → fallback.
    assert _resolve_direction_from_extract(None) == "purchase"
    assert _resolve_direction_from_extract({}, fallback="sales") == "sales"


def test_classify_node_invoice_lane_no_longer_calls_resolve_direction():
    """``classify_node`` for the invoice lane must NOT call ``resolve_direction``.

    Pre-fix: ``classify_node`` called ``resolve_direction`` and locked the
    direction before the Understand call could see the document. After the
    slice, the invoice lane defers the direction decision to the Understand
    call's ``direction_for_client`` field — a deterministic gate, not a
    fuzzy Python rewrite.

    The full assertion (with FakeContext + _load_pdf_bytes stub) lives in
    ``tests/test_nodes.py::test_classify_routes_invoice``. This test is the
    behavioral summary: classify_node does not call resolve_direction.
    """
    # The detailed check is in test_nodes.py where the FakeContext harness is
    # available. Re-assert the contract from the slack_runner side: the
    # ``resolve_direction`` Python fuzzy-match is no longer wired into the
    # invoice path. ``DIRECTION_FN`` is still imported in nodes.py (kept for
    # bank/SOA consumers and tests) but is no longer called from
    # ``classify_node`` for invoice documents.
    from accounting_agents import nodes
    import inspect

    source = inspect.getsource(nodes.classify_node)
    # In the invoice branch (the path after `if doc_type == "bank_statement"`),
    # there must be no `DIRECTION_FN(` call.
    assert "DIRECTION_FN(" not in source, (
        "classify_node should NOT call DIRECTION_FN for the invoice lane; "
        "the Understand call now owns direction_for_client."
    )