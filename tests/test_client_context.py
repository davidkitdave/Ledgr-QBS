"""Hermetic tests for ledgr_slack.client_context.

No external files, no Firestore, no GCP calls. Builds in-memory .xlsx workbooks
with openpyxl in tmp_path so every test is fully self-contained.
"""

from __future__ import annotations

from openpyxl import Workbook

from ledgr_slack.client_context import (
    client_context_from_state,
    load_client_setup,
    InMemoryClientStore,
)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _make_workbook(tmp_path, name="TestClient Setup.xlsx", include_sys_config=False):
    """Build a minimal in-memory xlsx with Category_Mapping + Entity_Memory."""
    wb = Workbook()
    # openpyxl creates a default "Sheet" — remove it
    default = wb.active
    wb.remove(default)

    if include_sys_config:
        ws_sc = wb.create_sheet("Sys_Config")
        ws_sc.append(["Key", "Value"])
        ws_sc.append(["CLIENT_ID", "from-sheet"])
        ws_sc.append(["REGION", "MALAYSIA"])

    # Category_Mapping — one enabled, one disabled
    ws_cm = wb.create_sheet("Category_Mapping")
    ws_cm.append(["Category", "Account Code", "Enabled", "Notes"])
    ws_cm.append(["Office Supplies", "6100", True, ""])
    ws_cm.append(["Travel", "6200", False, "disabled row"])

    # Entity_Memory
    ws_em = wb.create_sheet("Entity_Memory")
    ws_em.append([
        "Name", "Reg No / Tax ID", "Mapping Code",
        "Role (Debtor / Creditor)", "Tax Code",
    ])
    ws_em.append(["Acme Corp", "201234567A", "4000", "Creditor", "SR"])

    path = tmp_path / name
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# load_client_setup — no Sys_Config sheet
# --------------------------------------------------------------------------- #

class TestLoadClientSetupNoSysConfig:
    def test_only_enabled_category_mapping_rows_kept(self, tmp_path):
        path = _make_workbook(tmp_path)
        ctx = load_client_setup(path, client_id="test-client-1")

        assert "Office Supplies" in ctx.category_mapping
        assert ctx.category_mapping["Office Supplies"] == "6100"
        assert "Travel" not in ctx.category_mapping  # disabled row excluded

    def test_entity_memory_parses(self, tmp_path):
        path = _make_workbook(tmp_path)
        ctx = load_client_setup(path, client_id="test-client-1")

        assert len(ctx.entity_memory) == 1
        em = ctx.entity_memory[0]
        assert em.name == "Acme Corp"
        assert em.reg_no == "201234567A"
        assert em.mapping_code == "4000"
        assert em.role == "Creditor"
        assert em.tax_code == "SR"

    def test_profile_fields_are_empty_defaults(self, tmp_path):
        path = _make_workbook(tmp_path)
        ctx = load_client_setup(path, client_id="test-client-1")

        assert ctx.region == ""
        assert ctx.accounting_software == "QBS Ledger"
        assert ctx.base_currency == ""
        assert ctx.tax_registered is None

    def test_client_id_from_param(self, tmp_path):
        path = _make_workbook(tmp_path)
        ctx = load_client_setup(path, client_id="explicit-id-42")
        assert ctx.client_id == "explicit-id-42"

    def test_fye_month_defaults_to_none(self, tmp_path):
        path = _make_workbook(tmp_path)
        ctx = load_client_setup(path, client_id="test-client-1")
        assert ctx.fye_month is None


# --------------------------------------------------------------------------- #
# to_state / client_context_from_state round-trip
# --------------------------------------------------------------------------- #

class TestStateRoundTrip:
    def test_to_state_includes_fye_month_key(self, tmp_path):
        path = _make_workbook(tmp_path)
        ctx = load_client_setup(path, client_id="c1")
        state = ctx.to_state()
        assert "fye_month" in state
        assert state["fye_month"] is None

    def test_round_trip_preserves_fye_month(self, tmp_path):
        path = _make_workbook(tmp_path)
        ctx = load_client_setup(path, client_id="c1")
        ctx.fye_month = 3  # simulate onboarding-set value

        state = ctx.to_state()
        assert state["fye_month"] == 3

        ctx2 = client_context_from_state(state)
        assert ctx2.fye_month == 3

    def test_round_trip_preserves_category_mapping(self, tmp_path):
        path = _make_workbook(tmp_path)
        ctx = load_client_setup(path, client_id="c1")
        ctx2 = client_context_from_state(ctx.to_state())

        assert "Office Supplies" in ctx2.category_mapping

    def test_round_trip_preserves_entity_memory(self, tmp_path):
        path = _make_workbook(tmp_path)
        ctx = load_client_setup(path, client_id="c1")
        ctx2 = client_context_from_state(ctx.to_state())

        assert len(ctx2.entity_memory) == 1
        assert ctx2.entity_memory[0].name == "Acme Corp"


# --------------------------------------------------------------------------- #
# InMemoryClientStore.from_setup_dir — client_id from parent folder name
# --------------------------------------------------------------------------- #

class TestInMemoryClientStoreFromSetupDir:
    def test_keyed_by_parent_folder_name(self, tmp_path):
        client_dir = tmp_path / "acme-client"
        client_dir.mkdir()
        _make_workbook(client_dir, name="Acme Client Setup.xlsx")

        store = InMemoryClientStore.from_setup_dir(tmp_path)
        ctx = store.get("acme-client")

        assert ctx is not None
        assert ctx.client_id == "acme-client"
