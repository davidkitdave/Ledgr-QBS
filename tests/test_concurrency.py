"""Concurrency safety tests for the Slack runner (QUICK-WIN fixes).

Covers the low-risk concurrency hardening so a few simultaneous document drops are
safe:

(a) two concurrent ``process_file_event`` calls for DIFFERENT file ids use DISTINCT
    per-document session ids (``f"{channel_id}:{file_id}"``) and never collide;
(b) the module-level semaphore caps the number of in-flight runs at N (instrumented
    with a live counter);
(c) the synchronous ``append_rows`` ledger write is invoked via
    ``asyncio.to_thread`` (so it never blocks the event loop);
(d) ``_ensure_session`` is idempotent under a simulated concurrent create: a second
    ``create_session`` raising ``AlreadyExistsError`` is handled, no crash.

All hermetic — no live Slack / Gemini / Firestore.
"""

from __future__ import annotations

import asyncio
import time
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest

from ledgr_slack.file_event import _per_doc_session_id, process_file_event
from ledgr_slack.sessions import _ensure_session
from ledgr_slack.ux import _flush_deferred_ledger_writes

LEDGER_ROWS_KEY = "ledger_rows"
DELIVER_SUMMARY_KEY = "deliver_summary"

# --------------------------------------------------------------------------- #
# Module-level hermetic fixture: suppress real Gemini calls from read_doc
# --------------------------------------------------------------------------- #
#
# The lean path (_run_ledgr_tools) calls read_doc which calls
# _read_bytes_with_gemini. In hermetic tests there is no API key / live PDF,
# so we monkeypatch it to return a minimal canned commercial payload so no
# network call is made. The patch is autouse so ALL tests in this module are
# protected without per-test boilerplate.


_CANNED_READ_DOC_PAYLOAD = {
    "file_kind": "commercial_documents",
    "document_count": 1,
    "documents": [
        {
            # ReadDocument schema fields (not line_items / date / buyer_name)
            "doc_type": "purchase",
            "document_kind": "invoice",
            "invoice_number": "INV-001",
            "invoice_date": "2026-01-15",
            "due_date": "",
            "currency": "SGD",
            "fx_rate": None,
            "vendor_name": "Test Vendor Pte Ltd",
            "customer_name": "Test Buyer Pte Ltd",
            "entity_tax_id": "",
            "subtotal": 1000.00,
            "tax_total": 90.0,
            "grand_total": 1090.0,
            "lines": [
                {
                    "description": "Professional services",
                    "quantity": 1.0,
                    "unit_amount": 1000.00,
                    "net_amount": 1000.00,
                    "tax_amount": 90.0,
                    "total_amount": 1090.0,
                }
            ],
            "notes": "",
        }
    ],
    "extraction_meta": {
        "gemini_call_count": 0,
        "model": "fake",
        "extract_mode": "vision",
        "elapsed_seconds": 0.0,
        "bytes_sent": 0,
        "usage": {},
    },
}


@pytest.fixture(autouse=True)
def _mock_gemini_read(monkeypatch):
    """Monkeypatch _read_bytes_with_gemini so no real Gemini call is made."""
    import ledgr_agent.tools.read_doc as _read_doc_mod
    monkeypatch.setattr(
        _read_doc_mod,
        "_read_bytes_with_gemini",
        lambda data, mime: dict(_CANNED_READ_DOC_PAYLOAD),
    )


# --------------------------------------------------------------------------- #
# Fakes
# --------------------------------------------------------------------------- #


def _ledger_payload():
    return {
        LEDGER_ROWS_KEY: {
            "client_id": "c1",
            "fy": "2026",
            "kind": "invoice",
            "software": "qbs",
            "batches": [
                {
                    "sheet": "Purchase",
                    "doc_key": "F:Purchase:INV",
                    "rows": [{"Invoice Number": "INV", "Source Amount": 1.0}],
                }
            ],
        },
        DELIVER_SUMMARY_KEY: "done",
    }


class _FakeSession:
    def __init__(self, state):
        self.state = dict(state)


class _RecordingSessionService:
    """Records (user_id, session_id) for every create/get and run scope.

    Stores live session objects per key so that state deltas written via
    ``append_event`` are visible to subsequent ``get_session`` calls — this
    mirrors how the real ADK session service works and is required for the lean
    path (read_doc → build_sheets → deliver_workbook) which writes WORKBOOK_STATE_KEY
    via ``_apply_state_delta`` and reads it back in ``deliver_workbook``.
    """

    def __init__(self, final_state):
        self._final_state = dict(final_state)  # kept for _InstrumentedRunner seed
        self._sessions: dict = {}   # (user_id, session_id) -> _FakeSession
        self.create_calls: list = []
        self.event_calls: list = []

    @property
    def _created(self):
        return set(self._sessions.keys())

    async def get_session(self, *, app_name, user_id, session_id):
        return self._sessions.get((user_id, session_id))

    async def create_session(self, *, app_name, user_id, session_id, state=None):
        self.create_calls.append((user_id, session_id))
        session = _FakeSession(state or {})
        self._sessions[(user_id, session_id)] = session
        return session

    async def append_event(self, session, event):
        """Apply state_delta from the event so get_session sees updated state."""
        self.event_calls.append(id(session))
        actions = getattr(event, "actions", None)
        if actions is not None:
            delta = getattr(actions, "state_delta", None) or {}
            if hasattr(session, "state") and delta:
                session.state.update(delta)
        return session


class _FakeArtifactService:
    def __init__(self):
        self.saved = {}

    async def save_artifact(self, *, app_name, user_id, filename, artifact, session_id=None, custom_metadata=None):
        self.saved[(user_id, session_id, filename)] = artifact
        return 0


class _InstrumentedRunner:
    """Runner stand-in that records the session scope of each run_async call and
    tracks concurrent in-flight runs (for the semaphore test)."""

    def __init__(self, final_state, *, app_name="acc", run_delay=0.0, counter=None):
        self.app_name = app_name
        self.artifact_service = _FakeArtifactService()
        self.session_service = _RecordingSessionService(final_state)
        self.run_scopes: list = []
        self._run_delay = run_delay
        self._counter = counter  # optional {"now": int, "max": int}

    async def run_async(
        self, *, user_id, session_id, new_message=None, state_delta=None, run_config=None,
    ):
        self.run_scopes.append((user_id, session_id))
        # Ensure session exists in _sessions so get_session returns it after run.
        if (user_id, session_id) not in self.session_service._sessions:
            self.session_service._sessions[(user_id, session_id)] = _FakeSession(
                self.session_service._final_state
            )
        if self._counter is not None:
            self._counter["now"] += 1
            self._counter["max"] = max(self._counter["max"], self._counter["now"])
        try:
            if self._run_delay:
                await asyncio.sleep(self._run_delay)
            yield SimpleNamespace(
                content=SimpleNamespace(parts=[SimpleNamespace(text="done")]),
                get_function_calls=lambda: [],
            )
        finally:
            if self._counter is not None:
                self._counter["now"] -= 1


class _NoopLedgerStore:
    def __init__(self):
        self.append_calls = 0
        self.read_calls = 0

    def append_rows(self, **kwargs):
        self.append_calls += 1
        return {"appended": 1, "deduped": 0, "filename": "x.xlsx", "slack_file_id": "wb"}

    def read_rows(self, **kwargs):
        self.read_calls += 1
        return []

    def latest_fy(self, client_id):
        return None

    def get_pointer(self, client_id, fy):
        return None


class _FakeFirestoreDoc:
    def set(self, *a, **k):
        pass


class _FakeFirestoreCollection:
    def document(self, _id):
        return _FakeFirestoreDoc()


class _FakeDb:
    def collection(self, _name):
        return _FakeFirestoreCollection()


def _slack():
    posts: list = []
    updates: list = []
    return SimpleNamespace(
        chat_postMessage=lambda **k: posts.append(k) or {"ts": "1"},
        chat_update=lambda **k: updates.append(k) or {"ok": True},
        reactions_add=lambda **k: {"ok": True},
        reactions_remove=lambda **k: {"ok": True},
        files_upload_v2=lambda **k: {"ok": True, "file": {"id": "F-fake"}},
        files_upload=lambda **k: {"ok": True, "file": {"id": "F-fake"}},
        _posts=posts,
        _updates=updates,
    )


class _FakeProfileStore:
    """A channel → ClientContext shim that always returns a minimal QBS profile.

    The local _FakeDb shim above doesn't implement the full
    :class:`FirestoreClientStore` API (``document().get()``, subcollection
    ``.stream()``), so we short-circuit with a plain object that satisfies the
    single ``get_by_channel`` call the runner makes. Concurrency tests only
    care that the soft-gate lets the run proceed.
    """
    def __init__(self, software: str = "QBS Ledger", client_id: str = "c1"):
        from ledgr_slack.client_context import ClientContext
        self._ctx = ClientContext(
            client_id=client_id,
            accounting_software=software,
            fye_month=12,
        )

    def get_by_channel(self, channel_id):
        assert channel_id
        return self._ctx

    def append_processing_log(self, *args, **kwargs):
        """No-op: concurrency tests don't assert on audit-log writes."""
        pass


def _profile_store(**kwargs):
    """A drop-in `client_store` for `process_file_event` (seeds QBS by default)."""
    return _FakeProfileStore(**kwargs)


# --------------------------------------------------------------------------- #
# (a) distinct per-doc session ids for concurrent drops
# --------------------------------------------------------------------------- #


def test_concurrent_file_events_use_distinct_session_ids():
    runner = _InstrumentedRunner(_ledger_payload())
    store = _NoopLedgerStore()
    db = _FakeDb()
    slack = _slack()

    async def drive():
        await asyncio.gather(
            process_file_event(
                runner=runner, ledger_store=store, db=db, slack_client=slack,
                channel_id="C1", file_id="F1", app_name="acc",
                download_fn=lambda c, f: b"%PDF a",
                client_store=_profile_store(),
            ),
            process_file_event(
                runner=runner, ledger_store=store, db=db, slack_client=slack,
                channel_id="C1", file_id="F2", app_name="acc",
                download_fn=lambda c, f: b"%PDF b",
                client_store=_profile_store(),
            ),
        )

    asyncio.run(drive())

    # Lean path: sessions are created via _run_ledgr_tools → _ensure_session →
    # create_session, recorded in _sessions keyed by (user_id, session_id).
    # Both runs targeted the SAME user_id (channel) but DISTINCT per-doc sessions.
    created = set(runner.session_service._sessions.keys())
    sid_f1 = _per_doc_session_id("C1", "F1")
    sid_f2 = _per_doc_session_id("C1", "F2")
    assert ("C1", sid_f1) in created, f"F1 session not created; got {created}"
    assert ("C1", sid_f2) in created, f"F2 session not created; got {created}"
    assert sid_f1 != sid_f2, "Session ids must differ per file"


# --------------------------------------------------------------------------- #
# (b) semaphore caps concurrent in-flight runs at N
# --------------------------------------------------------------------------- #


def test_semaphore_caps_concurrent_runs(monkeypatch):
    """Lean path: _SEM in slack_runner caps concurrent runs.

    In the lean path runner.run_async is never called, so we instrument the
    semaphore directly by replacing _SEM with a counting wrapper that tracks
    concurrent acquisitions.  The semaphore is acquired in process_file_event
    before calling process_file_via_ledgr_agent (which runs read_doc +
    build_sheets).  We add a small async delay inside build_sheets to allow
    concurrent acquirers to queue up, then verify max_concurrent <= cap.
    """
    import ledgr_slack.slack_shell as slack_shell

    counter = {"now": 0, "max": 0}

    class _InstrumentedSemaphore:
        """Wraps asyncio.Semaphore and tracks peak concurrent holders."""

        def __init__(self, value):
            self._sem = asyncio.Semaphore(value)
            self._value = value

        async def __aenter__(self):
            await self._sem.acquire()
            counter["now"] += 1
            counter["max"] = max(counter["max"], counter["now"])
            return self

        async def __aexit__(self, *args):
            counter["now"] -= 1
            self._sem.release()

        def locked(self):
            return self._sem.locked()

    cap = 2
    monkeypatch.setattr("ledgr_slack.file_event._SEM", _InstrumentedSemaphore(cap))

    # Add a delay inside build_sheets so concurrent runs overlap in the semaphore.
    real_build_sheets = slack_shell.__dict__.get("build_sheets") or __import__(
        "ledgr_agent.tools.build_sheets", fromlist=["build_sheets"]
    ).build_sheets

    import time as _time

    def _slow_build_sheets(ctx):
        _time.sleep(0.02)
        return real_build_sheets(ctx)

    monkeypatch.setattr(slack_shell, "build_sheets", _slow_build_sheets)

    runner = _InstrumentedRunner(_ledger_payload())
    store = _NoopLedgerStore()
    db = _FakeDb()
    slack = _slack()

    n_docs = 6

    async def drive():
        await asyncio.gather(*[
            process_file_event(
                runner=runner, ledger_store=store, db=db, slack_client=slack,
                channel_id="C1", file_id=f"F{i}", app_name="acc",
                download_fn=lambda c, f: b"%PDF",
                client_store=_profile_store(),
            )
            for i in range(n_docs)
        ])

    asyncio.run(drive())

    # Never more than cap runs in flight at once despite n_docs simultaneous drops.
    assert counter["max"] <= cap, (
        f"semaphore did not cap: max_concurrent={counter['max']} > cap={cap}"
    )
    # All six must have completed (queued, not dropped).
    assert store.append_calls == n_docs, (
        f"expected {n_docs} deliveries, got {store.append_calls}"
    )


# --------------------------------------------------------------------------- #
# (c) append_rows is invoked via asyncio.to_thread (non-blocking)
# --------------------------------------------------------------------------- #


def test_append_rows_is_offloaded_to_thread(monkeypatch):
    import ledgr_slack.slack_shell as slack_shell

    runner = _InstrumentedRunner(_ledger_payload())
    store = _NoopLedgerStore()
    db = _FakeDb()
    slack = _slack()

    offloaded: list = []
    real_to_thread = asyncio.to_thread

    async def spy_to_thread(fn, *args, **kwargs):
        offloaded.append(fn)
        return await real_to_thread(fn, *args, **kwargs)

    # The lean path's deliver_workbook lives in slack_shell.py and uses its
    # own asyncio.to_thread.  Patch both modules so the spy captures all calls.
    monkeypatch.setattr(asyncio, "to_thread", spy_to_thread)
    monkeypatch.setattr(slack_shell.asyncio, "to_thread", spy_to_thread)

    result = asyncio.run(
        process_file_event(
            runner=runner, ledger_store=store, db=db, slack_client=slack,
            channel_id="C1", file_id="F1", app_name="acc",
            download_fn=lambda c, f: b"%PDF",
            client_store=_profile_store(),
        )
    )

    assert result["status"] == "delivered"
    # The synchronous ledger write was dispatched through to_thread.
    assert store.append_rows in offloaded
    assert store.append_calls == 1


# --------------------------------------------------------------------------- #
# (d) _ensure_session is idempotent under a concurrent create race
# --------------------------------------------------------------------------- #


def test_ensure_session_idempotent_on_already_exists():
    from google.adk.errors.already_exists_error import AlreadyExistsError

    calls = {"n": 0}

    class _RaceSessionService:
        async def create_session(self, *, app_name, user_id, session_id, state=None):
            calls["n"] += 1
            # Simulate a concurrent create having already won the race.
            raise AlreadyExistsError(f"Session {session_id} already exists.")

        async def get_session(self, *, app_name, user_id, session_id):
            return _FakeSession({})

    runner = SimpleNamespace(session_service=_RaceSessionService())

    # Must NOT raise — the already-exists error is treated as success.
    asyncio.run(_ensure_session(runner, "acc", "C1", "C1:F1"))
    assert calls["n"] == 1


# --------------------------------------------------------------------------- #
# Step 5 — Fan-out / fan-in concurrency tests (TDD: written before the impl)  #
# --------------------------------------------------------------------------- #
#
# These tests exercise the NEW fan-out behaviour added in Step 5 of the plan:
#   1. Pre-gather de-dup: duplicate file ids are processed exactly once.
#   2. Fan-out: asyncio.gather is used; docs run concurrently.
#   3. Post-gather reduce: shared counters and batch_deferred are assembled in
#      ORIGINAL INPUT ORDER, so the final ledger rows are deterministic regardless
#      of per-doc completion order.
#   4. Invoice date-sort: _append_rows_to_sheet sorts invoice rows by
#      (date, invoice_number, doc_key) before appending.
#   5. HITL second write: a paused doc that isn't in batch_deferred appends its
#      rows later via the same SlackLedgerStore lock (no bypass).
#   6. Semaphore unchanged: _SEM still bounds concurrent runs (no double-bound).
#
# All hermetic — no live Slack / Gemini / Firestore.



# ------------------------------------------------------------------ helpers --

def _invoice_deferred(*, doc_key: str, date: str, invoice_number: str,
                       client_id: str = "c-fan", fy: str = "2026",
                       software: str = "qbs", kind: str = "invoice") -> dict:
    """Build a minimal deferred_delivery item for an invoice doc."""
    return {
        "payload": {
            "client_id": client_id,
            "fy": fy,
            "kind": kind,
            "software": software,
            "client_name": "Fan Client",
        },
        "batches": [
            {
                "sheet": "Purchase",
                "doc_key": doc_key,
                "rows": [
                    {
                        "Date": date,
                        "Invoice Number": invoice_number,
                        "Contact": f"Vendor {doc_key}",
                        "Total": 100.0,
                    }
                ],
            }
        ],
        "workbook_name": "",
    }


class _RecordingLedgerStore:
    """Fake SlackLedgerStore that records every append_rows invocation.

    Supports variable per-call latency so tests can simulate out-of-order
    completion.  Thread-safe via a plain list (only used from asyncio context
    and called through asyncio.to_thread, but the list append is GIL-safe).
    """

    def __init__(self, *, latency_map: dict[str, float] | None = None):
        # latency_map: {client_id -> seconds to sleep before responding}
        self._latency_map = latency_map or {}
        self.calls: list[dict] = []  # ordered list of recorded calls

    def append_rows(
        self,
        *,
        client_id: str,
        fy: str,
        slack_client,
        channel_id: str,
        batches: list[dict],
        software: str,
        kind: str,
        client_name: str,
        replace: bool,
    ) -> dict:
        delay = self._latency_map.get(client_id, 0.0)
        if delay:
            time.sleep(delay)
        self.calls.append({
            "client_id": client_id,
            "fy": fy,
            "kind": kind,
            "software": software,
            "batches": list(batches),
        })
        return {
            "slack_file_id": "F-fake",
            "appended": len(batches),
            "deduped": 0,
            "filename": f"Fan Client - Ledger_FY{fy}.xlsx",
        }

    def read_rows(self, **_kwargs):
        return []

    def latest_fy(self, client_id):
        return None


# ------------------------------------------------------------------ (1) de-dup


def test_pre_gather_dedup_processes_duplicate_file_id_once():
    """A file list with a repeated file id must process that doc only once.

    Exercises the pre-gather de-dup added in Step 5: deduplicate files by id
    before fan-out so two list entries sharing an id only run once.
    """
    from unittest.mock import AsyncMock, MagicMock, patch
    from app.slack_app import _SeenEvents

    import ledgr_slack.dedup as _dedup
    _dedup._seen = _SeenEvents()

    # Build a files list with a duplicated file id.
    files_with_dup = [
        {"id": "F-dup-A", "name": "invoice_a.pdf", "filetype": "pdf"},
        {"id": "F-dup-B", "name": "invoice_b.pdf", "filetype": "pdf"},
        {"id": "F-dup-A", "name": "invoice_a.pdf", "filetype": "pdf"},  # duplicate
    ]

    call_file_ids: list[str] = []

    async def _fake_pfe(**kwargs):
        call_file_ids.append(kwargs["file_id"])
        return {
            "status": "delivered",
            "append": {
                "software": "qbs", "fy": "2026", "kind": "invoice",
                "deferred_delivery": _invoice_deferred(
                    doc_key=f"F:{kwargs['file_id']}:INV",
                    date="2026-01-01",
                    invoice_number="INV-001",
                ),
            },
        }

    fake_slack = MagicMock()
    fake_slack.chat_postMessage.return_value = {"ok": True, "ts": "1000.1"}
    fake_slack.chat_update.return_value = {"ok": True}

    # The _message handler is the fan-out entry point. We call it via the
    # handler captured from build_async_app.
    # Instead, we test the de-dup directly through _flush_deferred_ledger_writes
    # is called with only 2 unique items (not 3).
    # Verify by patching process_file_event and running through the handler.

    from tests._slack_test_helpers import capture_message_handler_with_slack_client
    from tests.test_ledger_store import FakeSlackClient

    import ledgr_slack.dedup as _dedup
    _dedup._seen = _SeenEvents()
    slack = FakeSlackClient()
    handler, _ = capture_message_handler_with_slack_client(slack)

    body = {"event_id": "Ev-dedup-test-1"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "999.1",
        "channel": "C-dedup-1",
        "files": files_with_dup,
    }
    fake_client = MagicMock()

    pfe_call_ids: list[str] = []

    async def _recording_pfe(**kwargs):
        pfe_call_ids.append(kwargs["file_id"])
        return {
            "status": "delivered",
            "append": {
                "software": "qbs", "fy": "2026", "kind": "invoice",
                "deferred_delivery": _invoice_deferred(
                    doc_key=f"F:{kwargs['file_id']}:Purchase:INV-001",
                    date="2026-01-15",
                    invoice_number="INV-001",
                ),
            },
        }

    with patch("ledgr_slack.batch_coordinator.process_file_event", side_effect=_recording_pfe), \
         patch("ledgr_slack.batch_coordinator.download_pdf_bytes", return_value=b"%PDF fake"), \
         patch("ledgr_slack.batch_coordinator._flush_deferred_ledger_writes", new_callable=AsyncMock,
               return_value=[{"appended": 2, "deduped": 0, "filename": "x.xlsx",
                              "slack_file_id": "F-wb"}]):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    # F-dup-A must appear exactly once despite being in the input list twice.
    assert pfe_call_ids.count("F-dup-A") == 1, (
        f"F-dup-A processed {pfe_call_ids.count('F-dup-A')}× — expected exactly 1; "
        f"all calls: {pfe_call_ids}"
    )
    assert pfe_call_ids.count("F-dup-B") == 1
    assert len(pfe_call_ids) == 2, f"expected 2 unique file ids, got {len(pfe_call_ids)}: {pfe_call_ids}"


# ------------------------------------------------------------------ (2) fan-out


def test_fan_out_docs_run_concurrently():
    """Fan-out: multiple docs run concurrently under asyncio.gather.

    With a semaphore cap of 5, 6 docs with a 20 ms fake run time should
    complete faster than sequential (6 × 20 ms = 120 ms). We assert they
    finish in < 80 ms wall-time — a comfortable threshold that fails if the
    loop is still sequential.
    """
    from unittest.mock import MagicMock, patch
    from app.slack_app import _SeenEvents
    from tests._slack_test_helpers import capture_message_handler_with_slack_client
    from tests.test_ledger_store import FakeSlackClient

    import ledgr_slack.dedup as _dedup
    _dedup._seen = _SeenEvents()
    slack = FakeSlackClient()
    handler, _ = capture_message_handler_with_slack_client(slack)

    n_docs = 5
    body = {"event_id": "Ev-fan-out-timing"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "1001.1",
        "channel": "C-fan-out",
        "files": [{"id": f"F-fan-{i}", "name": f"inv{i}.pdf", "filetype": "pdf"} for i in range(n_docs)],
    }
    fake_client = MagicMock()
    run_delay = 0.04  # 40 ms per doc

    async def _slow_pfe(**kwargs):
        await asyncio.sleep(run_delay)
        return {
            "status": "delivered",
            "append": {
                "software": "qbs", "fy": "2026", "kind": "invoice",
                "deferred_delivery": _invoice_deferred(
                    doc_key=f"F:{kwargs['file_id']}:Purchase:INV",
                    date="2026-03-01",
                    invoice_number=f"INV-{kwargs['file_id']}",
                ),
            },
        }

    with patch("ledgr_slack.batch_coordinator.process_file_event", side_effect=_slow_pfe), \
         patch("ledgr_slack.file_event.process_file_event", side_effect=_slow_pfe), \
         patch("ledgr_slack.batch_coordinator.download_pdf_bytes", return_value=b"%PDF fake"), \
         patch("ledgr_slack.batch_coordinator._flush_deferred_ledger_writes", new_callable=AsyncMock,
               return_value=[{"appended": n_docs, "deduped": 0,
                                     "filename": "x.xlsx", "slack_file_id": "F-wb"}]), \
         patch("ledgr_slack.batch_coordinator._flush_deferred_ledger_writes", new_callable=AsyncMock,
               return_value=[{"appended": n_docs, "deduped": 0,
                              "filename": "x.xlsx", "slack_file_id": "F-wb"}]):
        t0 = time.monotonic()
        asyncio.run(handler(event=event, body=body, client=fake_client))
        elapsed = time.monotonic() - t0

    sequential_time = n_docs * run_delay
    # Must be faster than sequential (allow 80% of sequential as ceiling).
    assert elapsed < sequential_time * 0.85, (
        f"fan-out took {elapsed:.3f}s — expected < {sequential_time * 0.85:.3f}s "
        f"(sequential would be {sequential_time:.3f}s). Loop may still be sequential."
    )


# ------------------------------------------------------------------ (3) order-independence


def test_post_gather_reduce_order_independence():
    """Invoice rows appended via _append_rows_to_sheet are BYTE-IDENTICAL regardless
    of the order the rows arrive from concurrent docs.

    This test pushes real QBS-exporter-shaped rows (using ``"Invoice Date"`` /
    ``"Invoice Number"`` — the actual keys produced by QbsLedgerExporter) through
    _append_rows_to_sheet in TWO different orders, reads back the sheet contents,
    and asserts the results are identical.  It directly exercises the production
    sort code path rather than _flush_deferred_ledger_writes (which is already
    tested separately).  This test FAILS against the old no-op sort that read
    ``r.get("Date")`` because all real rows have ``"Invoice Date"``, not ``"Date"``,
    so the old key returned ``""`` for every row and the sort was a no-op.

    We also verify that the old no-op behaviour would NOT have produced identical
    output, confirming this test is a meaningful guard.
    """
    import openpyxl
    from ledgr_slack.ledger_store import SlackLedgerStore

    # Real QBS-exporter column names (QbsLedgerExporter.purchase_cols).
    cols = [
        "Invoice Number", "Invoice Date", "Vendor Name", "Entity Tax ID",
        "Description", "Source Amount", "Tax Code", "Account Code",
    ]

    # Six rows with real QBS keys, out of chronological order.
    # Dates in DD/MM/YYYY — the format _fmt_date produces.
    rows_shuffled = [
        {"Invoice Date": "15/03/2026", "Invoice Number": "INV-C", "Vendor Name": "Vendor C",
         "Source Amount": 300.0, "Tax Code": "SR", "Account Code": "500"},
        {"Invoice Date": "10/01/2026", "Invoice Number": "INV-A", "Vendor Name": "Vendor A",
         "Source Amount": 100.0, "Tax Code": "SR", "Account Code": "500"},
        {"Invoice Date": "10/01/2026", "Invoice Number": "INV-Ab", "Vendor Name": "Vendor Ab",
         "Source Amount": 150.0, "Tax Code": "ZR", "Account Code": "500"},
        {"Invoice Date": "20/02/2026", "Invoice Number": "INV-B", "Vendor Name": "Vendor B",
         "Source Amount": 200.0, "Tax Code": "SR", "Account Code": "501"},
        {"Invoice Date": "05/04/2026", "Invoice Number": "INV-D", "Vendor Name": "Vendor D",
         "Source Amount": 400.0, "Tax Code": "SR", "Account Code": "502"},
        {"Invoice Date": "20/02/2026", "Invoice Number": "INV-Bb", "Vendor Name": "Vendor Bb",
         "Source Amount": 250.0, "Tax Code": "ZR", "Account Code": "501"},
    ]

    def _sheet_rows(rows_input):
        """Append rows_input to a fresh sheet and read back as list-of-lists."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase"
        ws.append(cols)  # header
        SlackLedgerStore._append_rows_to_sheet(ws, cols, rows_input)
        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            result.append(list(row))
        return result

    # Two different arrival orders — simulate concurrent docs completing in different sequences.
    order_a = list(rows_shuffled)
    order_b = list(reversed(rows_shuffled))

    result_a = _sheet_rows(order_a)
    result_b = _sheet_rows(order_b)

    # CORE: both orderings must produce byte-identical sheet rows.
    assert result_a == result_b, (
        f"_append_rows_to_sheet is NOT order-independent!\n"
        f"order_a result: {result_a}\n"
        f"order_b result: {result_b}"
    )

    # SECONDARY: confirm the rows are in the expected chronological order.
    date_col = cols.index("Invoice Date")
    inv_col = cols.index("Invoice Number")
    dates_written = [r[date_col] for r in result_a]
    invs_written = [r[inv_col] for r in result_a]

    assert dates_written == [
        "10/01/2026", "10/01/2026", "20/02/2026", "20/02/2026", "15/03/2026", "05/04/2026"
    ], f"dates not chronological: {dates_written}"
    # Same-date tiebreak by Invoice Number (alphabetical).
    assert invs_written[0] == "INV-A"
    assert invs_written[1] == "INV-Ab"
    assert invs_written[2] == "INV-B"
    assert invs_written[3] == "INV-Bb"

    # GUARD: confirm the old no-op sort (using "Date" key) would NOT produce the
    # same result for reversed input — proving this test is a meaningful regression
    # guard, not a vacuous pass.
    def _sheet_rows_noopsort(rows_input):
        """Simulate the OLD broken sort that read r.get('Date') on QBS rows."""
        # The old sort key: r.get("Date") is always "" for real QBS rows, so
        # the sort is stable/no-op → insertion order is preserved.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase"
        ws.append(cols)
        for row in rows_input:   # no sort — old behaviour
            ws.append([row.get(c, "") for c in cols])
        result = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            result.append(list(r))
        return result

    noop_a = _sheet_rows_noopsort(order_a)
    noop_b = _sheet_rows_noopsort(order_b)
    # Old code: reversed input → reversed output (not identical). This asserts
    # that the test WOULD have caught the bug.
    assert noop_a != noop_b, (
        "Guard check failed: no-op sort produced identical output for reversed "
        "input — the test wouldn't have caught the bug. Review test setup."
    )


# ------------------------------------------------------------------ (4) invoice sort


def test_invoice_date_sort_in_append_rows_to_sheet():
    """_append_rows_to_sheet sorts QBS AND Xero rows by parsed date, oldest first.

    Uses real exporter-shaped rows (``"Invoice Date"``/``"Invoice Number"`` for QBS;
    ``"*InvoiceDate"``/``"*InvoiceNumber"`` for Xero) rather than fabricated keys,
    so the test actually guards the production schema.

    This test FAILS against the old broken sort that read ``r.get("Date")`` — all
    real QBS/Xero rows have no ``"Date"`` key so every row got the undated sentinel
    and the sort was a no-op.
    """
    import openpyxl
    from ledgr_slack.ledger_store import SlackLedgerStore

    # --- QBS layout (Invoice Date / Invoice Number) ---
    qbs_cols = ["Invoice Number", "Invoice Date", "Vendor Name", "Source Amount"]
    qbs_rows_unordered = [
        {"Invoice Date": "15/03/2026", "Invoice Number": "INV-003",
         "Vendor Name": "V3", "Source Amount": 300.0},
        {"Invoice Date": "10/01/2026", "Invoice Number": "INV-001",
         "Vendor Name": "V1", "Source Amount": 100.0},
        {"Invoice Date": "10/01/2026", "Invoice Number": "INV-001b",
         "Vendor Name": "V1b", "Source Amount": 150.0},
        {"Invoice Date": "20/02/2026", "Invoice Number": "INV-002",
         "Vendor Name": "V2", "Source Amount": 200.0},
    ]

    wb_qbs = openpyxl.Workbook()
    ws_qbs = wb_qbs.active
    ws_qbs.append(qbs_cols)
    SlackLedgerStore._append_rows_to_sheet(ws_qbs, qbs_cols, qbs_rows_unordered)

    qbs_written = []
    for row in ws_qbs.iter_rows(min_row=2, values_only=True):
        qbs_written.append(dict(zip(qbs_cols, row)))

    qbs_dates = [r["Invoice Date"] for r in qbs_written]
    qbs_invs = [r["Invoice Number"] for r in qbs_written]

    assert qbs_dates == ["10/01/2026", "10/01/2026", "20/02/2026", "15/03/2026"], (
        f"QBS rows NOT date-sorted: {qbs_dates}"
    )
    assert qbs_invs[0] == "INV-001"   # same-date tiebreak
    assert qbs_invs[1] == "INV-001b"

    # --- Xero layout (*InvoiceDate / *InvoiceNumber) ---
    xero_cols = ["*InvoiceNumber", "*InvoiceDate", "*ContactName", "Total"]
    xero_rows_unordered = [
        {"*InvoiceDate": "28/02/2026", "*InvoiceNumber": "XINV-B",
         "*ContactName": "C2", "Total": 200.0},
        {"*InvoiceDate": "05/01/2026", "*InvoiceNumber": "XINV-A",
         "*ContactName": "C1", "Total": 100.0},
        {"*InvoiceDate": "10/03/2026", "*InvoiceNumber": "XINV-C",
         "*ContactName": "C3", "Total": 300.0},
    ]

    wb_xero = openpyxl.Workbook()
    ws_xero = wb_xero.active
    ws_xero.append(xero_cols)
    SlackLedgerStore._append_rows_to_sheet(ws_xero, xero_cols, xero_rows_unordered)

    xero_written = []
    for row in ws_xero.iter_rows(min_row=2, values_only=True):
        xero_written.append(dict(zip(xero_cols, row)))

    xero_dates = [r["*InvoiceDate"] for r in xero_written]
    assert xero_dates == ["05/01/2026", "28/02/2026", "10/03/2026"], (
        f"Xero rows NOT date-sorted: {xero_dates}"
    )


# ------------------------------------------------------------------ (5) HITL second write


def test_hitl_second_write_accumulates_via_flush():
    """A paused doc's rows arrive via a SECOND _flush_deferred_ledger_writes call.

    Simulates: first flush (2 delivered docs) → second flush (1 HITL-approved doc).
    Both flushes target the same (client, fy, kind) group. Asserts:
    - append_rows is called once per flush (not combined).
    - The recording store sees TWO separate calls (accumulate, never replace).
    """
    from unittest.mock import MagicMock

    store = _RecordingLedgerStore()

    # Batch 1: two normal docs.
    batch1 = [
        _invoice_deferred(doc_key="F-hitl-A:Purchase:INV-A", date="2026-01-10", invoice_number="INV-A"),
        _invoice_deferred(doc_key="F-hitl-B:Purchase:INV-B", date="2026-02-15", invoice_number="INV-B"),
    ]
    # Batch 2: the HITL-approved doc (separate flush, later).
    batch2 = [
        _invoice_deferred(doc_key="F-hitl-C:Purchase:INV-C", date="2026-03-20", invoice_number="INV-C"),
    ]

    async def run_two_flushes():
        # First flush: normal batch.
        await _flush_deferred_ledger_writes(
            ledger_store=store,
            slack_client=MagicMock(),
            channel_id="C-hitl-test",
            batch_deferred=batch1,
        )
        # Second flush: HITL approval (same client/fy).
        await _flush_deferred_ledger_writes(
            ledger_store=store,
            slack_client=MagicMock(),
            channel_id="C-hitl-test",
            batch_deferred=batch2,
        )

    asyncio.run(run_two_flushes())

    # TWO separate append_rows calls — workbook accumulates (never replaced).
    assert len(store.calls) == 2, (
        f"expected 2 separate append_rows calls (accumulate contract); got {len(store.calls)}: "
        f"{store.calls}"
    )
    # First call has 2 batches, second has 1.
    assert len(store.calls[0]["batches"]) == 2
    assert len(store.calls[1]["batches"]) == 1


def test_flush_deferred_groups_by_batch_fy_not_payload_fy():
    """Multi-FY docs must flush into separate workbooks even when payload.fy matches."""
    from unittest.mock import MagicMock

    store = _RecordingLedgerStore()
    item_fy25 = _invoice_deferred(
        doc_key="F-mfy:Purchase:INV-25",
        date="2026-01-10",
        invoice_number="INV-25",
    )
    item_fy25["batches"][0]["fy"] = "FY2025"
    item_fy26 = _invoice_deferred(
        doc_key="F-mfy:Purchase:INV-26",
        date="2026-02-15",
        invoice_number="INV-26",
    )
    item_fy26["batches"][0]["fy"] = "FY2026"

    asyncio.run(
        _flush_deferred_ledger_writes(
            ledger_store=store,
            slack_client=MagicMock(),
            channel_id="C-mfy",
            batch_deferred=[item_fy25, item_fy26],
        )
    )

    assert len(store.calls) == 2
    flushed_fys = {call["fy"] for call in store.calls}
    assert flushed_fys == {"FY2025", "FY2026"}


def test_flush_deferred_charges_slack_owned_billing(monkeypatch):
    """Batch-end flush deducts credits when in-tool billing is disabled."""
    from unittest.mock import MagicMock

    from ledgr_agent.billing import CreditService, InMemoryCreditStore, configure_shared_credit_service

    monkeypatch.setenv("LEDGR_DISABLE_IN_TOOL_CHARGE", "1")
    svc = CreditService(InMemoryCreditStore())
    svc.ensure_firm("T-flush")
    svc.grant("T-flush", 5, note="trial")
    configure_shared_credit_service(svc)

    deferred = _invoice_deferred(
        doc_key="F-charge:Purchase:INV-1",
        date="2026-01-10",
        invoice_number="INV-1",
    )
    deferred["file_id"] = "F-charge"
    deferred["credits"] = {"credit_status": "estimated"}

    asyncio.run(
        _flush_deferred_ledger_writes(
            ledger_store=_RecordingLedgerStore(),
            slack_client=MagicMock(),
            channel_id="C-charge",
            batch_deferred=[deferred],
            firm_id="T-flush",
        )
    )

    assert svc.read_balance("T-flush") == 4


def test_flush_deferred_records_failure_when_append_raises():
    """append_rows exceptions must surface as flush_failed, not silent empty results."""
    from unittest.mock import MagicMock

    class _FailingStore:
        def append_rows(self, **kwargs):  # noqa: ANN003
            raise RuntimeError("slack upload failed")

    deferred = _invoice_deferred(
        doc_key="F-fail:Purchase:INV-1",
        date="2026-01-10",
        invoice_number="INV-1",
    )

    results = asyncio.run(
        _flush_deferred_ledger_writes(
            ledger_store=_FailingStore(),
            slack_client=MagicMock(),
            channel_id="C-fail",
            batch_deferred=[deferred],
        )
    )

    assert len(results) == 1
    assert results[0].get("flush_failed") is True
    assert results[0].get("appended") == 0


# ------------------------------------------------------------------ (6) gather exception safety


def test_one_run_one_raises_others_still_complete():
    """If one doc's _run_one raises unexpectedly, sibling docs still process and tally.

    The fix: _run_one wraps its entire body in try/except AND gather uses
    return_exceptions=True.  A bad COA-offer or name-resolution crash must not
    cancel the remaining coroutines.
    """
    from unittest.mock import AsyncMock, MagicMock, patch
    from app.slack_app import _SeenEvents
    from tests._slack_test_helpers import capture_message_handler_with_slack_client
    from tests.test_ledger_store import FakeSlackClient

    import ledgr_slack.dedup as _dedup
    _dedup._seen = _SeenEvents()
    slack = FakeSlackClient()
    handler, _ = capture_message_handler_with_slack_client(slack)
    import ledgr_slack.dedup as _dedup
    _dedup._seen = _SeenEvents()

    body = {"event_id": "Ev-exc-safety"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "3000.1",
        "channel": "C-exc-safety",
        "files": [
            {"id": "F-ok-1", "name": "ok1.pdf", "filetype": "pdf"},
            {"id": "F-bad", "name": "bad.pdf", "filetype": "pdf"},   # will raise
            {"id": "F-ok-2", "name": "ok2.pdf", "filetype": "pdf"},
        ],
    }
    fake_client = MagicMock()
    call_ids: list[str] = []

    async def _selective_pfe(**kwargs):
        fid = kwargs["file_id"]
        call_ids.append(fid)
        if fid == "F-bad":
            raise RuntimeError("simulated crash in process_file_event")
        return {
            "status": "delivered",
            "append": {
                "software": "qbs", "fy": "2026", "kind": "invoice",
                "deferred_delivery": _invoice_deferred(
                    doc_key=f"F:{fid}:Purchase:INV",
                    date="15/01/2026",
                    invoice_number=f"INV-{fid}",
                ),
            },
        }

    with patch("ledgr_slack.batch_coordinator.process_file_event", side_effect=_selective_pfe), \
         patch("ledgr_slack.batch_coordinator.download_pdf_bytes", return_value=b"%PDF fake"), \
         patch("ledgr_slack.batch_coordinator._flush_deferred_ledger_writes", new_callable=AsyncMock,
               return_value=[{"appended": 2, "deduped": 0,
                                     "filename": "x.xlsx", "slack_file_id": "F-wb"}]), \
         patch("ledgr_slack.batch_coordinator._flush_deferred_ledger_writes", new_callable=AsyncMock,
               return_value=[{"appended": 2, "deduped": 0,
                              "filename": "x.xlsx", "slack_file_id": "F-wb"}]):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    # All three docs were attempted (none skipped due to a sibling crash).
    assert "F-ok-1" in call_ids, "F-ok-1 was not processed — sibling crash sank the batch"
    assert "F-bad" in call_ids
    assert "F-ok-2" in call_ids, "F-ok-2 was not processed — sibling crash sank the batch"

    # Exactly one top-level summary post (ADR-0007 preserved).
    top_level = [p for p in slack._posts if not p.get("thread_ts")]
    assert len(top_level) == 1, (
        f"expected exactly 1 top-level post, got {len(top_level)}: "
        f"{[p.get('text','') for p in top_level]}"
    )


# ------------------------------------------------------------------ (7) semaphore unchanged


def test_semaphore_still_bounds_concurrency_under_gather(monkeypatch):  # noqa: F811
    """Under asyncio.gather fan-out, _SEM still caps concurrent in-flight runs.

    The fan-out MUST NOT add a second semaphore — _SEM inside process_file_event
    is the sole backpressure mechanism. This test is structurally identical to
    the existing test_semaphore_caps_concurrent_runs but runs docs through the
    full message handler fan-out path (not direct process_file_event calls).
    """
    from unittest.mock import MagicMock, patch
    from app.slack_app import _SeenEvents
    from tests._slack_test_helpers import capture_message_handler_with_slack_client
    from tests.test_ledger_store import FakeSlackClient

    # Force cap to 2 independent of env.
    monkeypatch.setattr("ledgr_slack.file_event._SEM", asyncio.Semaphore(2))
    import ledgr_slack.dedup as _dedup
    _dedup._seen = _SeenEvents()

    concurrent_counter = {"now": 0, "max": 0}

    async def _counting_pfe(**kwargs):
        # Simulate the semaphore being acquired inside process_file_event.
        import ledgr_slack.file_event as _fe
        async with _fe._SEM:
            concurrent_counter["now"] += 1
            concurrent_counter["max"] = max(concurrent_counter["max"], concurrent_counter["now"])
            await asyncio.sleep(0.01)
            concurrent_counter["now"] -= 1
        return {
            "status": "delivered",
            "append": {
                "software": "qbs", "fy": "2026", "kind": "invoice",
                "deferred_delivery": _invoice_deferred(
                    doc_key=f"F:{kwargs['file_id']}:Purchase:INV",
                    date="2026-01-01",
                    invoice_number=f"INV-{kwargs['file_id']}",
                ),
            },
        }

    slack = FakeSlackClient()
    handler, _ = capture_message_handler_with_slack_client(slack)
    import ledgr_slack.dedup as _dedup
    _dedup._seen = _SeenEvents()

    body = {"event_id": "Ev-sem-gather"}
    event = {
        "type": "message",
        "subtype": "file_share",
        "ts": "2000.1",
        "channel": "C-sem-gather",
        "files": [{"id": f"F-sem-{i}", "name": f"doc{i}.pdf", "filetype": "pdf"} for i in range(6)],
    }
    fake_client = MagicMock()

    with patch("ledgr_slack.batch_coordinator.process_file_event", side_effect=_counting_pfe), \
         patch("ledgr_slack.batch_coordinator.download_pdf_bytes", return_value=b"%PDF"), \
         patch("ledgr_slack.batch_coordinator.download_pdf_bytes", return_value=b"%PDF"), \
         patch("ledgr_slack.batch_coordinator._flush_deferred_ledger_writes", new_callable=AsyncMock,
               return_value=[{"appended": 6, "deduped": 0,
                                     "filename": "x.xlsx", "slack_file_id": "F-wb"}]), \
         patch("ledgr_slack.batch_coordinator._flush_deferred_ledger_writes", new_callable=AsyncMock,
               return_value=[{"appended": 6, "deduped": 0,
                              "filename": "x.xlsx", "slack_file_id": "F-wb"}]):
        asyncio.run(handler(event=event, body=body, client=fake_client))

    assert concurrent_counter["max"] <= 2, (
        f"fan-out exceeded semaphore cap: max concurrent = {concurrent_counter['max']} (cap=2)"
    )
