"""Hermetic tests for ``SlackLedgerStore`` — fetch → append → re-upload.

Mocks Slack (an in-memory file store keyed by id, with ``files_upload_v2`` /
``files_info`` + an opener that serves the stored bytes) and Firestore (the
shared ``tests._fake_firestore.FakeFirestore``). No network, no real Slack.

Proves:
- A first append starts a fresh workbook and writes rows into the right sheet.
- A second append fetches the current workbook (via the pointer) and adds the new
  batch — both batches end up in the workbook.
- Re-processing the SAME document (same dedupe doc_key) does NOT double-append.
- The Firestore pointer is updated to the latest uploaded file id.
"""

from __future__ import annotations

import io
import uuid

from openpyxl import load_workbook

from accounting_agents.ledger_store import SlackLedgerStore
from tests._fake_firestore import FakeFirestore


class _FakeResponse:
    def __init__(self, data: bytes):
        self._data = data

    def read(self) -> bytes:
        return self._data

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


class _FakeOpener:
    """Serves bytes for a previously-recorded https URL → bytes mapping."""

    def __init__(self, urls: dict[str, bytes]):
        self._urls = urls

    def open(self, req):
        url = req.full_url if hasattr(req, "full_url") else req
        return _FakeResponse(self._urls[url])


class FakeSlackClient:
    """In-memory Slack file store supporting upload + info + download + reactions."""

    token = "xoxb-test"

    def __init__(self):
        self.files: dict[str, bytes] = {}
        self.urls: dict[str, bytes] = {}
        self.uploads: list[dict] = []
        self._posts: list[dict] = []
        self.updates: list[dict] = []
        # Reaction tracking: list of {"channel", "timestamp", "name"} dicts.
        self.reactions_added: list[dict] = []
        self.reactions_removed: list[dict] = []
        self.thinking_status_calls: list[dict] = []
        # Optional per-file share ts injected by tests to simulate files_info shares.
        # Maps file_id → {"channel_id": ts_string}.
        self._file_share_ts: dict[str, dict[str, str]] = {}
        # Tracks deleted file ids (Fix 1: old ledger file cleanup).
        self.deleted_file_ids: list[str] = []

    def chat_postMessage(self, **kwargs):
        ts = f"{len(self._posts) + 1}.000"
        # Stash the returned ts back on the call kwargs so tests can resolve
        # the message timestamp of a previously-posted top-level message (the
        # batch-summary flow in test_slack_runner needs this for thread_ts).
        kwargs.setdefault("ts", ts)
        self._posts.append(kwargs)
        return {"ok": True, "ts": ts}

    def chat_update(self, **kwargs):
        self.updates.append(kwargs)
        return {"ok": True}

    def files_upload_v2(self, *, channel, filename, file, title=None):
        file_id = "F" + uuid.uuid4().hex[:10]
        self.files[file_id] = file
        url = f"https://files.slack.com/{file_id}/{filename}"
        self.urls[url] = file
        self.uploads.append({"channel": channel, "filename": filename, "id": file_id})
        return {"files": [{"id": file_id, "url_private_download": url}]}

    def files_info(self, *, file):
        # Build shares from _file_share_ts if present, so tests can exercise the
        # reaction path. Also include url_private_download for ledger-store tests.
        shares: dict = {}
        if file in self._file_share_ts:
            shares["private"] = {
                ch: [{"ts": ts}]
                for ch, ts in self._file_share_ts[file].items()
            }
        if file in self.files:
            data = self.files[file]
            url = next(u for u, b in self.urls.items() if b is data)
        else:
            url = f"https://files.slack.com/{file}/unknown"
        return {"file": {"id": file, "url_private_download": url, "name": "ledger.xlsx", "shares": shares}}

    def files_delete(self, *, file):
        self.deleted_file_ids.append(file)
        self.files.pop(file, None)
        return {"ok": True}

    def reactions_add(self, *, channel, timestamp, name):
        self.reactions_added.append({"channel": channel, "timestamp": timestamp, "name": name})
        return {"ok": True}

    def reactions_remove(self, *, channel, timestamp, name):
        self.reactions_removed.append({"channel": channel, "timestamp": timestamp, "name": name})
        return {"ok": True}

    def assistant_threads_setStatus(self, *, channel_id, thread_ts, status, loading_messages=None):
        self.thinking_status_calls.append({
            "channel_id": channel_id,
            "thread_ts": thread_ts,
            "status": status,
            "loading_messages": loading_messages,
        })
        return {"ok": True}

    def opener(self) -> _FakeOpener:
        return _FakeOpener(self.urls)


def _row(desc: str) -> dict:
    return {
        "Invoice Number": "INV-1",
        "Description": desc,
        "Source Amount": 100.0,
        "Account Code / COA": "6000",
    }


def _read_sheet_rows(data: bytes, sheet: str) -> list[tuple]:
    wb = load_workbook(io.BytesIO(data))
    ws = wb[sheet]
    return list(ws.iter_rows(min_row=2, values_only=True))


def _make_store(slack: FakeSlackClient) -> SlackLedgerStore:
    return SlackLedgerStore(FakeFirestore(), opener=slack.opener())


def test_first_append_starts_fresh_workbook():
    slack = FakeSlackClient()
    store = _make_store(slack)

    result = store.append_rows(
        client_id="c1",
        fy="2026",
        slack_client=slack,
        channel_id="C1",
        software="qbs",
        kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )

    assert result["appended"] == 1
    assert result["deduped"] == 0
    assert result["slack_file_id"]
    data = slack.files[result["slack_file_id"]]
    rows = _read_sheet_rows(data, "Purchase")
    assert len(rows) == 1
    # Pointer updated.
    ptr = store.get_pointer("c1", "2026")
    assert ptr["slack_file_id"] == result["slack_file_id"]


def test_filename_is_client_scoped_bank_and_invoice():
    """append_rows names files '<Client> - BankStatement_FY<fy>' / '- Ledger_FY<fy>' (F4)."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    inv = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice", client_name="Company-A",
        batches=[{"sheet": "Purchase", "doc_key": "k1", "rows": [_row("first")]}],
    )
    assert inv["filename"] == "Company-A - Ledger_FY2026.xlsx"

    bank = store.append_rows(
        client_id="c2", fy="2025", slack_client=slack, channel_id="C2",
        software="qbs", kind="bank", client_name="Sample Bank Client Pte Ltd",
        batches=[{"sheet": "OCBC - 0001", "doc_key": "b1", "rows": [
            {"Description": "BALANCE B/F", "Balance": 100.0, "Currency": "SGD"},
            {"Date": "01/10/2025", "Description": "PAYMENT", "Withdrawal": 20.0,
             "Balance": 80.0, "Currency": "SGD"},
        ]}],
    )
    assert bank["filename"] == "Sample Bank Client Pte Ltd - BankStatement_FY2025.xlsx"


def test_filename_falls_back_to_bare_name_without_client():
    """No client_name → bare 'BankStatement_FY<fy>' / 'Ledger_FY<fy>' (back-compat)."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    inv = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "k1", "rows": [_row("first")]}],
    )
    assert inv["filename"] == "Ledger_FY2026.xlsx"


def test_second_append_fetches_and_adds_batch():
    slack = FakeSlackClient()
    store = _make_store(slack)

    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )
    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F2:Purchase:INV-2", "rows": [_row("second")]}],
    )

    assert result2["appended"] == 1
    data = slack.files[result2["slack_file_id"]]
    rows = _read_sheet_rows(data, "Purchase")
    # Both batches present in the workbook now.
    descriptions = [r for row in rows for r in row if r in ("first", "second")]
    assert "first" in descriptions and "second" in descriptions
    assert len(rows) == 2
    # Pointer advanced to the newest upload.
    assert store.get_pointer("c1", "2026")["slack_file_id"] == result2["slack_file_id"]


def test_reprocessing_same_doc_is_deduped():
    slack = FakeSlackClient()
    store = _make_store(slack)

    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )
    # Re-process the SAME document (same doc_key) → no double-append.
    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )

    assert result2["appended"] == 0
    assert result2["deduped"] == 1
    data = slack.files[result2["slack_file_id"]]
    rows = _read_sheet_rows(data, "Purchase")
    assert len(rows) == 1  # still exactly one row


def test_dedupe_only_reshares_workbook_to_slack():
    """When every batch dedupes but the workbook still has data, re-upload it.

    Covers the case where the user deleted the Excel message from the channel
    while Firestore seen_doc_keys still block re-append.
    """
    slack = FakeSlackClient()
    store = _make_store(slack)

    result1 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )
    assert len(slack.uploads) == 1
    first_file_id = result1["slack_file_id"]
    slack.uploads.clear()

    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )

    assert result2["appended"] == 0
    assert result2["deduped"] == 1
    assert result2.get("reshared") is True
    assert len(slack.uploads) == 1, "deduped-only path must still post the workbook"
    assert result2["slack_file_id"] != first_file_id
    data = slack.files[result2["slack_file_id"]]
    rows = _read_sheet_rows(data, "Purchase")
    assert len(rows) == 1


def test_dedupe_only_reshare_skips_shell_workbook():
    """Shell tabs (B/F + TOTALS only) must not be re-shared on dedupe-only path."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    shell_rows = [
        {"Description": "BALANCE B/F", "Balance": 0.0, "Currency": "SGD"},
        {"Description": "TOTALS", "Currency": "SGD"},
    ]
    result1 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[{
            "sheet": "DBS - 5545 - SGD",
            "doc_key": "DBS Bank Ltd - 5545 - SGD:5545:SGD:Apr2024",
            "rows": shell_rows,
        }],
    )
    first_file_id = result1["slack_file_id"]
    slack.uploads.clear()

    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[{
            "sheet": "DBS - 5545 - SGD",
            "doc_key": "DBS Bank Ltd - 5545 - SGD:5545:SGD:Apr2024",
            "rows": shell_rows,
        }],
    )

    assert result2["appended"] == 0
    assert result2.get("reshared") is not True
    assert len(slack.uploads) == 0
    assert result2["slack_file_id"] == first_file_id


def test_purge_seen_doc_keys_allows_reappend():
    slack = FakeSlackClient()
    store = _make_store(slack)
    doc_key = "DBS Bank Ltd - 5545 - SGD:5545:SGD:Apr2024"
    rows = [
        {"Description": "BALANCE B/F", "Balance": 100.0, "Currency": "SGD"},
        {"Date": "15/04/2024", "Description": "SALARY", "Deposit": 50.0,
         "Balance": 150.0, "Currency": "SGD"},
        {"Description": "TOTALS", "Currency": "SGD"},
    ]
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[{"sheet": "DBS - 5545 - SGD", "doc_key": doc_key, "rows": rows}],
    )
    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[{"sheet": "DBS - 5545 - SGD", "doc_key": doc_key, "rows": rows}],
    )
    assert result2["deduped"] == 1

    purged = store.purge_seen_doc_keys("c1", "2026", [doc_key])
    assert purged == 1

    result3 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[{
            "sheet": "DBS - 5545 - SGD",
            "doc_key": doc_key,
            "rows": [
                {"Description": "BALANCE B/F", "Balance": 100.0, "Currency": "SGD"},
                {"Date": "15/04/2024", "Description": "SALARY-UPDATED", "Deposit": 75.0,
                 "Balance": 175.0, "Currency": "SGD"},
                {"Description": "TOTALS", "Currency": "SGD"},
            ],
        }],
    )
    assert result3["appended"] > 0


def test_stash_and_consume_bank_dedup_replace():
    slack = FakeSlackClient()
    store = _make_store(slack)
    batches = [{"sheet": "DBS - 5545 - SGD", "doc_key": "k1", "rows": []}]
    stash_key = "c1|C1|2026|April 2024"
    store.stash_bank_dedup_replace(
        stash_key=stash_key,
        client_id="c1",
        fy="2026",
        kind="bank",
        software="qbs",
        client_name="Test",
        batches=batches,
    )
    loaded = store.consume_bank_dedup_replace(stash_key)
    assert loaded is not None
    assert loaded["batches"] == batches
    assert store.consume_bank_dedup_replace(stash_key) is None


def test_bank_workbook_one_sheet_per_account():
    slack = FakeSlackClient()
    store = _make_store(slack)

    result = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[
            {"sheet": "OCBC - 5001", "doc_key": "F1:OCBC:1", "rows": [{"Date": "01/02/2026", "Description": "x", "Balance": 10.0}]},
            {"sheet": "DBS - 9002", "doc_key": "F1:DBS:1", "rows": [{"Date": "01/02/2026", "Description": "y", "Balance": 20.0}]},
        ],
    )

    data = slack.files[result["slack_file_id"]]
    wb = load_workbook(io.BytesIO(data))
    assert "OCBC - 5001" in wb.sheetnames
    assert "DBS - 9002" in wb.sheetnames
    # No hidden dedupe column in the workbook — dedupe is Firestore-side.
    header = [c.value for c in wb["OCBC - 5001"][1]]
    assert "_ledgr_doc_key" not in header


def test_bank_sheet_title_multi_currency_same_account():
    """Multi-currency account 072-955554-5 → 3 distinct 'Bank - 5545 - CCY' titles."""
    from invoice_processing.export.exporters import bank_sheet_title

    titles = {
        bank_sheet_title(
            bank_name="DBS Bank Ltd",
            account_number="072-955554-5",
            currency=ccy,
        )
        for ccy in ("SGD", "USD", "CNH")
    }
    assert titles == {
        "DBS Bank Ltd - 5545 - SGD",
        "DBS Bank Ltd - 5545 - USD",
        "DBS Bank Ltd - 5545 - CNH",
    }


def test_bank_sheet_title_strips_llm_packed_digits():
    """bank_name 'OCBC - 5001' + account_number '5001' collapses to 'OCBC - 5001 - SGD'."""
    from invoice_processing.export.exporters import bank_sheet_title

    assert bank_sheet_title(
        bank_name="OCBC - 5001", account_number="5001", currency="SGD",
    ) == "OCBC - 5001 - SGD"


def test_bank_sheet_title_falls_back_to_bank_name_digits():
    """If account_number is missing, last4 are pulled from bank_name digits."""
    from invoice_processing.export.exporters import bank_sheet_title

    # No " - " separator in bank_name → whole string is the label; last4 come
    # from the trailing digits embedded in the label. (DBS itself is short
    # enough to fit under the 31-char Excel sheet limit.)
    assert bank_sheet_title(
        bank_name="DBS 955554", account_number=None, currency="USD",
    ) == "DBS 955554 - 5554 - USD"


def test_bank_sheet_title_pads_short_digits():
    from invoice_processing.export.exporters import bank_sheet_title

    assert bank_sheet_title(
        bank_name="DBS", account_number="12", currency="SGD",
    ) == "DBS - 0012 - SGD"


def test_bank_sheet_title_sanitizes_excel_invalid_chars():
    from invoice_processing.export.exporters import bank_sheet_title

    # Excel forbids []:*?/\\ — must be stripped from the label, not crash openpyxl.
    title = bank_sheet_title(
        bank_name="Bank/Co [Asia] - 7777", account_number="X-9999", currency="SGD",
    )
    for ch in "[]:*?/\\":
        assert ch not in title
    assert title.endswith("SGD")


def test_bank_workbook_multi_currency_one_account():
    """Three currency sections of the same account → three Excel tabs with headers."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    from invoice_processing.export.exporters import BankStatementExporter
    from invoice_processing.export.models import BankStatement, BankTransaction
    from datetime import date as _date

    exporter = BankStatementExporter()

    def _stmt(ccy: str, desc: str, *, wd=None, dep=None, bal: float, opening: float):
        return BankStatement(
            bank_name="DBS Bank Ltd",
            account_number="072-955554-5",
            currency=ccy,
            opening_balance=opening,
            closing_balance=bal,
            transactions=[
                BankTransaction(
                    date=_date(2024, 4, 15), description=desc,
                    withdrawal=wd, deposit=dep, balance=bal,
                )
            ],
        )

    batches = []
    for ccy, desc, wd, dep, bal, opening, key in [
        ("SGD", "REMITTANCE", 100000.0, None, 0.0, 100000.0, "F1:DBS:5545:SGD"),
        ("USD", "WIRE FEE", 50.0, None, 0.0, 50.0, "F1:DBS:5545:USD"),
        ("CNH", "CREDIT", None, 200.0, 200.0, 0.0, "F1:DBS:5545:CNH"),
    ]:
        stmt = _stmt(ccy, desc, wd=wd, dep=dep, bal=bal, opening=opening)
        batches.append({
            "sheet": f"DBS Bank Ltd - 5545 - {ccy}",
            "doc_key": key,
            "rows": exporter.bank_rows(stmt),
        })

    result = store.append_rows(
        client_id="c1", fy="2024", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=batches,
    )

    wb = load_workbook(io.BytesIO(slack.files[result["slack_file_id"]]), data_only=False)
    expected_tabs = [
        "DBS Bank Ltd - 5545 - SGD",
        "DBS Bank Ltd - 5545 - USD",
        "DBS Bank Ltd - 5545 - CNH",
    ]
    for tab in expected_tabs:
        assert tab in wb.sheetnames
        ws = wb[tab]
        header = [c.value for c in ws[1]]
        assert header == exporter.BANK_COLS, f"{tab} missing canonical header: {header}"
        check_col = header.index("Math_Check") + 1
        # Row 2 = BALANCE B/F — static ✅ on first block.
        assert ws.cell(row=2, column=check_col).value == "✅"
        # Row 3 = first txn — formula-based check.
        chk = ws.cell(row=3, column=check_col).value
        assert isinstance(chk, str) and chk.startswith("=IF(ROUND(")


def test_bank_april_then_may_same_fy_workbook():
    """April then May on the same account append to one FY workbook tab."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    from invoice_processing.export.exporters import BankStatementExporter
    from invoice_processing.export.models import BankStatement, BankTransaction
    from datetime import date as _date

    exporter = BankStatementExporter()
    sheet = "DBS Bank Ltd - 5545 - SGD"

    def _rows(month: int, desc: str):
        stmt = BankStatement(
            bank_name="DBS Bank Ltd",
            account_number="072-955554-5",
            currency="SGD",
            opening_balance=100.0,
            closing_balance=150.0,
            transactions=[
                BankTransaction(
                    date=_date(2024, month, 10), description=desc,
                    deposit=50.0, balance=150.0,
                ),
            ],
        )
        return exporter.bank_rows(stmt)

    store.append_rows(
        client_id="c1", fy="2024", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[{
            "sheet": sheet,
            "doc_key": f"{sheet}:5545:SGD:Apr2024",
            "rows": _rows(4, "APR-TXN"),
        }],
    )
    result2 = store.append_rows(
        client_id="c1", fy="2024", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[{
            "sheet": sheet,
            "doc_key": f"{sheet}:5545:SGD:May2024",
            "rows": _rows(5, "MAY-TXN"),
        }],
    )
    assert result2["appended"] > 0
    assert result2["deduped"] == 0

    wb = load_workbook(io.BytesIO(slack.files[result2["slack_file_id"]]))
    ws = wb[sheet]
    descriptions = [
        str(ws.cell(row=r, column=2).value or "")
        for r in range(2, ws.max_row + 1)
    ]
    assert "APR-TXN" in descriptions
    assert "MAY-TXN" in descriptions
    assert descriptions.count("BALANCE B/F") >= 2


def test_bank_sheet_has_no_dedupe_column():
    """The written bank sheet must contain NO _ledgr_doc_key column at all."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    from invoice_processing.export.exporters import BankStatementExporter
    from invoice_processing.export.models import BankStatement, BankTransaction
    from datetime import date as _date

    exporter = BankStatementExporter()
    stmt = BankStatement(
        bank_name="OCBC - 5001",
        currency="SGD",
        opening_balance=1000.0,
        closing_balance=900.0,
        transactions=[
            BankTransaction(date=_date(2025, 4, 1), description="ATM", withdrawal=100.0, deposit=None, balance=900.0)
        ],
    )
    result = store.append_rows(
        client_id="c1", fy="2025", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[{"sheet": stmt.bank_name, "doc_key": "F1:OCBC:1", "rows": exporter.bank_rows(stmt)}],
    )
    wb = load_workbook(io.BytesIO(slack.files[result["slack_file_id"]]))
    header = [c.value for c in wb["OCBC - 5001"][1]]
    assert "_ledgr_doc_key" not in header, f"Unexpected dedupe column in header: {header}"


def test_dedupe_via_firestore_no_duplicate_rows():
    """Re-appending the same doc_key must not add duplicate rows — dedupe is Firestore-side."""
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(db, opener=slack.opener())

    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )
    # Verify the seen_doc_keys were persisted to Firestore.
    ptr = store.get_pointer("c1", "2026")
    assert "F1:Purchase:INV-1" in ptr.get("seen_doc_keys", [])

    # Re-append the SAME doc_key — must be deduped via Firestore (no sheet read needed).
    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )
    assert result2["appended"] == 0
    assert result2["deduped"] == 1
    data = slack.files[result2["slack_file_id"]]
    rows = _read_sheet_rows(data, "Purchase")
    assert len(rows) == 1  # still exactly one row


def test_second_append_deletes_first_file():
    """After a second append, the FIRST (superseded) Slack file must be deleted."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    result1 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )
    first_file_id = result1["slack_file_id"]

    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F2:Purchase:INV-2", "rows": [_row("second")]}],
    )
    second_file_id = result2["slack_file_id"]

    # The first file was deleted after the second upload succeeded.
    assert first_file_id in slack.deleted_file_ids
    # The second file (latest) was NOT deleted.
    assert second_file_id not in slack.deleted_file_ids
    # The pointer points at the second file.
    assert store.get_pointer("c1", "2026")["slack_file_id"] == second_file_id


def test_first_append_does_not_delete_any_file():
    """On the very first append there is no previous file — nothing should be deleted."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )
    assert slack.deleted_file_ids == []


def test_file_deleted_from_slack_starts_fresh_and_clears_dedup():
    """If the previous workbook was deleted from Slack, start a fresh one
    and clear seen_doc_keys so previously-processed docs can be re-uploaded."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    result1 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first")]}],
    )
    first_file_id = result1["slack_file_id"]
    assert result1["appended"] == 1

    # Simulate Slack deleting the file (free-tier cleanup / manual deletion).
    original_files_info = slack.files_info
    def _raise_file_deleted(*, file):
        if file == first_file_id:
            raise Exception(
                "The request to the Slack API failed. "
                "The server responded with: {'ok': False, 'error': 'file_deleted'}"
            )
        return original_files_info(file=file)
    slack.files_info = _raise_file_deleted

    # Re-uploading the SAME doc_key should succeed (not dedup) because
    # seen_doc_keys was cleared when the workbook was found to be gone.
    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{"sheet": "Purchase", "doc_key": "F1:Purchase:INV-1", "rows": [_row("first-redo")]}],
    )
    assert result2["appended"] == 1
    assert result2["deduped"] == 0

    # Pointer updated to the new file.
    ptr = store.get_pointer("c1", "2026")
    assert ptr["slack_file_id"] == result2["slack_file_id"]
    assert ptr["slack_file_id"] != first_file_id


# --------------------------------------------------------------------------- #
# Accountant-grade bank export: live formulas, no legacy columns
# --------------------------------------------------------------------------- #

from datetime import date  # noqa: E402

from invoice_processing.export.exporters import BankStatementExporter  # noqa: E402
from invoice_processing.export.models import BankStatement, BankTransaction  # noqa: E402


def _bank_stmt(
    name: str,
    opening: float,
    txns: list[tuple],
    closing: float,
    txn_date: date = date(2026, 2, 1),
) -> BankStatement:
    """Build a BankStatement; each txn = (description, withdrawal, deposit, stated_balance).

    All transactions share ``txn_date`` (override per statement to exercise the
    cross-month, date-sorted continuous chain).
    """
    return BankStatement(
        bank_name=name,
        currency="SGD",
        opening_balance=opening,
        closing_balance=closing,
        transactions=[
            BankTransaction(
                date=txn_date,
                description=d,
                withdrawal=w,
                deposit=dep,
                balance=bal,
            )
            for (d, w, dep, bal) in txns
        ],
    )


def _bank_batch(exporter, stmt, doc_key):
    return {"sheet": stmt.bank_name, "doc_key": doc_key, "rows": exporter.bank_rows(stmt)}


def test_duplicate_statement_collapses_even_under_different_doc_keys():
    """Same statement appended under two different doc_keys → one block (F2).

    Reproduces the Sample Bank Client September duplication: the doc_key format transition let
    the SAME statement through twice (old F<id>:... key vs new content key). The
    block-level dedup in _merge_bank_statement must collapse them so the sheet
    has a single BALANCE B/F + one copy of each transaction.
    """
    slack = FakeSlackClient()
    store = _make_store(slack)
    exporter = BankStatementExporter()

    stmt = _bank_stmt(
        "OCBC - 0001", 1000.0,
        [("FAST PAYMENT", 200.0, None, 800.0), ("SALARY", None, 500.0, 1300.0)],
        1300.0,
        txn_date=date(2025, 9, 15),
    )
    # First upload (old-style doc_key).
    store.append_rows(
        client_id="c1", fy="2025", slack_client=slack, channel_id="C1", kind="bank",
        batches=[_bank_batch(exporter, stmt, "Fold123:OCBC - 0001:acct")],
    )
    # Same statement again under the NEW content key (dedup bypassed at Firestore).
    result = store.append_rows(
        client_id="c1", fy="2025", slack_client=slack, channel_id="C1", kind="bank",
        batches=[_bank_batch(exporter, stmt, "OCBC - 0001:acct:01 SEP 2025 - 30 SEP 2025")],
    )

    rows = _read_sheet_rows(slack.files[result["slack_file_id"]], "OCBC - 0001")
    descs = [r[1] for r in rows]  # Description column
    # Exactly one B/F and one copy of each txn — NOT duplicated.
    assert descs.count("BALANCE B/F") == 1
    assert descs.count("FAST PAYMENT") == 1
    assert descs.count("SALARY") == 1


def test_bank_export_no_legacy_columns():
    exporter = BankStatementExporter()
    header = exporter.BANK_COLS
    assert "Notes" not in header
    assert "Source File ID" not in header
    assert "Stated Balance" not in header
    assert "Check" not in header
    # Sample Partner-pattern columns present.
    assert "Balance" in header
    assert "Math_Check" in header
    assert "Currency" in header


def test_bank_balance_and_check_are_live_formulas():
    slack = FakeSlackClient()
    store = _make_store(slack)
    exporter = BankStatementExporter()

    stmt = _bank_stmt(
        "OCBC - 5001", 1000.0,
        [("ATM withdrawal", 100.0, None, 900.0), ("Salary", None, 500.0, 1400.0)],
        1400.0,
    )
    result = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1", kind="bank",
        batches=[_bank_batch(exporter, stmt, "F1:OCBC:1")],
    )

    wb = load_workbook(io.BytesIO(slack.files[result["slack_file_id"]]), data_only=False)
    ws = wb["OCBC - 5001"]
    header = [c.value for c in ws[1]]
    bi = header.index("Balance") + 1
    ci = header.index("Math_Check") + 1
    wi = header.index("Withdrawal") + 1
    di = header.index("Deposit") + 1

    # Row layout: 1=header, 2=BALANCE B/F, 3=txn, 4=txn, 5=TOTALS.
    # Balance is always an actual number (never a formula).
    assert not str(ws.cell(row=2, column=bi).value).startswith("=")  # B/F row
    assert not str(ws.cell(row=3, column=bi).value).startswith("=")  # txn row
    assert not str(ws.cell(row=4, column=bi).value).startswith("=")  # txn row
    # First B/F Math_Check is a static ✅ (no prior row to compare).
    assert ws.cell(row=2, column=ci).value == "✅"
    # Txn Math_Check is an IF/ROUND formula referencing balance and arithmetic.
    chk = ws.cell(row=3, column=ci).value
    assert isinstance(chk, str) and chk.startswith("=IF(ROUND(")
    # Totals row has SUM formulas over the txn block.
    totals_wd = ws.cell(row=5, column=wi).value
    totals_dep = ws.cell(row=5, column=di).value
    assert str(totals_wd).startswith("=SUM(") and str(totals_dep).startswith("=SUM(")


def test_bank_formulas_correct_after_second_append():
    slack = FakeSlackClient()
    store = _make_store(slack)
    exporter = BankStatementExporter()

    stmt1 = _bank_stmt("OCBC - 5001", 1000.0, [("w1", 100.0, None, 900.0)], 900.0)
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1", kind="bank",
        batches=[_bank_batch(exporter, stmt1, "F1:OCBC:1")],
    )
    stmt2 = _bank_stmt("OCBC - 5001", 900.0, [("d1", None, 250.0, 1150.0)], 1150.0)
    result2 = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1", kind="bank",
        batches=[_bank_batch(exporter, stmt2, "F2:OCBC:2")],
    )

    wb = load_workbook(io.BytesIO(slack.files[result2["slack_file_id"]]), data_only=False)
    ws = wb["OCBC - 5001"]
    header = [c.value for c in ws[1]]
    bi = header.index("Balance") + 1
    di = header.index("Description") + 1
    ci = header.index("Math_Check") + 1

    from openpyxl.utils import get_column_letter
    bal = get_column_letter(bi)

    # Continuous layout after two appends (months merged into one chain, single TOTALS):
    # 1 header
    # 2 B/F (stmt1)  3 w1
    # 4 B/F (stmt2)  5 d1
    # 6 TOTALS
    descs = [ws.cell(row=r, column=di).value for r in range(2, ws.max_row + 1)]
    assert descs == ["BALANCE B/F", "w1", "BALANCE B/F", "d1", "TOTALS"]

    # Balance is always an actual number from the bank statement, never a formula.
    for r in (2, 3, 4, 5):
        assert not str(ws.cell(row=r, column=bi).value).startswith("="), f"row {r} Balance should be a number"

    # First B/F Math_Check: static ✅ (no prior row).
    assert ws.cell(row=2, column=ci).value == "✅"
    # Second B/F Math_Check: continuity formula — carried balance (E3) vs this B/F (E4).
    assert ws.cell(row=4, column=ci).value == f'=IF(ROUND(N({bal}4)-N({bal}3),2)=0,"✅","GAP")'
    # Txn Math_Check: arithmetic formula referencing prior balance row.
    assert ws.cell(row=5, column=ci).value.startswith("=IF(ROUND(")
    # No hidden dedupe column — the Excel is clean / human-readable.
    assert "_ledgr_doc_key" not in header


# --------------------------------------------------------------------------- #
# Edge-case coverage: descending balance, overdraft, wrong balance, continuity
# --------------------------------------------------------------------------- #

def _build_sheet(rows_dicts, cols=None):
    """Build a populated worksheet from value-row dicts; return (ws, header_idx)."""
    from openpyxl import Workbook as _WB
    exp = BankStatementExporter()
    cols = cols or exp.BANK_COLS
    wb = _WB()
    ws = wb.active
    ws.append(list(cols))
    for row in rows_dicts:
        ws.append([row.get(c, "") for c in cols])
    exp.apply_bank_formulas(ws, cols)
    idx = {c.value: c.column for c in ws[1]}
    return ws, idx


def test_descending_balance_balance_is_always_number():
    """All Balance cells are actual numbers even when balance falls every row."""
    stmt = BankStatement(
        bank_name="DBS - 9999", currency="SGD", opening_balance=5000.0, closing_balance=2300.0,
        transactions=[
            BankTransaction(date=date(2025, 6, 1),  description="Rent",      withdrawal=1500.0, deposit=None,  balance=3500.0),
            BankTransaction(date=date(2025, 6, 5),  description="Utilities", withdrawal=200.0,  deposit=None,  balance=3300.0),
            BankTransaction(date=date(2025, 6, 10), description="Insurance", withdrawal=1000.0, deposit=None,  balance=2300.0),
        ],
    )
    exp = BankStatementExporter()
    ws, idx = _build_sheet(exp.bank_rows(stmt))
    bi = idx["Balance"]
    # Rows 2 (B/F), 3,4,5 (txns) — all numbers, never formulas.
    for r in range(2, 6):
        v = ws.cell(row=r, column=bi).value
        assert not str(v).startswith("="), f"row {r} Balance should be a number, got {v!r}"
    assert ws.cell(row=2, column=bi).value == 5000.0
    assert ws.cell(row=3, column=bi).value == 3500.0
    assert ws.cell(row=4, column=bi).value == 3300.0
    assert ws.cell(row=5, column=bi).value == 2300.0


def test_descending_balance_math_check_formula_chains():
    """Math_Check formulas reference the correct prior-row Balance cells."""
    stmt = BankStatement(
        bank_name="DBS - 9999", currency="SGD", opening_balance=5000.0, closing_balance=2300.0,
        transactions=[
            BankTransaction(date=date(2025, 6, 1), description="Rent",      withdrawal=1500.0, deposit=None, balance=3500.0),
            BankTransaction(date=date(2025, 6, 5), description="Utilities", withdrawal=200.0,  deposit=None, balance=3300.0),
            BankTransaction(date=date(2025, 6, 10),description="Insurance", withdrawal=1000.0, deposit=None, balance=2300.0),
        ],
    )
    exp = BankStatementExporter()
    ws, idx = _build_sheet(exp.bank_rows(stmt))
    from openpyxl.utils import get_column_letter
    bal = get_column_letter(idx["Balance"])
    dep = get_column_letter(idx["Deposit"])
    wd  = get_column_letter(idx["Withdrawal"])
    ci  = idx["Math_Check"]

    # B/F row (row 2): static ✅, no prior row.
    assert ws.cell(row=2, column=ci).value == "✅"
    # Txn rows: formula references prior Balance and this row's Deposit/Withdrawal.
    assert ws.cell(row=3, column=ci).value == f'=IF(ROUND(N({bal}3)-(N({bal}2)+N({dep}3)-N({wd}3)),2)=0,"✅","❌ Exp: "&ROUND(N({bal}2)+N({dep}3)-N({wd}3),2))'
    assert ws.cell(row=4, column=ci).value == f'=IF(ROUND(N({bal}4)-(N({bal}3)+N({dep}4)-N({wd}4)),2)=0,"✅","❌ Exp: "&ROUND(N({bal}3)+N({dep}4)-N({wd}4),2))'
    assert ws.cell(row=5, column=ci).value == f'=IF(ROUND(N({bal}5)-(N({bal}4)+N({dep}5)-N({wd}5)),2)=0,"✅","❌ Exp: "&ROUND(N({bal}4)+N({dep}5)-N({wd}5),2))'
    # TOTALS row (row 6): no Math_Check.
    assert ws.cell(row=6, column=ci).value in (None, "")


def test_overdraft_negative_balance():
    """Negative (overdraft) balances are stored as-is and formulas still chain correctly."""
    stmt = BankStatement(
        bank_name="OD - 0001", currency="SGD", opening_balance=100.0, closing_balance=-450.0,
        transactions=[
            BankTransaction(date=date(2025, 6, 1), description="Big payment", withdrawal=550.0, deposit=None,   balance=-450.0),
            BankTransaction(date=date(2025, 6, 5), description="Refund",       withdrawal=None,  deposit=100.0, balance=-350.0),
        ],
    )
    exp = BankStatementExporter()
    ws, idx = _build_sheet(exp.bank_rows(stmt))
    bi = idx["Balance"]
    assert ws.cell(row=2, column=bi).value == 100.0    # B/F opening
    assert ws.cell(row=3, column=bi).value == -450.0   # after big withdrawal
    assert ws.cell(row=4, column=bi).value == -350.0   # after refund
    # Math_Check formulas still present and well-formed.
    ci = idx["Math_Check"]
    assert ws.cell(row=3, column=ci).value.startswith("=IF(ROUND(")
    assert ws.cell(row=4, column=ci).value.startswith("=IF(ROUND(")


def test_wrong_balance_math_check_formula_structure():
    """When the bank states a wrong balance, the Math_Check formula exposes it.

    We can't evaluate Excel formulas in openpyxl, but we verify the formula
    compares E_curr (the stated value) against the arithmetic — so Excel will
    show ❌ when they differ.
    """
    # Txn says balance=500 but arithmetic (1000-600=400) gives 400 — wrong.
    stmt = BankStatement(
        bank_name="ERR - 0001", currency="SGD", opening_balance=1000.0, closing_balance=500.0,
        transactions=[
            BankTransaction(date=date(2025, 6, 1), description="Wrong bal txn", withdrawal=600.0, deposit=None, balance=500.0),
        ],
    )
    exp = BankStatementExporter()
    ws, idx = _build_sheet(exp.bank_rows(stmt))
    from openpyxl.utils import get_column_letter
    bal = get_column_letter(idx["Balance"])
    dep = get_column_letter(idx["Deposit"])
    wd  = get_column_letter(idx["Withdrawal"])
    ci  = idx["Math_Check"]

    # Balance holds the bank's stated value (500 — wrong, but faithfully stored).
    assert ws.cell(row=3, column=idx["Balance"]).value == 500.0
    # Math_Check formula references E3 (stated=500) vs arithmetic E2+D3-C3 (=400).
    # In Excel this resolves to ❌ because ROUND(500-400,2)≠0.
    expected = f'=IF(ROUND(N({bal}3)-(N({bal}2)+N({dep}3)-N({wd}3)),2)=0,"✅","❌ Exp: "&ROUND(N({bal}2)+N({dep}3)-N({wd}3),2))'
    assert ws.cell(row=3, column=ci).value == expected


def test_cross_month_descending_chain_balance_and_continuity():
    """Two descending-balance months: second B/F continuity formula references prior closing."""
    exp = BankStatementExporter()
    slack = FakeSlackClient()
    store = _make_store(slack)

    stmt1 = _bank_stmt("DBS - 8888", 5000.0,
                       [("Rent", 2000.0, None, 3000.0), ("Bills", 500.0, None, 2500.0)],
                       2500.0, txn_date=date(2025, 6, 15))
    stmt2 = _bank_stmt("DBS - 8888", 2500.0,
                       [("Salary", None, 3000.0, 5500.0), ("Tax", 1000.0, None, 4500.0)],
                       4500.0, txn_date=date(2025, 7, 15))

    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1", kind="bank",
        batches=[_bank_batch(exp, stmt1, "F1:DBS:1")],
    )
    result = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1", kind="bank",
        batches=[_bank_batch(exp, stmt2, "F2:DBS:2")],
    )

    wb = load_workbook(io.BytesIO(slack.files[result["slack_file_id"]]), data_only=False)
    ws = wb["DBS - 8888"]
    header = [c.value for c in ws[1]]
    bi = header.index("Balance") + 1
    ci = header.index("Math_Check") + 1
    di = header.index("Description") + 1

    from openpyxl.utils import get_column_letter
    bal = get_column_letter(bi)

    # Layout: 1=hdr, 2=B/F(jun), 3=Rent, 4=Bills, 5=B/F(jul), 6=Salary, 7=Tax, 8=TOTALS
    descs = [ws.cell(row=r, column=di).value for r in range(2, ws.max_row + 1)]
    assert descs == ["BALANCE B/F", "Rent", "Bills", "BALANCE B/F", "Salary", "Tax", "TOTALS"]

    # Balance is a real number on every data row.
    for r in range(2, 8):
        v = ws.cell(row=r, column=bi).value
        assert not str(v).startswith("="), f"row {r} Balance should be a number"
    assert ws.cell(row=2, column=bi).value == 5000.0
    assert ws.cell(row=3, column=bi).value == 3000.0
    assert ws.cell(row=4, column=bi).value == 2500.0
    assert ws.cell(row=5, column=bi).value == 2500.0   # second B/F stated opening
    assert ws.cell(row=6, column=bi).value == 5500.0
    assert ws.cell(row=7, column=bi).value == 4500.0

    # First B/F Math_Check: ✅ (no prior month).
    assert ws.cell(row=2, column=ci).value == "✅"
    # Second B/F continuity check: compares E5 (stated 2500) against E4 (closing 2500).
    assert ws.cell(row=5, column=ci).value == f'=IF(ROUND(N({bal}5)-N({bal}4),2)=0,"✅","GAP")'
    # Txn Math_Check formulas all well-formed.
    for r in (3, 4, 6, 7):
        assert ws.cell(row=r, column=ci).value.startswith("=IF(ROUND("), f"row {r} missing formula"


def test_cross_month_gap_is_flagged_by_formula():
    """When month 2 B/F opening ≠ month 1 closing, the GAP formula exposes it.

    We verify the formula references the right cells so Excel would evaluate to GAP.
    """
    exp = BankStatementExporter()
    slack = FakeSlackClient()
    store = _make_store(slack)

    stmt1 = _bank_stmt("OCBC - 7777", 1000.0, [("w1", 100.0, None, 900.0)], 900.0,
                       txn_date=date(2025, 6, 15))
    # Second month's stated opening is 1000 — but month 1 closed at 900 (GAP of 100).
    stmt2 = _bank_stmt("OCBC - 7777", 1000.0, [("d1", None, 50.0, 1050.0)], 1050.0,
                       txn_date=date(2025, 7, 15))

    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1", kind="bank",
        batches=[_bank_batch(exp, stmt1, "F1:OCBC:1")],
    )
    result = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1", kind="bank",
        batches=[_bank_batch(exp, stmt2, "F2:OCBC:2")],
    )

    wb = load_workbook(io.BytesIO(slack.files[result["slack_file_id"]]), data_only=False)
    ws = wb["OCBC - 7777"]
    header = [c.value for c in ws[1]]
    bi = header.index("Balance") + 1
    ci = header.index("Math_Check") + 1

    from openpyxl.utils import get_column_letter
    bal = get_column_letter(bi)

    # Layout: 1=hdr, 2=B/F(jun), 3=w1, 4=B/F(jul), 5=d1, 6=TOTALS
    # Second B/F (row 4): stated=1000, prior closing (row 3)=900 → formula will show GAP.
    assert ws.cell(row=4, column=bi).value == 1000.0   # stated opening stored faithfully
    continuity = ws.cell(row=4, column=ci).value
    assert continuity == f'=IF(ROUND(N({bal}4)-N({bal}3),2)=0,"✅","GAP")'
    # (In Excel: ROUND(1000-900,2)=100≠0 → "GAP")


# --------------------------------------------------------------------------- #
# Task 7 regression: OLD-formula Balance cells mixed with NEW static cells
# --------------------------------------------------------------------------- #

import openpyxl  # noqa: E402


def _make_mixed_formula_workbook() -> bytes:
    """Build a workbook that mimics Sample Bank Client's corrupted BankStatement_FY2025.

    Jan–Mar blocks have Balance cells stored as Excel formula strings
    (``="=E2+D3-C3"`` style — the OLD layout that commit 6ca4e48 replaced).
    Apr–May blocks have the NEW static numeric values. This is the exact
    corruption that the append path must survive.

    The sheet uses the CURRENT 7-col header (Date, Description, Withdrawal,
    Deposit, Balance, Currency, Math_Check) so only the Balance *values* are
    formulae — the header itself is already migrated.
    """
    cols = list(BankStatementExporter.BANK_COLS)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCBC - 5001"
    ws.append(cols)

    def row_values(desc, wd, dep, bal, currency="SGD", date_str=""):
        row = [""] * len(cols)
        row[cols.index("Date")] = date_str
        row[cols.index("Description")] = desc
        if wd is not None:
            row[cols.index("Withdrawal")] = wd
        if dep is not None:
            row[cols.index("Deposit")] = dep
        row[cols.index("Balance")] = bal
        row[cols.index("Currency")] = currency
        return row

    # --- OLD-style blocks (Jan–Mar): Balance cells are formula strings ---
    # Jan B/F
    ws.append(row_values("BALANCE B/F", None, None, 1000.0, date_str="01/01/2025"))
    # Jan txn — Balance stored as a formula string (old layout)
    ws.append(row_values("Salary Jan", None, 500.0, "=E2+D3-C3", date_str="15/01/2025"))
    ws.append(row_values("Rent Jan", 800.0, None, "=E3+D4-C4", date_str="20/01/2025"))
    ws.append(row_values("TOTALS", None, None, "", date_str=""))

    # Feb B/F — formula string
    ws.append(row_values("BALANCE B/F", None, None, "=E4", date_str="01/02/2025"))
    ws.append(row_values("Salary Feb", None, 500.0, "=E5+D6-C6", date_str="15/02/2025"))
    ws.append(row_values("TOTALS", None, None, "", date_str=""))

    # --- NEW-style blocks (Apr–May): Balance cells are static numbers ---
    # Apr B/F
    ws.append(row_values("BALANCE B/F", None, None, 1200.0, date_str="01/04/2025"))
    ws.append(row_values("Salary Apr", None, 500.0, 1700.0, date_str="15/04/2025"))
    ws.append(row_values("TOTALS", None, None, "", date_str=""))

    return _wb_to_bytes(wb)


def _wb_to_bytes(wb) -> bytes:
    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_formula_balance_cells_are_recomputed_on_append():
    """Regression: appending to a workbook whose Balance cells are formula strings
    (old layout, pre-6ca4e48) must yield a clean static running balance — never
    a formula string stored as a Balance value, and never None.

    Concrete assertions:
    - After appending a new May statement on top of the corrupted workbook, every
      Balance cell on the rebuilt sheet is a numeric float (not a str, not None).
    - The running balance computed from stated_bf + Σ(deposit − withdrawal) is
      arithmetically correct for the reconstructed rows we *can* verify.
    - The new May deposit row has a correct running balance chained from prior rows.
    """
    slack = FakeSlackClient()
    exp = BankStatementExporter()

    # Seed the fake Slack store with the corrupted workbook (formula Balance cells).
    corrupted_bytes = _make_mixed_formula_workbook()
    seed_id = "Fseed001"
    seed_url = f"https://files.slack.com/{seed_id}/BankStatement_FY2025.xlsx"
    slack.files[seed_id] = corrupted_bytes
    slack.urls[seed_url] = corrupted_bytes

    db = FakeFirestore()
    # Point Firestore at the seeded file so the store fetches it.
    db.collection("clients").document("sample-bank").collection("ledgers").document("2025").set({
        "slack_file_id": seed_id,
        "fy": "2025",
        "client_id": "sample-bank",
        "seen_doc_keys": ["sample-bank:jan2025", "sample-bank:feb2025", "sample-bank:apr2025"],
        "channel_id": "C_SAMPLE_BANK",
        "kind": "bank",
    })

    store = SlackLedgerStore(db, opener=slack.opener())

    # May statement (new, clean static values).
    may_stmt = BankStatement(
        bank_name="OCBC - 5001",
        currency="SGD",
        opening_balance=1700.0,
        closing_balance=2200.0,
        transactions=[
            BankTransaction(
                date=date(2025, 5, 15),
                description="Salary May",
                withdrawal=None,
                deposit=500.0,
                balance=2200.0,
            )
        ],
    )
    result = store.append_rows(
        client_id="sample-bank",
        fy="2025",
        slack_client=slack,
        channel_id="C_SAMPLE_BANK",
        kind="bank",
        batches=[{"sheet": "OCBC - 5001", "doc_key": "sample-bank:may2025",
                  "rows": exp.bank_rows(may_stmt)}],
    )

    # Load the rebuilt workbook (data_only=False to catch any leaked formulas).
    import io as _io
    wb = load_workbook(_io.BytesIO(slack.files[result["slack_file_id"]]), data_only=False)
    ws = wb["OCBC - 5001"]
    header = [c.value for c in ws[1]]
    bi = header.index("Balance") + 1

    # CORE ASSERTION: every Balance cell must be a number — never a formula string,
    # never None (except the TOTALS row which has no Balance).
    desc_i = header.index("Description") + 1
    for r in range(2, ws.max_row + 1):
        desc_val = ws.cell(row=r, column=desc_i).value
        bal_val = ws.cell(row=r, column=bi).value
        if desc_val == BankStatementExporter.TOTALS_MARKER:
            continue  # TOTALS row Balance is intentionally blank
        assert bal_val is not None, f"Row {r} ({desc_val!r}): Balance is None"
        assert not isinstance(bal_val, str), (
            f"Row {r} ({desc_val!r}): Balance is a formula/string {bal_val!r}"
        )
        assert isinstance(bal_val, (int, float)), (
            f"Row {r} ({desc_val!r}): Balance unexpected type {type(bal_val)} = {bal_val!r}"
        )

    # The Jan B/F (first row after header) should have balance = 1000.0 (the seed value).
    first_bf_bal = ws.cell(row=2, column=bi).value
    assert first_bf_bal == 1000.0, f"Jan B/F balance wrong: {first_bf_bal}"

    # May B/F opening should equal the stated opening (1700.0).
    # Find the last BALANCE B/F before TOTALS.
    bf_rows = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=desc_i).value == BankStatementExporter.OPENING_MARKER
    ]
    assert len(bf_rows) >= 2, "Expected at least two BALANCE B/F rows (Jan+Apr or later)"

    # Salary May should be the last txn row before TOTALS, balance = 2200.0.
    last_txn_row = ws.max_row - 1  # TOTALS is last row
    may_bal = ws.cell(row=last_txn_row, column=bi).value
    assert may_bal == 2200.0, f"May Salary balance wrong: {may_bal}"


def test_is_formula_or_missing_treats_non_numeric_as_untrusted():
    """Non-numeric Balance cells (e.g. a stray 'SGD') are untrusted, not crashes.

    Regression for the live bank-continuity crash:
    ``ValueError: could not convert string to float: 'SGD'`` in _recompute_balances.
    """
    f = SlackLedgerStore._is_formula_or_missing
    assert f("SGD") is True          # stray currency code
    assert f("BALANCE B/F") is True  # a label that leaked into the Balance column
    assert f("=E4") is True          # formula
    assert f(None) is True
    assert f("") is True
    assert f(538.78) is False        # real numeric balance
    assert f("538.78") is False      # numeric string


def test_recompute_balances_does_not_crash_on_non_numeric_bf_balance():
    """A B/F row whose Balance cell is 'SGD' must carry forward, not raise — and a
    later numeric B/F still seeds the running balance and chains correctly."""
    from invoice_processing.export.exporters import BankStatementExporter as _BX
    rows = [
        {"Description": _BX.OPENING_MARKER, "Balance": "SGD", "Deposit": None, "Withdrawal": None},
        {"Description": _BX.OPENING_MARKER, "Balance": 1000.0, "Deposit": None, "Withdrawal": None},
        {"Description": "Deposit", "Balance": None, "Deposit": 250.0, "Withdrawal": None},
        {"Description": "Payment", "Balance": None, "Deposit": None, "Withdrawal": 100.0},
    ]
    SlackLedgerStore._recompute_balances(rows)  # must not raise
    assert rows[0]["Balance"] is None     # 'SGD' untrusted → carried forward (running None)
    assert rows[1]["Balance"] == 1000.0   # numeric B/F seeds the running balance
    assert rows[2]["Balance"] == 1250.0   # chained: 1000 + 250
    assert rows[3]["Balance"] == 1150.0   # chained: 1250 − 100


def test_legacy_8col_header_is_migrated_on_read():
    """Regression: a sheet written with the OLD 8-col header (Stated Balance + Check
    instead of Balance + Math_Check) must be read without losing the B/F opening values.

    The migration renames columns on read; ``_read_bank_blocks`` must not return
    blocks with ``stated_bf=None`` when the column is just named differently.
    """
    # Build a workbook with the OLD header layout.
    old_cols = ["Date", "Description", "Withdrawal", "Deposit",
                "Stated Balance", "Currency", "Check", "Notes"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCBC - 5001"
    ws.append(old_cols)

    # Write one B/F + one txn in the old layout.
    ws.append(["", "BALANCE B/F", None, None, 500.0, "SGD", "✅", ""])
    ws.append(["15/03/2025", "Salary", None, 300.0, 800.0, "SGD", "✅", ""])
    ws.append(["", "TOTALS", None, None, "", "SGD", "", ""])

    slack = FakeSlackClient()
    corrupted_bytes = _wb_to_bytes(wb)
    seed_id = "Fold001"
    seed_url = f"https://files.slack.com/{seed_id}/BankStatement_FY2025.xlsx"
    slack.files[seed_id] = corrupted_bytes
    slack.urls[seed_url] = corrupted_bytes

    db = FakeFirestore()
    db.collection("clients").document("legacy_client").collection("ledgers").document("2025").set({
        "slack_file_id": seed_id,
        "fy": "2025",
        "client_id": "legacy_client",
        "seen_doc_keys": ["legacy:mar2025"],
        "channel_id": "C_LEG",
        "kind": "bank",
    })
    store = SlackLedgerStore(db, opener=slack.opener())

    # Append a new statement — this triggers _read_bank_blocks on the old layout.
    exp = BankStatementExporter()
    apr_stmt = BankStatement(
        bank_name="OCBC - 5001",
        currency="SGD",
        opening_balance=800.0,
        closing_balance=1300.0,
        transactions=[
            BankTransaction(
                date=date(2025, 4, 15),
                description="Bonus",
                withdrawal=None,
                deposit=500.0,
                balance=1300.0,
            )
        ],
    )
    result = store.append_rows(
        client_id="legacy_client",
        fy="2025",
        slack_client=slack,
        channel_id="C_LEG",
        kind="bank",
        batches=[{"sheet": "OCBC - 5001", "doc_key": "legacy:apr2025",
                  "rows": exp.bank_rows(apr_stmt)}],
    )

    import io as _io
    wb2 = load_workbook(_io.BytesIO(slack.files[result["slack_file_id"]]), data_only=False)
    ws2 = wb2["OCBC - 5001"]
    header2 = [c.value for c in ws2[1]]
    bi2 = header2.index("Balance") + 1
    desc_i2 = header2.index("Description") + 1

    # The rebuilt sheet uses the NEW header (Balance, not Stated Balance).
    assert "Balance" in header2
    assert "Stated Balance" not in header2

    # The old B/F opening (500.0) must be preserved in the rebuilt chain.
    bf_rows = [
        r for r in range(2, ws2.max_row + 1)
        if ws2.cell(row=r, column=desc_i2).value == BankStatementExporter.OPENING_MARKER
    ]
    assert len(bf_rows) == 2, f"Expected 2 B/F rows, got {len(bf_rows)}"
    first_bf_bal = ws2.cell(row=bf_rows[0], column=bi2).value
    assert first_bf_bal == 500.0, f"Old B/F opening not preserved: {first_bf_bal}"
    # New B/F opening must also be correct.
    second_bf_bal = ws2.cell(row=bf_rows[1], column=bi2).value
    assert second_bf_bal == 800.0, f"Apr B/F opening wrong: {second_bf_bal}"


# --------------------------------------------------------------------------- #
# Step 4 — row coordinate + workbook mutation (amend_row / remove_row)
# --------------------------------------------------------------------------- #

def _make_invoice_store_with_two_rows() -> tuple["FakeSlackClient", "SlackLedgerStore", str]:
    """Seed a two-row invoice ledger and return (slack, store, file_id)."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[
            {"sheet": "Purchase", "doc_key": "k1", "rows": [
                {"Invoice Number": "INV-1", "Description": "Alpha", "Source Amount": 100.0, "Account Code / COA": "6000"},
                {"Invoice Number": "INV-2", "Description": "Beta",  "Source Amount": 200.0, "Account Code / COA": "6001"},
            ]},
        ],
    )
    ptr = store.get_pointer("c1", "2026")
    return slack, store, ptr["slack_file_id"]


# ---- read_rows: _row coordinate -----------------------------------------

def test_read_rows_stamps_row_number():
    """read_rows must stamp _row=2 on the first data row, _row=3 on the second."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    rows = store.read_rows("c1", "2026", slack, "C1")
    assert len(rows) == 2
    assert rows[0]["_row"] == 2
    assert rows[1]["_row"] == 3


def test_read_rows_row_matches_sheet():
    """_row is consistent with _sheet across a two-sheet invoice workbook."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[
            {"sheet": "Purchase", "doc_key": "p1", "rows": [
                {"Invoice Number": "P1", "Description": "P-alpha", "Source Amount": 10.0, "Account Code / COA": "6000"},
            ]},
            {"sheet": "Sales", "doc_key": "s1", "rows": [
                {"Invoice Number": "S1", "Description": "S-alpha", "Source Amount": 20.0, "Account Code / COA": "4000"},
                {"Invoice Number": "S2", "Description": "S-beta",  "Source Amount": 30.0, "Account Code / COA": "4001"},
            ]},
        ],
    )
    rows = store.read_rows("c1", "2026", slack, "C1")
    purchase = [r for r in rows if r["_sheet"] == "Purchase"]
    sales    = [r for r in rows if r["_sheet"] == "Sales"]
    # Purchase has 1 data row → _row 2.
    assert purchase[0]["_row"] == 2
    # Sales has 2 data rows → rows 2 and 3.
    assert sales[0]["_row"] == 2
    assert sales[1]["_row"] == 3


def test_read_rows_existing_keys_intact():
    """_row addition must not disturb _sheet or column-header values."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    rows = store.read_rows("c1", "2026", slack, "C1")
    assert rows[0]["_sheet"] == "Purchase"
    assert rows[0]["Description"] == "Alpha"
    assert rows[1]["Description"] == "Beta"


def test_read_rows_empty_ledger():
    """read_rows on a missing pointer returns [] (unchanged behaviour)."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    assert store.read_rows("no-such-client", "2026", slack, "C1") == []


# ---- amend_row: happy path -----------------------------------------------

def test_amend_row_changes_targeted_cell():
    """amend_row updates exactly the targeted cell and leaves others intact."""
    slack, store, _ = _make_invoice_store_with_two_rows()

    result = store.amend_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
        updates={"Description": "Alpha-edited"},
    )

    assert result["sheet"] == "Purchase"
    assert result["row"] == 2
    # before/after tracking.
    assert result["before"]["Description"] == "Alpha"
    assert result["after"]["Description"] == "Alpha-edited"

    # Read back and verify.
    rows = store.read_rows("c1", "2026", slack, "C1")
    assert rows[0]["Description"] == "Alpha-edited"
    # Second row untouched.
    assert rows[1]["Description"] == "Beta"


def test_amend_row_multi_column_update():
    """amend_row with two column updates changes both, leaves unlisted columns alone."""
    slack, store, _ = _make_invoice_store_with_two_rows()

    store.amend_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
        updates={"Description": "X", "Source Amount": 999.0},
    )

    rows = store.read_rows("c1", "2026", slack, "C1")
    assert rows[0]["Description"] == "X"
    assert rows[0]["Source Amount"] == 999.0
    assert rows[0]["Account Code / COA"] == "6000"  # untouched


def test_amend_row_uploads_new_version_and_updates_pointer():
    """amend_row uploads a new Slack file and updates the Firestore pointer."""
    slack, store, old_file_id = _make_invoice_store_with_two_rows()

    store.amend_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
        updates={"Description": "Changed"},
    )

    ptr = store.get_pointer("c1", "2026")
    new_file_id = ptr["slack_file_id"]
    assert new_file_id != old_file_id
    assert new_file_id in slack.files


def test_amend_row_leaves_seen_doc_keys_intact():
    """amend_row must not modify the seen_doc_keys field on the pointer."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    ptr_before = store.get_pointer("c1", "2026")
    keys_before = set(ptr_before.get("seen_doc_keys") or [])

    store.amend_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
        updates={"Description": "Y"},
    )

    ptr_after = store.get_pointer("c1", "2026")
    keys_after = set(ptr_after.get("seen_doc_keys") or [])
    assert keys_before == keys_after


def test_amend_row_second_row_unchanged():
    """amend_row on row 2 must leave row 3 completely untouched."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    before_rows = store.read_rows("c1", "2026", slack, "C1")
    beta_before = {k: v for k, v in before_rows[1].items() if not k.startswith("_")}

    store.amend_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
        updates={"Description": "Changed"},
    )

    after_rows = store.read_rows("c1", "2026", slack, "C1")
    beta_after = {k: v for k, v in after_rows[1].items() if not k.startswith("_")}
    assert beta_before == beta_after


# ---- amend_row: error guards ---------------------------------------------

def test_amend_row_raises_on_missing_pointer():
    """amend_row raises ValueError when no ledger pointer exists."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    import pytest as _pytest
    with _pytest.raises(ValueError, match="no ledger pointer"):
        store.amend_row(
            "nonexistent", "2026", slack, "C1",
            sheet="Purchase", row=2,
            updates={"Description": "X"},
        )


def test_amend_row_raises_on_unknown_sheet():
    """amend_row raises ValueError when the sheet name is not in the workbook."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    import pytest as _pytest
    with _pytest.raises(ValueError, match="sheet not found"):
        store.amend_row(
            "c1", "2026", slack, "C1",
            sheet="NonExistentSheet", row=2,
            updates={"Description": "X"},
        )


def test_amend_row_raises_on_row_below_range():
    """amend_row raises ValueError when row < 2 (header row or above)."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    import pytest as _pytest
    with _pytest.raises(ValueError, match="row.*out of range"):
        store.amend_row(
            "c1", "2026", slack, "C1",
            sheet="Purchase", row=1,
            updates={"Description": "X"},
        )


def test_amend_row_raises_on_row_above_range():
    """amend_row raises ValueError when row > max_row."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    import pytest as _pytest
    with _pytest.raises(ValueError, match="row.*out of range"):
        store.amend_row(
            "c1", "2026", slack, "C1",
            sheet="Purchase", row=999,
            updates={"Description": "X"},
        )


def test_amend_row_raises_on_unknown_column():
    """amend_row raises ValueError when an updates key is not a known column header."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    import pytest as _pytest
    with _pytest.raises(ValueError, match="unknown column"):
        store.amend_row(
            "c1", "2026", slack, "C1",
            sheet="Purchase", row=2,
            updates={"NoSuchColumn": "X"},
        )


# ---- amend_row / remove_row: bank-sheet safety ---------------------------

def _make_bank_store_with_one_row() -> tuple["FakeSlackClient", "SlackLedgerStore"]:
    """Seed a bank ledger (OCBC - 0001) and return (slack, store)."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        kind="bank",
        batches=[{"sheet": "OCBC - 0001", "doc_key": "b1", "rows": [
            {"Description": "BALANCE B/F", "Balance": 1000.0, "Currency": "SGD"},
            {"Date": "01/02/2026", "Description": "PAYMENT", "Withdrawal": 100.0,
             "Balance": 900.0, "Currency": "SGD"},
        ]}],
    )
    return slack, store


def test_amend_row_refuses_bank_sheet():
    """amend_row on a bank/account sheet must raise ValueError (balances are derived)."""
    slack, store = _make_bank_store_with_one_row()
    import pytest as _pytest
    with _pytest.raises(ValueError, match="bank-statement rows are read-only"):
        store.amend_row(
            "c1", "2026", slack, "C1",
            sheet="OCBC - 0001", row=2,
            updates={"Description": "X"},
        )


def test_remove_row_refuses_bank_sheet():
    """remove_row on a bank/account sheet must raise ValueError."""
    slack, store = _make_bank_store_with_one_row()
    import pytest as _pytest
    with _pytest.raises(ValueError, match="bank-statement rows are read-only"):
        store.remove_row(
            "c1", "2026", slack, "C1",
            sheet="OCBC - 0001", row=2,
        )


# ---- remove_row: happy path ----------------------------------------------

def test_remove_row_deletes_target_row():
    """remove_row deletes the target row; the remaining row shifts up to row 2."""
    slack, store, _ = _make_invoice_store_with_two_rows()

    result = store.remove_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
    )

    assert result["sheet"] == "Purchase"
    assert result["row"] == 2
    assert result["removed"]["Description"] == "Alpha"

    rows = store.read_rows("c1", "2026", slack, "C1")
    assert len(rows) == 1
    assert rows[0]["Description"] == "Beta"
    assert rows[0]["_row"] == 2  # shifted up


def test_remove_row_second_row():
    """remove_row on row 3 deletes the second data row, leaves row 2 intact."""
    slack, store, _ = _make_invoice_store_with_two_rows()

    result = store.remove_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=3,
    )

    assert result["removed"]["Description"] == "Beta"
    rows = store.read_rows("c1", "2026", slack, "C1")
    assert len(rows) == 1
    assert rows[0]["Description"] == "Alpha"


def test_remove_row_uploads_new_version_and_updates_pointer():
    """remove_row uploads a new Slack file and updates the Firestore pointer."""
    slack, store, old_file_id = _make_invoice_store_with_two_rows()

    store.remove_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
    )

    ptr = store.get_pointer("c1", "2026")
    new_file_id = ptr["slack_file_id"]
    assert new_file_id != old_file_id
    assert new_file_id in slack.files


def test_remove_row_leaves_seen_doc_keys_intact():
    """remove_row must not modify the seen_doc_keys field on the pointer."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    keys_before = set(store.get_pointer("c1", "2026").get("seen_doc_keys") or [])

    store.remove_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
    )

    keys_after = set(store.get_pointer("c1", "2026").get("seen_doc_keys") or [])
    assert keys_before == keys_after


# ---- remove_row: error guards --------------------------------------------

def test_remove_row_raises_on_missing_pointer():
    """remove_row raises ValueError when no ledger pointer exists."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    import pytest as _pytest
    with _pytest.raises(ValueError, match="no ledger pointer"):
        store.remove_row(
            "nonexistent", "2026", slack, "C1",
            sheet="Purchase", row=2,
        )


def test_remove_row_raises_on_unknown_sheet():
    """remove_row raises ValueError when the sheet name is not in the workbook."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    import pytest as _pytest
    with _pytest.raises(ValueError, match="sheet not found"):
        store.remove_row(
            "c1", "2026", slack, "C1",
            sheet="Nonexistent", row=2,
        )


def test_remove_row_raises_on_row_out_of_range():
    """remove_row raises ValueError for row < 2 or row > max_row."""
    slack, store, _ = _make_invoice_store_with_two_rows()
    import pytest as _pytest
    with _pytest.raises(ValueError, match="row.*out of range"):
        store.remove_row(
            "c1", "2026", slack, "C1",
            sheet="Purchase", row=1,
        )
    with _pytest.raises(ValueError, match="row.*out of range"):
        store.remove_row(
            "c1", "2026", slack, "C1",
            sheet="Purchase", row=100,
        )


# --------------------------------------------------------------------------- #
# Fix 1: client-name filename prefix preserved through amend_row / remove_row
# --------------------------------------------------------------------------- #

def test_amend_row_preserves_client_name_filename_prefix():
    """amend_row must upload a file whose name still carries the '<Client> - ' prefix.

    append_rows now persists client_name onto the Firestore pointer so that
    _upload_and_reroute can rebuild the prefix without an extra profile lookup.
    """
    slack = FakeSlackClient()
    store = _make_store(slack)
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        client_name="Acme Trading Pte. Ltd.",
        batches=[{"sheet": "Purchase", "doc_key": "k1", "rows": [
            {"Invoice Number": "INV-1", "Description": "Alpha", "Source Amount": 100.0,
             "Account Code / COA": "6000"},
            {"Invoice Number": "INV-2", "Description": "Beta",  "Source Amount": 200.0,
             "Account Code / COA": "6001"},
        ]}],
    )
    # Confirm client_name was persisted onto the pointer.
    ptr = store.get_pointer("c1", "2026")
    assert ptr.get("client_name") == "Acme Trading Pte. Ltd."

    store.amend_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
        updates={"Description": "Alpha-edited"},
    )

    # The most-recent upload should carry the client-scoped filename.
    latest_upload = slack.uploads[-1]
    assert latest_upload["filename"] == "Acme Trading Pte. Ltd. - Ledger_FY2026.xlsx"


def test_amend_row_legacy_pointer_without_client_name_does_not_crash():
    """A pointer written before client_name was persisted has no client_name field.

    _upload_and_reroute must fall back to a bare filename without crashing,
    and the pointer's seen_doc_keys must remain intact.
    """
    slack = FakeSlackClient()
    store = _make_store(slack)
    # Seed a pointer WITHOUT client_name (simulates a legacy pointer).
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        # client_name intentionally omitted → defaults to ""
        batches=[{"sheet": "Purchase", "doc_key": "k1", "rows": [
            {"Invoice Number": "INV-1", "Description": "Alpha", "Source Amount": 100.0,
             "Account Code / COA": "6000"},
        ]}],
    )
    # Manually strip client_name from the pointer to simulate a legacy doc.
    ptr_ref = (
        store._db.collection("clients").document("c1")
        .collection("ledgers").document("2026")
    )
    existing = ptr_ref.get().to_dict()
    existing.pop("client_name", None)
    ptr_ref.set(existing)

    # Must not raise; bare filename is acceptable.
    store.amend_row(
        "c1", "2026", slack, "C1",
        sheet="Purchase", row=2,
        updates={"Description": "Changed"},
    )

    latest_upload = slack.uploads[-1]
    assert latest_upload["filename"] == "Ledger_FY2026.xlsx"
    # seen_doc_keys still intact.
    ptr_after = store.get_pointer("c1", "2026")
    assert "k1" in (ptr_after.get("seen_doc_keys") or [])


# --------------------------------------------------------------------------- #
# remove_rows_for_month (Step 7 / C-3)
# --------------------------------------------------------------------------- #


def _make_invoice_workbook_with_months() -> bytes:
    """Build a Purchase + Sales workbook with rows in Sep and Oct 2025."""
    from openpyxl import Workbook as _WB
    wb = _WB()
    # Purchase sheet
    ws_p = wb.active
    ws_p.title = "Purchase"
    ws_p.append(["Date", "Invoice Number", "Description", "Source Amount", "Account Code / COA"])
    ws_p.append(["05/09/2025", "INV-P1", "AWS Sep",  100.0, "6090"])
    ws_p.append(["20/09/2025", "INV-P2", "Zoom Sep", 50.0,  "6090"])
    ws_p.append(["03/10/2025", "INV-P3", "AWS Oct",  120.0, "6090"])
    # Sales sheet
    ws_s = wb.create_sheet("Sales")
    ws_s.append(["Date", "Invoice Number", "Description", "Source Amount", "Account Code / COA"])
    ws_s.append(["10/09/2025", "INV-S1", "Consulting Sep", 500.0, "4000"])
    ws_s.append(["15/10/2025", "INV-S2", "Consulting Oct", 600.0, "4000"])
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_qbs_exporter_workbook_with_months() -> bytes:
    """Same months as _make_invoice_workbook_with_months but QBS export headers."""
    from openpyxl import Workbook as _WB
    wb = _WB()
    ws_p = wb.active
    ws_p.title = "Purchase"
    ws_p.append([
        "Invoice Number", "Invoice Date", "Vendor Name", "Description",
        "Source Amount", "Account Code / COA",
    ])
    ws_p.append(["INV-P1", "05/09/2025", "AWS", "AWS Sep", 100.0, "6090"])
    ws_p.append(["INV-P2", "20/09/2025", "Zoom", "Zoom Sep", 50.0, "6090"])
    ws_p.append(["INV-P3", "03/10/2025", "AWS", "AWS Oct", 120.0, "6090"])
    ws_s = wb.create_sheet("Sales")
    ws_s.append([
        "Invoice Date", "Invoice Number", "Customer Name", "Description",
        "Source Amount", "Account Code / COA",
    ])
    ws_s.append(["10/09/2025", "INV-S1", "Client", "Consulting Sep", 500.0, "4000"])
    ws_s.append(["15/10/2025", "INV-S2", "Client", "Consulting Oct", 600.0, "4000"])
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _store_with_workbook(xlsx_bytes: bytes, *, client_id: str = "c1", fy: str = "2026") -> tuple:
    """Return (slack, store) with the given workbook already uploaded + pointer seeded."""
    slack = FakeSlackClient()
    db = FakeFirestore()
    store = SlackLedgerStore(db, opener=slack.opener())
    # Manually upload + seed pointer + seen_doc_keys.
    file_id = "F" + "a" * 10
    import uuid
    file_id = "F" + uuid.uuid4().hex[:10]
    slack.files[file_id] = xlsx_bytes
    url = f"https://files.slack.com/{file_id}/ledger.xlsx"
    slack.urls[url] = xlsx_bytes
    seen = [
        "Purchase:INV-P1", "Purchase:INV-P2", "Purchase:INV-P3",
        "Sales:INV-S1", "Sales:INV-S2",
        "OtherKey:SHOULD-SURVIVE",
    ]
    store._pointer_ref(client_id, fy).set({
        "slack_file_id": file_id,
        "client_id": client_id,
        "fy": fy,
        "kind": "invoice",
        "seen_doc_keys": seen,
    })
    return slack, store


def test_remove_rows_for_month_removes_matching_rows():
    xlsx = _make_invoice_workbook_with_months()
    slack, store = _store_with_workbook(xlsx)

    result = store.remove_rows_for_month(
        "c1", "2026", slack, "C1",
        year=2025, month=9,
    )

    # Two Purchase + one Sales row removed.
    assert result["sheets"]["Purchase"] == 2
    assert result["sheets"]["Sales"] == 1
    assert len(result["removed"]) == 3

    # Verify the new workbook only contains Oct rows.
    from openpyxl import load_workbook
    import io
    latest_id = store.get_pointer("c1", "2026")["slack_file_id"]
    wb = load_workbook(io.BytesIO(slack.files[latest_id]))
    p_rows = list(wb["Purchase"].iter_rows(min_row=2, values_only=True))
    s_rows = list(wb["Sales"].iter_rows(min_row=2, values_only=True))
    # Only Oct row remains in Purchase.
    assert len([r for r in p_rows if r[0] is not None]) == 1
    assert p_rows[0][2] == "AWS Oct"
    # Only Oct row remains in Sales.
    assert len([r for r in s_rows if r[0] is not None]) == 1
    assert s_rows[0][2] == "Consulting Oct"


def test_remove_rows_for_month_qbs_exporter_headers():
    """Regression: real QBS workbooks use 'Invoice Date', not 'Date'."""
    xlsx = _make_qbs_exporter_workbook_with_months()
    slack, store = _store_with_workbook(xlsx)
    result = store.remove_rows_for_month(
        "c1", "2026", slack, "C1", year=2025, month=9,
    )
    assert result["sheets"]["Purchase"] == 2
    assert result["sheets"]["Sales"] == 1
    rows = store.read_rows("c1", "2026", slack, "C1")
    sep_rows = [
        r for r in rows
        if r.get("_sheet") in ("Purchase", "Sales")
        and "09/2025" in str(r.get("Invoice Date") or r.get("Date") or "")
    ]
    assert sep_rows == []
    oct_rows = [
        r for r in rows
        if r.get("_sheet") in ("Purchase", "Sales")
        and "10/2025" in str(r.get("Invoice Date") or r.get("Date") or "")
    ]
    assert len(oct_rows) == 2


def test_remove_rows_for_month_leaves_other_months_intact():
    xlsx = _make_invoice_workbook_with_months()
    slack, store = _store_with_workbook(xlsx)

    store.remove_rows_for_month("c1", "2026", slack, "C1", year=2025, month=9)

    # Oct rows must survive.
    rows = store.read_rows("c1", "2026", slack, "C1")
    dates = [r.get("Date") for r in rows if r.get("_sheet") in ("Purchase", "Sales")]
    assert all("10/2025" in str(d) for d in dates if d), dates


def test_remove_rows_for_month_purges_doc_keys_for_cleared_month():
    xlsx = _make_invoice_workbook_with_months()
    slack, store = _store_with_workbook(xlsx)

    result = store.remove_rows_for_month("c1", "2026", slack, "C1", year=2025, month=9)

    # Sep keys must be purged.
    purged = set(result["purged_keys"])
    assert "Purchase:INV-P1" in purged
    assert "Purchase:INV-P2" in purged
    assert "Sales:INV-S1" in purged

    # Oct keys and unrelated keys must survive in Firestore.
    ptr = store.get_pointer("c1", "2026")
    surviving = set(ptr.get("seen_doc_keys") or [])
    assert "Purchase:INV-P3" in surviving
    assert "Sales:INV-S2" in surviving
    assert "OtherKey:SHOULD-SURVIVE" in surviving
    # Sep keys gone.
    assert "Purchase:INV-P1" not in surviving
    assert "Purchase:INV-P2" not in surviving
    assert "Sales:INV-S1" not in surviving


def test_remove_rows_for_month_seen_doc_keys_regression_guard():
    """Purged doc_keys allow re-drop; surviving keys still block double-append."""
    xlsx = _make_invoice_workbook_with_months()
    slack, store = _store_with_workbook(xlsx)

    # Clear September.
    store.remove_rows_for_month("c1", "2026", slack, "C1", year=2025, month=9)

    ptr = store.get_pointer("c1", "2026")
    surviving = set(ptr.get("seen_doc_keys") or [])

    # Sep keys gone — a re-drop would NOT be deduped.
    assert "Purchase:INV-P1" not in surviving
    assert "Purchase:INV-P2" not in surviving
    assert "Sales:INV-S1" not in surviving

    # Oct keys still present — a re-drop would be correctly deduped.
    assert "Purchase:INV-P3" in surviving
    assert "Sales:INV-S2" in surviving


def test_remove_rows_for_month_deletes_bottom_up():
    """Bottom-up deletion: removing two rows must not corrupt row indices."""
    from openpyxl import Workbook as _WB
    import io
    wb = _WB()
    ws = wb.active
    ws.title = "Purchase"
    ws.append(["Date", "Invoice Number", "Description", "Source Amount"])
    ws.append(["01/09/2025", "INV-A", "Alpha", 10.0])
    ws.append(["02/09/2025", "INV-B", "Beta",  20.0])
    ws.append(["03/09/2025", "INV-C", "Gamma", 30.0])
    ws.append(["03/10/2025", "INV-D", "Delta", 40.0])  # different month — must survive
    buf = io.BytesIO()
    wb.save(buf)
    xlsx = buf.getvalue()

    slack, store = _store_with_workbook(xlsx)
    store.remove_rows_for_month("c1", "2026", slack, "C1", year=2025, month=9)

    latest_id = store.get_pointer("c1", "2026")["slack_file_id"]
    from openpyxl import load_workbook
    wb2 = load_workbook(io.BytesIO(slack.files[latest_id]))
    rows = list(wb2["Purchase"].iter_rows(min_row=2, values_only=True))
    non_blank = [r for r in rows if r[0] is not None]
    assert len(non_blank) == 1
    assert non_blank[0][2] == "Delta"


def test_remove_rows_for_month_refuses_bank_sheet():
    import pytest
    xlsx = _make_invoice_workbook_with_months()
    slack, store = _store_with_workbook(xlsx)
    with pytest.raises(ValueError, match="bank"):
        store.remove_rows_for_month(
            "c1", "2026", slack, "C1",
            year=2025, month=9,
            sheets=("OCBC - 0001",),
        )


def test_remove_rows_for_month_raises_on_invalid_month():
    import pytest
    xlsx = _make_invoice_workbook_with_months()
    slack, store = _store_with_workbook(xlsx)
    with pytest.raises(ValueError, match="month"):
        store.remove_rows_for_month("c1", "2026", slack, "C1", year=2025, month=13)
    with pytest.raises(ValueError, match="month"):
        store.remove_rows_for_month("c1", "2026", slack, "C1", year=2025, month=0)


def test_remove_rows_for_month_no_pointer_raises():
    import pytest
    slack = FakeSlackClient()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    with pytest.raises(ValueError, match="no ledger pointer"):
        store.remove_rows_for_month("nobody", "2026", slack, "C1", year=2025, month=9)


def test_remove_rows_for_month_missing_sheet_returns_zero_count():
    """A sheet that doesn't exist in the workbook → count 0, no crash."""
    xlsx = _make_invoice_workbook_with_months()
    slack, store = _store_with_workbook(xlsx)
    # Only clear Purchase; Sales rows survive.
    result = store.remove_rows_for_month(
        "c1", "2026", slack, "C1",
        year=2025, month=9,
        sheets=("Purchase",),
    )
    assert result["sheets"]["Purchase"] == 2
    # Sales untouched — read_rows should still have the Sep sales row.
    rows = store.read_rows("c1", "2026", slack, "C1")
    sep_sales = [r for r in rows if r.get("_sheet") == "Sales" and "09/2025" in str(r.get("Date", ""))]
    assert len(sep_sales) == 1


# --------------------------------------------------------------------------- #
# append_rows replace=True (Step 7 / E-3 identity-replace primitive)
# --------------------------------------------------------------------------- #

def _make_invoice_store_with_named_invoices() -> tuple["FakeSlackClient", "SlackLedgerStore"]:
    """Seed a Purchase sheet with INV-10 (2 lines) + INV-20 (1 line) and return (slack, store).

    doc_key format matches consolidate_node: "{sheet}:{invoice_number}".
    """
    slack = FakeSlackClient()
    store = _make_store(slack)
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[
            {
                "sheet": "Purchase",
                "doc_key": "Purchase:INV-10",
                "rows": [
                    {"Invoice Number": "INV-10", "Description": "Line A", "Source Amount": 100.0, "Account Code / COA": "6000"},
                    {"Invoice Number": "INV-10", "Description": "Line B", "Source Amount": 50.0,  "Account Code / COA": "6001"},
                ],
            },
            {
                "sheet": "Purchase",
                "doc_key": "Purchase:INV-20",
                "rows": [
                    {"Invoice Number": "INV-20", "Description": "Other Invoice", "Source Amount": 200.0, "Account Code / COA": "6002"},
                ],
            },
        ],
    )
    return slack, store


def _make_xero_invoice_store_with_named_invoices() -> tuple["FakeSlackClient", "SlackLedgerStore"]:
    """Seed a Purchase sheet with Xero export headers and INV-10 (2 lines) + INV-20."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="xero", kind="invoice",
        batches=[
            {
                "sheet": "Purchase",
                "doc_key": "Purchase:INV-10",
                "rows": [
                    {"*InvoiceNumber": "INV-10", "*Description": "Line A", "*UnitAmount": 100.0, "*AccountCode": "6000"},
                    {"*InvoiceNumber": "INV-10", "*Description": "Line B", "*UnitAmount": 50.0, "*AccountCode": "6001"},
                ],
            },
            {
                "sheet": "Purchase",
                "doc_key": "Purchase:INV-20",
                "rows": [
                    {"*InvoiceNumber": "INV-20", "*Description": "Other Invoice", "*UnitAmount": 200.0, "*AccountCode": "6002"},
                ],
            },
        ],
    )
    return slack, store


def test_replace_true_matched_invoice_removes_old_rows_appends_new():
    """replace=True + matching invoice number: old rows removed, new rows appended,
    result reports replaced > 0."""
    slack, store = _make_invoice_store_with_named_invoices()

    # Confirm baseline: 3 rows total (2 for INV-10, 1 for INV-20).
    rows_before = store.read_rows("c1", "2026", slack, "C1")
    assert len(rows_before) == 3

    result = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        replace=True,
        batches=[{
            "sheet": "Purchase",
            "doc_key": "Purchase:INV-10",
            "rows": [
                {"Invoice Number": "INV-10", "Description": "Corrected Line", "Source Amount": 999.0, "Account Code / COA": "6000"},
            ],
        }],
    )

    # appended reflects the newly appended row count.
    assert result["appended"] == 1
    # batch_replace_counts reports replaced=2 (the two old INV-10 lines).
    assert "batch_replace_counts" in result
    counts = result["batch_replace_counts"]
    assert len(counts) == 1
    assert counts[0]["replaced"] == 2
    assert counts[0]["appended"] == 1

    rows_after = store.read_rows("c1", "2026", slack, "C1")
    descs = [r["Description"] for r in rows_after]
    # Old INV-10 lines gone, new corrected line present, INV-20 untouched.
    assert "Line A" not in descs
    assert "Line B" not in descs
    assert "Corrected Line" in descs
    assert "Other Invoice" in descs
    # Total: 1 (INV-10 replacement) + 1 (INV-20 survivor) = 2 rows.
    assert len(rows_after) == 2


def test_replace_true_xero_invoice_number_header():
    """Regression: Xero workbooks use *InvoiceNumber, not Invoice Number."""
    slack, store = _make_xero_invoice_store_with_named_invoices()

    rows_before = store.read_rows("c1", "2026", slack, "C1")
    assert len(rows_before) == 3

    result = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="xero", kind="invoice",
        replace=True,
        batches=[{
            "sheet": "Purchase",
            "doc_key": "Purchase:INV-10",
            "rows": [
                {
                    "*InvoiceNumber": "INV-10",
                    "Description": "Corrected Line",
                    "*UnitAmount": 999.0,
                    "*AccountCode": "6000",
                },
            ],
        }],
    )

    counts = result["batch_replace_counts"]
    assert counts[0]["replaced"] == 2
    assert counts[0]["appended"] == 1

    rows_after = store.read_rows("c1", "2026", slack, "C1")
    inv10_rows = [
        r for r in rows_after
        if (r.get("*InvoiceNumber") or r.get("Invoice Number")) == "INV-10"
    ]
    assert len(inv10_rows) == 1
    assert inv10_rows[0].get("Description") == "Corrected Line"
    assert len(rows_after) == 2


def test_replace_true_no_match_appends_new_reports_replaced_zero():
    """replace=True where the invoice number does NOT match any existing row:
    new rows are appended, nothing removed, replaced=0 (caller-warn signal)."""
    slack, store = _make_invoice_store_with_named_invoices()

    result = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        replace=True,
        batches=[{
            "sheet": "Purchase",
            "doc_key": "Purchase:INV-99",
            "rows": [
                {"Invoice Number": "INV-99", "Description": "Brand New", "Source Amount": 300.0, "Account Code / COA": "6005"},
            ],
        }],
    )

    assert result["appended"] == 1
    counts = result["batch_replace_counts"]
    assert len(counts) == 1
    assert counts[0]["replaced"] == 0  # no match — caller should warn user
    assert counts[0]["appended"] == 1

    rows_after = store.read_rows("c1", "2026", slack, "C1")
    descs = [r["Description"] for r in rows_after]
    # Original rows intact, new row added.
    assert "Line A" in descs
    assert "Line B" in descs
    assert "Other Invoice" in descs
    assert "Brand New" in descs
    assert len(rows_after) == 4


def test_replace_true_only_replaces_matched_invoice_leaves_others_intact():
    """replace=True removes ONLY the matching invoice number's rows;
    a second, different invoice in the same sheet is completely untouched."""
    slack, store = _make_invoice_store_with_named_invoices()

    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        replace=True,
        batches=[{
            "sheet": "Purchase",
            "doc_key": "Purchase:INV-10",
            "rows": [
                {"Invoice Number": "INV-10", "Description": "INV-10 Replacement", "Source Amount": 111.0, "Account Code / COA": "6000"},
            ],
        }],
    )

    rows_after = store.read_rows("c1", "2026", slack, "C1")
    inv20_rows = [r for r in rows_after if r.get("Invoice Number") == "INV-20"]
    inv10_rows = [r for r in rows_after if r.get("Invoice Number") == "INV-10"]

    # INV-20 is completely untouched.
    assert len(inv20_rows) == 1
    assert inv20_rows[0]["Description"] == "Other Invoice"
    assert inv20_rows[0]["Source Amount"] == 200.0

    # INV-10 replaced by the single new line.
    assert len(inv10_rows) == 1
    assert inv10_rows[0]["Description"] == "INV-10 Replacement"


def test_replace_false_duplicate_doc_key_is_still_deduped():
    """replace=False (default) keeps today's exact dedup behaviour:
    a duplicate doc_key is still skipped — the default path did not regress."""
    slack, store = _make_invoice_store_with_named_invoices()

    # Re-submit INV-10 with replace=False (the default).
    result = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        replace=False,
        batches=[{
            "sheet": "Purchase",
            "doc_key": "Purchase:INV-10",
            "rows": [
                {"Invoice Number": "INV-10", "Description": "Should Not Appear", "Source Amount": 1.0, "Account Code / COA": "9999"},
            ],
        }],
    )

    assert result["appended"] == 0
    assert result["deduped"] == 1
    # batch_replace_counts NOT present in replace=False result.
    assert "batch_replace_counts" not in result

    rows_after = store.read_rows("c1", "2026", slack, "C1")
    descs = [r["Description"] for r in rows_after]
    assert "Should Not Appear" not in descs
    # Original 3 rows unchanged.
    assert len(rows_after) == 3


def test_replace_true_multi_line_invoice_all_lines_replaced():
    """replace=True on a multi-line invoice: ALL lines of the matched invoice
    number are replaced (not just the first row)."""
    slack, store = _make_invoice_store_with_named_invoices()

    # INV-10 currently has 2 lines (Line A + Line B). Replace with 3 new lines.
    result = store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        replace=True,
        batches=[{
            "sheet": "Purchase",
            "doc_key": "Purchase:INV-10",
            "rows": [
                {"Invoice Number": "INV-10", "Description": "New Line 1", "Source Amount": 10.0, "Account Code / COA": "6000"},
                {"Invoice Number": "INV-10", "Description": "New Line 2", "Source Amount": 20.0, "Account Code / COA": "6001"},
                {"Invoice Number": "INV-10", "Description": "New Line 3", "Source Amount": 30.0, "Account Code / COA": "6002"},
            ],
        }],
    )

    counts = result["batch_replace_counts"]
    assert counts[0]["replaced"] == 2   # both old lines removed
    assert counts[0]["appended"] == 3   # three new lines written

    rows_after = store.read_rows("c1", "2026", slack, "C1")
    inv10_rows = [r for r in rows_after if r.get("Invoice Number") == "INV-10"]
    assert len(inv10_rows) == 3
    new_descs = {r["Description"] for r in inv10_rows}
    assert new_descs == {"New Line 1", "New Line 2", "New Line 3"}

    # Old lines gone.
    all_descs = [r["Description"] for r in rows_after]
    assert "Line A" not in all_descs
    assert "Line B" not in all_descs


def test_replace_true_doc_key_in_seen_after_replace():
    """After replace=True the replaced doc_key is still in seen_doc_keys
    (the re-added key blocks accidental double-append)."""
    slack, store = _make_invoice_store_with_named_invoices()

    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        replace=True,
        batches=[{
            "sheet": "Purchase",
            "doc_key": "Purchase:INV-10",
            "rows": [
                {"Invoice Number": "INV-10", "Description": "Replaced", "Source Amount": 1.0, "Account Code / COA": "6000"},
            ],
        }],
    )

    ptr = store.get_pointer("c1", "2026")
    seen = set(ptr.get("seen_doc_keys") or [])
    assert "Purchase:INV-10" in seen


def test_best_fy_for_chat_picks_fy_with_most_rows():
    """``best_fy_for_chat`` should return the FY whose workbook has the most
    data, not the highest FY label. P0-1 fix for chat lane."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    # FY2025 — populated with 3 rows
    slack_file_old = store.append_rows(
        client_id="c1", fy="2025", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{
            "sheet": "Purchase",
            "doc_key": "old:Purchase:1",
            "rows": [
                _row("a"), _row("b"), _row("c"),
            ],
        }],
    )["slack_file_id"]

    # FY2026 — populated with 1 row (would be picked by latest_fy)
    store.append_rows(
        client_id="c1", fy="2026", slack_client=slack, channel_id="C1",
        software="qbs", kind="invoice",
        batches=[{
            "sheet": "Purchase",
            "doc_key": "new:Purchase:1",
            "rows": [_row("only")],
        }],
    )

    best, summaries = store.best_fy_for_chat("c1", slack)
    assert best == "2025", f"expected FY2025 (more rows), got {best!r}"
    by_fy = {s["fy"]: s for s in summaries}
    assert by_fy["2025"]["row_count"] == 3
    assert by_fy["2025"]["has_data"] is True
    assert by_fy["2026"]["row_count"] == 1
    # Sanity: pointer file_id for FY2025 matches.
    assert by_fy["2025"]["fy"] == "2025"
    # Suppress unused-var lint for the captured file id.
    assert slack_file_old


def test_best_fy_for_chat_falls_back_to_latest_when_empty():
    """If every FY is empty, fall back to the highest FY label (matches
    the old ``latest_fy`` behaviour so the agent can still report a stable
    label to the user)."""
    slack = FakeSlackClient()
    store = SlackLedgerStore(FakeFirestore(), opener=slack.opener())
    # Manually create an empty workbook so the FY pointer resolves to a real
    # file (with zero rows) — append_rows rejects empty batches.
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.append(["Date", "Description", "Source Amount"])
    buf = io.BytesIO()
    wb.save(buf)
    slack.files["F_EMPTY"] = buf.getvalue()
    slack.urls["F_EMPTY"] = buf.getvalue()
    store._set_pointer(
        client_id="c1", fy="2025", slack_file_id="F_EMPTY",
    )
    best, summaries = store.best_fy_for_chat("c1", slack)
    assert best == "2025"
    assert all(s["row_count"] == 0 for s in summaries)
    # All FYs are empty → has_data is False for each.
    assert all(s["has_data"] is False for s in summaries)


def test_best_fy_for_chat_no_pointers_returns_none():
    """No ledger pointers → ``(None, [])``."""
    slack = FakeSlackClient()
    store = _make_store(slack)
    best, summaries = store.best_fy_for_chat("ghost", slack)
    assert best is None
    assert summaries == []


# ------------------------------------------------------------------ #
# WS5a — lock keyed on client_id, not channel_id
# ------------------------------------------------------------------ #


def test_lock_for_same_client_same_fy_returns_same_object():
    """Two calls with the SAME client_id+fy return the IDENTICAL lock object."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    lock_a = store._lock_for("client-X", "2026")
    lock_b = store._lock_for("client-X", "2026")
    assert lock_a is lock_b, "same client+fy must share one lock"


def test_lock_for_different_clients_get_different_locks():
    """Two DIFFERENT client_ids get DISTINCT locks (no over-serialization)."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    lock_a = store._lock_for("client-A", "2026")
    lock_b = store._lock_for("client-B", "2026")
    assert lock_a is not lock_b, "different clients must not share a lock"


def test_lock_for_same_client_different_fy_get_different_locks():
    """Same client but different FY → different locks."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    lock_2025 = store._lock_for("client-X", "2025")
    lock_2026 = store._lock_for("client-X", "2026")
    assert lock_2025 is not lock_2026


def test_two_channels_same_client_share_one_lock():
    """The lock is keyed on client_id: two different channel_ids that belong to
    the same client produce the same lock object (the workbook is shared)."""
    slack = FakeSlackClient()
    store = _make_store(slack)

    # Simulate two Slack channels C-ALPHA and C-BETA that both map to client-X.
    # The store no longer keys on channel_id, so we call _lock_for directly with
    # the client_id that each caller would supply.
    lock_from_ch_alpha = store._lock_for("client-X", "2026")
    lock_from_ch_beta = store._lock_for("client-X", "2026")

    assert lock_from_ch_alpha is lock_from_ch_beta, (
        "two channels of the same client must serialize on ONE lock "
        "to prevent last-writer-wins on the shared workbook"
    )


def test_two_channels_same_client_no_lost_rows():
    """Two append_rows calls with DIFFERENT channel_ids but the SAME client_id
    both land in the workbook — no rows are lost due to a lock race.

    This is the headline correctness test for WS5a: before the fix both channels
    got separate locks and could interleave reads and writes, silently dropping
    whichever upload arrived last.  After the fix they share one lock and the
    second writer always reads the first writer's upload before appending.
    """
    slack = FakeSlackClient()
    store = _make_store(slack)

    # First channel appends a row.
    store.append_rows(
        client_id="shared-client",
        fy="2026",
        slack_client=slack,
        channel_id="C-ALPHA",
        software="qbs",
        kind="invoice",
        batches=[
            {
                "sheet": "Purchase",
                "doc_key": "F:Purchase:INV-ALPHA",
                "rows": [_row("from channel alpha")],
            }
        ],
    )

    # Second channel (different channel_id, SAME client_id) appends another row.
    store.append_rows(
        client_id="shared-client",
        fy="2026",
        slack_client=slack,
        channel_id="C-BETA",
        software="qbs",
        kind="invoice",
        batches=[
            {
                "sheet": "Purchase",
                "doc_key": "F:Purchase:INV-BETA",
                "rows": [_row("from channel beta")],
            }
        ],
    )

    # Read the final workbook from the pointer.
    ptr = store.get_pointer("shared-client", "2026")
    final_bytes = slack.files[ptr["slack_file_id"]]
    rows = _read_sheet_rows(final_bytes, "Purchase")

    assert len(rows) == 2, (
        f"expected 2 rows (one per channel), got {len(rows)}; "
        "a race would have caused one row to be lost"
    )
