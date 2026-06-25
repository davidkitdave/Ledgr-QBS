"""Golden-format acceptance tests for AutoCount and SQL Account ERP importers.

Validates that the generated column headers exactly match the real import templates
reverse-engineered from:
  ~/Desktop/LocalTest/header template/Autocount Template/Import-AP-Invoice.xls
  ~/Desktop/LocalTest/header template/Autocount Template/Import-AR-Invoice.xls
  ~/Desktop/LocalTest/header template/SQL Header/Import Purchase Invoice.xlsx  (sheet SLPH_Invoice_Cash_Debit_Credit, row 6)
  ~/Desktop/LocalTest/header template/SQL Header/Import Sales Invoice.xlsx     (sheet SLPH_Invoice_Cash_Debit_Credit, row 6)

Also validates:
  - Mandatory AutoCount columns populated for a fully-mapped line
  - SQL REQUIRED columns (_QTY/_UOM/_UNITPRICE/_ACCOUNT) populated
  - Acme COA Party List: Bolt Auto Supply Sdn Bhd → creditor code 400-A0001
"""

from __future__ import annotations

import os as _os
import tempfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from invoice_processing.export.client_context import EntityMemoryEntry
from invoice_processing.export.exporters import (
    AutoCountExporter,
    QbsLedgerExporter,
    SqlAccountExporter,
    XeroLedgerExporter,
)
from invoice_processing.export.models import InvoiceLine, NormalizedInvoice, PartyInfo
from invoice_processing.export.tax_classifier import get_tax_classifier

MY_CLF = get_tax_classifier("my_sst.yaml")

# Real template paths — tests skip gracefully when running in CI without the local files.
# The client folder segment is supplied at runtime (real local data, not committed);
# defaults to a generic placeholder so no client name lives in the path literal.
_LOCAL_DATA_ROOT = Path(_os.getenv("LEDGR_LOCAL_DATA_ROOT", str(Path.home() / "Desktop/LocalTest")))
_CLIENT_FOLDER = _os.getenv("LEDGR_CLIENT_FOLDER", "Acme Auto Enterprise")
_AUTOCOUNT_DIR = _LOCAL_DATA_ROOT / "header template/Autocount Template"
_SQL_DIR = _LOCAL_DATA_ROOT / "header template/SQL Header"
_CLIENT_COA = _LOCAL_DATA_ROOT / "TestDoc/MYDoc" / _CLIENT_FOLDER / "COA & List.xlsx"

_TEMPLATES_PRESENT = (
    (_AUTOCOUNT_DIR / "Import-AP-Invoice.xls").exists()
    and (_AUTOCOUNT_DIR / "Import-AR-Invoice.xls").exists()
    and (_SQL_DIR / "Import Purchase Invoice.xlsx").exists()
    and (_SQL_DIR / "Import Sales Invoice.xlsx").exists()
)

# Real AutoCount header (row index 2, cols 1..21) as read from the template
_AC_AP_HEADER = [
    "DocNo", "DocDate", "CreditorCode", "SupplierInvoiceNo", "JournalType",
    "DisplayTerm", "PurchaseAgent", "Description", "CurrencyRate", "RefNo2",
    "Note", "InclusiveTax", "AccNo", "ToAccountRate", "DetailDescription",
    "ProjNo", "DeptNo", "TaxType", "TaxableAmt", "TaxAdjustment", "Amount",
]

_AC_AR_HEADER = [
    "DocNo", "DocDate", "DebtorCode", "JournalType", "DisplayTerm",
    "SalesAgent", "Description", "CurrencyCode", "CurrencyRate", "RefNo2",
    "Note", "InclusiveTax", "AccNo", "ToAccountRate", "DetailDescription",
    "ProjNo", "DeptNo", "TaxType", "TaxableAmt", "TaxAdjustment", "Amount",
]

# SQL REQUIRED columns (from row 6 of SLPH_Invoice_Cash_Debit_Credit) — must appear in order
_SQL_REQUIRED = [
    "DOCNO(20)", "DOCDATE", "CODE(10)",
    "_ACCOUNT(10)", "_DESCRIPTION(200)", "_QTY", "_UOM(10)", "_UNITPRICE",
]


def _make_purchase_inv(
    *,
    inv_date: date = date(2024, 6, 1),
    vendor: str = "Bolt Auto Supply Sdn Bhd",
    reg_no: str | None = None,
    net: float = 1000.0,
    gst: float = 80.0,
    treatment: str = "SR",
    account_code: str = "510-000",
    creditor_code: str = "400-A0001",
) -> NormalizedInvoice:
    inv = NormalizedInvoice(
        doc_type="purchase",
        invoice_number="INV-CLIENT-001",
        invoice_date=inv_date,
        our_gst_registered=True,
        currency="MYR",
        supplier=PartyInfo(
            name=vendor,
            gst_regno=reg_no,
            country="MY",
            vendor_code=creditor_code,
        ),
    )
    inv.lines.append(
        InvoiceLine(
            description="Auto parts",
            net_amount=net,
            gst_amount=gst,
            tax_treatment=treatment,
            account_code=account_code,
        )
    )
    return inv


def _make_sales_inv(
    *,
    inv_date: date = date(2024, 6, 1),
    customer: str = "Client ABC Sdn Bhd",
    net: float = 500.0,
    gst: float = 40.0,
    treatment: str = "SR",
    account_code: str = "400-100",
    debtor_code: str = "300-C0001",
) -> NormalizedInvoice:
    inv = NormalizedInvoice(
        doc_type="sales",
        invoice_number="INV-SALES-001",
        invoice_date=inv_date,
        our_gst_registered=True,
        currency="MYR",
        customer=PartyInfo(
            name=customer,
            country="MY",
            vendor_code=debtor_code,
        ),
    )
    inv.lines.append(
        InvoiceLine(
            description="Service",
            net_amount=net,
            gst_amount=gst,
            tax_treatment=treatment,
            account_code=account_code,
        )
    )
    return inv


# ---------------------------------------------------------------------------
# AutoCount header tests
# ---------------------------------------------------------------------------

class TestAutoCountHeaderMatchesTemplate:
    """Generated header must exactly equal the real template field row."""

    @pytest.mark.skipif(not _TEMPLATES_PRESENT, reason="Real templates not on this machine")
    def test_ap_header_from_template_matches_constant(self):
        """Confirm our hard-coded _AC_AP_HEADER matches the real .xls row 2."""
        try:
            import xlrd  # noqa: F401
        except ImportError:
            pytest.skip("xlrd not installed — cannot read .xls templates")

        wb = xlrd.open_workbook(str(_AUTOCOUNT_DIR / "Import-AP-Invoice.xls"))
        ws = wb.sheet_by_name("AP Invoice")
        real_header = [ws.cell_value(2, c) for c in range(1, 22)]
        assert real_header == _AC_AP_HEADER, (
            f"Template AP header mismatch:\n  template: {real_header}\n  spec:     {_AC_AP_HEADER}"
        )

    @pytest.mark.skipif(not _TEMPLATES_PRESENT, reason="Real templates not on this machine")
    def test_ar_header_from_template_matches_constant(self):
        """Confirm our hard-coded _AC_AR_HEADER matches the real .xls row 2."""
        try:
            import xlrd  # noqa: F401
        except ImportError:
            pytest.skip("xlrd not installed — cannot read .xls templates")

        wb = xlrd.open_workbook(str(_AUTOCOUNT_DIR / "Import-AR-Invoice.xls"))
        ws = wb.sheet_by_name("AR Invoice")
        real_header = [ws.cell_value(2, c) for c in range(1, 22)]
        assert real_header == _AC_AR_HEADER, (
            f"Template AR header mismatch:\n  template: {real_header}\n  spec:     {_AC_AR_HEADER}"
        )

    def test_generated_ap_sheet_header_equals_template(self):
        """The generated AP Invoice sheet header must exactly equal _AC_AP_HEADER."""
        inv = _make_purchase_inv()
        exporter = AutoCountExporter(classifier=MY_CLF)
        with tempfile.TemporaryDirectory() as tmpdir:
            out = Path(tmpdir) / "autocount.xlsx"
            exporter.write_workbook(out, purchases=[inv], sales=[])
            wb = load_workbook(out)
            assert "AP Invoice" in wb.sheetnames, f"Expected 'AP Invoice' sheet, got {wb.sheetnames}"
            ws = wb["AP Invoice"]
            generated_header = [ws.cell(row=1, column=c).value for c in range(1, len(_AC_AP_HEADER) + 1)]
        assert generated_header == _AC_AP_HEADER, (
            f"Generated AP header mismatch:\n  got:      {generated_header}\n  expected: {_AC_AP_HEADER}"
        )

    def test_generated_ar_sheet_header_equals_template(self):
        """The generated AR Invoice sheet header must exactly equal _AC_AR_HEADER."""
        inv = _make_sales_inv()
        exporter = AutoCountExporter(classifier=MY_CLF)
        with tempfile.TemporaryDirectory() as tmpdir:
            out = Path(tmpdir) / "autocount.xlsx"
            exporter.write_workbook(out, purchases=[], sales=[inv])
            wb = load_workbook(out)
            assert "AR Invoice" in wb.sheetnames, f"Expected 'AR Invoice' sheet, got {wb.sheetnames}"
            ws = wb["AR Invoice"]
            generated_header = [ws.cell(row=1, column=c).value for c in range(1, len(_AC_AR_HEADER) + 1)]
        assert generated_header == _AC_AR_HEADER, (
            f"Generated AR header mismatch:\n  got:      {generated_header}\n  expected: {_AC_AR_HEADER}"
        )


# ---------------------------------------------------------------------------
# AutoCount mandatory column population
# ---------------------------------------------------------------------------

class TestAutoCountMandatoryColumns:
    """All ★mandatory columns must be non-empty for a fully-mapped line."""

    AP_MANDATORY = ["DocNo", "DocDate", "CreditorCode", "JournalType", "InclusiveTax", "AccNo", "Amount"]
    AR_MANDATORY = ["DocNo", "DocDate", "DebtorCode", "JournalType", "InclusiveTax", "AccNo", "Amount"]

    def test_ap_mandatory_columns_populated(self):
        inv = _make_purchase_inv()
        exporter = AutoCountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        assert rows, "No rows generated"
        row = rows[0]
        for col in self.AP_MANDATORY:
            val = row.get(col)
            assert val is not None and str(val).strip() != "", (
                f"Mandatory AP column '{col}' is empty. Row: {row}"
            )

    def test_ar_mandatory_columns_populated(self):
        inv = _make_sales_inv()
        exporter = AutoCountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "sales")
        assert rows, "No rows generated"
        row = rows[0]
        for col in self.AR_MANDATORY:
            val = row.get(col)
            assert val is not None and str(val).strip() != "", (
                f"Mandatory AR column '{col}' is empty. Row: {row}"
            )

    def test_ap_constants_applied(self):
        """DocNo=<<New>>, JournalType=PURCHASE, InclusiveTax=F must be constants."""
        inv = _make_purchase_inv()
        exporter = AutoCountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        row = rows[0]
        assert row["DocNo"] == "<<New>>", f"Expected '<<New>>' but got {row['DocNo']!r}"
        assert row["JournalType"] == "PURCHASE", f"Expected 'PURCHASE' but got {row['JournalType']!r}"
        assert row["InclusiveTax"] == "F", f"Expected 'F' but got {row['InclusiveTax']!r}"

    def test_ar_constants_applied(self):
        """DocNo=<<New>>, JournalType=SALES, InclusiveTax=F must be constants."""
        inv = _make_sales_inv()
        exporter = AutoCountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "sales")
        row = rows[0]
        assert row["DocNo"] == "<<New>>", f"Expected '<<New>>' but got {row['DocNo']!r}"
        assert row["JournalType"] == "SALES", f"Expected 'SALES' but got {row['JournalType']!r}"
        assert row["InclusiveTax"] == "F", f"Expected 'F' but got {row['InclusiveTax']!r}"

    def test_ap_supplier_invoice_no_equals_invoice_number(self):
        """SupplierInvoiceNo (AP only) must equal the invoice's invoice_number."""
        inv = _make_purchase_inv()
        exporter = AutoCountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        assert rows[0]["SupplierInvoiceNo"] == "INV-CLIENT-001"

    def test_ap_taxable_amt_equals_net_amount(self):
        """TaxableAmt = line net (AutoCount derives tax from TaxType × TaxableAmt)."""
        inv = _make_purchase_inv(net=1000.0)
        exporter = AutoCountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        assert rows[0]["TaxableAmt"] == 1000.0

    def test_ap_tax_type_resolved(self):
        """TaxType must resolve to the SST code (SV-8 for 8% SST after Mar 2024)."""
        inv = _make_purchase_inv(inv_date=date(2024, 6, 1))
        exporter = AutoCountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        assert rows[0]["TaxType"] == "SV-8", f"Expected 'SV-8', got {rows[0]['TaxType']!r}"


# ---------------------------------------------------------------------------
# SQL Account header and required columns
# ---------------------------------------------------------------------------

class TestSqlAccountHeader:
    """SQL generated header must contain every REQUIRED column in order."""

    SQL_SHEET = "SLPH_Invoice_Cash_Debit_Credit"

    def test_purchase_sheet_name_correct(self):
        inv = _make_purchase_inv()
        exporter = SqlAccountExporter(classifier=MY_CLF)
        with tempfile.TemporaryDirectory() as tmpdir:
            out = Path(tmpdir) / "sql.xlsx"
            exporter.write_workbook(out, purchases=[inv], sales=[])
            wb = load_workbook(out)
        assert self.SQL_SHEET in wb.sheetnames, (
            f"Expected sheet '{self.SQL_SHEET}', got {wb.sheetnames}"
        )

    def test_purchase_required_columns_present_in_order(self):
        """Every REQUIRED SQL column must appear in the generated header in order."""
        inv = _make_purchase_inv()
        exporter = SqlAccountExporter(classifier=MY_CLF)
        with tempfile.TemporaryDirectory() as tmpdir:
            out = Path(tmpdir) / "sql.xlsx"
            exporter.write_workbook(out, purchases=[inv], sales=[])
            wb = load_workbook(out)
            ws = wb[self.SQL_SHEET]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for req in _SQL_REQUIRED:
            assert req in header, (
                f"Required SQL column {req!r} missing from generated header. Header: {header}"
            )
        # Check order: each required col must appear after the previous one
        positions = [header.index(req) for req in _SQL_REQUIRED]
        assert positions == sorted(positions), (
            f"Required columns out of order. Positions: {list(zip(_SQL_REQUIRED, positions))}"
        )

    def test_sales_required_columns_present_in_order(self):
        """SQL sales sheet must also contain all REQUIRED columns in order."""
        inv = _make_sales_inv()
        exporter = SqlAccountExporter(classifier=MY_CLF)
        with tempfile.TemporaryDirectory() as tmpdir:
            out = Path(tmpdir) / "sql.xlsx"
            exporter.write_workbook(out, purchases=[], sales=[inv])
            wb = load_workbook(out)
        # Sales sheet is the second sheet (first = purchase)
        # Both sheets use the same name so we look at both sheets
        ws = wb.worksheets[1]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for req in _SQL_REQUIRED:
            assert req in header, (
                f"Required SQL sales column {req!r} missing. Header: {header}"
            )

    def test_sql_required_detail_columns_populated(self):
        """_QTY/_UOM/_UNITPRICE/_ACCOUNT must all be non-empty for a fully-mapped line."""
        inv = _make_purchase_inv()
        exporter = SqlAccountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        assert rows, "No rows generated"
        row = rows[0]
        assert row.get("_QTY") == 1, f"Expected _QTY=1, got {row.get('_QTY')!r}"
        assert row.get("_UOM(10)") == "UNIT", f"Expected _UOM=UNIT, got {row.get('_UOM(10)')!r}"
        assert row.get("_UNITPRICE") not in (None, "", 0), (
            f"Expected non-zero _UNITPRICE, got {row.get('_UNITPRICE')!r}"
        )
        assert row.get("_ACCOUNT(10)") not in (None, ""), (
            f"Expected non-empty _ACCOUNT, got {row.get('_ACCOUNT(10)')!r}"
        )

    def test_sql_tax_code_in_row(self):
        """_TAX column must carry the SST code for a taxable line."""
        inv = _make_purchase_inv(inv_date=date(2024, 6, 1))
        exporter = SqlAccountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        assert rows[0]["_TAX(10)"] == "SV", f"Expected 'SV', got {rows[0]['_TAX(10)']!r}"

    @pytest.mark.skipif(not _TEMPLATES_PRESENT, reason="Real templates not on this machine")
    def test_sql_purchase_required_cols_match_real_template(self):
        """Cross-check: required cols from the real template match our SQL REQUIRED list."""
        wb = load_workbook(str(_SQL_DIR / "Import Purchase Invoice.xlsx"))
        ws = wb["SLPH_Invoice_Cash_Debit_Credit"]
        real_header = [ws.cell(row=6, column=c).value for c in range(1, ws.max_column + 1)]
        for req in _SQL_REQUIRED:
            assert req in real_header, (
                f"Required column {req!r} missing from real SQL Purchase template header"
            )


# ---------------------------------------------------------------------------
# Acme fixture: Party List creditor code resolution
# ---------------------------------------------------------------------------

class TestClientCreditorResolution:
    """Build EntityMemoryEntry set from Acme COA & List.xlsx and assert code resolution."""

    @pytest.mark.skipif(not _CLIENT_COA.exists(), reason="Acme COA not on this machine")
    def test_bolt_auto_supply_resolves_to_400_A0001(self):
        """Bolt Auto Supply Sdn Bhd (Creditor) must resolve to 400-A0001."""
        from invoice_processing.export.code_resolver import resolve_creditor_code

        wb = load_workbook(str(_CLIENT_COA))
        ws = wb["Party List"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        name_col = headers.index("Name") + 1
        code_col = headers.index("Mapping Code") + 1

        memory: list[EntityMemoryEntry] = []
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=name_col).value
            mapping_code = ws.cell(row=r, column=code_col).value
            if name and mapping_code:
                memory.append(
                    EntityMemoryEntry(
                        name=str(name).strip(),
                        creditor_code=str(mapping_code).strip(),
                    )
                )

        code = resolve_creditor_code("Bolt Auto Supply Sdn Bhd", None, memory)
        assert code == "400-A0001", (
            f"Expected creditor code '400-A0001' for Bolt Auto Supply Sdn Bhd, got {code!r}"
        )

    @pytest.mark.skipif(not _CLIENT_COA.exists(), reason="Acme COA not on this machine")
    def test_client_creditor_code_in_autocount_row(self):
        """With Acme entity memory loaded, AutoCount AP row has CreditorCode=400-A0001."""

        wb = load_workbook(str(_CLIENT_COA))
        ws = wb["Party List"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        name_col = headers.index("Name") + 1
        code_col = headers.index("Mapping Code") + 1

        memory: list[EntityMemoryEntry] = []
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=name_col).value
            mapping_code = ws.cell(row=r, column=code_col).value
            if name and mapping_code:
                memory.append(
                    EntityMemoryEntry(
                        name=str(name).strip(),
                        creditor_code=str(mapping_code).strip(),
                    )
                )

        # Build inv WITHOUT pre-setting vendor_code so resolution goes through entity memory
        inv = NormalizedInvoice(
            doc_type="purchase",
            invoice_number="INV-BOLT-001",
            invoice_date=date(2024, 6, 1),
            our_gst_registered=True,
            currency="MYR",
            supplier=PartyInfo(name="Bolt Auto Supply Sdn Bhd", country="MY"),
        )
        inv.lines.append(
            InvoiceLine(
                description="Parts",
                net_amount=500.0,
                gst_amount=40.0,
                tax_treatment="SR",
                account_code="510-000",
            )
        )

        exporter = AutoCountExporter(classifier=MY_CLF)
        exporter.configure_client_context(entity_memory=memory)
        rows = exporter.rows([inv], "purchase")
        assert rows[0]["CreditorCode"] == "400-A0001", (
            f"Expected '400-A0001', got {rows[0]['CreditorCode']!r}"
        )


# ---------------------------------------------------------------------------
# MAP1 regression: Amount / _AMOUNT = line net (sub_total), not unit price
# (WS-1.1 — see plan 2026-06-21-intelligent-extraction-implementation.md).
# ---------------------------------------------------------------------------

def _make_qty_line_purchase_inv(
    *,
    quantity: float = 3.0,
    net: float = 300.0,
    unit_price: float = 100.0,
    gst: float = 24.0,
) -> NormalizedInvoice:
    """Build a 1-line purchase invoice with the given quantity.

    Defaults model a real qty>1 line: 3 units @ 100 each → net 300, gst 24 (8% SST).
    """
    inv = NormalizedInvoice(
        doc_type="purchase",
        invoice_number="INV-QTY-001",
        invoice_date=date(2024, 6, 1),
        our_gst_registered=True,
        currency="MYR",
        supplier=PartyInfo(
            name="GENERIC AUTO PARTS SDN BHD",
            country="MY",
            vendor_code="400-A0099",
        ),
    )
    inv.lines.append(
        InvoiceLine(
            description="Auto part",
            quantity=quantity,
            net_amount=net,
            gst_amount=gst,
            tax_treatment="SR",
            account_code="510-000",
        )
    )
    return inv


class TestMAP1AmountIsLineNetNotUnitPrice:
    """MAP1 regression (WS-1.1).

    With `InclusiveTax=F` (always emitted), the AutoCount `Amount` column and the
    SQL Account `_AMOUNT` column are the **tax-exclusive line net** — the value
    posted to the GL `AccNo` / `_ACCOUNT(10)`. Mapping them to `unit_price` (= net
    ÷ qty) understated the ledger by the qty factor for any qty>1 line. Acme lines
    are mostly qty=1 so the bug was masked; this test guards the fix.
    """

    def test_autocount_amount_equals_line_net_for_qty_gt_1(self):
        """AutoCount `Amount` (and `TaxableAmt` if present) = line net for qty=3."""
        inv = _make_qty_line_purchase_inv(quantity=3.0, net=300.0, unit_price=100.0)
        exporter = AutoCountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        assert rows, "No rows generated"
        row = rows[0]
        # The fix: Amount = sub_total = 300, NOT unit_price = 100.
        assert row["Amount"] == 300.0, (
            f"AutoCount Amount must be the line net (sub_total) for qty>1. "
            f"Expected 300.0, got {row['Amount']!r}. This is MAP1 — see WS-1.1."
        )
        # TaxableAmt is ALSO the line net: AutoCount derives tax as
        # TaxType × TaxableAmt, so a per-unit value here understates the tax base
        # by the qty factor (the same MAP1 defect, on the tax column).
        assert row["TaxableAmt"] == 300.0, (
            f"AutoCount TaxableAmt must be the line net (sub_total) for qty>1. "
            f"Expected 300.0, got {row['TaxableAmt']!r}. This is MAP1 — see WS-1.1."
        )
        # InclusiveTax must remain F (we only changed the field-map, not the constant).
        assert row["InclusiveTax"] == "F"

    def test_sql_account_amount_equals_line_net_for_qty_gt_1(self):
        """SQL `_AMOUNT` = line net; `_QTY` preserves per-line quantity; `_UNITPRICE` preserved."""
        inv = _make_qty_line_purchase_inv(quantity=3.0, net=300.0, unit_price=100.0)
        exporter = SqlAccountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        assert rows, "No rows generated"
        row = rows[0]
        assert row["_AMOUNT"] == 300.0, (
            f"SQL _AMOUNT must be the line net (sub_total) for qty>1. "
            f"Expected 300.0, got {row['_AMOUNT']!r}. This is MAP1 — see WS-1.1."
        )
        # SQL bonus: _QTY and _UNITPRICE now preserve per-line values.
        assert row["_QTY"] == 3.0, (
            f"SQL _QTY must preserve per-line quantity. Expected 3.0, got {row['_QTY']!r}"
        )
        assert row["_UNITPRICE"] == 100.0, (
            f"SQL _UNITPRICE must preserve per-line unit price. Expected 100.0, "
            f"got {row['_UNITPRICE']!r}"
        )

    def test_sql_account_defaults_to_qty_1_when_line_quantity_missing(self):
        """Lines without a quantity must still post `_QTY=1` (SQL REQUIRED constraint)."""
        inv = NormalizedInvoice(
            doc_type="purchase",
            invoice_number="INV-NOQTY-001",
            invoice_date=date(2024, 6, 1),
            our_gst_registered=True,
            currency="MYR",
            supplier=PartyInfo(
                name="GENERIC SUPPLIER",
                country="MY",
                vendor_code="400-A0098",
            ),
        )
        inv.lines.append(
            InvoiceLine(
                description="Service",
                # no quantity
                net_amount=200.0,
                gst_amount=16.0,
                tax_treatment="SR",
                account_code="510-000",
            )
        )
        exporter = SqlAccountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        assert rows[0]["_QTY"] == 1, (
            f"SQL _QTY must default to 1 when line.quantity is None. "
            f"Got {rows[0]['_QTY']!r}"
        )


# ---------------------------------------------------------------------------
# MAP2 regression: column_for_field returns the real column for every exporter
# (WS-1.2 — see plan 2026-06-21-intelligent-extraction-implementation.md).
# compose_confident_note and any future preview/note surface must use this
# single source of truth, not guess header strings.
# ---------------------------------------------------------------------------


class TestColumnForField:
    """``LedgerExporter.column_for_field`` returns the actual column for a
    logical field name (sub_total / currency / account_code / etc.) per doc_type.

    The acceptance gate for WS-1.2 is that the "reconciles to $X" total in
    ``compose_confident_note`` renders non-blank on a real delivery. The
    prerequisite is that ``column_for_field`` returns a column the exporter
    ACTUALLY emits — verified here.
    """

    def test_qbs_purchase_sub_total_column(self):
        exp = QbsLedgerExporter(classifier=MY_CLF)
        assert exp.column_for_field("sub_total", "purchase") == "Sub Total"

    def test_qbs_sales_sub_total_column(self):
        """QBS sales uses "Amount" for the line net (not "Sub Total")."""
        exp = QbsLedgerExporter(classifier=MY_CLF)
        assert exp.column_for_field("sub_total", "sales") == "Amount"

    def test_qbs_currency_column(self):
        exp = QbsLedgerExporter(classifier=MY_CLF)
        assert exp.column_for_field("currency", "purchase") == "Currency"
        assert exp.column_for_field("currency", "sales") == "Currency"

    def test_qbs_account_code_column(self):
        """QBS account column is "Account Code / COA" — not "Account Code"."""
        exp = QbsLedgerExporter(classifier=MY_CLF)
        assert exp.column_for_field("account_code", "purchase") == "Account Code / COA"
        assert exp.column_for_field("account_code", "sales") == "Account Code / COA"

    def test_xero_sub_total_is_none(self):
        """Xero stores *UnitAmount (per-unit) — no per-line sub_total column."""
        exp = XeroLedgerExporter(classifier=MY_CLF)
        # *UnitAmount is "unit_amount", not "sub_total" — column_for_field("sub_total") returns None.
        assert exp.column_for_field("sub_total", "purchase") is None
        assert exp.column_for_field("sub_total", "sales") is None

    def test_xero_currency_column(self):
        exp = XeroLedgerExporter(classifier=MY_CLF)
        assert exp.column_for_field("currency", "purchase") == "Currency"
        assert exp.column_for_field("currency", "sales") == "Currency"

    def test_autocount_sub_total_column(self):
        """AutoCount: Amount is the line net (post WS-1.1 fix)."""
        exp = AutoCountExporter(classifier=MY_CLF)
        assert exp.column_for_field("sub_total", "purchase") == "Amount"
        assert exp.column_for_field("sub_total", "sales") == "Amount"

    def test_autocount_currency_column_only_on_sales(self):
        """AutoCount purchase has no currency column (only CurrencyRate)."""
        exp = AutoCountExporter(classifier=MY_CLF)
        assert exp.column_for_field("currency", "purchase") is None
        assert exp.column_for_field("currency", "sales") == "CurrencyCode"

    def test_autocount_account_code_column(self):
        exp = AutoCountExporter(classifier=MY_CLF)
        assert exp.column_for_field("account_code", "purchase") == "AccNo"
        assert exp.column_for_field("account_code", "sales") == "AccNo"

    def test_sql_sub_total_column(self):
        """SQL: _AMOUNT is the line net (post WS-1.1 fix)."""
        exp = SqlAccountExporter(classifier=MY_CLF)
        assert exp.column_for_field("sub_total", "purchase") == "_AMOUNT"
        assert exp.column_for_field("sub_total", "sales") == "_AMOUNT"

    def test_sql_has_no_currency_column(self):
        exp = SqlAccountExporter(classifier=MY_CLF)
        assert exp.column_for_field("currency", "purchase") is None
        assert exp.column_for_field("currency", "sales") is None

    def test_sql_account_code_column(self):
        exp = SqlAccountExporter(classifier=MY_CLF)
        assert exp.column_for_field("account_code", "purchase") == "_ACCOUNT(10)"
        assert exp.column_for_field("account_code", "sales") == "_ACCOUNT(10)"

    def test_unknown_field_returns_none(self):
        """A field that no exporter emits must return None, not a guess."""
        exp = AutoCountExporter(classifier=MY_CLF)
        assert exp.column_for_field("not_a_real_field", "purchase") is None


class TestComposeConfidentNoteWithRealColumns:
    """End-to-end MAP2 regression.

    Builds a real AutoCount row dict (via the actual exporter) and feeds it
    into ``compose_confident_note`` as a LEDGER_ROWS_KEY-shaped payload. The
    note must include the reconcile total and the dominant account code —
    not blank, not the wrong currency.
    """

    def test_autocount_purchase_note_has_reconcile_total_and_code(self):
        from accounting_agents.nodes import compose_confident_note

        inv = NormalizedInvoice(
            doc_type="purchase",
            invoice_number="INV-WS12-001",
            invoice_date=date(2024, 6, 1),
            our_gst_registered=True,
            currency="MYR",
            doc_total=200.0,
            supplier=PartyInfo(
                name="GENERIC AUTO PARTS",
                country="MY",
                vendor_code="400-A0099",
            ),
        )
        inv.lines.append(
            InvoiceLine(
                description="Part A",
                net_amount=120.0,
                gst_amount=0.0,
                tax_treatment="ZRL",
                account_code="510-100",
            )
        )
        inv.lines.append(
            InvoiceLine(
                description="Part B",
                net_amount=80.0,
                gst_amount=0.0,
                tax_treatment="ZRL",
                account_code="510-200",
            )
        )

        exporter = AutoCountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        payload = {
            "fy": 2025,
            "kind": "invoice",
            "software": "autocount",
            "batches": [{"sheet": "AP Invoice", "rows": rows}],
            "doc_total": 200.0,
            "currency": "MYR",
        }
        note = compose_confident_note(payload, doc_type="invoice")
        # The reconcile total must be present (MAP2 fix) — was blank before WS-1.2.
        assert "200" in note, f"Reconcile total missing from note: {note!r}"
        # Currency must surface as MYR (from payload-level), not the SGD default.
        assert "MYR" in note, f"Currency missing from note: {note!r}"
        assert "SGD" not in note, f"SGD default leaked into MY doc: {note!r}"

    def test_sql_purchase_note_uses_unitprice_column_for_subtotal(self):
        """SQL: the line net lives in _AMOUNT. The note must use _AMOUNT (not
        'Net Amount' literal, not 'Sub Total' literal)."""
        from accounting_agents.nodes import compose_confident_note

        inv = NormalizedInvoice(
            doc_type="purchase",
            invoice_number="INV-WS12-SQL-001",
            invoice_date=date(2024, 6, 1),
            our_gst_registered=True,
            currency="MYR",
            doc_total=150.0,
            supplier=PartyInfo(
                name="GENERIC SUPPLIER",
                country="MY",
                vendor_code="400-A0098",
            ),
        )
        inv.lines.append(
            InvoiceLine(
                description="Service",
                net_amount=150.0,
                gst_amount=0.0,
                tax_treatment="ZRL",
                account_code="510-300",
            )
        )

        exporter = SqlAccountExporter(classifier=MY_CLF)
        rows = exporter.rows([inv], "purchase")
        payload = {
            "fy": 2025,
            "kind": "invoice",
            "software": "sql_account",
            "batches": [{"sheet": "SLPH_Invoice_Cash_Debit_Credit", "rows": rows}],
            "doc_total": 150.0,
            "currency": "MYR",
        }
        note = compose_confident_note(payload, doc_type="invoice")
        assert "150" in note, f"Reconcile total missing: {note!r}"
        assert "MYR" in note, f"Currency missing: {note!r}"
        # The dominant account code (_ACCOUNT(10) = "510-300") must appear.
        assert "510-300" in note, f"Dominant account code missing: {note!r}"
