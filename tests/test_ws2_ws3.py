"""Deterministic tests for WS2 (schema-complete extraction or flag) and WS3
(per-client export format correctness).

No LLM / network calls: every test builds the real model classes directly and
asserts on exporter row dicts, validator behaviour, and total carry-through.
Prompt quality is validated by eval, not here.
"""

from __future__ import annotations

from datetime import date

import openpyxl

from invoice_processing.classify.document_classifier import ClassificationResult
from invoice_processing.export.categorizer import categorize_invoice
from invoice_processing.export.client_context import ClientContext, CoaAccount
from invoice_processing.export.exporters import (
    QbsLedgerExporter,
    XeroLedgerExporter,
    validate_required_fields,
)
from invoice_processing.export.models import InvoiceLine, NormalizedInvoice, PartyInfo
from invoice_processing.extract.invoice_extractor import (
    ExtractedInvoice,
    ExtractedLine,
    to_normalized,
)
from invoice_processing.pipeline import process_document


# --------------------------------------------------------------------------- #
# Builders
# --------------------------------------------------------------------------- #

def _two_line_invoice(*, doc_total=None, due_date=date(2025, 2, 15), account_codes=("500", "500")):
    """A 2-line purchase invoice: SR 1000+90 and ZR 1000+0 (Σ = 2090)."""
    return NormalizedInvoice(
        doc_type="purchase",
        invoice_number="INV-9",
        invoice_date=date(2025, 1, 15),
        due_date=due_date,
        currency="SGD",
        supplier=PartyInfo(name="Acme Pte Ltd", country="SG", gst_regno="200012345A"),
        doc_total=doc_total,
        lines=[
            InvoiceLine(
                description="Standard-rated services",
                net_amount=1000.0,
                gst_amount=90.0,
                account_code=account_codes[0],
                tax_keyword="SR",
            ),
            InvoiceLine(
                description="Zero-rated services",
                net_amount=1000.0,
                gst_amount=0.0,
                account_code=account_codes[1],
                tax_keyword="ZR",
            ),
        ],
    )


class TestNonRegisteredGstAbsorption:
    """Non-GST-registered client: all lines No Tax; input GST absorbed into amounts."""

    def _sr_line_invoice(self, *, our_gst_registered: bool) -> NormalizedInvoice:
        return NormalizedInvoice(
            doc_type="purchase",
            invoice_number="INV-ABS",
            invoice_date=date(2025, 1, 15),
            currency="SGD",
            supplier=PartyInfo(name="Supplier Pte Ltd", gst_regno="200012345A"),
            our_gst_registered=our_gst_registered,
            doc_total=109.0,
            lines=[
                InvoiceLine(
                    description="Office supplies",
                    net_amount=100.0,
                    gst_amount=9.0,
                    account_code="500",
                ),
            ],
        )

    def test_xero_non_registered_absorbs_gst(self):
        inv = self._sr_line_invoice(our_gst_registered=False)
        rows = XeroLedgerExporter().rows([inv], "purchase")
        assert len(rows) == 1
        row = rows[0]
        assert row["*TaxType"] == "No Tax"
        assert row["TaxAmount"] == 0.0
        assert float(row["*UnitAmount"]) == 109.0
        assert float(row["Total"]) == 109.0

    def test_xero_registered_splits_sr_tax(self):
        inv = self._sr_line_invoice(our_gst_registered=True)
        rows = XeroLedgerExporter().rows([inv], "purchase")
        row = rows[0]
        assert row["*TaxType"] == "SR"
        assert row["TaxAmount"] == 9.0
        assert float(row["*UnitAmount"]) == 100.0

    def test_qbs_non_registered_absorbs_gst(self):
        inv = self._sr_line_invoice(our_gst_registered=False)
        rows = QbsLedgerExporter().rows([inv], "purchase")
        row = rows[0]
        assert row["Sub Total"] == 109.0
        assert row["Tax Amount"] == 0.0
        assert row["Total Amount"] == 109.0

    def test_telco_split_registered_keeps_sr_and_zr(self):
        inv = _two_line_invoice(doc_total=2090.0, account_codes=("500", "500"))
        inv.our_gst_registered = True
        rows = XeroLedgerExporter().rows([inv], "purchase")
        assert [r["*TaxType"] for r in rows] == ["SR", "ZR"]
        assert [r["TaxAmount"] for r in rows] == [90.0, 0.0]

    def test_telco_split_non_registered_all_no_tax_gross(self):
        inv = _two_line_invoice(doc_total=2090.0, account_codes=("500", "500"))
        inv.our_gst_registered = False
        rows = XeroLedgerExporter().rows([inv], "purchase")
        assert [r["*TaxType"] for r in rows] == ["No Tax", "No Tax"]
        assert [r["TaxAmount"] for r in rows] == [0.0, 0.0]
        assert float(rows[0]["*UnitAmount"]) == 1090.0  # 1000 + 90 absorbed
        assert float(rows[1]["*UnitAmount"]) == 1000.0


# =========================================================================== #
# WS3 — Xero Total per-line rule
# =========================================================================== #

class TestXeroTotalPerLine:
    def test_every_row_shows_invoice_total_fallback(self):
        """No doc_total -> Total = Σ line net + Σ line tax on every row (2090)."""
        inv = _two_line_invoice(doc_total=None)
        rows = XeroLedgerExporter().rows([inv], "purchase")
        assert len(rows) == 2
        assert [r["Total"] for r in rows] == [2090.0, 2090.0]

    def test_every_row_prefers_authoritative_doc_total(self):
        """doc_total present -> that authoritative value on every row."""
        inv = _two_line_invoice(doc_total=2090.0)
        rows = XeroLedgerExporter().rows([inv], "purchase")
        assert [r["Total"] for r in rows] == [2090.0, 2090.0]

    def test_doc_total_overrides_line_sum(self):
        """A doc_total that differs from Σ lines is still used verbatim (reconcile
        guard flags the mismatch separately)."""
        inv = _two_line_invoice(doc_total=2000.0)
        rows = XeroLedgerExporter().rows([inv], "purchase")
        assert [r["Total"] for r in rows] == [2000.0, 2000.0]

    def test_sales_rows_also_carry_invoice_total(self):
        inv = _two_line_invoice(doc_total=2090.0)
        inv.doc_type = "sales"
        inv.customer = PartyInfo(name="Buyer Pte Ltd", country="SG")
        rows = XeroLedgerExporter().rows([inv], "sales")
        assert [r["Total"] for r in rows] == [2090.0, 2090.0]


# =========================================================================== #
# WS3 — QBS unchanged + header order matches the demo templates
# =========================================================================== #

class TestXeroUnitAmountTiesOnDiscount:
    """*UnitAmount = effective (post-discount) amount so Qty×UnitAmount = line net."""

    def test_unit_amount_uses_net_not_sticker(self):
        inv = NormalizedInvoice(
            doc_type="purchase",
            invoice_number="INV-D",
            invoice_date=date(2025, 1, 15),
            currency="USD",
            supplier=PartyInfo(name="Gymsales Software Pty Ltd"),
            doc_total=220.0,
            lines=[
                # sticker 275, discounted to 220
                InvoiceLine(description="Essential", quantity=1, unit_amount=275.0,
                            net_amount=220.0, gst_amount=0.0, account_code="6-4500"),
                # sticker 150, 100% discount -> 0
                InvoiceLine(description="Stand Alone App", quantity=1, unit_amount=150.0,
                            net_amount=0.0, gst_amount=0.0, account_code="6-4500"),
            ],
        )
        rows = XeroLedgerExporter().rows([inv], "purchase")
        units = [float(r["*UnitAmount"]) for r in rows]
        qtys = [float(r["*Quantity"]) for r in rows]
        assert units == [220.0, 0.0]                                  # net/qty, not sticker
        assert sum(q * u for q, u in zip(qtys, units)) == 220.0       # ties to the line nets
        assert all(float(r["Total"]) == 220.0 for r in rows)         # invoice total per row


class TestQbsExporterUnchanged:
    def test_per_line_total_is_net_plus_tax(self):
        """QBS Total Amount stays per-line net+tax (NOT the invoice total)."""
        inv = _two_line_invoice(doc_total=2090.0)
        rows = QbsLedgerExporter().rows([inv], "purchase")
        assert [r["Total Amount"] for r in rows] == [1090.0, 1000.0]
        assert [r["Sub Total"] for r in rows] == [1000.0, 1000.0]

    def test_sales_per_line_total_unchanged(self):
        inv = _two_line_invoice(doc_total=2090.0)
        inv.doc_type = "sales"
        inv.customer = PartyInfo(name="Buyer Pte Ltd", country="SG")
        rows = QbsLedgerExporter().rows([inv], "sales")
        assert [r["Total"] for r in rows] == [1090.0, 1000.0]
        assert [r["Amount"] for r in rows] == [1000.0, 1000.0]


class TestHeaderOrderMatchesTemplates:
    """Exporter column order must match the in-repo demo .xlsx templates."""

    DEMO = "invoice_processing/data/export_demo"

    def _template_headers(self, filename):
        wb = openpyxl.load_workbook(f"{self.DEMO}/{filename}")
        return {
            ws.title: [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for ws in wb.worksheets
        }

    def test_qbs_headers_match(self):
        hdr = self._template_headers("qbs_ledger_FY2025.xlsx")
        exp = QbsLedgerExporter()
        assert exp.purchase_cols == hdr["Purchase"]
        assert exp.sales_cols == hdr["Sales"]

    def test_xero_headers_match(self):
        hdr = self._template_headers("xero_ledger_FY2025.xlsx")
        exp = XeroLedgerExporter()
        assert exp.purchase_cols == hdr["Purchase"]
        assert exp.sales_cols == hdr["Sales"]
        # The Total column the WS3 rule populates is part of the template order.
        assert "Total" in exp.purchase_cols
        assert "Total" in exp.sales_cols


# =========================================================================== #
# WS2 step 1 — doc total carried from ExtractedInvoice into NormalizedInvoice
# =========================================================================== #

class TestToNormalizedCarriesDocTotal:
    def _extracted(self):
        return ExtractedInvoice(
            doc_type="invoice",
            invoice_number="INV-1",
            invoice_date="2025-01-15",
            due_date="2025-02-15",
            currency="SGD",
            issuer_name="Acme Pte Ltd",
            issuer_gst_regno="200012345A",
            bill_to_name="Buyer Pte Ltd",
            lines=[
                ExtractedLine(description="SR", net_amount=1000.0, gst_amount=90.0, tax_label="SR"),
                ExtractedLine(description="ZR", net_amount=1000.0, gst_amount=0.0, tax_label="ZR"),
            ],
            subtotal=2000.0,
            gst_total=90.0,
            total=2090.0,
            issuer_tax_system="NONE",
        )

    def test_doc_totals_carried(self):
        norm = to_normalized(self._extracted(), direction="purchase")
        assert norm.doc_subtotal == 2000.0
        assert norm.doc_gst_total == 90.0
        assert norm.doc_total == 2090.0

    def test_absent_totals_become_none(self):
        ex = self._extracted()
        ex.subtotal = None
        ex.gst_total = None
        ex.total = None
        norm = to_normalized(ex, direction="purchase")
        assert norm.doc_subtotal is None
        assert norm.doc_gst_total is None
        assert norm.doc_total is None

    def test_carried_total_drives_xero_total_column(self):
        """End-to-end: extracted total -> normalized -> Xero Total column."""
        norm = to_normalized(self._extracted(), direction="purchase")
        rows = XeroLedgerExporter().rows([norm], "purchase")
        assert [r["Total"] for r in rows] == [2090.0, 2090.0]


# =========================================================================== #
# WS2 step 3+4 — required_fields + validator
# =========================================================================== #

class TestRequiredFields:
    def test_xero_required_are_star_columns(self):
        exp = XeroLedgerExporter()
        assert all(c.startswith("*") for c in exp.required_fields("purchase"))
        assert "*DueDate" in exp.required_fields("purchase")
        assert "*AccountCode" in exp.required_fields("sales")

    def test_qbs_required_purchase(self):
        req = QbsLedgerExporter().required_fields("purchase")
        assert req == [
            "Invoice Number", "Invoice Date", "Vendor Name", "Sub Total", "Total Amount",
            "Account Code / COA",
        ]

    def test_qbs_required_sales(self):
        req = QbsLedgerExporter().required_fields("sales")
        assert req == [
            "Invoice Number", "Invoice Date", "Customer Name", "Amount", "Total",
            "Account Code / COA",
        ]


class TestValidateRequiredFields:
    def test_complete_invoice_has_no_missing(self):
        inv = _two_line_invoice(due_date=date(2025, 2, 15), account_codes=("500", "500"))
        assert validate_required_fields(inv, XeroLedgerExporter(), "purchase") == []
        assert validate_required_fields(inv, QbsLedgerExporter(), "purchase") == []

    def test_missing_due_date_not_flagged_for_xero(self):
        """Founder rule: a missing due_date is filled from invoice_date on export,
        so it must NOT be flagged for review (the fallback is intended, silent)."""
        inv = _two_line_invoice(due_date=None, account_codes=("500", "500"))
        missing = validate_required_fields(inv, XeroLedgerExporter(), "purchase")
        assert "due_date" not in missing
        # the exported *DueDate cell falls back to invoice_date on every row
        rows = XeroLedgerExporter().rows([inv], "purchase")
        assert all(r["*DueDate"] and r["*DueDate"] == r["*InvoiceDate"] for r in rows)

    def test_missing_due_date_not_required_for_qbs(self):
        inv = _two_line_invoice(due_date=None, account_codes=("500", "500"))
        assert validate_required_fields(inv, QbsLedgerExporter(), "purchase") == []

    def test_missing_line_account_code_flagged_with_line_number(self):
        inv = _two_line_invoice(due_date=date(2025, 2, 15), account_codes=("500", None))
        for exp in (XeroLedgerExporter(), QbsLedgerExporter()):
            missing = validate_required_fields(inv, exp, "purchase")
            assert "account_code (line 2)" in missing

    def test_missing_invoice_level_field_has_no_line_suffix(self):
        inv = _two_line_invoice(due_date=date(2025, 2, 15))
        inv.invoice_number = None
        missing = validate_required_fields(inv, QbsLedgerExporter(), "purchase")
        assert "invoice_number" in missing
        assert "invoice_number (line 1)" not in missing

    def test_combined_missing_fields(self):
        # due_date missing is silent (invoice_date fallback); the genuine gap is the line account_code
        inv = _two_line_invoice(due_date=None, account_codes=("500", None))
        missing = validate_required_fields(inv, XeroLedgerExporter(), "purchase")
        assert "due_date" not in missing
        assert "account_code (line 2)" in missing


# =========================================================================== #
# WS2 step 4 — validation wired into process_document (flag, don't drop)
# =========================================================================== #

def _xero_client() -> ClientContext:
    """Xero client whose entity memory maps the supplier to account code 500."""
    return ClientContext(
        client_id="xero-client",
        client_name="Test Client Pte Ltd",
        fye_month=3,
        accounting_software="Xero Ledger",
        base_currency="SGD",
        tax_registered=True,
        coa=[CoaAccount(code="500", description="Office", account_type="Expense", keywords="office")],
    )


def _categorize_no_llm(inv, *, coa, category_mapping, entity_memory, **_kw):
    return categorize_invoice(
        inv, coa=coa, category_mapping=category_mapping, entity_memory=entity_memory, use_llm=False
    )


def _classify_invoice(path, **_kw) -> ClassificationResult:
    return ClassificationResult(
        doc_type="invoice", issuer_name="Acme Supplier", bill_to_name="Test Client Pte Ltd",
        currency="SGD", total_amount=109.0, confidence=0.99, reason="stub",
    )


def _direction_purchase(cls, *, client_name=None, **_kw) -> str:
    return "purchase"


def _extract_missing_due_date(path, **_kw) -> ExtractedInvoice:
    """An otherwise-complete invoice with NO due_date and NO payment terms."""
    return ExtractedInvoice(
        doc_type="invoice",
        invoice_number="INV-001",
        invoice_date="2025-01-15",
        due_date=None,
        currency="SGD",
        issuer_name="Acme Supplier",
        issuer_gst_regno="200012345A",
        bill_to_name="Test Client Pte Ltd",
        lines=[ExtractedLine(description="Office supplies", net_amount=100.0, gst_amount=9.0, tax_label="SR")],
        subtotal=100.0,
        gst_total=9.0,
        total=109.0,
        issuer_tax_system="NONE",
    )


def _extract_complete(path, **_kw) -> ExtractedInvoice:
    ex = _extract_missing_due_date(path)
    ex.due_date = "2025-02-15"
    return ex


class TestPipelineFlagsMissingRequiredFields:
    def _bank_stub(self, path, **_kw):
        return None, "stub"

    def test_missing_due_date_not_flagged_for_xero(self, tmp_path):
        # Founder rule: missing due_date falls back to invoice_date on export -> NOT flagged.
        p = tmp_path / "inv.pdf"
        p.write_bytes(b"%PDF stub")
        doc = process_document(
            p, _xero_client(),
            classify_fn=_classify_invoice,
            direction_fn=_direction_purchase,
            extract_fn=_extract_missing_due_date,
            bank_fn=self._bank_stub,
            categorize_fn=_categorize_no_llm,
        )
        assert not doc.note.startswith("ERROR")
        assert doc.normalized is not None
        assert doc.reconciled is True
        assert "needs review" not in doc.note

    def test_missing_required_field_flags_doc_for_xero(self, tmp_path):
        # A genuinely missing Xero *required field (invoice_number) IS flagged - not dropped.
        def _extract_no_invoice_number(path, **_kw) -> ExtractedInvoice:
            ex = _extract_complete(path)
            ex.invoice_number = None
            return ex
        p = tmp_path / "inv.pdf"
        p.write_bytes(b"%PDF stub")
        doc = process_document(
            p, _xero_client(),
            classify_fn=_classify_invoice,
            direction_fn=_direction_purchase,
            extract_fn=_extract_no_invoice_number,
            bank_fn=self._bank_stub,
            categorize_fn=_categorize_no_llm,
        )
        assert not doc.note.startswith("ERROR")
        assert doc.normalized is not None          # flagged, not dropped
        assert doc.reconciled is False
        assert "needs review: missing" in doc.note

    def test_complete_invoice_not_flagged(self, tmp_path):
        p = tmp_path / "inv.pdf"
        p.write_bytes(b"%PDF stub")
        doc = process_document(
            p, _xero_client(),
            classify_fn=_classify_invoice,
            direction_fn=_direction_purchase,
            extract_fn=_extract_complete,
            bank_fn=self._bank_stub,
            categorize_fn=_categorize_no_llm,
        )
        assert not doc.note.startswith("ERROR")
        assert doc.reconciled is True
        assert "needs review" not in doc.note


# =========================================================================== #
# M1 — zero-line invoice is flagged, not silently passed through
# =========================================================================== #

class TestZeroLineInvoiceFlagged:
    def test_validate_zero_lines_returns_flag(self):
        """An invoice with no lines → ['no line items extracted'], not []."""
        inv = NormalizedInvoice(
            doc_type="purchase",
            invoice_number="INV-0",
            invoice_date=date(2025, 1, 15),
            due_date=date(2025, 2, 15),
            supplier=PartyInfo(name="Acme", country="SG"),
            lines=[],
        )
        for exp in (XeroLedgerExporter(), QbsLedgerExporter()):
            missing = validate_required_fields(inv, exp, "purchase")
            assert missing == ["no line items extracted"], f"{exp.__class__.__name__}: {missing}"

    def test_pipeline_flags_zero_line_invoice(self, tmp_path):
        """process_document with a zero-line extraction → reconciled=False + note."""
        def _extract_no_lines(path, **_kw) -> ExtractedInvoice:
            return ExtractedInvoice(
                doc_type="invoice",
                invoice_number="INV-0",
                invoice_date="2025-01-15",
                due_date="2025-02-15",
                currency="SGD",
                issuer_name="Acme Supplier",
                bill_to_name="Test Client Pte Ltd",
                lines=[],
                subtotal=0.0,
                gst_total=0.0,
                total=0.0,
                issuer_tax_system="NONE",
            )

        p = tmp_path / "empty.pdf"
        p.write_bytes(b"%PDF stub")
        doc = process_document(
            p, _xero_client(),
            classify_fn=_classify_invoice,
            direction_fn=_direction_purchase,
            extract_fn=_extract_no_lines,
            bank_fn=lambda path, **_: (None, "stub"),
            categorize_fn=_categorize_no_llm,
        )
        assert not doc.note.startswith("ERROR")
        assert doc.normalized is not None          # row preserved, not dropped
        assert doc.reconciled is False
        assert "no line items extracted" in doc.note


# =========================================================================== #
# L1 — write_workbook round-trip: Total cell equals invoice total on both rows
# =========================================================================== #

class TestXeroWriteWorkbookTotalCells:
    def test_total_cells_in_xlsx(self, tmp_path):
        """write_workbook → reload with openpyxl → Total column = invoice total on
        every data row (not just the row-dict level checked in other tests)."""
        inv = _two_line_invoice(doc_total=2090.0, due_date=date(2025, 2, 15),
                                account_codes=("500", "500"))
        path = tmp_path / "xero_test.xlsx"
        exp = XeroLedgerExporter()
        exp.write_workbook(path, purchases=[inv])

        wb = openpyxl.load_workbook(path)
        ws = wb["Purchase"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        total_col_idx = headers.index("Total") + 1      # 1-based

        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 2, f"Expected 2 data rows, got {len(data_rows)}"
        for i, row in enumerate(data_rows, start=1):
            cell_val = row[total_col_idx - 1]
            assert cell_val == 2090.0, (
                f"Purchase row {i}: Total cell = {cell_val!r}, expected 2090.0"
            )


# =========================================================================== #
# L2 — process_batch: flagged + clean invoice both appear in the workbook
# =========================================================================== #

class TestBatchFlaggedAndCleanBothWritten:
    """One invoice missing a required field (flagged) and one complete (clean) for
    the same FY workbook — both must produce rows; the flagged one carries the flag."""

    def _bank_stub(self, path, **_kw):
        return None, "stub"

    def test_both_rows_written_flag_carried(self, tmp_path):
        from io import BytesIO
        from invoice_processing.pipeline import process_batch

        flagged_p = tmp_path / "flagged_inv.pdf"
        clean_p   = tmp_path / "clean_inv.pdf"
        flagged_p.write_bytes(b"%PDF stub")
        clean_p.write_bytes(b"%PDF stub")

        def _extract_flagged(path, **_kw) -> ExtractedInvoice:
            # Missing invoice_number → Xero *required field → flagged (missing due_date alone no longer flags)
            return ExtractedInvoice(
                doc_type="invoice", invoice_number=None,
                invoice_date="2025-01-15", due_date="2025-02-15",
                currency="SGD", issuer_name="Acme Supplier",
                issuer_gst_regno="200012345A",
                bill_to_name="Test Client Pte Ltd",
                lines=[ExtractedLine(description="Office supplies",
                                    net_amount=100.0, gst_amount=9.0, tax_label="SR")],
                subtotal=100.0, gst_total=9.0, total=109.0,
                issuer_tax_system="NONE",
            )

        def _extract_clean(path, **_kw) -> ExtractedInvoice:
            ex = _extract_flagged(path)
            ex.invoice_number = "INV-C"
            ex.due_date = "2025-02-15"
            return ex

        call_seq = iter([_extract_flagged, _extract_clean])

        def _extract_dispatch(path, **_kw):
            return next(call_seq)(path)

        result = process_batch(
            [str(flagged_p), str(clean_p)],
            _xero_client(),
            classify_fn=_classify_invoice,
            direction_fn=_direction_purchase,
            extract_fn=_extract_dispatch,
            bank_fn=self._bank_stub,
            categorize_fn=_categorize_no_llm,
        )

        # Both docs processed without ERROR
        assert len(result.docs) == 2
        assert len(result.errors) == 0
        assert all(not d.note.startswith("ERROR") for d in result.docs)

        # Both land in the same FY workbook
        assert "Ledger_FY2025.xlsx" in result.workbooks

        # Both produce rows in the Purchase sheet
        wb = openpyxl.load_workbook(BytesIO(result.workbooks["Ledger_FY2025.xlsx"]))
        data_rows = list(wb["Purchase"].iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 2, f"Expected 2 data rows, got {len(data_rows)}"

        # The flagged doc carries reconciled=False + review note
        flagged_doc = next(d for d in result.docs if "flagged" in d.path)
        clean_doc   = next(d for d in result.docs if "clean" in d.path)
        assert flagged_doc.reconciled is False
        assert "needs review" in flagged_doc.note
        assert clean_doc.reconciled is True
        assert "needs review" not in clean_doc.note
